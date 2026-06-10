# -*- coding: utf-8 -*-
"""ملف مؤشرات أداء درب 2026 — ملف استراتيجي موحّد (محطات وعقار) بمعادلات حقيقية وحماية.
   تدرّج صاعد: مؤشرات الإدارات/الموظفين ← الخارطة التنفيذية ← مستهدفات 2026 ← الركائز الخمسية."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.protection import WorkbookProtection

PWD = "darb2026"
MONTHS = ["يناير","فبراير","مارس","أبريل","مايو","يونيو","يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]

# ===== هوية درب: برتقالي + رمادي + أبيض =====
ORANGE="F47A21"          # برتقالي درب (لون التميّز)
ORANGE_LT="FDE3D1"       # برتقالي فاتح للتنبيهات
NAVY="58595B"            # رمادي داكن (العناوين الرئيسية)
BLUE="808285"            # رمادي درب (العناوين الفرعية)
STEEL="A7A9AC"           # رمادي فاتح (رؤوس الجداول)
AXIS="E6E7E8"            # رمادي فاتح جداً
GREEN_IN="C6EFCE"; GREY="F2F2F2"; GOLD=ORANGE_LT; WHITE="FFFFFF"; RED_IN="F8CBAD"
thin=Side(style="thin", color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def font(sz=11,b=False,color="000000"): return Font(name="Arial",size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
def center(wrap=True): return Alignment(horizontal="center",vertical="center",wrap_text=wrap)
def right(wrap=True):  return Alignment(horizontal="right", vertical="center",wrap_text=wrap)
LOCKED=Protection(locked=True); UNLOCKED=Protection(locked=False)
FMT={"pct":"0%","pct1":"0.0%","int":"#,##0","num1":"0.0","rial":'#,##0 "ر.س"'}
def style_cell(c,*,value=None,f=None,fillc=None,align=None,fmt=None,lock=True,border=True):
    if value is not None: c.value=value
    c.font=f or font()
    if fillc: c.fill=fill(fillc)
    c.alignment=align or right()
    if fmt: c.number_format=fmt
    c.protection=LOCKED if lock else UNLOCKED
    if border: c.border=BORDER
    return c
def protect(ws):
    ws.protection.sheet=True; ws.protection.password=PWD
    ws.protection.formatCells=False; ws.protection.selectLockedCells=True
    ws.protection.selectUnlockedCells=True; ws.protection.sort=True; ws.protection.autoFilter=True
    ws.sheet_view.rightToLeft=True

# ===================== طبقات الربط الاستراتيجي =====================
PL_GROW,PL_INNO,PL_SUST="النمو","الابتكار","الاستدامة"
PILLARS=[PL_GROW,PL_INNO,PL_SUST]
P_GROW1="التوسع في المواقع"; P_FRAN="الامتياز التجاري"; P_REACH="زيادة وصول العملاء"
P_NEW="إطلاق مشاريع جديدة"; P_DESIGN="تطوير التصاميم والهوية"; P_TECH="التحول التقني"
P_SAHAT="ساحات درب"; P_TANKI="تطبيق «تانكي»"
P_CX="تعزيز تجربة العملاء"; P_QUAL="جودة التشغيل"; P_LEASE="منظومة التأجير"; P_PEOPLE="الاستثمار في الموظفين"

wb=Workbook()

# ===================== 1) دليل الاستخدام =====================
ws=wb.active; ws.title="دليل الاستخدام"; ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=3; ws.column_dimensions["B"].width=110
style_cell(ws["B1"],value="درب  ·  Darb",f=font(22,True,ORANGE),align=right(False),border=False)
ws.row_dimensions[1].height=30
style_cell(ws["B2"],value="درب · المنظومة الاستراتيجية لمؤشرات الأداء 2026 — قطاع المحطات والعقار",
           f=font(17,True,NAVY),align=right(False),border=False)
guide=[("",""),
 ("🧭 الفكرة","ملف واحد يتدرّج من فوق لتحت: الركائز الخمسية ← مستهدفات 2026 ← الخارطة التنفيذية ← مؤشرات كل إدارة ← تقييم كل موظف. كل المؤشرات تغذّي نفس اللوحة عبر «قاعدة المؤشرات»."),
 ("🟩 خلايا الإدخال","الخلايا الخضراء فقط قابلة للتعديل — يدخل فيها المسؤول أرقامه الشهرية. باقي الخلايا (المؤشر/المستهدف/المعادلة) مقفولة ومحمية."),
 ("🔢 المعادلات","نسبة التحقيق وYTD والحالة والدرجات تُحسب آلياً بمعادلات حقيقية، وتصعد تلقائياً للمشروع والركيزة."),
 ("🔒 الحماية",f"كل شيت محمي بكلمة مرور: {PWD}  — لفك الحماية مؤقتاً: مراجعة ← حماية الورقة ← أدخل كلمة المرور."),
 ("👤 الصلاحيات","عند التوزيع: كل مسؤول يُمنح صلاحية الإدخال على نطاق خلاياه الخضراء فقط (Allow Edit Ranges / صلاحيات SharePoint). الهيكل والمستهدفات مقفولة للجميع."),
 ("📊 الحالة","✅ محقق = 100%+   |   🟡 قريب = 85%–99%   |   🔴 تحت الهدف = أقل من 85%   |   — = بانتظار إدخال / مستهدف وصفي."),
 ("🏢 الإدارات","الامتياز التجاري · التشغيل والمحطات · الاستثمار · العقار — كلها في نفس الملف وتغذّي نفس اللوحة التنفيذية."),
]
r=3
for a,b in guide:
    style_cell(ws.cell(r,2,(a+"  —  "+b) if a else b), f=font(11,bool(a)),align=right(True),border=False)
    ws.row_dimensions[r].height=30 if b else 8; r+=1
protect(ws)

# ===================== 2) الاستراتيجية =====================
ws=wb.create_sheet("الاستراتيجية"); ws.sheet_view.showGridLines=False
for i,w in enumerate([3,26,42,30],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.merge_cells("B2:D2")
style_cell(ws["B2"],value="ركائز واستراتيجية درب 2026 — قطاع المحطات والعقار",f=font(16,True,WHITE),fillc=NAVY,align=center(False),border=False)
ws.row_dimensions[2].height=32
for c,h in enumerate(["","الركيزة (5 سنوات) / الهدف","الوصف","مستهدف 2026"],1):
    style_cell(ws.cell(4,c,h),f=font(11,True,WHITE),fillc=BLUE,align=center())
strat=[("النمو","ركيزة","التوسع في المواقع وزيادة الوصول للعملاء وتنمية الإيرادات.",""),
 ("التوسع في المواقع","","افتتاح وتشغيل محطات جديدة.","التشغيل: 85 محطة · الامتياز: 150 محطة"),
 ("الامتياز والعقود","","تنمية الإيرادات عبر الامتياز والاستثمار.","الامتياز: 167 عقد · الاستثمار: 190 عقد"),
 ("زيادة وصول العملاء","","رفع المبيعات والتغطية الجغرافية.","718.8M لتر · 13 منطقة · 59 مدينة"),
 ("الابتكار","ركيزة","مشاريع جديدة وتحول تقني وتطوير الهوية.",""),
 ("الهوية والتصاميم","","تطبيق الهوية المؤسسية في المحطات.","التزام 55% بالهوية"),
 ("التحول التقني وساحات درب","","أتمتة المحطات وتطوير ساحات درب وتطبيق تانكي.","أتمتة 90% · تأجير ساحة درب ≥80%"),
 ("الاستدامة","ركيزة","جودة التشغيل وتجربة العميل والاستثمار في الموظفين.",""),
 ("تجربة العملاء","","رفع الرضا وجودة التجربة.","قوقل ≥4.5–4.7 · CSAT ≥90%"),
 ("جودة التشغيل","","كفاءة وجودة عمليات المحطات.","جاهزية 99% · سلامة 100%"),
 ("منظومة التأجير (العقار)","","الإشغال والتحصيل واستدامة المستأجرين.","إشغال ≥60% · 246 علامة · تحصيل 95%+"),
 ("الاستثمار في الموظفين","","التطوير والاستبقاء.","استبقاء الشركاء 90%+"),]
r=5
for name,kind,desc,tgt in strat:
    isr=kind=="ركيزة"
    style_cell(ws.cell(r,2,name),f=font(12,True,WHITE if isr else NAVY),fillc=BLUE if isr else None,align=right(True))
    style_cell(ws.cell(r,3,desc),f=font(11,False,WHITE if isr else "000000"),fillc=BLUE if isr else None,align=right(True))
    style_cell(ws.cell(r,4,tgt),f=font(11,True,WHITE if isr else BLUE),fillc=BLUE if isr else None,align=right(True))
    style_cell(ws.cell(r,1,""),fillc=BLUE if isr else None); ws.row_dimensions[r].height=26; r+=1
protect(ws)

# ===================== بيانات مؤشرات الإدارات =====================
# سجل المؤشر: (المحور, الاسم, الوحدة, القطبية, التجميع, المستهدف, نص المستهدف, التنسيق, الركيزة, المشروع)
AXF=["النمو والتوسع","المبيعات والأداء التجاري","الأداء المالي","التشغيل والجودة",
     "الهوية والتسويق","التوسع الإقليمي","تجربة العملاء والشركاء","استدامة الشراكة والامتثال"]
KPIS=[
 (0,"عدد محطات الامتياز المشغّلة (تراكمي)","محطة","↑","LAST",150,"150 محطة","int"),
 (0,"عدد عقود الامتياز الجديدة الموقّعة","عقد","↑","SUM",167,"167 عقد","int"),
 (1,"إجمالي مبيعات وقود محطات الامتياز","لتر","↑","SUM",718838077,"718.8M لتر","int"),
 (1,"نسبة المحطات التي حققت مستهدف المبيعات","%","↑","AVG",0.90,"90%+","pct"),
 (1,"متوسط مبيعات المحطة الواحدة شهرياً","لتر/محطة","↑","AVG",None,"+10% نمو شهري","int"),
 (1,"نسبة تأجير وحدات الامتياز","%","↑","AVG",0.70,"70%","pct"),
 (2,"إجمالي إيرادات إدارة الامتياز","ر.س","↑","SUM",None,"تصاعدي مستمر","rial"),
 (2,"رفع هامش الربحية للامتياز","هللة/لتر","↑","AVG",1.8,"+1.8 هللة/لتر","num1"),
 (2,"نسبة التحصيل المالي من شركاء الامتياز","%","↑","AVG",0.95,"95%+","pct"),
 (2,"متوسط أيام التحصيل (DSO)","يوم","↓","AVG",30,"≤ 30 يوم","num1"),
 (2,"إجمالي المستحقات المتأخرة (+30 يوم)","ر.س","↓","LAST",None,"تراجع مستمر","rial"),
 (2,"نسبة الشركاء الملتزمين بمواعيد السداد","%","↑","AVG",0.90,"90%+","pct"),
 (2,"معدل دوران رأس المال (أيام الدورة)","يوم","↓","AVG",5,"≤ 5 أيام","num1"),
 (3,"نسبة جاهزية محطات الامتياز (صفر انقطاع)","%","↑","AVG",0.99,"99%+","pct"),
 (3,"عدد حالات انقطاع الوقود الموثّقة","حالة","↓","SUM",0,"0 حالة","int"),
 (3,"نسبة إغلاق ملاحظات الجودة الميدانية","%","↑","AVG",0.95,"95%+","pct"),
 (3,"متوسط درجة تقييم جودة المحطة (قائمة التحقق)","درجة/100","↑","AVG",85,"85+","num1"),
 (3,"نسبة تنفيذ الزيارات الميدانية المجدولة","%","↑","AVG",1.0,"100%","pct"),
 (3,"نسبة المحطات المستوفية لمعايير السلامة","%","↑","AVG",1.0,"100%","pct"),
 (3,"متوسط وقت معالجة البلاغات التشغيلية","ساعة","↓","AVG",24,"≤ 24 ساعة","num1"),
 (3,"نسبة أتمتة محطات الامتياز","%","↑","LAST",0.90,"90%","pct"),
 (3,"نسبة تطبيق المعيار الشامل داخل المحطات","%","↑","LAST",0.30,"30%","pct"),
 (4,"نسبة التزام المحطات بالهوية المؤسسية","%","↑","LAST",0.55,"55%","pct"),
 (5,"نسبة العقود التي تغطي المناطق الـ13","%","↑","LAST",1.0,"100%","pct"),
 (5,"نسبة العقود في الـ12 مدينة الرئيسية","%","↓","LAST",0.40,"< 40%","pct"),
 (5,"نسبة العقود خارج المدن الرئيسية","%","↑","LAST",0.60,"> 60%","pct"),
 (5,"نسبة العقود على الطرق الإقليمية","%","↑","LAST",0.25,"> 25%","pct"),
 (6,"متوسط تقييم المحطات على خرائط قوقل","نجمة/5","↑","AVG",4.7,"≥ 4.7 نجمة","num1"),
 (6,"معدل رضا شركاء الامتياز (CSAT)","%","↑","AVG",0.90,"90%+","pct"),
 (6,"نسبة حل الشكاوى خلال المدة المحددة","%","↑","AVG",0.95,"95%+","pct"),
 (6,"نسبة انخفاض الشكاوى التشغيلية والإدارية","%","↑","AVG",None,"-30%","pct"),
 (7,"نسبة استدامة أصحاب الامتياز (Retention)","%","↑","LAST",0.90,"90%","pct"),
 (7,"نسبة زيادة عدد أصحاب الامتياز","%","↑","LAST",1.0,"100% نمو","pct"),
 (7,"نسبة اكتمال المستندات النظامية قبل التوقيع","%","↑","AVG",1.0,"100%","pct"),
 (7,"متوسط مدة إغلاق عقد الامتياز","أسبوع","↓","AVG",1,"≤ أسبوع","num1"),
 (7,"نسبة إنجاز التقارير الدورية في موعدها","%","↑","AVG",1.0,"100%","pct"),
]
LINK=[(PL_GROW,P_GROW1),(PL_GROW,P_FRAN),(PL_GROW,P_REACH),(PL_GROW,P_REACH),(PL_GROW,P_REACH),
 (PL_SUST,P_LEASE),(PL_GROW,P_FRAN),(PL_GROW,P_FRAN),(PL_SUST,P_QUAL),(PL_SUST,P_QUAL),
 (PL_SUST,P_QUAL),(PL_SUST,P_QUAL),(PL_SUST,P_QUAL),(PL_SUST,P_QUAL),(PL_SUST,P_QUAL),
 (PL_SUST,P_QUAL),(PL_SUST,P_QUAL),(PL_SUST,P_QUAL),(PL_SUST,P_QUAL),(PL_SUST,P_QUAL),
 (PL_INNO,P_TECH),(PL_INNO,P_TECH),(PL_INNO,P_DESIGN),
 (PL_GROW,P_GROW1),(PL_GROW,P_GROW1),(PL_GROW,P_GROW1),(PL_GROW,P_GROW1),
 (PL_SUST,P_CX),(PL_SUST,P_CX),(PL_SUST,P_CX),(PL_SUST,P_CX),
 (PL_SUST,P_CX),(PL_SUST,P_CX),(PL_SUST,P_QUAL),(PL_GROW,P_FRAN),(PL_SUST,P_QUAL)]
FRAN=[(AXF[a],n,u,p,g,t,tt,f,LINK[i][0],LINK[i][1]) for i,(a,n,u,p,g,t,tt,f) in enumerate(KPIS)]

OPS=[
 ("النمو","عدد المحطات المشغّلة (تراكمي)","محطة","↑","LAST",85,"85 محطة","int",PL_GROW,P_GROW1),
 ("الربحية","نمو مبيعات بنزين 91 مقارنة بالعام السابق","%","↑","AVG",None,"يُحدد لاحقاً","pct",PL_GROW,P_REACH),
 ("الربحية","نمو مبيعات بنزين 95 مقارنة بالعام السابق","%","↑","AVG",None,"يُحدد لاحقاً","pct",PL_GROW,P_REACH),
 ("الربحية","نمو مبيعات بنزين 98 مقارنة بالعام السابق","%","↑","AVG",None,"يُحدد لاحقاً","pct",PL_GROW,P_REACH),
 ("الربحية","نمو مبيعات الديزل مقارنة بالعام السابق","%","↑","AVG",None,"يُحدد لاحقاً","pct",PL_GROW,P_REACH),
 ("حماية الإيراد","نسبة جاهزية المحطة الشاملة (Uptime)","%","↑","AVG",0.99,"99%+","pct",PL_SUST,P_QUAL),
 ("حماية الإيراد","نسبة المضخات العاملة","%","↑","AVG",0.97,"≥ 97%","pct",PL_SUST,P_QUAL),
 ("ضغط التكلفة","تكلفة التشغيل لكل لتر مباع","هللة/لتر","↓","AVG",None,"خفض ≥5% سنوياً","num1",PL_SUST,P_QUAL),
 ("حماية الأصول","نسبة الصيانة الوقائية المنفذة في موعدها","%","↑","AVG",0.95,"≥ 95%","pct",PL_SUST,P_QUAL),
 ("تجربة العميل","نسبة انخفاض الشكاوى مقارنة بالشهر السابق","%","↑","AVG",None,"−30%","pct",PL_SUST,P_CX),
 ("تجربة العميل","تقييم Google Maps للمحطات","نجمة/5","↑","AVG",4.5,"≥ 4.5","num1",PL_SUST,P_CX),
 ("تجربة العميل","سرعة تعبئة الوقود","دقيقة","↓","AVG",3,"≤ 3 دقائق","num1",PL_SUST,P_CX),
 ("تجربة العميل","دقة التعبئة بدون أخطاء","%","↑","AVG",0.99,"≥ 99%","pct",PL_SUST,P_QUAL),
 ("تجربة العميل","تطبيق المعيار الشامل في المحطات","%","↑","LAST",0.80,"≥ 80%","pct",PL_INNO,P_TECH),
 ("السلامة","الالتزام بمعايير السلامة داخل المحطة","%","↑","AVG",1.0,"100%","pct",PL_SUST,P_QUAL),
 ("البيئة","التشجير داخل المحطة (معيار الوزارة)","%","↑","LAST",1.0,"100%","pct",PL_SUST,P_QUAL),
]
INV=[
 ("نمو المحفظة","إجمالي عقود الاستثمار الجديدة (جميع النماذج)","عقد","↑","SUM",190,"190 عقد","int",PL_GROW,P_GROW1),
 ("نمو المحفظة","عقود الاستثمار العادي الجديدة","عقد","↑","SUM",90,"90 عقد","int",PL_GROW,P_GROW1),
 ("نمو المحفظة","عقود الاستحواذ الجديدة","عقد","↑","SUM",100,"100 عقد","int",PL_GROW,P_GROW1),
 ("نمو المحفظة","المحطات المستحوَذ عليها المشغّلة (تراكمي)","محطة","↑","LAST",120,"120 محطة","int",PL_GROW,P_GROW1),
 ("نمو المحفظة","عقود تأجير الأرض الجديدة","عقد","↑","SUM",11,"11 عقد","int",PL_GROW,P_GROW1),
 ("نمو المحفظة","عقود إيجار المحطة الجديدة","عقد","↑","SUM",12,"12 عقد","int",PL_GROW,P_GROW1),
 ("نمو المحفظة","عقود تشغيل المحطة الجديدة","عقد","↑","SUM",68,"68 عقد","int",PL_GROW,P_GROW1),
 ("الربحية وجودة الموقع","نسبة المواقع المعتمدة ذات العائد المجدي","%","↑","AVG",0.80,"80%+","pct",PL_SUST,P_QUAL),
 ("الربحية وجودة الموقع","متوسط فترة استرداد رأس المال","سنة","↓","AVG",5,"≤ 5 سنوات","num1",PL_SUST,P_QUAL),
 ("الربحية وجودة الموقع","متوسط قيمة عقد الإيجار السنوي (أرض)","ر.س","↑","AVG",None,"يُحدد لاحقاً","rial",PL_GROW,P_GROW1),
 ("التوسع الإقليمي","نسبة العقود التي تغطي المناطق الـ13","%","↑","LAST",1.0,"100%","pct",PL_GROW,P_GROW1),
 ("التوسع الإقليمي","نسبة العقود خارج المدن الرئيسية","%","↑","LAST",0.60,"> 60%","pct",PL_GROW,P_GROW1),
 ("التوسع الإقليمي","نسبة العقود في الـ12 مدينة الرئيسية","%","↓","LAST",0.40,"< 40%","pct",PL_GROW,P_GROW1),
 ("التوسع الإقليمي","نسبة العقود على الطرق الإقليمية","%","↑","LAST",0.25,"> 25%","pct",PL_GROW,P_GROW1),
 ("كفاءة الدورة","متوسط مدة إغلاق عقد الاستثمار","أسبوع","↓","AVG",1,"≤ أسبوع","num1",PL_GROW,P_GROW1),
 ("كفاءة الدورة","نسبة العقود المتحوّلة لمحطات مشغّلة","%","↑","AVG",0.30,"≥ 30%","pct",PL_GROW,P_GROW1),
 ("كفاءة الدورة","نسبة تخفيض إيجار المواقع عالية التكاليف","%","↑","AVG",0.30,"≥ 30%","pct",PL_SUST,P_QUAL),
 ("الامتثال","نسبة اكتمال المستندات النظامية قبل التوقيع","%","↑","AVG",1.0,"100%","pct",PL_SUST,P_QUAL),
 ("الامتثال","نسبة العقود المؤرشفة إلكترونياً خلال 48 ساعة","%","↑","AVG",0.95,"95%+","pct",PL_SUST,P_QUAL),
 ("جودة التعاقد","جودة التفاوض مع المستثمرين (توافق مع المستهدفات)","%","↑","AVG",0.98,"98%","pct",PL_GROW,P_GROW1),
 ("جودة التعاقد","سرعة إنجاز إجراءات التعاقد","أسبوع","↓","AVG",1,"≤ أسبوع","num1",PL_GROW,P_GROW1),
 ("جودة التعاقد","مدى رضا المستثمرين وشركاء الاستثمار","%","↑","AVG",0.80,"80%+","pct",PL_SUST,P_CX),
 ("جودة التعاقد","نسبة اكتمال أرشفة ملفات العملاء","%","↑","AVG",1.0,"100%","pct",PL_SUST,P_QUAL),
 ("جودة التعاقد","نسبة إنجاز التقارير الدورية في موعدها","%","↑","AVG",1.0,"100%","pct",PL_SUST,P_QUAL),
]
RE=[
 ("الإدارة","نمو الإيرادات الإيجارية","%","↑","AVG",0.08,"> 8% سنوياً","pct",PL_SUST,P_LEASE),
 ("الإدارة","نسبة الالتزام بالعقود","%","↑","AVG",1.0,"100%","pct",PL_SUST,P_LEASE),
 ("الإدارة","نسبة التحصيل","%","↑","AVG",0.95,"> 95%","pct",PL_SUST,P_LEASE),
 ("التأجير","نسبة الإشغال","%","↑","AVG",0.60,"≥ 60%","pct",PL_SUST,P_LEASE),
 ("التأجير","الوحدات المؤجَّرة شهرياً","وحدة","↑","AVG",68,"68 شهرياً","int",PL_SUST,P_LEASE),
 ("التأجير","العملاء المحتملون (LEADS)","عميل","↑","AVG",680,"544–680","int",PL_GROW,P_REACH),
 ("التأجير","معدل التحويل","%","↑","AVG",0.20,"15–25%","pct",PL_GROW,P_REACH),
 ("التأجير","زيادة عدد العلامات التجارية","علامة","↑","LAST",246,"246 علامة","int",PL_SUST,P_LEASE),
 ("الإبداع والتكامل","نسبة تأجير ساحة درب","%","↑","LAST",0.80,"≥ 80%","pct",PL_INNO,P_SAHAT),
 ("الإبداع والتكامل","علامات تجارية جديدة (ساحة درب)","علامة","↑","SUM",6,"6 علامات","int",PL_INNO,P_SAHAT),
 ("الإبداع والتكامل","الشراكات الاستراتيجية","شراكة","↑","SUM",5,"5 شراكات","int",PL_INNO,P_NEW),
 ("إدارة الأملاك","نسبة تجديد العقود","%","↑","AVG",1.0,"100%","pct",PL_SUST,P_LEASE),
 ("تنسيق المشاريع","متوسط زمن تسليم الوحدة","يوم","↓","AVG",7,"≤ 7 أيام","num1",PL_SUST,P_QUAL),
 ("تنسيق المشاريع","نسبة إغلاق الطلبات خلال 3 أسابيع","%","↑","AVG",0.90,"> 90%","pct",PL_SUST,P_QUAL),
 ("الجودة والاستدامة","استدامة المستأجرين (لم يُنهَ مبكراً)","%","↑","AVG",0.80,"≥ 80%","pct",PL_SUST,P_CX),
 ("الجودة والاستدامة","معدل رضا المستأجرين (استبيان ربعي)","%","↑","AVG",0.90,"≥ 90%","pct",PL_SUST,P_CX),
]

DIG=[  # إدارة التقنية الرقمية — تُغلق فجوتَي «تطبيق تانكي» و«التحول التقني»
 ("منتج تانكي","عدد مستخدمي تطبيق «تانكي» النشطين شهرياً (MAU)","مستخدم","↑","LAST",None,"يُحدد لاحقاً","int",PL_INNO,P_TANKI),
 ("منتج تانكي","نمو تنزيلات تطبيق تانكي","%","↑","AVG",None,"يُحدد لاحقاً","pct",PL_INNO,P_TANKI),
 ("منتج تانكي","نسبة المعاملات الرقمية عبر تانكي","%","↑","AVG",None,"يُحدد لاحقاً","pct",PL_INNO,P_TANKI),
 ("منتج تانكي","معدل رضا مستخدمي تطبيق تانكي","نجمة/5","↑","AVG",4.5,"≥ 4.5","num1",PL_INNO,P_TANKI),
 ("التحول التقني","نسبة أتمتة العمليات الأساسية (شركة)","%","↑","LAST",None,"يُحدد لاحقاً","pct",PL_INNO,P_TECH),
 ("التحول التقني","نسبة جاهزية الأنظمة (Uptime)","%","↑","AVG",0.995,"≥ 99.5%","pct",PL_INNO,P_TECH),
 ("التحول التقني","متوسط زمن حل أعطال الأنظمة","ساعة","↓","AVG",4,"≤ 4 ساعات","num1",PL_INNO,P_TECH),
 ("التحول التقني","نسبة المشاريع التقنية المنجزة في الوقت","%","↑","AVG",0.90,"≥ 90%","pct",PL_INNO,P_TECH),
]
HR=[  # إدارة الموارد البشرية — تُغلق فجوة «الاستثمار في الموظفين»
 ("الاستبقاء","معدل استبقاء الموظفين (Retention)","%","↑","LAST",0.90,"≥ 90%","pct",PL_SUST,P_PEOPLE),
 ("الاستبقاء","معدل دوران الموظفين (Turnover)","%","↓","AVG",0.10,"≤ 10%","pct",PL_SUST,P_PEOPLE),
 ("التطوير","متوسط ساعات التدريب لكل موظف","ساعة","↑","SUM",40,"≥ 40 ساعة","num1",PL_SUST,P_PEOPLE),
 ("التطوير","نسبة الموظفين المشمولين بخطة تطوير","%","↑","LAST",0.80,"≥ 80%","pct",PL_SUST,P_PEOPLE),
 ("الرضا والانتماء","مؤشر رضا الموظفين (eNPS)","نقطة","↑","AVG",30,"≥ 30","num1",PL_SUST,P_PEOPLE),
 ("الاستقطاب","نسبة شغل الشواغر خلال 30 يوم","%","↑","AVG",0.85,"≥ 85%","pct",PL_SUST,P_PEOPLE),
 ("الأداء","نسبة إنجاز تقييمات الأداء في موعدها","%","↑","AVG",1.0,"100%","pct",PL_SUST,P_PEOPLE),
 ("الامتثال","نسبة السعودة","%","↑","LAST",None,"حسب النطاق","pct",PL_SUST,P_PEOPLE),
]

# ===================== باني ورقة المؤشرات (محرك موحّد) =====================
LEAD_KW=["عدد","زيارات","فرص","تحويل","مبيعات","نمو","LEADS","تنزيلات","تدريب","طلب","حملات","استقطاب"]
def classify(nm):
    return "🔵 قائد" if any(k in nm for k in LEAD_KW) else "🟣 لاحق"
def build_kpi_sheet(name, title, records, tagged=True):
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
    col_w=[4,20,40,10,8,9,12,16]+[9]*12+[12,12,14,16,22]+[12,10,9,14]
    for i,w in enumerate(col_w,1): ws.column_dimensions[get_column_letter(i)].width=w
    last_col="AC"
    ws.merge_cells(f"A1:{last_col}1")
    style_cell(ws["A1"],value=title,f=font(15,True,WHITE),fillc=NAVY,align=center(False),border=False)
    ws.row_dimensions[1].height=30
    ws.merge_cells(f"A2:{last_col}2")
    style_cell(ws["A2"],value="🟩 الأخضر = إدخال (الأرقام الشهرية + خط الأساس + الوزن)  ·  🔵 قائد/🟣 لاحق  ·  الدرجة الموزونة والحالة تلقائية",
               f=font(10,False,NAVY),fillc=GOLD,align=center(False),border=False)
    header=["#","المحور","المؤشر","الوحدة","القطبية","التجميع","المستهدف","نص المستهدف"]+MONTHS+["YTD","نسبة التحقيق","الحالة","الركيزة (5س)","المشروع (الخارطة)","خط الأساس 2025","النوع","الوزن %","الدرجة الموزونة"]
    for c,h in enumerate(header,1): style_cell(ws.cell(3,c,h),f=font(10,True,WHITE),fillc=BLUE,align=center())
    ws.row_dimensions[3].height=34; ws.freeze_panes="I4"
    dv=DataValidation(type="decimal",operator="greaterThanOrEqual",formula1="0",allow_blank=True)
    dv.error="أدخل رقماً موجباً فقط"; dv.errorTitle="قيمة غير صحيحة"; ws.add_data_validation(dv)
    START=4; w_default=round(1.0/len(records),4)
    for idx,rec in enumerate(records):
        axis,nm,unit,pol,agg,tgt,ttxt,fmt,pillar,project=rec
        r=START+idx
        style_cell(ws.cell(r,1,idx+1),f=font(10,True),align=center())
        style_cell(ws.cell(r,2,axis),f=font(10),align=right(True))
        style_cell(ws.cell(r,3,nm),f=font(10),align=right(True))
        style_cell(ws.cell(r,4,unit),align=center())
        style_cell(ws.cell(r,5,pol),align=center())
        style_cell(ws.cell(r,6,agg),align=center())
        style_cell(ws.cell(r,7,tgt if tgt is not None else ""),fmt=FMT[fmt] if tgt is not None else None,align=center())
        style_cell(ws.cell(r,8,ttxt),f=font(9,color="595959"),align=center())
        for m in range(12):
            cc=ws.cell(r,9+m); style_cell(cc,fillc=GREEN_IN,fmt=FMT[fmt],align=center(False),lock=False); dv.add(cc)
        rng=f"$I{r}:$T{r}"
        ws.cell(r,21).value=(f'=IF($F{r}="SUM",IF(COUNT({rng})=0,"",SUM({rng})),'
            f'IF($F{r}="AVG",IFERROR(AVERAGE({rng}),""),'
            f'IF($F{r}="LAST",IFERROR(LOOKUP(2,1/({rng}<>""),{rng}),""),"")))')
        style_cell(ws.cell(r,21),f=font(10,True),fmt=FMT[fmt],align=center())
        ws.cell(r,22).value=(f'=IF($G{r}="","",IF($U{r}="","",'
            f'IF($G{r}=0,IF($U{r}<=0,1,0),'
            f'IF($E{r}="↓",IFERROR($G{r}/$U{r},0),IFERROR($U{r}/$G{r},0)))))')
        style_cell(ws.cell(r,22),f=font(10,True,NAVY),fmt="0%",align=center())
        ws.cell(r,23).value=f'=IF($V{r}="","—",IF($V{r}>=1,"✅ محقق",IF($V{r}>=0.85,"🟡 قريب","🔴 تحت الهدف")))'
        style_cell(ws.cell(r,23),f=font(10,True),align=center())
        style_cell(ws.cell(r,24,pillar if tagged else ""),f=font(10,True,BLUE),align=center())
        style_cell(ws.cell(r,25,project if tagged else ""),f=font(10),align=right(True))
        # خانات الحوكمة
        bc=ws.cell(r,26); style_cell(bc,fillc=GREEN_IN,fmt=FMT[fmt],align=center(),lock=False); dv.add(bc)
        style_cell(ws.cell(r,27,classify(nm)),f=font(10),align=center())
        wc=ws.cell(r,28,w_default); style_cell(wc,fillc=GREEN_IN,fmt="0%",align=center(),lock=False)
        ws.cell(r,29).value=f'=IF($V{r}="","",$AB{r}*MIN($V{r},1))'
        style_cell(ws.cell(r,29),f=font(10,True,NAVY),fmt="0%",align=center())
        ws.row_dimensions[r].height=26
    END=START+len(records)-1
    TOT=END+1
    ws.merge_cells(start_row=TOT,start_column=2,end_row=TOT,end_column=27)
    style_cell(ws.cell(TOT,2,"الإجمالي الموزون للإدارة"),f=font(11,True,WHITE),fillc=NAVY,align=center())
    style_cell(ws.cell(TOT,1,""),fillc=NAVY)
    style_cell(ws.cell(TOT,28,f"=SUM(AB{START}:AB{END})"),f=font(11,True,"F47A21"),fillc=NAVY,fmt="0%",align=center())
    style_cell(ws.cell(TOT,29,f"=SUM(AC{START}:AC{END})"),f=font(12,True,"F47A21"),fillc=NAVY,fmt="0%",align=center())
    protect(ws)
    return START,END

# الإدارات الأربع
DEPTS=[("الامتياز التجاري","الامتياز · المؤشرات","إدارة الامتياز التجاري والمبيعات · مؤشرات الأداء 2026  (36 مؤشر / 8 محاور)",FRAN),
       ("التشغيل والمحطات","التشغيل · المؤشرات","إدارة التشغيل والمحطات · مؤشرات الأداء 2026  (16 مؤشر / 7 محاور)",OPS),
       ("الاستثمار","الاستثمار · المؤشرات","إدارة الاستثمار · مؤشرات الأداء 2026  (24 مؤشر / 6 محاور)",INV),
       ("العقار","العقار · المؤشرات","إدارة العقار · مؤشرات الأداء 2026  (16 مؤشر / 6 محاور)",RE),
       ("التقنية الرقمية","التقنية الرقمية · المؤشرات","إدارة التقنية الرقمية · مؤشرات الأداء 2026  (8 مؤشرات) — تانكي والتحول التقني",DIG),
       ("الموارد البشرية","الموارد البشرية · المؤشرات","إدارة الموارد البشرية · مؤشرات الأداء 2026  (8 مؤشرات) — الاستثمار في الموظفين",HR)]
DEPT_RANGES=[]
for dep,sh,title,recs in DEPTS:
    s,e=build_kpi_sheet(sh,title,recs)
    DEPT_RANGES.append((dep,sh,s,e))

# ===================== قاعدة المؤشرات (تجميع موحّد) =====================
cons=wb.create_sheet("قاعدة المؤشرات"); cons.sheet_view.showGridLines=False
for i,w in enumerate([4,22,42,16,26,12,14],1): cons.column_dimensions[get_column_letter(i)].width=w
cons.merge_cells("A1:G1")
style_cell(cons["A1"],value="قاعدة المؤشرات الموحّدة — مصدر تغذية اللوحة والخارطة والركائز (آلي)",f=font(14,True,WHITE),fillc=NAVY,align=center(False),border=False)
cons.merge_cells("A2:G2")
style_cell(cons["A2"],value="تُحدَّث آلياً من أوراق المؤشرات — لا تُعدَّل يدوياً",f=font(10,False,NAVY),fillc=GOLD,align=center(False),border=False)
for c,h in enumerate(["#","الإدارة","المؤشر","الركيزة","المشروع","نسبة التحقيق","الحالة"],1):
    style_cell(cons.cell(3,c,h),f=font(10,True,WHITE),fillc=BLUE,align=center())
cr=4; seq=1
for dep,sh,s,e in DEPT_RANGES:
    for rr in range(s,e+1):
        style_cell(cons.cell(cr,1,seq),f=font(9),align=center())
        style_cell(cons.cell(cr,2,dep),f=font(9),align=right())
        style_cell(cons.cell(cr,3,f"='{sh}'!C{rr}"),f=font(9),align=right(True))
        style_cell(cons.cell(cr,4,f"='{sh}'!X{rr}"),f=font(9),align=center())
        style_cell(cons.cell(cr,5,f"='{sh}'!Y{rr}"),f=font(9),align=right(True))
        style_cell(cons.cell(cr,6,f"='{sh}'!V{rr}"),f=font(9,True,NAVY),fmt="0%",align=center())
        style_cell(cons.cell(cr,7,f"='{sh}'!W{rr}"),f=font(9),align=center())
        cr+=1; seq+=1
CEND=cr-1
protect(cons)
CB=f"'قاعدة المؤشرات'!$B$4:$B${CEND}"   # الإدارة
CC=f"'قاعدة المؤشرات'!$D$4:$D${CEND}"   # الركيزة
CD=f"'قاعدة المؤشرات'!$E$4:$E${CEND}"   # المشروع
CV=f"'قاعدة المؤشرات'!$F$4:$F${CEND}"   # نسبة التحقيق
CS=f"'قاعدة المؤشرات'!$G$4:$G${CEND}"   # الحالة

# ===================== لوحة الإدارة التنفيذية =====================
ws=wb.create_sheet("لوحة الإدارة التنفيذية"); ws.sheet_view.showGridLines=False
for i,w in enumerate([3,30,16,16,16,16],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.merge_cells("B2:F2")
style_cell(ws["B2"],value="درب · لوحة الإدارة التنفيذية 2026 — المحطات والعقار",f=font(18,True,WHITE),fillc=NAVY,align=center(False),border=False)
ws.row_dimensions[2].height=38
ws.merge_cells("B4:C4"); style_cell(ws["B4"],value="مؤشر الإنجاز العام للشركة",f=font(11,True,WHITE),fillc=BLUE,align=center())
ws.merge_cells("B5:C6"); style_cell(ws["B5"],value=f'=IFERROR(AVERAGE({CV}),0)',f=font(30,True,NAVY),fmt="0%",align=center())
for i,(title,key,clr) in enumerate([("✅ محقق","✅ محقق",GREEN_IN),("🟡 قريب","🟡 قريب",GOLD),("🔴 تحت الهدف","🔴 تحت الهدف",RED_IN)]):
    L=get_column_letter(4+i)
    style_cell(ws[f"{L}4"],value=title,f=font(11,True),fillc=clr,align=center())
    style_cell(ws[f"{L}5"],value=f'=COUNTIF({CS},"{key}")',f=font(22,True,NAVY),align=center()); ws.merge_cells(f"{L}5:{L}6")
ws.row_dimensions[4].height=22; ws.row_dimensions[5].height=30
# الركائز (5 سنوات) — أعمدة H..J
for col,w in zip(["H","I","J"],[26,14,14]): ws.column_dimensions[col].width=w
style_cell(ws["H4"],value="الأداء حسب الركائز (5 سنوات)",f=font(12,True,WHITE),fillc=NAVY,align=center()); ws.merge_cells("H4:J4")
for c,h in zip([8,9,10],["الركيزة","الإنجاز","الحالة"]): style_cell(ws.cell(5,c,h),f=font(11,True,WHITE),fillc=STEEL,align=center())
for i,pl in enumerate(PILLARS):
    r=6+i
    style_cell(ws.cell(r,8,pl),f=font(11,True),align=right())
    style_cell(ws.cell(r,9,f'=IFERROR(AVERAGEIFS({CV},{CC},"{pl}"),"—")'),f=font(11,True,NAVY),fmt="0%",align=center())
    style_cell(ws.cell(r,10,f'=IF(ISNUMBER(I{r}),IF(I{r}>=1,"✅",IF(I{r}>=0.85,"🟡","🔴")),"—")'),f=font(11,True),align=center())
# جدول الإدارات
dr=9
style_cell(ws.cell(dr,2,"الأداء حسب الإدارة (تغذّي نفس اللوحة)"),f=font(12,True,WHITE),fillc=BLUE,align=center()); ws.merge_cells(start_row=dr,start_column=2,end_row=dr,end_column=4)
for c,h in zip([2,3,4],["الإدارة","الإنجاز","الحالة"]): style_cell(ws.cell(dr+1,c,h),f=font(11,True,WHITE),fillc=STEEL,align=center())
for i,(dep,sh,s,e) in enumerate(DEPT_RANGES):
    r=dr+2+i
    style_cell(ws.cell(r,2,dep),f=font(10),align=right())
    style_cell(ws.cell(r,3,f'=IFERROR(AVERAGEIFS({CV},{CB},"{dep}"),"—")'),f=font(10,True,NAVY),fmt="0%",align=center())
    style_cell(ws.cell(r,4,f'=IF(ISNUMBER(C{r}),IF(C{r}>=1,"✅ محقق",IF(C{r}>=0.85,"🟡 قريب","🔴 تحت الهدف")),"—")'),f=font(10,True),align=center())
protect(ws)

# ===================== الخارطة التنفيذية =====================
ws=wb.create_sheet("الخارطة التنفيذية"); ws.sheet_view.showGridLines=False
for i,w in enumerate([4,15,26,30,40,14,18],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.merge_cells("A1:G1")
style_cell(ws["A1"],value="الخارطة التنفيذية 2026 — مشاريع الشركة وربطها بالمؤشرات (المحطات والعقار)",f=font(15,True,WHITE),fillc=NAVY,align=center(False),border=False)
ws.row_dimensions[1].height=30
ws.merge_cells("A2:G2")
style_cell(ws["A2"],value="المصدر: «رؤية وأهداف 2026» — سحور درب  ·  التغطية: ✅ مرتبط 🟡 جزئي 🔴 فجوة  ·  أداء المؤشرات آلي من قاعدة المؤشرات",f=font(9,False,NAVY),fillc=GOLD,align=center(False),border=False)
for c,h in enumerate(["#","الركيزة","المشروع","الإدارة المسؤولة","المؤشرات المرتبطة","التغطية","أداء المؤشرات (مباشر)"],1):
    style_cell(ws.cell(3,c,h),f=font(10,True,WHITE),fillc=BLUE,align=center())
ws.row_dimensions[3].height=30
STMAP={"ok":("✅ مرتبط",GREEN_IN),"partial":("🟡 جزئي",GOLD),"gap":("🔴 فجوة",RED_IN)}
PROJ=[
 (PL_GROW,P_GROW1,"التشغيل + الامتياز + الاستثمار","محطات مشغّلة (85+150) · 190 عقد استثماري","ok"),
 (PL_GROW,P_FRAN,"الامتياز التجاري","عقود الامتياز (167) · الإيرادات والهامش","ok"),
 (PL_GROW,P_REACH,"الامتياز + التشغيل + العقار","مبيعات الوقود · نمو المبيعات · LEADS","ok"),
 (PL_INNO,P_NEW,"العقار (جزئي) — مقترح: إدارة تطوير","الشراكات الاستراتيجية (5) فقط","partial"),
 (PL_INNO,P_DESIGN,"غير محددة — مقترح: التسويق والعلامة","نسبة التزام المحطات بالهوية (55%) فقط","partial"),
 (PL_INNO,P_TECH,"التقنية الرقمية + التشغيل","أتمتة العمليات · جاهزية الأنظمة · المعيار الشامل","ok"),
 (PL_INNO,P_SAHAT,"العقار","تأجير ساحة درب (≥80%) · 6 علامات جديدة","ok"),
 (PL_INNO,P_TANKI,"التقنية الرقمية","MAU · التنزيلات · المعاملات الرقمية · الرضا","ok"),
 (PL_SUST,P_CX,"الامتياز + التشغيل + العقار + الاستثمار","CSAT · قوقل · الشكاوى · رضا المستأجرين","ok"),
 (PL_SUST,P_QUAL,"التشغيل + الامتياز + الاستثمار","جاهزية 99% · السلامة 100% · الصيانة","ok"),
 (PL_SUST,P_LEASE,"العقار","الإشغال (≥60%) · 68 وحدة/شهر · التحصيل","ok"),
 (PL_SUST,P_PEOPLE,"الموارد البشرية","الاستبقاء · التدريب · eNPS · السعودة","ok"),
]
groups={}
for p in PROJ: groups[p[0]]=groups.get(p[0],0)+1
r=4; i=0; done={}
for pillar,proj,dept,kpis,stt in PROJ:
    i+=1; sttxt,clr=STMAP[stt]
    style_cell(ws.cell(r,1,i),f=font(10,True),align=center())
    if pillar not in done:
        ws.merge_cells(start_row=r,start_column=2,end_row=r+groups[pillar]-1,end_column=2)
        style_cell(ws.cell(r,2,pillar),f=font(12,True,WHITE),fillc=BLUE,align=center()); done[pillar]=True
    style_cell(ws.cell(r,3,proj),f=font(11,True,NAVY),align=right(True))
    style_cell(ws.cell(r,4,dept),f=font(10),fillc=(RED_IN if "غير محددة" in dept else None),align=right(True))
    style_cell(ws.cell(r,5,kpis),f=font(10),align=right(True))
    style_cell(ws.cell(r,6,sttxt),f=font(11,True),fillc=clr,align=center())
    style_cell(ws.cell(r,7,f'=IF(COUNTIFS({CD},"{proj}")=0,"—",IFERROR(AVERAGEIFS({CV},{CD},"{proj}"),"—"))'),f=font(11,True,NAVY),fmt="0%",align=center())
    ws.row_dimensions[r].height=30; r+=1
r+=1; ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
style_cell(ws.cell(r,2,"✅ أُغلقت فجوات تانكي/التحول التقني/الموظفين بإضافة إدارتَي التقنية الرقمية والموارد البشرية · المتبقّي:"),f=font(11,True,WHITE),fillc="538135",align=right()); r+=1
for g in ["🟡 إطلاق مشاريع جديدة — يغذّيه مؤشر الشراكات الاستراتيجية فقط؛ يحتاج مالكاً (إدارة تطوير/PMO) ومؤشرات أوضح.",
          "🟡 تطوير التصاميم والهوية — مغطّى بمؤشر التزام الهوية فقط؛ يحتاج إدارة تسويق/علامة مالكة.",
          "📌 توصية: تثبيت خطوط الأساس 2025 والأوزان (الأعمدة الجديدة) لكل مؤشر لتدقيق المستهدفات."]:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
    style_cell(ws.cell(r,2,g),f=font(10),align=right(True)); ws.row_dimensions[r].height=24; r+=1
protect(ws)

# ===================== نماذج تقييم الامتياز (5) =====================
SCALE=[("ممتاز","90% – 100%","تجاوز المستهدف"),("جيد جداً","80% – 89%","حقق المستهدف تقريباً"),
       ("جيد","70% – 79%","حقق معظم المستهدفات"),("مقبول","60% – 69%","يحتاج تحسين"),("ضعيف","أقل من 60%","يحتاج خطة تطوير")]
def eval_form(sheet,role,code,rows):
    ws=wb.create_sheet(sheet); ws.sheet_view.showGridLines=False
    for i,w in enumerate([4,40,10,34,14,14,22],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.merge_cells("A1:G1"); style_cell(ws["A1"],value=f"إدارة الامتياز التجاري · نموذج تقييم الأداء — {role}",f=font(14,True,WHITE),fillc=NAVY,align=center(False),border=False)
    ws.row_dimensions[1].height=28
    ws.merge_cells("A2:G2"); style_cell(ws["A2"],value=f"الكود: {code}  ·  ربع سنوي  ·  🟩 أدخل «نسبة التحقيق» فقط — الدرجة تُحسب تلقائياً",f=font(10,False,NAVY),fillc=GOLD,align=center(False),border=False)
    rr=3
    for a,b in [("اسم الموظف:","القسم: الامتياز التجاري"),("المدير المباشر:","الربع / السنة:"),("تاريخ التقييم:","حالة التقييم:")]:
        style_cell(ws.cell(rr,1,a),f=font(10,True),align=right())
        style_cell(ws.cell(rr,2,""),fillc=GREEN_IN,lock=False); ws.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=3)
        style_cell(ws.cell(rr,4,b),f=font(10,True),align=right())
        style_cell(ws.cell(rr,5,""),fillc=GREEN_IN,lock=False); ws.merge_cells(start_row=rr,start_column=5,end_row=rr,end_column=7); rr+=1
    for c,h in enumerate(["#","مؤشر الأداء","الوزن %","المستهدف","نسبة التحقيق","الدرجة المحققة","ملاحظات"],1):
        style_cell(ws.cell(rr,c,h),f=font(10,True,WHITE),fillc=BLUE,align=center())
    ws.row_dimensions[rr].height=30; first=rr+1
    dvp=DataValidation(type="decimal",operator="between",formula1="0",formula2="1.2",allow_blank=True)
    dvp.error="أدخل نسبة بين 0% و120%"; dvp.errorTitle="نسبة التحقيق"; ws.add_data_validation(dvp)
    rr=first
    for i,(nm,w,tgt) in enumerate(rows):
        style_cell(ws.cell(rr,1,i+1),f=font(10,True),align=center())
        style_cell(ws.cell(rr,2,nm),f=font(10),align=right(True))
        style_cell(ws.cell(rr,3,w/100.0),fmt="0%",align=center())
        style_cell(ws.cell(rr,4,tgt),f=font(9,color="595959"),align=right(True))
        ac=ws.cell(rr,5); style_cell(ac,fillc=GREEN_IN,fmt="0%",align=center(),lock=False); dvp.add(ac)
        style_cell(ws.cell(rr,6,f"=C{rr}*MIN(E{rr},1)"),f=font(10,True,NAVY),fmt="0%",align=center())
        style_cell(ws.cell(rr,7,""),fillc=GREEN_IN,lock=False); ws.row_dimensions[rr].height=28; rr+=1
    last=rr-1
    style_cell(ws.cell(rr,2,"المجموع الكلي"),f=font(11,True,WHITE),fillc=NAVY,align=center())
    style_cell(ws.cell(rr,3,f"=SUM(C{first}:C{last})"),f=font(11,True,WHITE),fillc=NAVY,fmt="0%",align=center())
    style_cell(ws.cell(rr,4,""),fillc=NAVY)
    style_cell(ws.cell(rr,5,"الدرجة:"),f=font(11,True,WHITE),fillc=NAVY,align=center())
    style_cell(ws.cell(rr,6,f"=SUM(F{first}:F{last})"),f=font(13,True,"F47A21"),fillc=NAVY,fmt="0%",align=center())
    tc=f"F{rr}"; style_cell(ws.cell(rr,1,""),fillc=NAVY); style_cell(ws.cell(rr,7,""),fillc=NAVY); rr+=1
    style_cell(ws.cell(rr,2,"التقدير"),f=font(11,True,WHITE),fillc=BLUE,align=center())
    style_cell(ws.cell(rr,3,f'=IF({tc}>=0.9,"ممتاز",IF({tc}>=0.8,"جيد جداً",IF({tc}>=0.7,"جيد",IF({tc}>=0.6,"مقبول","ضعيف"))))'),f=font(11,True,NAVY),align=center()); ws.merge_cells(start_row=rr,start_column=3,end_row=rr,end_column=4); rr+=2
    style_cell(ws.cell(rr,2,"سلّم التقييم"),f=font(11,True,WHITE),fillc=BLUE,align=center()); ws.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=4); rr+=1
    for g,p,m in SCALE:
        style_cell(ws.cell(rr,2,g),f=font(10,True),align=center()); style_cell(ws.cell(rr,3,p),f=font(10),align=center())
        style_cell(ws.cell(rr,4,m),f=font(10),align=right()); rr+=1
    protect(ws)

eval_form("الامتياز · مدير الإدارة","مدير إدارة الامتياز التجاري","FM-01",[
 ("عدد محطات الامتياز الجديدة",15,"100%+ من المحطات المستهدفة"),("إجمالي مبيعات محطات الامتياز",15,"100%+ من مستهدف المبيعات الربعي"),
 ("استمرارية تزويد الوقود",10,"0 حالات انقطاع"),("نسبة التحصيل المالي",15,"100%+ من المستحقات الربعية"),
 ("الالتزام بالهوية المؤسسية",15,"70%+ من المحطات ملتزمة"),("انخفاض عدد الشكاوى",10,"خفض 20%+ عن الربع السابق"),
 ("تقييمات خرائط قوقل",10,"4.7 – 5"),("انتظام وجودة التقارير",10,"100% في موعدها")])
eval_form("الامتياز · مسؤول الامتياز","مسؤول الامتياز التجاري","FM-02",[
 ("المتابعة مع عملاء الامتياز",30,"متابعة مستمرة 100%"),("عدم انقطاع المحطات من الوقود",35,"0 حالات انقطاع"),
 ("متابعة طلبات وتوريد الوقود",10,"100% متابعة يومية"),("إجمالي مبيعات المحطات",10,"تحقيق هدف المبيعات"),
 ("التحصيل المالي",10,"100% خلال الربع"),("التقارير الدورية الشاملة",5,"100% في مواعيدها")])
eval_form("الامتياز · المراقب الميداني","المراقب الميداني","FM-03",[
 ("عدد الزيارات الميدانية المنفذة",30,"100% من الزيارات المجدولة"),("متابعة تطبيق الهوية المؤسسية",20,"100% دون ملاحظات متكررة"),
 ("تقارير الجودة الميدانية والـChecklist",20,"100% رفع التقارير"),("متابعة العلامات التجارية",10,"100% حسب الجدول"),
 ("تطبيق معايير السلامة الميدانية",10,"100%"),("متابعة الشكاوى وإغلاقها",10,"الاستجابة خلال 24 ساعة")])
eval_form("الامتياز · مسؤول التعاقدات","مسؤول التعاقدات","FM-04",[
 ("جودة التفاوض مع العملاء",20,"98% توافق مع المستهدفات"),("اكتمال المستندات النظامية قبل التوقيع",15,"100%"),
 ("سرعة إنجاز إجراءات العقد",30,"خلال يومي عمل 100%"),("أرشفة العقود الإلكترونية",10,"خلال 48 ساعة 100%"),
 ("متابعة تنفيذ بنود العقد",10,"100%"),("مدى رضا العملاء",10,"95% رضا"),("تحديث قاعدة بيانات العملاء",5,"100% دوري")])
eval_form("الامتياز · الموظف الإداري","الموظف الإداري","FM-05",[
 ("دقة وسرعة إعداد التقارير الدورية",30,"100% في الوقت المحدد"),("متابعة الشكاوى وحالات العملاء",30,"100% حتى الإغلاق"),
 ("متابعة عمولات تأجير الوحدات",20,"100%"),("تحديث بيانات العملاء والمحطات",10,"100% دقة"),("إعداد المراسلات الرسمية",10,"100% في الوقت المحدد")])

# ===================== مرجع مؤشرات موظفي الامتياز =====================
ws=wb.create_sheet("الامتياز · الموظفون"); ws.sheet_view.showGridLines=False
for i,w in enumerate([3,22,46,12,4],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.merge_cells("B2:E2"); style_cell(ws["B2"],value="إدارة الامتياز · مؤشرات كل موظف (مرجع الربط)",f=font(15,True,WHITE),fillc=NAVY,align=center(False),border=False); ws.row_dimensions[2].height=28
EMP=[("علاء عزت","أخصائي امتياز — عقود واستقطاب",[("●","عدد عقود الامتياز الجديدة الموقّعة","🔴 رئيسي"),("●","عدد محطات الامتياز المشغّلة (تراكمي)","🔴 رئيسي"),("●","نسبة تحويل الفرص إلى عقود موقّعة","🔴 رئيسي"),("●","نسبة استدامة أصحاب الامتياز","🔴 رئيسي")]),
 ("أنس أبو طوق","مراقب ميداني — هوية وبنية تحتية",[("●","نسبة جاهزية محطات الامتياز","🔴 رئيسي"),("●","نسبة إغلاق ملاحظات الجودة الميدانية","🔴 رئيسي"),("●","نسبة المحطات المستوفية لمعايير السلامة","🔴 رئيسي"),("●","نسبة التزام المحطات بالهوية المؤسسية","🔴 رئيسي")]),
 ("حمدي عبد الجليل","أخصائي امتياز — زيارات واستقطاب",[("●","نسبة تنفيذ الزيارات الميدانية المجدولة","🔴 رئيسي"),("○","عدد عقود الامتياز الجديدة الموقّعة","🔴 رئيسي"),("○","متوسط درجة تقييم جودة المحطة","🔴 رئيسي")]),
 ("رغد الجندي","أخصائية امتياز — تسويق ومالية",[("●","إجمالي إيرادات إدارة الامتياز","🔴 رئيسي"),("●","نسبة التحصيل المالي من شركاء الامتياز","🔴 رئيسي"),("●","متوسط تقييم المحطات على خرائط قوقل","🔴 رئيسي"),("●","عدد الحملات التسويقية الميدانية","🟠 مهم")]),
 ("إدريس الوائلي","أخصائي امتياز — تقارير وتنسيق مالي",[("●","متوسط أيام التحصيل (DSO)","🟠 مهم"),("●","إجمالي المستحقات المتأخرة (+30 يوم)","🟠 مهم"),("●","نسبة إنجاز التقارير الدورية في موعدها","🟠 مهم")]),
 ("أحمد رمضان","أخصائي امتياز — Dynamics ومتابعة نظام",[("○","نسبة اكتمال المستندات النظامية قبل التوقيع","🟠 مهم"),("○","نسبة إنجاز التقارير الدورية في موعدها","🟠 مهم")]),
 ("عبدالسلام الوالي","أخصائي امتياز — عمليات ومخزون",[("●","عدد حالات انقطاع الوقود الموثّقة","🔴 رئيسي"),("●","متوسط وقت معالجة البلاغات التشغيلية","🟠 مهم")]),
 ("المعتصم المصعبي","أخصائي امتياز — عمليات ومبيعات",[("●","متوسط مبيعات المحطة الواحدة شهرياً","🟠 مهم"),("○","إجمالي مبيعات وقود محطات الامتياز","🔴 رئيسي")]),
 ("يحيى العدله","أخصائي امتياز — تسجيل وتوثيق",[("●","نسبة اكتمال المستندات النظامية قبل التوقيع","🟠 مهم"),("○","نسبة حل الشكاوى خلال المدة المحددة","🟠 مهم")]),
 ("إسماعيل سندي","أخصائي امتياز — تواصل عملاء وتأجير",[("●","معدل رضا شركاء الامتياز (CSAT)","🔴 رئيسي"),("●","نسبة حل الشكاوى خلال المدة المحددة","🟠 مهم"),("●","نسبة انخفاض الشكاوى التشغيلية والإدارية","🟠 مهم")]),]
r=4
for nm,role,kpis in EMP:
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
    style_cell(ws.cell(r,2,f"{nm}  —  {role}  ·  {len(kpis)} مؤشر"),f=font(12,True,WHITE),fillc=BLUE,align=right()); r+=1
    for c,h in zip([2,3,4],["المؤشر","الأولوية","●/○"]): style_cell(ws.cell(r,c if c!=4 else 4,h),f=font(10,True,WHITE),fillc=STEEL,align=center())
    r+=1
    for mark,kpi,pri in kpis:
        style_cell(ws.cell(r,2,kpi),f=font(10),align=right(True)); style_cell(ws.cell(r,3,pri),f=font(10),align=center()); style_cell(ws.cell(r,4,mark),f=font(11,True),align=center()); r+=1
    r+=1
protect(ws)

# ===================== أوراق أدوار التشغيل (3) =====================
OPS_MGR=[("الرقابة المالية","تنفيذ جرد الوردية مرتين يومياً","%","↑","AVG",1.0,"100%","pct","",""),
 ("الرقابة المالية","تسليم حصيلة المبيعات في الموعد","%","↑","AVG",1.0,"100%","pct","",""),
 ("الرقابة المالية","التطابق بين POS والنقدية","%","↑","AVG",1.0,"100%","pct","",""),
 ("الرقابة المالية","حالات عدم تطابق غير مبررة","حالة","↓","SUM",0,"0 حالة","int","",""),
 ("توريد الوقود","دقة قياس مستوى الخزان يومياً","%","↑","AVG",1.0,"100%","pct","",""),
 ("توريد الوقود","طلب توريد الوقود قبل نفاده","%","↑","AVG",1.0,"100%","pct","",""),
 ("توريد الوقود","حالات انقطاع الوقود","حالة","↓","SUM",0,"0 حالة","int","",""),
 ("توريد الوقود","توثيق تفريغ الوقود (OP-F-36)","%","↑","AVG",1.0,"100%","pct","",""),
 ("إدارة الفريق","الالتزام بزي الفريق ومظهره","%","↑","AVG",1.0,"100%","pct","",""),
 ("إدارة الفريق","توزيع الزي بعهدة موقّعة","%","↑","AVG",1.0,"100%","pct","",""),
 ("إدارة الفريق","تجديد الوثائق والتراخيص مبكراً","%","↑","AVG",1.0,"100%","pct","",""),
 ("إدارة الفريق","الاستجابة لبلاغات الطوارئ والأعطال","دقيقة","↓","AVG",15,"≤ 15 د","num1","",""),
 ("جودة الخدمة","نظافة المرافق يومياً","%","↑","AVG",1.0,"100%","pct","",""),
 ("جودة الخدمة","معالجة شكاوى العملاء خلال 24 ساعة","%","↑","AVG",0.90,"90%+","pct","",""),]
OPS_SUP=[("الزيارات الميدانية","تنفيذ الزيارات المجدولة كاملاً","%","↑","AVG",1.0,"100%","pct","",""),
 ("الزيارات الميدانية","رفع تقرير الزيارة خلال 24 ساعة","%","↑","AVG",1.0,"100%","pct","",""),
 ("الزيارات الميدانية","التحقق من جرد الوردية أثناء الزيارة","%","↑","AVG",1.0,"100%","pct","",""),
 ("الزيارات الميدانية","إغلاق ملاحظات الزيارة خلال أسبوع","%","↑","AVG",0.90,"90%+","pct","",""),
 ("الرقابة والتوريد","التحقق من تطابق POS والنقدية","%","↑","AVG",1.0,"100%","pct","",""),
 ("الرقابة والتوريد","متابعة طلبات التوريد وتسليم الوقود","%","↑","AVG",1.0,"100%","pct","",""),
 ("الرقابة والتوريد","مراجعة نسبة جاهزية المضخات","%","↑","AVG",0.97,"≥ 97%","pct","",""),
 ("متابعة الموظفين","تقييم أداء الموظفين بنموذج OP-F-51","%","↑","AVG",1.0,"100%","pct","",""),
 ("متابعة الموظفين","متابعة الالتزام بالزي والمظهر","%","↑","AVG",1.0,"100%","pct","",""),
 ("متابعة الموظفين","حالات مخالفة السلامة دون إجراء","حالة","↓","SUM",0,"0 حالة","int","",""),
 ("التقارير","رفع التقارير الشهرية في موعدها","%","↑","AVG",1.0,"100%","pct","",""),
 ("التقارير","متابعة تجديد تراخيص المحطات","%","↑","AVG",1.0,"100%","pct","",""),]
OPS_FUEL=[("الخدمة والسرعة","وقت تعبئة الوقود للعميل","دقيقة","↓","AVG",5,"≤ 5 دقائق","num1","",""),
 ("الخدمة والسرعة","الترحيب بالعميل عند الوصول","%","↑","AVG",1.0,"100%","pct","",""),
 ("الخدمة والسرعة","إعادة تركيب غطاء الخزان","%","↑","AVG",1.0,"100%","pct","",""),
 ("الخدمة والسرعة","عرض إيصال الدفع على العميل","%","↑","AVG",1.0,"100%","pct","",""),
 ("المظهر والالتزام","الالتزام بالزي الرسمي كاملاً","%","↑","AVG",1.0,"100%","pct","",""),
 ("المظهر والالتزام","النظافة الشخصية والمظهر العام","%","↑","AVG",1.0,"100%","pct","",""),
 ("المظهر والالتزام","حالات استخدام الجوال أثناء العمل","حالة","↓","SUM",0,"0 حالة","int","",""),
 ("الدقة والسلامة","دقة نوع الوقود المُضخ","%","↑","AVG",1.0,"100%","pct","",""),
 ("الدقة والسلامة","إغلاق المضخة عند اكتمال الكمية","%","↑","AVG",1.0,"100%","pct","",""),
 ("الدقة والسلامة","الالتزام بإجراءات السلامة عند الضخ","%","↑","AVG",1.0,"100%","pct","",""),]
build_kpi_sheet("التشغيل · مدير المحطة","مؤشرات الأداء — مدير المحطة (شهري · SOP-OM-02/13)",OPS_MGR,tagged=False)
build_kpi_sheet("التشغيل · مشرف ميداني","مؤشرات الأداء — مشرف ميداني (شهري · TQM-M-02)",OPS_SUP,tagged=False)
build_kpi_sheet("التشغيل · عامل التعبئة","مؤشرات الأداء — عامل تعبئة الوقود (شهري · SOP-OM-08)",OPS_FUEL,tagged=False)

# ===================== المؤشرات الحيوية (North Star + Vital Few) =====================
SH_F='الامتياز · المؤشرات'; SH_I='الاستثمار · المؤشرات'; SH_R='العقار · المؤشرات'
SH_D='التقنية الرقمية · المؤشرات'; SH_H='الموارد البشرية · المؤشرات'
ws=wb.create_sheet("المؤشرات الحيوية"); ws.sheet_view.showGridLines=False
for i,w in enumerate([3,16,18,40,16,14,12,14],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.merge_cells("B2:H2")
style_cell(ws["B2"],value="المؤشرات الحيوية — النجمة الشمالية والمؤشرات الاستراتيجية القليلة",f=font(17,True,WHITE),fillc=NAVY,align=center(False),border=False)
ws.row_dimensions[2].height=34
# النجمة الشمالية
ws.merge_cells("B4:D4"); style_cell(ws["B4"],value="🌟 النجمة الشمالية — إجمالي مبيعات الوقود (لتر)",f=font(12,True,WHITE),fillc=ORANGE,align=center())
ws.merge_cells("B5:C6"); style_cell(ws["B5"],value=f"='{SH_F}'!U6",f=font(26,True,ORANGE),fmt="#,##0",align=center())
style_cell(ws["D5"],value="المستهدف 2026",f=font(10,True),align=center())
style_cell(ws["D6"],value=f"='{SH_F}'!H6",f=font(12,True,NAVY),align=center())
ws.merge_cells("E4:H4"); style_cell(ws["E4"],value="نسبة تحقيق النجمة الشمالية",f=font(12,True,WHITE),fillc=BLUE,align=center())
ws.merge_cells("E5:H6"); style_cell(ws["E5"],value=f"='{SH_F}'!V6",f=font(28,True,NAVY),fmt="0%",align=center())
ws.row_dimensions[5].height=28
# جدول المؤشرات الحيوية
hr=8
for c,h in zip([2,3,4,5,6,7,8],["الركيزة","الإدارة","المؤشر","المستهدف","YTD","الإنجاز","الحالة"]):
    style_cell(ws.cell(hr,c,h),f=font(11,True,WHITE),fillc=BLUE,align=center())
ws.row_dimensions[hr].height=26
VITAL=[(PL_GROW,"الامتياز",SH_F,6),(PL_GROW,"الامتياز",SH_F,4),(PL_GROW,"الامتياز",SH_F,5),(PL_GROW,"الاستثمار",SH_I,4),
 (PL_INNO,"الامتياز",SH_F,24),(PL_INNO,"العقار",SH_R,12),(PL_INNO,"التقنية الرقمية",SH_D,4),
 (PL_SUST,"الامتياز",SH_F,12),(PL_SUST,"الامتياز",SH_F,17),(PL_SUST,"الامتياز",SH_F,32),
 (PL_SUST,"العقار",SH_R,7),(PL_SUST,"العقار",SH_R,19),(PL_SUST,"الاستثمار",SH_I,11),(PL_SUST,"الموارد البشرية",SH_H,4)]
vg={}
for v in VITAL: vg[v[0]]=vg.get(v[0],0)+1
r=hr+1; doneV={}
for pillar,dep,sh,row in VITAL:
    if pillar not in doneV:
        ws.merge_cells(start_row=r,start_column=2,end_row=r+vg[pillar]-1,end_column=2)
        style_cell(ws.cell(r,2,pillar),f=font(12,True,WHITE),fillc=STEEL,align=center()); doneV[pillar]=True
    style_cell(ws.cell(r,3,dep),f=font(10),align=center())
    style_cell(ws.cell(r,4,f"='{sh}'!C{row}"),f=font(10),align=right(True))
    style_cell(ws.cell(r,5,f"='{sh}'!H{row}"),f=font(10),align=center())
    style_cell(ws.cell(r,6,f"='{sh}'!U{row}"),f=font(10,True),align=center())
    style_cell(ws.cell(r,7,f"='{sh}'!V{row}"),f=font(10,True,NAVY),fmt="0%",align=center())
    style_cell(ws.cell(r,8,f"='{sh}'!W{row}"),f=font(10,True),align=center())
    ws.row_dimensions[r].height=24; r+=1
protect(ws)

# ===================== الحوكمة (RACI + إيقاع + طموحات) =====================
ws=wb.create_sheet("الحوكمة"); ws.sheet_view.showGridLines=False
for i,w in enumerate([3,26,26,26,30],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.merge_cells("B2:E2"); style_cell(ws["B2"],value="حوكمة الأداء — الإيقاع والمسؤوليات وطموحات النمو",f=font(16,True,WHITE),fillc=NAVY,align=center(False),border=False)
ws.row_dimensions[2].height=30
r=4
style_cell(ws.cell(r,2,"⏱️ إيقاع المراجعة (Cadence)"),f=font(12,True,WHITE),fillc=BLUE,align=right()); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5); r+=1
for c,h in zip([2,3,4,5],["الدورية","النشاط","المسؤول","المخرج"]): style_cell(ws.cell(r,c,h),f=font(10,True,WHITE),fillc=STEEL,align=center())
r+=1
for row in [("شهري","مراجعة تشغيلية لمؤشرات كل إدارة","مدير الإدارة","إغلاق الأرقام + خطة تصحيح"),
            ("ربع سنوي (QBR)","مراجعة استراتيجية + تقييم الأفراد + معايرة المستهدفات","الإدارة التنفيذية","قرارات + تحديث الأوزان"),
            ("نصف سنوي","مراجعة الخارطة التنفيذية والمشاريع","مكتب الاستراتيجية","تقدّم المشاريع + المخاطر"),
            ("سنوي","إعادة ضبط OKR وربطها بالركائز الخمسية","الرئيس التنفيذي","خطة العام القادم")]:
    for c,v in zip([2,3,4,5],row): style_cell(ws.cell(r,c,v),f=font(10),align=right(True))
    ws.row_dimensions[r].height=30; r+=1
r+=1
style_cell(ws.cell(r,2,"👥 مصفوفة RACI"),f=font(12,True,WHITE),fillc=BLUE,align=right()); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5); r+=1
style_cell(ws.cell(r,2,"R = منفّذ · A = مُساءل · C = مُستشار · I = مُطّلع"),f=font(10,False,NAVY),fillc=GOLD,align=right()); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5); r+=1
for c,h in zip([2,3,4,5],["المهمة","المسؤول (A)","المنفّذ (R)","يُستشار/يُطّلع"]): style_cell(ws.cell(r,c,h),f=font(10,True,WHITE),fillc=STEEL,align=center())
r+=1
for row in [("إدخال الأرقام الشهرية","مدير الإدارة","المسؤول عن المؤشر","المشرف الميداني (C)"),
            ("اعتماد المستهدفات والأوزان","الإدارة التنفيذية","مكتب الاستراتيجية","مدراء الإدارات (C)"),
            ("تقييم أداء الموظف","المدير المباشر","المدير المباشر","الموارد البشرية (I)"),
            ("تجميع لوحة التنفيذيين","مكتب الاستراتيجية","محلل الأداء","الإدارة التنفيذية (I)")]:
    for c,v in zip([2,3,4,5],row): style_cell(ws.cell(r,c,v),f=font(10),align=right(True))
    ws.row_dimensions[r].height=26; r+=1
r+=1
style_cell(ws.cell(r,2,"🚀 طموحات 3 سنوات (2026 → 2028) — الجسر بين الرؤية الخمسية والسنة"),f=font(12,True,WHITE),fillc=BLUE,align=right()); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5); r+=1
for c,h in zip([2,3,4,5],["الركيزة","2026","2027","2028"]): style_cell(ws.cell(r,c,h),f=font(10,True,WHITE),fillc=STEEL,align=center())
r+=1
for row in [("النمو — المحطات المشغّلة","85 + 150","≈ 320","≈ 450"),
            ("النمو — العقود الجديدة","167 + 190","تصاعدي","تصاعدي"),
            ("الابتكار — أتمتة/تانكي","إطلاق وتثبيت","توسّع","نضج رقمي"),
            ("الاستدامة — رضا وجودة","CSAT 90% · جاهزية 99%","الحفاظ والتحسين","تميّز تشغيلي")]:
    for c,v in zip([2,3,4,5],row): style_cell(ws.cell(r,c,v),f=font(10),align=right(True if c==2 else False));
    ws.row_dimensions[r].height=26; r+=1
style_cell(ws.cell(r+1,2,"📌 الأرقام التقديرية للطموحات الثلاثية تحتاج اعتماد الإدارة التنفيذية وربطها بخطوط الأساس 2025."),f=font(9,False,"808080"),align=right()); ws.merge_cells(start_row=r+1,start_column=2,end_row=r+1,end_column=5)
protect(ws)

# ===================== ترتيب الأوراق =====================
order=["دليل الاستخدام","الاستراتيجية","المؤشرات الحيوية","الخارطة التنفيذية","لوحة الإدارة التنفيذية","الحوكمة",
 "الامتياز · المؤشرات","الامتياز · الموظفون","الامتياز · مدير الإدارة","الامتياز · مسؤول الامتياز",
 "الامتياز · المراقب الميداني","الامتياز · مسؤول التعاقدات","الامتياز · الموظف الإداري",
 "التشغيل · المؤشرات","التشغيل · مدير المحطة","التشغيل · مشرف ميداني","التشغيل · عامل التعبئة",
 "الاستثمار · المؤشرات","العقار · المؤشرات","التقنية الرقمية · المؤشرات","الموارد البشرية · المؤشرات","قاعدة المؤشرات"]
for idx,nm in enumerate(order):
    cur=wb.sheetnames.index(nm); wb.move_sheet(nm, idx-cur)

wb.security=WorkbookProtection(workbookPassword=PWD, lockStructure=True)
out="درب-مؤشرات-الأداء-2026.xlsx"; wb.save(out)
print("saved:",out,os.path.getsize(out),"bytes  | sheets:",len(wb.sheetnames))
print("consolidated rows:",CEND-3)
