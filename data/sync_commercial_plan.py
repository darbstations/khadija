# -*- coding: utf-8 -*-
"""مزامنة ورقة نموذج العامل داخل الخطة التجارية مع النسخة المعتمدة"""
# يُشغَّل من جذر المستودع:  PYTHONPATH=data python3 data/sync_commercial_plan.py
from openpyxl import load_workbook
import worker_model as W
import competition_model as C
import rebrand as R
import financial_model as FIN

P = "docs/commercial-plan.xlsx"
ST = W.stations()
FIVE = C.five()
wb = load_workbook(P)
old = wb.sheetnames.index("نموذج العامل")
comp = wb.sheetnames.index("المنافسون")
print("قبل:", " · ".join(wb.sheetnames))

for name in ("نموذج العامل", W.DATA_SHEET, "مطابقة الوردية مع الطلب",
             C.SHEET, C.RIVALS, FIN.SHEET):
    if name in wb.sheetnames:
        del wb[name]

# المنافسون يُعاد بناؤها (أُضيفت درب باي وأُعيد تصنيف Fleet Plus)
# الأداء المالي يسبق كل شيء — هو إطار قراءة الخطة كلها
FIN.build_financial(wb, idx=1)
comp = wb.sheetnames.index(C.RIVALS) if C.RIVALS in wb.sheetnames else comp + 1
C.build_rivals(wb, idx=comp)
# والسيطرة الميدانية تليها مباشرة — فهي شاهدها
C.build_control(wb, FIVE, idx=comp + 1)
# نموذج العامل حُذف للتو، فيُثبَّت موضعه بجاره الباقي لا باسمه
old = wb.sheetnames.index("تجربة الشرايع")
end = W.build_data(wb, ST)                  # تُضاف في النهاية
W.build_worker(wb, ST, end, idx=old)
W.build_shift(wb, ST, idx=old + 1)

# الأوراق القديمة (بُنيت بسكربت سابق) تُعاد لألوان الهوية
BUILT = {"المنافسون", C.SHEET, "نموذج العامل", "مطابقة الوردية مع الطلب", W.DATA_SHEET}
done = R.rebrand_all(wb, skip=BUILT)
print("أُعيد تلوين:", " · ".join(f"{k} ({v})" for k, v in done.items() if v))

wb.save(P)
print("بعد:", " · ".join(wb.sheetnames))
