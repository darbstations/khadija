# -*- coding: utf-8 -*-
"""ورقة سجل الوحدات التأجيرية — من سجل المحطات الرسمي (١٨٦ محطة)"""
import csv
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule, DataBarRule
import worker_model as W
import brand as B

SHEET = "سجل الوحدات التأجيرية"
TYPES = [("kiosk", "كشك"), ("shop", "محل"), ("drive", "درايف ثرو"),
         ("carwash", "خدمات سيارات + مغسلة"), ("market", "سوبرماركت"),
         ("other", "مرافق أخرى")]
CATS = ["مشغّلة", "تحت التنفيذ", "امتياز"]
CATNOTE = {
    "مشغّلة": "الشبكة العاملة — إشغال صحي ولا أزمة شغور",
    "تحت التنفيذ": "لم تفتح بعد — الشغور طبيعي، وهنا فرصة التأجير المسبق",
    "امتياز": "الأسوأ إشغالاً — ومعظمها محلات لم تُسوَّق",
}


def units(path="outlets/data/units-registry.csv"):
    return list(csv.DictReader(open(path, encoding="utf-8")))


def _n(r, k):
    try: return int(float(r[k] or 0))
    except (ValueError, TypeError): return 0


def build_units(wb, U, idx=None):
    NC = 11
    ws = wb.create_sheet(SHEET) if idx is None else wb.create_sheet(SHEET, idx)
    W.setup(ws, [4, 20, 10, 11, 11, 11, 11, 11, 11, 11, 40], freeze="A5")
    W.title(ws, "سجل الوحدات التأجيرية — ١٨٦ محطة",
            "من سجل المحطات الرسمي · الوحدة = كشك أو محل أو درايف ثرو أو مغسلة أو سوبرماركت", NC)
    W.rule(ws, 3, NC)

    # ── ① الملخص حسب الفئة
    r = W.band(ws, 4, NC, "① الصورة الكاملة — ثلاث فئات لا واحدة")
    W.header(ws, r, ["#", "الفئة", "محطات", "الوحدات", "المؤجر", "المحجوز",
                     "الشاغر", "الإشغال"] + [""] * 2 + ["القراءة"])
    c0 = r + 1
    for i, cat in enumerate(CATS):
        g = [x for x in U if x["category"] == cat]
        n = sum(_n(x, "total_n") for x in g); le = sum(_n(x, "total_leased") for x in g)
        hd = sum(_n(x, "total_held") for x in g); vc = sum(_n(x, "total_vacant") for x in g)
        rr = c0 + i
        for j, (v, fm) in enumerate(zip(
                [i + 1, cat, len(g), n, le, hd, vc, le / n if n else 0],
                [None, None, "0", "#,##0", "#,##0", "#,##0", "#,##0", "0%"]), 1):
            c = ws.cell(rr, j, v); c.border = W.BOX; c.alignment = W.CTR
            c.font = W.BOLD if j == 2 else W.BLACK
            if fm: c.number_format = fm
        ws.cell(rr, NC, CATNOTE[cat]).alignment = W.WRAP
        ws.cell(rr, NC).border = W.BOX
        ws.row_dimensions[rr].height = 22
    cend = c0 + len(CATS) - 1
    tr = cend + 1
    ws.cell(tr, 2, "المجموع").font = W.BOLD
    for j in (3, 4, 5, 6, 7):
        ws.cell(tr, j, f"=SUM({chr(64+j)}{c0}:{chr(64+j)}{cend})").number_format = "#,##0"
    ws.cell(tr, 8, f"=IFERROR(E{tr}/D{tr},\"\")").number_format = "0%"
    for j in range(1, NC + 1):
        ws.cell(tr, j).border = W.BOX; ws.cell(tr, j).fill = W.FS
        ws.cell(tr, j).font = W.BOLD; ws.cell(tr, j).alignment = W.CTR
    ws.conditional_formatting.add(f"H{c0}:H{cend}", CellIsRule(
        operator="greaterThanOrEqual", formula=["0.7"], fill=W.OK,
        font=Font(name=W.F, size=10, bold=True, color=B.D_GOOD)))
    ws.conditional_formatting.add(f"H{c0}:H{cend}", CellIsRule(
        operator="lessThan", formula=["0.3"], fill=W.BAD,
        font=Font(name=W.F, size=10, bold=True, color=B.D_BAD)))
    r = tr + 2

    # ── ② حسب نوع الوحدة
    r = W.band(ws, r, NC, "② أي نوع من الوحدات لا يُؤجَّر")
    W.header(ws, r, ["#", "نوع الوحدة"] + [f"{c} — العدد" for c in CATS]
             + ["الإجمالي", "المؤجر", "الشاغر", "الإشغال"] + [""] + ["الدلالة"])
    t0 = r + 1
    TNOTE = {"محل": "المنتج العالق — أضعف إشغال في كل الفئات، وهو أكثر الوحدات عدداً",
             "كشك": "الأعلى إشغالاً في المشغّلة",
             "سوبرماركت": "إشغال جيد حيث تعمل المحطة",
             "خدمات سيارات + مغسلة": "معيار الافتتاح الإلزامي — الطلب عليه قائم",
             "درايف ثرو": "عدد محدود وإشغال متوسط",
             "مرافق أخرى": "صرافات ومرافق — عدد هامشي"}
    for i, (key, ar) in enumerate(TYPES):
        rr = t0 + i
        per = [sum(_n(x, key + "_n") for x in U if x["category"] == c) for c in CATS]
        n = sum(_n(x, key + "_n") for x in U)
        le = sum(_n(x, key + "_leased") for x in U)
        vc = sum(_n(x, key + "_vacant") for x in U)
        vals = [i + 1, ar] + per + [n, le, vc, le / n if n else 0]
        fmts = [None, None, "#,##0", "#,##0", "#,##0", "#,##0", "#,##0", "#,##0", "0%"]
        for j, (v, fm) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(rr, j, v); c.border = W.BOX; c.alignment = W.CTR
            c.font = W.BOLD if j == 2 else W.BLACK
            if fm: c.number_format = fm
        ws.cell(rr, NC, TNOTE.get(ar, "")).alignment = W.WRAP
        ws.cell(rr, NC).border = W.BOX
        ws.row_dimensions[rr].height = 21
    tend = t0 + len(TYPES) - 1
    ws.conditional_formatting.add(f"I{t0}:I{tend}", CellIsRule(
        operator="lessThan", formula=["0.3"], fill=W.BAD,
        font=Font(name=W.F, size=10, bold=True, color=B.D_BAD)))
    ws.conditional_formatting.add(f"H{t0}:H{tend}", DataBarRule(
        start_type="num", start_value=0, end_type="max", color=B.ORANGE))
    r = tend + 2

    # ── ③ أكبر الفرص
    r = W.band(ws, r, NC, "③ أكبر الفرص — أعلى المحطات بالوحدات الشاغرة")
    W.header(ws, r, ["#", "المحطة", "الكود", "المدينة", "الفئة", "التعاقد",
                     "الوحدات", "المؤجر", "الشاغر", "الإشغال", "ملاحظة"])
    top = sorted(U, key=lambda x: -_n(x, "total_vacant"))[:12]
    o0 = r + 1
    for i, x in enumerate(top):
        rr = o0 + i
        n = _n(x, "total_n"); le = _n(x, "total_leased"); vc = _n(x, "total_vacant")
        vals = [i + 1, x["name"][:28], x["code"], x["city"], x["category"],
                x["contract"][:16], n, le, vc, le / n if n else 0]
        fmts = [None] * 6 + ["#,##0", "#,##0", "#,##0", "0%"]
        for j, (v, fm) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(rr, j, v); c.border = W.BOX; c.alignment = W.CTR
            c.font = W.BOLD if j == 2 else W.BLACK
            if fm: c.number_format = fm
        note = "شواغر محل بالأساس" if _n(x, "shop_vacant") > vc / 2 else ""
        ws.cell(rr, NC, note).alignment = W.WRAP; ws.cell(rr, NC).border = W.BOX
        ws.row_dimensions[rr].height = 20
    oend = o0 + len(top) - 1
    ws.conditional_formatting.add(f"I{o0}:I{oend}", DataBarRule(
        start_type="num", start_value=0, end_type="max", color=B.ORANGE))
    r = oend + 2

    shops_n = sum(_n(x, "shop_n") for x in U)
    shops_l = sum(_n(x, "shop_leased") for x in U)
    W.bullets(ws, r, NC, [
        ("ما يصححه هذا السجل", [
            "الشبكة العاملة إشغالها ٧٩٪ لا ٢٥٪ — لا توجد أزمة شغور في المحطات المفتوحة",
            "الشغور الحقيقي في مكانين: تحت التنفيذ (٦٨٩ وحدة) والامتياز (٢٦٩ وحدة)",
            f"المحل هو المنتج العالق: {shops_n:,} وحدة و{shops_l:,} مؤجرة فقط "
            f"({shops_l/shops_n*100:.0f}٪) — وكل نوع آخر فوق ٤٥٪",
            "محلات الامتياز ١٤٦ وحدة و٥ مؤجرة (٣٪) — أسوأ رقم في السجل كله",
        ], W.NAVY),
        ("ما يعنيه للخطة", [
            "التأجير المسبق قبل الافتتاح هو الفرصة الأكبر — ٦٨٩ وحدة في ٤٣ محطة",
            "٥٧٢ محلاً شاغراً هي بالضبط مخزون خطة الإكسسوارات (مساحة مقابل نسبة)",
            "شبكة الوسطاء تُوجَّه للامتياز وتحت التنفيذ لا للمحطات العاملة",
            "معيار الافتتاح (مغسلة وسوبرماركت) مسنود: إشغالهما ٧٦٪ و٨٥٪ حيث تعمل المحطة",
        ], W.NAVY),
        ("حدود السجل", [
            "لا يحمل قيمة إيجارية لكل وحدة — فلا يمكن تسعير الفرصة بالريال منه وحده",
            "«المحجوز» غير معرَّف: هل هو تفاوض قائم أم عقد لم يُوقَّع؟",
            "٩١ محطة امتياز هنا مقابل ٩٢ في تقرير يوليو — فرق محطة واحدة غير مُفسَّر",
        ], "C00000"),
    ])
    return ws
