# -*- coding: utf-8 -*-
"""بناة أوراق نموذج العامل — تُستخدم في الملف المستقل وفي الخطة التجارية معاً"""
import csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule, ColorScaleRule

import brand as B

# الأسماء القديمة محفوظة كواجهة — القيم صارت من هوية درب
F = B.FONT; NAVY = B.BGRAY; GRAY = B.INK2
ORANGE = B.ORANGE
FH = PatternFill("solid", fgColor=B.BGRAY); FS = PatternFill("solid", fgColor=B.T_BAND)
FI = PatternFill("solid", fgColor=B.T_GOLD); SECT = PatternFill("solid", fgColor=B.T_ORANGE)
OK = PatternFill("solid", fgColor=B.T_GOOD); WARN = PatternFill("solid", fgColor=B.T_GOLD)
BAD = PatternFill("solid", fgColor=B.T_BAD); CALC = PatternFill("solid", fgColor=B.T_NEUTRAL)
BLUE = Font(name=F, size=10, color=B.BLUE); BLACK = Font(name=F, size=10, color=B.INK)
GREEN = Font(name=F, size=10, color=B.GOOD)
HEAD = Font(name=F, size=10, bold=True, color="FFFFFF")
TITLE = Font(name=F, size=16, bold=True, color=B.BGRAY)
SUB = Font(name=F, size=10, color=B.INK2)
BOLD = Font(name=F, size=10, bold=True, color=B.INK)
H2 = Font(name=F, size=12, bold=True, color=B.ORANGE)
SMALL = Font(name=F, size=9, color=B.INK3)
thin = Side(style="thin", color=B.LINE2); BOX = Border(thin, thin, thin, thin)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="right", readingOrder=2)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
RGT = Alignment(horizontal="right", vertical="center", readingOrder=2)

DATA_SHEET = "بيانات الورديات"


def stations(path="outlets/data/station-shifts.csv"):
    return list(csv.DictReader(open(path, encoding="utf-8")))


def setup(ws, widths, freeze="A4"):
    ws.sheet_view.rightToLeft = True; ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    # الطباعة: عرضية بعرض صفحة واحدة — الأوراق أعرض من ورقة عمودية
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4


RULE = PatternFill("solid", fgColor=B.ORANGE)


def title(ws, t, sub, nc):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
    c = ws.cell(1, 1, t); c.font = TITLE; c.alignment = RGT; ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
    c = ws.cell(2, 1, sub); c.font = SUB; c.alignment = RGT; ws.row_dimensions[2].height = 16


def rule(ws, row, nc, h=3.5):
    """شريط الهوية البرتقالي"""
    for j in range(1, nc + 1):
        ws.cell(row, j).fill = RULE
    ws.row_dimensions[row].height = h


def header(ws, row, cols, h=30):
    for j, t in enumerate(cols, 1):
        c = ws.cell(row, j, t); c.font = HEAD; c.fill = FH; c.alignment = CTR; c.border = BOX
    ws.row_dimensions[row].height = h


def band(ws, row, nc, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row, 1, text); c.font = H2; c.fill = SECT; c.alignment = RGT; c.border = BOX
    ws.row_dimensions[row].height = 22
    return row + 1


def inp(ws, r, c, fmt="#,##0", val=None):
    cc = ws.cell(r, c, val); cc.font = BLUE; cc.fill = FI; cc.border = BOX
    cc.alignment = CTR; cc.number_format = fmt
    return cc


def calc(ws, r, c, f_, fmt="#,##0", link=False):
    cc = ws.cell(r, c, f_); cc.font = GREEN if link else BLACK; cc.fill = CALC
    cc.border = BOX; cc.alignment = CTR; cc.number_format = fmt
    return cc


def bullets(ws, row, nc, blocks):
    r = row
    for t, pts, col in blocks:
        c = ws.cell(r, 2, t); c.font = Font(name=F, size=10, bold=True, color=col); c.alignment = RGT
        r += 1
        for p in pts:
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=nc)
            c = ws.cell(r, 3, "•  " + p); c.font = Font(name=F, size=10); c.alignment = RGT
            ws.row_dimensions[r].height = 16; r += 1
        r += 1
    return r


# ══════════════════════════════════════════════════════════════════
def build_data(wb, ST, idx=None):
    """ورقة البيانات المرجعية — تُغذّي القوائم المنسدلة والتعبئة التلقائية"""
    dw = wb.create_sheet(DATA_SHEET) if idx is None else wb.create_sheet(DATA_SHEET, idx)
    cols = ["#", "المحطة", "لتر/زيارة", "زيارات/يوم", "عمال صباحية", "عمال مسائية",
            "حصة الصباح", "حصة المساء", "ساعات الذروة", "الطاقم المعياري", "الفعلي",
            "الكود", "اسم المقياس", "العقد", "المضخات"]
    setup(dw, [4, 17, 10, 11, 12, 12, 11, 11, 20, 12, 9, 9, 15, 9, 9], freeze="A4")
    title(dw, "بيانات المحطات المرجعية",
          "من تقرير المطابقة التشغيلية — يناير إلى يوليو ٢٠٢٦ (٢١٢ يوماً فعلياً)", len(cols))
    header(dw, 3, cols)
    for i, s in enumerate(ST):
        rr = 4 + i
        vals = [i + 1, s["name"], float(s["lpv"]), int(s["visits_day"]), int(s["day_w"]),
                int(s["eve_w"]), float(s["txn_day_share"]) / 100,
                float(s["txn_eve_share"]) / 100, s["peak_hours"], int(s["std"]),
                int(s["act"]), s["code"], s["fname"], s["contract"], int(s["pumps"])]
        fmts = [None, None, "0.00", "#,##0", "0", "0", "0.0%", "0.0%",
                None, "0", "0", None, None, None, "0"]
        for j, (v, fm) in enumerate(zip(vals, fmts), 1):
            c = dw.cell(rr, j, v); c.border = BOX; c.alignment = CTR
            c.font = BOLD if j == 2 else BLACK
            if fm: c.number_format = fm
        dw.row_dimensions[rr].height = 20
    end = 3 + len(ST)
    for k, txt in enumerate([
        "«حصة الصباح/المساء» من عدد الزيارات الفعلي — الوردية الصباحية ٦ ص إلى ٦ م والمسائية ٦ م إلى ٦ ص",
        "«اسم المقياس» هو الاسم المستخدم في ملف العمالة — يختلف عن اسم السجل في محطتين",
    ]):
        dw.cell(end + 2 + k, 2, txt).font = SMALL
        dw.cell(end + 2 + k, 2).alignment = RGT
    dw.cell(3, 17, "الورديات").font = BOLD
    for i, s in enumerate(["صباحية", "مسائية"]):
        dw.cell(4 + i, 17, s).alignment = CTR
    dw.cell(3, 18, "بوابة الجودة").font = BOLD
    for i, s in enumerate(["مستوفاة", "غير مستوفاة"]):
        dw.cell(4 + i, 18, s).alignment = CTR
    dw.column_dimensions["Q"].width = 12; dw.column_dimensions["R"].width = 14
    return end


# ══════════════════════════════════════════════════════════════════
def build_worker(wb, ST, DEND, name="نموذج العامل", idx=None):
    NC = 12
    ws = wb.create_sheet(name) if idx is None else wb.create_sheet(name, idx)
    setup(ws, [4, 17, 11, 14, 11, 12, 11, 10, 11, 13, 11, 32], freeze="C18")
    title(ws, "نموذج مبيعات العامل — مستهدف وحافز",
          "خط الأساس يُشتق من حصة الوردية الفعلية من الزيارات ÷ عدد عمالها — لا يُقدَّر يدوياً", NC)

    r = band(ws, 3, NC, "① مُدخلات المحطة — اختر المحطة والباقي يُعبَّأ من بيانات التشغيل")
    W0 = r
    (R_STA, R_MON, R_LPV, R_VIS, R_DW, R_EW, R_DSH, R_ESH,
     R_MRG, R_PCT, R_CAP, R_DAY, R_GRW, R_NETG) = (W0 + i for i in range(14))
    D = DATA_SHEET

    def look(col):
        return (f"=IFERROR(INDEX('{D}'!${col}$4:${col}${DEND},"
                f"MATCH($C${R_STA},'{D}'!$B$4:$B${DEND},0)),0)")

    WIN = [
        ("المحطة",                    "General", "list", None,  "اختر من القائمة ▼"),
        ("الشهر",                     "General", "in",   None,  "مثال: أغسطس ٢٠٢٦"),
        ("لتر لكل زيارة",             '0.00;-0.00;""', "C", None, "من بيانات المحطة"),
        ("زيارات المحطة (زيارة/يوم)", '#,##0;;""',     "D", None, "متوسط ٢١٢ يوماً"),
        ("عمال المضخات — صباحية",     '0;;""',         "E", None, "٦ ص – ٦ م"),
        ("عمال المضخات — مسائية",     '0;;""',         "F", None, "٦ م – ٦ ص"),
        ("حصة الصباح من الزيارات",    '0.0%;;""',      "G", None, "من الحركة الفعلية"),
        ("حصة المساء من الزيارات",    '0.0%;;""',      "H", None, "من الحركة الفعلية"),
        ("هامش المساهمة (هللة/لتر)",  "0.00",  "in", 13.36, "من قائمة الدخل"),
        ("نسبة الحافز من الهامش",     "0.0%",  "in", 0.15,  "قرار الإدارة"),
        ("سقف الحافز الشهري للعامل",  "#,##0", "in", 100,   "معتمد: ١٠٠ ريال"),
        ("أيام الشهر",                "0",     "in", 30,    ""),
        ("نسبة النمو المستهدف",       "0.0%",  "in", 0.05,  "حرّكها ← تتحرك كل المستهدفات"),
        ("يُشترط صافي محطة موجب",     "General", "in", "نعم", "يحمي نسبة الـ١٥٪"),
    ]
    for i, (lbl, fmt, kind, d, note) in enumerate(WIN):
        rr = W0 + i
        ws.cell(rr, 2, lbl).font = BOLD; ws.cell(rr, 2).alignment = RGT
        if kind in ("in", "list"):
            inp(ws, rr, 3, fmt, d)
        else:
            calc(ws, rr, 3, look(kind), fmt, link=True)
        ws.cell(rr, 5, note).font = SMALL; ws.cell(rr, 5).alignment = RGT

    ws.cell(R_STA, 4, f"=IFERROR(\"ذروة \"&INDEX('{D}'!$I$4:$I${DEND},"
                      f"MATCH($C${R_STA},'{D}'!$B$4:$B${DEND},0)),\"\")")
    ws.cell(R_STA, 4).font = SMALL; ws.cell(R_STA, 4).alignment = CTR

    dv_st = DataValidation(type="list", formula1=f"='{D}'!$B$4:$B${DEND}", allow_blank=True)
    dv_st.error = "اختر محطة من القائمة"; dv_st.errorTitle = "قيمة غير معتمدة"
    ws.add_data_validation(dv_st); dv_st.add(ws.cell(R_STA, 3))
    dv_yn = DataValidation(type="list", formula1='"نعم,لا"', allow_blank=False)
    ws.add_data_validation(dv_yn); dv_yn.add(ws.cell(R_NETG, 3))

    r = W0 + len(WIN) + 1
    r = band(ws, r, NC, "② بطاقة العامل — اختر الوردية فيُشتق خط الأساس · عبّئ الفعلي وبوابة الجودة")
    HR = r
    header(ws, HR, ["#", "اسم العامل", "الوردية", "خط الأساس المشتق", "تعديل معتمد",
                    "المستهدف", "الفعلي", "الفارق", "نسبة الإنجاز", "بوابة الجودة",
                    "الحافز (ريال)", "ملاحظة"])
    ROWS = 20
    L = HR + ROWS
    dv_sh = DataValidation(type="list", formula1=f"='{D}'!$Q$4:$Q$5", allow_blank=True)
    dv_gt = DataValidation(type="list", formula1=f"='{D}'!$R$4:$R$5", allow_blank=True)
    dv_gt.error = "اكتب: مستوفاة أو غير مستوفاة"; dv_gt.errorTitle = "قيمة غير معتمدة"
    ws.add_data_validation(dv_sh); ws.add_data_validation(dv_gt)

    for i in range(ROWS):
        rr = HR + 1 + i
        ws.cell(rr, 1, i + 1).alignment = CTR
        inp(ws, rr, 2, "General"); inp(ws, rr, 3, "General")
        calc(ws, rr, 4,
             f'=IF(OR($C{rr}="",$C${R_VIS}=0),"",ROUND(IF($C{rr}="صباحية",'
             f'$C${R_VIS}*$C${R_DSH}/MAX($C${R_DW},1),'
             f'$C${R_VIS}*$C${R_ESH}/MAX($C${R_EW},1)),0))', '#,##0;;""')
        inp(ws, rr, 5, '#,##0;;""')
        calc(ws, rr, 6, f'=IF($D{rr}="","",ROUND(IF($E{rr}="",$D{rr},$E{rr})*(1+$C${R_GRW}),0))')
        inp(ws, rr, 7, "#,##0")
        calc(ws, rr, 8, f'=IF(OR($G{rr}="",$F{rr}=""),"",$G{rr}-$F{rr})', "+#,##0;-#,##0")
        calc(ws, rr, 9, f'=IFERROR(IF($F{rr}="","",$G{rr}/$F{rr}),"")', "0%")
        c = ws.cell(rr, 10); c.font = BLUE; c.fill = FI; c.border = BOX; c.alignment = CTR
        calc(ws, rr, 11,
             f'=IF(OR($H{rr}="",$J{rr}<>"مستوفاة"),0,'
             f'IF(AND($C${R_NETG}="نعم",$H${L+1}<=0),0,'
             f'MIN(MAX($H{rr},0)*$C${R_LPV}*$C${R_MRG}/100*$C${R_PCT}*$C${R_DAY},$C${R_CAP})))')
        ws.cell(rr, 11).fill = FS
        inp(ws, rr, 12, "General")
        for j in range(1, NC + 1):
            ws.cell(rr, j).border = BOX
            if j != 12: ws.cell(rr, j).alignment = CTR
        ws.row_dimensions[rr].height = 19
        dv_sh.add(ws.cell(rr, 3)); dv_gt.add(ws.cell(rr, 10))

    ws.cell(L + 1, 2, "الإجمالي").font = BOLD
    for col in (4, 6, 7, 8, 11):
        ws.cell(L + 1, col, f"=SUM({get_column_letter(col)}{HR+1}:"
                            f"{get_column_letter(col)}{L})").number_format = "#,##0"
    ws.cell(L + 1, 9, f'=IFERROR($G{L+1}/$F{L+1},"")').number_format = "0%"
    ws.cell(L + 1, 3, f'=COUNTA($B{HR+1}:$B{L})&" عامل"')
    for j in range(1, NC + 1):
        ws.cell(L + 1, j).border = BOX; ws.cell(L + 1, j).fill = FS
        ws.cell(L + 1, j).font = BOLD; ws.cell(L + 1, j).alignment = CTR

    cf = ws.conditional_formatting
    cf.add(f"H{HR+1}:H{L}", CellIsRule(operator="greaterThan", formula=["0"], fill=OK,
                                       font=Font(name=F, size=10, color=B.D_GOOD)))
    cf.add(f"H{HR+1}:H{L}", CellIsRule(operator="lessThan", formula=["0"], fill=BAD,
                                       font=Font(name=F, size=10, color=B.D_BAD)))
    cf.add(f"I{HR+1}:I{L}", DataBarRule(start_type="num", start_value=0,
                                        end_type="num", end_value=1.3, color=B.ORANGE))
    cf.add(f"K{HR+1}:K{L}", CellIsRule(operator="greaterThan", formula=["0"], fill=OK,
                                       font=Font(name=F, size=10, bold=True, color=B.D_GOOD)))
    cf.add(f"J{HR+1}:J{L}", FormulaRule(formula=[f'$J{HR+1}="مستوفاة"'], fill=OK))
    cf.add(f"J{HR+1}:J{L}", FormulaRule(formula=[f'$J{HR+1}="غير مستوفاة"'], fill=BAD,
                                        font=Font(name=F, size=10, bold=True, color=B.D_BAD)))
    cf.add(f"B{HR+1}:B{L}", FormulaRule(
        formula=[f'AND($H{HR+1}>0,$J{HR+1}="غير مستوفاة")'], fill=WARN))
    cf.add(f"K{HR+1}:K{L}", FormulaRule(
        formula=[f'AND($K{HR+1}>0,$K{HR+1}=$C${R_CAP})'], fill=WARN,
        font=Font(name=F, size=10, bold=True, color=B.D_GOLD)))
    cf.add(f"E{HR+1}:E{L}", FormulaRule(
        formula=[f'AND($E{HR+1}<>"",$E{HR+1}<$D{HR+1})'], fill=BAD,
        font=Font(name=F, size=10, bold=True, color=B.D_BAD)))

    r = L + 3
    r = band(ws, r, NC, "③ ملخص المحطة — يتحدّث مع كل رقم تُدخله")
    SM = r
    header(ws, SM, ["#", "البند", "القيمة"] + [""] * 8 + ["الدلالة"])
    SUMR = [
        ("الزيادة الإجمالية (زيارة/يوم)", f"=$H${L+1}", "#,##0", "مجموع فروق العمال"),
        ("اللترات الإضافية شهرياً", f"=$C${SM+1}*$C${R_LPV}*$C${R_DAY}", "#,##0",
         "الزيادة × لتر/زيارة × الأيام"),
        ("الهامش الإضافي شهرياً (ريال)", f"=$C${SM+2}*$C${R_MRG}/100", "#,##0", "قبل خصم الحافز"),
        ("إجمالي الحافز شهرياً (ريال)", f"=$K${L+1}", "#,##0", "ما يُصرف للعمال"),
        ("صافي المكسب للشركة (ريال)", f"=$C${SM+3}-$C${SM+4}", "+#,##0;-#,##0",
         "الهامش ناقص الحافز"),
        ("نسبة الحافز من الهامش الفعلي", f'=IFERROR($C${SM+4}/$C${SM+3},"")', "0.0%",
         "تتجاوز المعتمدة إذا ربح بعض العمال وخسر غيرهم — الحافز فردي والهامش صافٍ"),
        ("عدد من بلغوا السقف", f'=COUNTIFS($K${HR+1}:$K${L},"="&$C${R_CAP})', "#,##0",
         "إن ارتفع العدد فالسقف يكبح التحفيز"),
        ("تعديلات يدوية على خط الأساس", f'=COUNTA($E${HR+1}:$E${L})', "#,##0",
         "كل تعديل يحتاج مبرراً مكتوباً"),
    ]
    for i, (lbl, f_, fmt, note) in enumerate(SUMR):
        rr = SM + 1 + i
        ws.cell(rr, 1, i + 1); ws.cell(rr, 2, lbl).font = BOLD
        ws.cell(rr, 3, f_).number_format = fmt
        ws.cell(rr, 12, note).alignment = WRAP
        for j in range(1, NC + 1):
            ws.cell(rr, j).border = BOX
            if j != 2: ws.cell(rr, j).font = BLACK
            if j != 12: ws.cell(rr, j).alignment = CTR
        ws.row_dimensions[rr].height = 20
    NET = SM + 5
    cf.add(f"C{NET}", CellIsRule(operator="greaterThan", formula=["0"], fill=OK,
                                 font=Font(name=F, size=10, bold=True, color=B.D_GOOD)))
    cf.add(f"C{NET}", CellIsRule(operator="lessThan", formula=["0"], fill=BAD,
                                 font=Font(name=F, size=10, bold=True, color=B.D_BAD)))
    cf.add(f"C{SM+8}", CellIsRule(operator="greaterThan", formula=["0"], fill=WARN,
                                  font=Font(name=F, size=10, bold=True, color=B.D_GOLD)))
    cf.add(f"C{SM+6}", FormulaRule(formula=[f"$C${SM+6}>$C${R_PCT}"], fill=BAD,
                                   font=Font(name=F, size=10, bold=True, color=B.D_BAD)))
    r2 = SM + len(SUMR) + 1
    ws.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=NC)
    ws.cell(r2, 2, "إذا تجاوزت النسبة الفعلية المعتمدة فالسبب أن الحافز يُصرف على مكسب كل عامل "
                   "بينما الشركة تقبض الصافي فقط — شرط «صافي محطة موجب» أعلاه يمنع الصرف "
                   "في شهر خاسر.").font = Font(name=F, size=9, color=B.D_BAD)
    ws.cell(r2, 2).alignment = RGT
    r = SM + len(SUMR) + 2

    r = band(ws, r, NC, "④ ماذا لو — حرّك نسبة النمو وقارن")
    SC = r
    header(ws, SC, ["#", "السيناريو", "نسبة النمو", "زيارة/يوم إضافية", "لترات شهرياً",
                    "الهامش الإضافي", "الحافز المتوقع", "صافي المكسب"] + [""] * 3 + ["الدلالة"])
    NWK = f'COUNTA($B${HR+1}:$B${L})'
    for i, (nm, g, note) in enumerate([
            ("متحفظ", 0.03, "زيادة يومية بسيطة تُدرَك بالانضباط وحده"),
            ("أساسي", 0.05, "النسبة المعتمدة في النموذج أعلاه"),
            ("طموح", 0.10, "يتطلب حملة ميدانية مصاحبة لا انضباطاً فقط")]):
        rr = SC + 1 + i
        ws.cell(rr, 1, i + 1); ws.cell(rr, 2, nm).font = BOLD
        inp(ws, rr, 3, "0.0%", g)
        ws.cell(rr, 4, f"=ROUND($D${L+1}*$C{rr},0)").number_format = "#,##0"
        ws.cell(rr, 5, f"=$D{rr}*$C${R_LPV}*$C${R_DAY}").number_format = "#,##0"
        ws.cell(rr, 6, f"=$E{rr}*$C${R_MRG}/100").number_format = "#,##0"
        ws.cell(rr, 7, f"=MIN($F{rr}*$C${R_PCT},{NWK}*$C${R_CAP})").number_format = "#,##0"
        ws.cell(rr, 8, f"=$F{rr}-$G{rr}").number_format = "+#,##0;-#,##0"
        ws.cell(rr, 12, note).alignment = WRAP
        for j in range(1, NC + 1):
            ws.cell(rr, j).border = BOX
            if j not in (2, 3): ws.cell(rr, j).font = BLACK
            if j != 12: ws.cell(rr, j).alignment = CTR
        ws.row_dimensions[rr].height = 20
    cf.add(f"H{SC+1}:H{SC+3}", ColorScaleRule(start_type="min", start_color="FFFFFF",
                                              end_type="max", end_color="C6E0B4"))
    cf.add(f"G{SC+1}:G{SC+3}", FormulaRule(
        formula=[f"$G{SC+1}>=({NWK})*$C${R_CAP}*0.99"], fill=WARN,
        font=Font(name=F, size=10, bold=True, color=B.D_GOLD)))
    ws.cell(SC + 4, 2, "الخانة البرتقالية تعني أن السقف بلغ حدّه — "
                       "الزيادة بعده تذهب كاملة للشركة").font = SMALL
    ws.cell(SC + 4, 2).alignment = RGT
    r = SC + 6

    r = band(ws, r, NC, "⑤ بوابة الجودة — تُقيَّم شهرياً لكل عامل")
    header(ws, r, ["#", "الشرط", "المقياس", "الحد"] + [""] * 7 + ["لماذا"])
    for i, (cond, msr, lim, why) in enumerate([
            ("لا شكوى مثبتة", "شكاوى العامل في الشهر", "صفر",
             "الحافز مكافأة أداء لا مكافأة حجم"),
            ("زمن التعبئة", "متوسط الزمن في الذروة", "≤ الحد المعتمد",
             "سرعة الخدمة هي ما يطلبه العميل"),
            ("الحضور في ساعات الذروة", "نسبة الحضور في ساعات المحطة", "100%",
             "ساعات الذروة معروفة لكل محطة — انظر خانة المحطة أعلاه"),
            ("النظافة والمظهر", "تقييم المشرف الأسبوعي", "مقبول فأعلى",
             "أول ما يراه العميل")]):
        rr = r + 1 + i
        ws.cell(rr, 1, i + 1); ws.cell(rr, 2, cond).font = BOLD
        ws.cell(rr, 3, msr); ws.cell(rr, 4, lim); ws.cell(rr, 12, why).alignment = WRAP
        for j in range(1, NC + 1):
            ws.cell(rr, j).border = BOX
            if j != 2: ws.cell(rr, j).font = BLACK
            if j != 12: ws.cell(rr, j).alignment = CTR
        ws.row_dimensions[rr].height = 21
    r += 6

    bullets(ws, r, NC, [
        ("كيف تستخدم الملف", [
            "اختر المحطة ← تُعبَّأ الزيارات وحصص الورديات وعدد العمال من بيانات التشغيل",
            "اختر وردية العامل ← يُشتق خط أساسه تلقائياً من حصة ورديته ÷ عدد عمالها",
            "لا تعبّئ إلا: الاسم · الوردية · الفعلي · بوابة الجودة",
            "الأصفر مُدخل · الرمادي محسوب · الأخضر مسحوب من ورقة البيانات",
        ], NAVY),
        ("لماذا يُشتق خط الأساس ولا يُقدَّر", [
            "خط الأساس المتساهل هو الطريقة الأولى لتفريغ الحافز من معناه",
            "عامل المساء يخدم زيارات أكثر من عامل الصباح بنفس العدد — الاشتقاق ينصفه",
            "خانة «تعديل معتمد» تُظهر أي تخفيض بالأحمر ويُحصى في الملخص",
        ], NAVY),
        ("قواعد الاحتساب", [
            "الحافز على الزيادة فوق المستهدف فقط — لا على الحجم الكلي",
            "لا يُصرف إطلاقاً إذا كانت بوابة الجودة غير مستوفاة",
            "سقف ١٠٠ ريال شهرياً لكل عامل يمنع المبالغة في شهر استثنائي",
            "الفارق السالب لا يُخصم — لكنه لا يُكافأ",
        ], NAVY),
        ("ما يُبطل النموذج", [
            "تعديل خط الأساس بالخفض دون مبرر مكتوب",
            "غياب قياس زمن الخدمة يفرّغ بوابة الجودة من معناها",
            "الصرف في شهر صافيه سالب — الحافز فردي والشركة تقبض الصافي",
        ], "C00000"),
    ])
    return ws


# ══════════════════════════════════════════════════════════════════
def build_shift(wb, ST, name="مطابقة الوردية مع الطلب", idx=None):
    MC = 12
    mw = wb.create_sheet(name) if idx is None else wb.create_sheet(name, idx)
    setup(mw, [4, 17, 11, 11, 10, 10, 11, 11, 11, 11, 10, 34], freeze="A4")
    title(mw, "مطابقة الوردية مع الطلب",
          "كل المحطات توزّع العمالة نصفين متساويين — بينما الطلب ليس نصفين", MC)
    header(mw, 3, ["#", "المحطة", "زيارات الصباح", "زيارات المساء", "عمال صباحاً",
                   "عمال مساءً", "حصة العمالة مساءً", "الفجوة", "حِمل عامل الصباح",
                   "حِمل عامل المساء", "الفارق", "القرار"])
    for i, s in enumerate(ST):
        rr = 4 + i
        vis = int(s["visits_day"]); dsh = float(s["txn_day_share"]) / 100
        esh = float(s["txn_eve_share"]) / 100
        dwk = int(s["day_w"]); ewk = int(s["eve_w"]); tot = dwk + ewk
        bd = vis * dsh / dwk; be = vis * esh / ewk
        staff_e = ewk / tot; move = tot * esh - ewk
        dec = (f"انقل {round(move)} عامل من الصباح إلى المساء" if abs(move) >= 0.5
               else "الفجوة أقل من عامل — عالجها بمواعيد الدخول لا بالعدد")
        vals = [i + 1, s["name"], round(vis * dsh), round(vis * esh), dwk, ewk,
                staff_e, esh - staff_e, round(bd), round(be), be / bd - 1, dec]
        fmts = [None, None, "#,##0", "#,##0", "0", "0", "0.0%", "+0.0%;-0.0%",
                "#,##0", "#,##0", "+0.0%;-0.0%", None]
        for j, (v, fm) in enumerate(zip(vals, fmts), 1):
            c = mw.cell(rr, j, v); c.border = BOX
            c.font = BOLD if j == 2 else BLACK
            c.alignment = WRAP if j == 12 else CTR
            if fm: c.number_format = fm
        mw.row_dimensions[rr].height = 22
    end = 3 + len(ST)
    mw.conditional_formatting.add(f"H4:H{end}", CellIsRule(
        operator="greaterThan", formula=["0.02"], fill=BAD,
        font=Font(name=F, size=10, bold=True, color=B.D_BAD)))
    mw.conditional_formatting.add(f"K4:K{end}", ColorScaleRule(
        start_type="min", start_color="FFFFFF", end_type="max", end_color=B.T_BAD))

    bullets(mw, end + 2, MC, [
        ("ما تقوله الأرقام", [
            "عامل المساء يخدم ٢٢–٢٨٪ زيارات أكثر من عامل الصباح في أربع محطات من خمس",
            "السبب أن العمالة موزّعة ٥٠/٥٠ بينما الطلب المسائي ٥٥–٥٦٪ من الزيارات",
            "المعيصم وحدها متوازنة — لأن ذروتها ٥ إلى ٧ مساءً لا بعد التاسعة",
        ], NAVY),
        ("لماذا يهم هذا في نموذج الحافز", [
            "خط أساس موحّد لكل العمال يظلم عامل المساء ويكافئ عامل الصباح على أقل جهد",
            "لذلك يُشتق خط الأساس من حصة الوردية الفعلية — لا من متوسط المحطة",
            "الاشتقاق يجعل ٥٪ نمو تعني الجهد نفسه على الورديتين",
        ], NAVY),
        ("حدود ما يمكن تنفيذه", [
            "النورية وحدها فيها عمالة تكفي لنقل عامل كامل (٢٠ عاملاً على المضخات)",
            "بقية المحطات فجوتها أقل من عامل — الحل تعديل مواعيد الدخول لا زيادة العدد",
            "تكديس الورديات بين ٨ و١٠ مساءً يغطي الذروة دون تكلفة إضافية",
        ], "C00000"),
    ])
    return mw
