# -*- coding: utf-8 -*-
"""ملف إكسل: الاستراتيجية التجارية لمنفذ العمرة الجديدة (MK007) — نموذج حيّ"""
import json, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

D = json.load(io.open('data.json', encoding='utf-8'))
S = D['stations'][0]                    # MK007
MONTHS = D['platform']['months']; NM = len(MONTHS)
FIN = S['finance']
OUT = 'استراتيجية-العمرة-الجديدة-MK007.xlsx'

F='Arial'; DARK='1B1C1E'; ORANGE='F18A2B'; ORANGE_L='FDEBD8'
GREEN='2E8B57'; GREEN_L='E3F4EA'; RED='C0392B'; RED_L='FBE6E4'
GOLD='B8860B'; GOLD_L='FFF6DC'; SKY='2874A6'; SKY_L='E4F0F8'
GREY='6B7280'; BAND='F7F8F9'; INK='1F2937'
fill=lambda c: PatternFill('solid', fgColor=c)
BLUE_F=Font(name=F,size=10,color='0000C0'); BLACK_F=Font(name=F,size=10,color=INK)
BOLD=lambda s=10,c=INK: Font(name=F,size=s,bold=True,color=c)
hair=Side(style='hair',color='C8CCD0'); BOX=Border(left=hair,right=hair,top=hair,bottom=hair)
MONEY='#,##0;(#,##0);"–"'; NUM='#,##0;(#,##0);"–"'; DEC2='#,##0.00;(#,##0.00);"–"'; PCT='0.0%'; PCT0='0%'

def sh(wb,name,cover=False):
    ws=wb.create_sheet(name)
    ws.sheet_view.rightToLeft=True; ws.sheet_view.showGridLines=False
    ws.sheet_properties.tabColor= DARK if cover else ORANGE
    ws.page_setup.orientation='landscape'; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr.fitToPage=True; ws.print_options.horizontalCentered=True
    return ws

def banner(ws,t,sub,n):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    c=ws.cell(1,1,t); c.font=Font(name=F,size=16,bold=True,color='FFFFFF'); c.fill=fill(DARK)
    c.alignment=Alignment(horizontal='right',vertical='center',indent=1); ws.row_dimensions[1].height=34
    for j in range(2,n+1): ws.cell(1,j).fill=fill(DARK)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=n)
    c=ws.cell(2,1,sub); c.font=Font(name=F,size=9,italic=True,color=GREY)
    c.alignment=Alignment(horizontal='right',vertical='center',indent=1); ws.row_dimensions[2].height=18
    return 4

def thead(ws,row,hs,ws_):
    for i,h in enumerate(hs,1):
        c=ws.cell(row,i,h); c.font=Font(name=F,size=9.5,bold=True,color='FFFFFF'); c.fill=fill(ORANGE)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=BOX
    ws.row_dimensions[row].height=30
    for i,w in enumerate(ws_,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes=ws.cell(row+1,1); return row+1

def put(ws,r,c,v,font=None,fmt=None,wrap=False,bg=None,al='right'):
    cell=ws.cell(r,c,v); cell.font=font or BLACK_F; cell.border=BOX
    cell.alignment=Alignment(horizontal=al,vertical='center',wrap_text=wrap,indent=1 if al=='right' else 0)
    if fmt: cell.number_format=fmt
    if bg: cell.fill=fill(bg)
    return cell

def band(ws,r0,r1,n):
    for r in range(r0,r1+1):
        if (r-r0)%2:
            for c in range(1,n+1):
                if ws.cell(r,c).fill.fgColor.rgb in (None,'00000000'): ws.cell(r,c).fill=fill(BAND)

def kpi(ws,row,col,span,val,lbl,color=ORANGE,fmt=None):
    ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
    c=ws.cell(row,col,val); c.font=Font(name=F,size=17,bold=True,color=color)
    c.alignment=Alignment(horizontal='center',vertical='center')
    if fmt: c.number_format=fmt
    ws.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+span-1)
    l=ws.cell(row+1,col,lbl); l.font=Font(name=F,size=8.5,color=GREY)
    l.alignment=Alignment(horizontal='center',vertical='top')
    top=Side(style='thick',color=color)
    for j in range(col,col+span):
        ws.cell(row,j).border=Border(top=top,left=hair,right=hair); ws.cell(row,j).fill=fill('FFFFFF')
        ws.cell(row+1,j).border=Border(left=hair,right=hair,bottom=hair); ws.cell(row+1,j).fill=fill('FFFFFF')
    ws.row_dimensions[row].height=28; ws.row_dimensions[row+1].height=16

def labels(ch,val=False,pct=False):
    d=DataLabelList(); d.showVal=val; d.showPercent=pct
    d.showSerName=False; d.showCatName=False; d.showLegendKey=False; d.showBubbleSize=False
    ch.dLbls=d; return ch

wb=Workbook(); wb.remove(wb.active)

# ═══════════ 1) نموذج محرّكات النمو (القلب — يُبنى أولاً لتُشير إليه بقية الأوراق) ═══════════
ws=sh(wb,'نموذج المحرّكات')
r=banner(ws,'نموذج محرّكات النمو — العمرة الجديدة','عدّل الخلايا الزرقاء (الافتراضات) ليعاد حساب الأثر بالكامل',5)
for i,w in enumerate([40,18,16,16,42],1): ws.column_dimensions[get_column_letter(i)].width=w

put(ws,r,1,'الافتراضات وخط الأساس',BOLD(11),bg=ORANGE_L)
for j in range(2,6): put(ws,r,j,'',bg=ORANGE_L)
ws.row_dimensions[r].height=20; r+=1
r=thead(ws,r,['البند','القيمة','الوحدة','','المصدر / الملاحظة'],[40,18,16,16,42])
A={}
def arow(lbl,val,unit,note,fmt=NUM,key=None):
    global r
    put(ws,r,1,lbl); put(ws,r,2,val,BLUE_F,fmt); put(ws,r,3,unit,Font(name=F,size=9,color=GREY),al='center')
    put(ws,r,4,''); put(ws,r,5,note,Font(name=F,size=9,color=GREY),wrap=True)
    if key: A[key]=r
    r+=1; return r-1
arow('حجم الوقود الحالي',FIN['liters'],'لتر/شهر','لوحة تحليلات درب — متوسط ٦ أشهر',NUM,'liters')
arow('حجم الوقود المستهدف',4500000,'لتر/شهر','مستهدف ١٢ شهراً — قابل للتعديل',NUM,'target')
arow('هامش الوقود',FIN['margin'],'ر.س/لتر','يُعتمد من المالية',DEC2,'margin')
arow('الزيارات',FIN['visits'],'زيارة/شهر','766,648 زيارة ÷ ٦ أشهر',NUM,'visits')
arow('الصافي الحالي',None,'ر.س/شهر','محسوب في «النموذج المالي» بالملف الشامل',MONEY,'net')
exp=FIN['salaries']+FIN['oprent']+FIN['utils']+FIN['maint']
nonfuel=FIN['units']*FIN['occ']/100*FIN['rent']+FIN['store']
ws.cell(A['net'],2).value=round(FIN['margin']*FIN['liters']+nonfuel-exp)
arow('إيجار مرساة القهوة المتوقّع',28000,'ر.س/شهر','قيمة الفرصة في خط الأنابيب',MONEY,'coffee')
r+=1

put(ws,r,1,'🚚 عقود الأساطيل — الحجم المؤمَّن',BOLD(11),bg=SKY_L)
for j in range(2,6): put(ws,r,j,'',bg=SKY_L)
ws.row_dimensions[r].height=20; r+=1
fh=thead(ws,r,['فئة الأسطول','الحجم لتر/شهر','الحالة','هامش ر.س/لتر','الجهات المستهدفة'],[40,18,16,16,42])
r=fh
dv=DataValidation(type='list',formula1='"فرصة,تواصل,عرض مقدّم,متعاقد"',allow_blank=True); ws.add_data_validation(dv)
for f_ in S['fleet']['items']:
    put(ws,r,1,f_['seg'],BOLD(9.5),wrap=True); put(ws,r,2,f_['liters'],BLUE_F,NUM)
    put(ws,r,3,f_['status'],BLUE_F,al='center'); dv.add(ws.cell(r,3))
    put(ws,r,4,f"=$B${A['margin']}",BLACK_F,DEC2)
    put(ws,r,5,f_['targets'],Font(name=F,size=9,color=GREY),wrap=True)
    ws.row_dimensions[r].height=38; r+=1
fl=r-1
band(ws,fh,fl,5)
put(ws,r,1,'إجمالي الحجم المؤمَّن',BOLD(),bg=SKY_L)
put(ws,r,2,f'=SUM(B{fh}:B{fl})',BOLD(11,SKY),NUM,bg=SKY_L)
for j in (3,4,5): put(ws,r,j,'',bg=SKY_L)
A['secured']=r; r+=1
put(ws,r,1,'منها متعاقد فعلياً',BOLD())
put(ws,r,2,f'=SUMIF(C{fh}:C{fl},"متعاقد",B{fh}:B{fl})',BLACK_F,NUM)
put(ws,r,5,'يرتفع كلما أُغلق عقد — غيّر الحالة أعلاه',Font(name=F,size=9,italic=True,color=GREY),wrap=True)
A['contracted']=r; r+=2

put(ws,r,1,'📊 احتساب الأثر',BOLD(11),bg=GREEN_L)
for j in range(2,6): put(ws,r,j,'',bg=GREEN_L)
ws.row_dimensions[r].height=20; r+=1
r=thead(ws,r,['البند','القيمة','الوحدة','','طريقة الاحتساب'],[40,18,16,16,42])
def crow(lbl,formula,unit,how,fmt=MONEY,key=None,bold=False,bg=None):
    global r
    fnt=BOLD(11) if bold else BLACK_F
    put(ws,r,1,lbl,fnt,bg=bg); put(ws,r,2,formula,fnt,fmt,bg=bg)
    put(ws,r,3,unit,Font(name=F,size=9,color=GREY),al='center',bg=bg); put(ws,r,4,'',bg=bg)
    put(ws,r,5,how,Font(name=F,size=9,color=GREY),wrap=True,bg=bg)
    if key: A[key]=r
    r+=1; return r-1
crow('فجوة الحجم',f"=$B${A['target']}-$B${A['liters']}",'لتر/شهر','المستهدف − الحالي',NUM,'gap')
crow('منها مؤمَّن بعقود',f"=$B${A['secured']}",'لتر/شهر','إجمالي الأساطيل',NUM,'sec2')
crow('نسبة التأمين من الفجوة',f"=IFERROR($B${A['secured']}/$B${A['gap']},\"\")",'٪','المؤمَّن ÷ الفجوة',PCT0,'seccov')
crow('يتبقّى استعادته من الحركة العابرة',f"=MAX(0,$B${A['gap']}-$B${A['secured']})",'لتر/شهر','الفجوة − المؤمَّن',NUM,'walkin')
r+=1
crow('أثر الوقود',f"=$B${A['gap']}*$B${A['margin']}",'ر.س/شهر','فجوة الحجم × الهامش',MONEY,'fuelimp')
crow('أثر مرساة القهوة',f"=$B${A['coffee']}",'ر.س/شهر','إيجار الوحدة الشاغرة',MONEY,'coffimp')
crow('إجمالي الأثر',f"=$B${A['fuelimp']}+$B${A['coffimp']}",'ر.س/شهر','الوقود + القهوة (بلا ازدواج: الأساطيل داخل فجوة الحجم)',MONEY,'total',bold=True,bg=GOLD_L)
r+=1
crow('الصافي المستهدف',f"=$B${A['net']}+$B${A['total']}",'ر.س/شهر','الصافي الحالي + الأثر',MONEY,'newnet',bold=True,bg=GREEN_L)
crow('نسبة الزيادة',f"=IFERROR($B${A['total']}/$B${A['net']},\"\")",'٪','الأثر ÷ الصافي الحالي',PCT,'uplift',bold=True,bg=GREEN_L)
crow('الربح لكل زيارة — الحالي',f"=IFERROR($B${A['net']}/$B${A['visits']},\"\")",'ر.س','الصافي ÷ الزيارات',DEC2,'ppv0')
crow('الربح لكل زيارة — المستهدف',f"=IFERROR($B${A['newnet']}/$B${A['visits']},\"\")",'ر.س','الصافي المستهدف ÷ الزيارات',DEC2,'ppv1',bold=True,bg=GREEN_L)
ws.cell(A['total'],2).comment=Comment('حجم الأساطيل جزء من فجوة الحجم — لا يُضاف فوقها تفادياً لازدواج الحساب.','القسم التجاري')
MODEL="'نموذج المحرّكات'"

# ═══════════ 2) الملخص التنفيذي ═══════════
ws=sh(wb,'الملخص التنفيذي',cover=True)
for i,w in enumerate([3,13,13,13,13,13,13,13,13,3],1): ws.column_dimensions[get_column_letter(i)].width=w
r=banner(ws,'الاستراتيجية التجارية — منفذ العمرة الجديدة (MK007)','١٢ شهراً · مقترح للاعتماد · الأرقام مرتبطة بورقة «نموذج المحرّكات»',10)
mo=S['fuel']['monthly']; peak=max(mo); last=mo[-1]
kpi(ws,r,2,2,FIN['liters'],'حجم الوقود (لتر/شهر)',ORANGE,NUM)
kpi(ws,r,4,2,FIN['visits'],'الزيارات / شهر',SKY,NUM)
kpi(ws,r,6,2,f"={MODEL}!B{A['net']}",'الصافي الحالي (ر.س/شهر)',GREEN,MONEY)
kpi(ws,r,8,2,f"={MODEL}!B{A['ppv0']}",'الربح لكل زيارة (ر.س)',ORANGE,DEC2)
r+=3
kpi(ws,r,2,2,(last-peak)/peak,'تراجع الحجم عن الذروة',RED,PCT0)
kpi(ws,r,4,2,f"={MODEL}!B{A['secured']}",'حجم قابل للتأمين بعقود (لتر)',SKY,NUM)
kpi(ws,r,6,2,f"={MODEL}!B{A['total']}",'الأثر المستهدف (ر.س/شهر)',GOLD,MONEY)
kpi(ws,r,8,2,f"={MODEL}!B{A['uplift']}",'الزيادة على الصافي',GREEN,PCT0)
r+=3
put(ws,r,2,'المسألة الاستراتيجية',BOLD(12,RED),bg=RED_L)
for j in range(3,10): put(ws,r,j,'',bg=RED_L)
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9); r+=1
txt=('المنفذ الأقوى في المحفظة — وحجمه يتآكل. أعلى تقييماً (4.8★) وأعلى ربحاً لكل زيارة، '
     'لكن حجم الوقود تراجع ٣٠٪ من ذروة مارس إلى يونيو بأثر ‎−222 ألف ر.س هامش شهرياً. '
     'السبب: كل الحجم تقريباً حركة عابرة غير متعاقد عليها، والعميل مجهول الهوية (التطبيق 0.6٪).')
put(ws,r,2,txt,Font(name=F,size=10,color=INK),wrap=True,bg=RED_L)
for j in range(3,10): put(ws,r,j,'',bg=RED_L)
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9); ws.row_dimensions[r].height=46; r+=2
put(ws,r,2,'الخيار الاستراتيجي:  حجم متعاقد · عميل معروف · زيارة مُسيَّلة',Font(name=F,size=13,bold=True,color='FFFFFF'),bg=ORANGE,al='center')
for j in range(3,10): put(ws,r,j,'',bg=ORANGE)
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9); ws.row_dimensions[r].height=30; r+=2

put(ws,r,2,'الاتجاه الشهري لحجم الوقود (مليون لتر)',BOLD(10)); r+=1
tr=r
put(ws,r,2,'الشهر',BOLD(9.5,'FFFFFF'),bg=ORANGE,al='center')
put(ws,r,3,'الفعلي',BOLD(9.5,'FFFFFF'),bg=ORANGE,al='center')
put(ws,r,4,'المستهدف',BOLD(9.5,'FFFFFF'),bg=ORANGE,al='center'); r+=1
for i,mn in enumerate(MONTHS):
    put(ws,r,2,mn,al='center'); put(ws,r,3,mo[i] if i<len(mo) else 0,BLUE_F,DEC2)
    put(ws,r,4,f"={MODEL}!$B${A['target']}/1000000",BLACK_F,DEC2); r+=1
band(ws,tr+1,r-1,4)
ch=LineChart()
ch.add_data(Reference(ws,min_col=3,max_col=4,min_row=tr,max_row=r-1),titles_from_data=True)
ch.set_categories(Reference(ws,min_col=2,min_row=tr+1,max_row=r-1))
ch.y_axis.numFmt='0.0'; ch.title='حجم الوقود: الفعلي مقابل المستهدف (مليون لتر)'; ch.width=17; ch.height=8; ch.style=2
ws.add_chart(ch,f'F{tr}')
r+=1
put(ws,r,2,'الأرقام تقديرات نموذجية للاعتماد — تُستبدل بأرقام المالية وسجل الأصول.',Font(name=F,size=9,italic=True,color=GOLD),bg=GOLD_L)
for j in range(3,10): put(ws,r,j,'',bg=GOLD_L)
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9)

# ═══════════ 3) خط الأساس ═══════════
ws=sh(wb,'خط الأساس')
r=banner(ws,'خط الأساس — الوضع الحالي','الأرقام الفعلية للمنفذ',4)
r=thead(ws,r,['البند','القيمة','الوحدة','قراءة'],[34,20,16,56])
d0=r
base=[('الإيراد',8270000,'ر.س/شهر','الأكبر في المحفظة'),
 ('حجم الوقود',FIN['liters'],'لتر/شهر','متراجع ٣٠٪ عن ذروة مارس'),
 ('الزيارات',FIN['visits'],'زيارة/شهر','قاعدة ضخمة غير مُستثمرة رقمياً'),
 ('متوسط الفاتورة',64.7,'ر.س','إنفاق أعلى من نظائرها'),
 ('متوسط التعبئة',30.6,'لتر','—'),
 ('دخل الوقود',round(FIN['margin']*FIN['liters']),'ر.س/شهر','هامش 0.15 ر.س/لتر'),
 ('دخل غير الوقود',round(nonfuel),'ر.س/شهر','تأجير + متجر وخدمات'),
 ('المصاريف',round(exp),'ر.س/شهر','رواتب + تشغيل + مرافق + صيانة'),
 ('تغطية غير الوقود للمصاريف',nonfuel/exp,'٪','المصاريف مغطاة بالكامل من غير الوقود ✅'),
 ('التقييم',4.8,'★ (1150)','الأعلى في المحيط — أصل تسويقي'),
 ('عدد الوحدات',sum(t.get('count',1) for t in S['tenants']['items']),'وحدة','13 مؤجّرة · 1 شاغرة'),
 ('مبيعات المستأجرين',sum((t.get('salesM') or [0])[-1] for t in S['tenants']['items'] if t['status']=='مؤجّرة'),'ر.س/شهر','الموثّقة — آخر شهر'),
 ('دخل الإيجار الموثّق',sum(t.get('rent',0) for t in S['tenants']['items'] if t['status']=='مؤجّرة'),'ر.س/شهر','نسبة الإيجار 14.8٪ — صحّية'),
 ('حصة التطبيق',0.006,'٪','القيد الملزم — العميل مجهول'),
 ('حصة النقدي',0.53,'٪','يعمّق مجهولية العميل'),
 ('المساحة',120000,'م²','تتيح توسّعاً في الوحدات')]
for lbl,v,u,note in base:
    put(ws,r,1,lbl,BOLD(9.5)); put(ws,r,2,v,BLUE_F,PCT if u=='٪' else (DEC2 if isinstance(v,float) and v<100 else NUM))
    put(ws,r,3,u,Font(name=F,size=9,color=GREY),al='center'); put(ws,r,4,note,Font(name=F,size=9.5,color=GREY),wrap=True)
    r+=1
band(ws,d0,r-1,4)
r+=1
put(ws,r,1,'مزيج الوقود',BOLD(11)); r+=1
mh=thead(ws,r,['المنتج','الحصة','لتر/شهر','قراءة'],[34,20,16,56]); r=mh
for m in S['fuel']['mix']:
    put(ws,r,1,m[0],BOLD(9.5)); put(ws,r,2,(m[1] or 0)/100,BLUE_F,PCT)
    put(ws,r,3,f"=B{r}*{MODEL}!$B${A['liters']}",BLACK_F,NUM)
    put(ws,r,4,'حصة الديزل المرتفعة تُثبت وجود أساطيل تشتري بلا عقد' if 'ديزل' in m[0] else '',Font(name=F,size=9.5,color=GREY),wrap=True)
    r+=1
band(ws,mh,r-1,4)

# ═══════════ 4) التشخيص ═══════════
ws=sh(wb,'التشخيص')
r=banner(ws,'التشخيص — أين المسألة فعلاً','المشكلة ليست جودة المحطة، بل أن حجمها غير مؤمَّن وعميلها غير معروف',4)
r=thead(ws,r,['#','الملاحظة','الدليل من البيانات','الأثر'],[7,34,44,40])
d0=r
diag=[('١','صفر لتر متعاقد عليه','حصة الديزل 28.1٪ تُثبت وجود أساطيل تشتري — بلا عقد','تقلّب الحجم الذي نراه'),
 ('٢','العميل مجهول الهوية','نقدي ٥٣٪ · التطبيق 0.6٪','لا أداة لاستعادة العميل أو استهدافه'),
 ('٣','فجوة مرساة القهوة','بترومين لديه تيم هورتنز (1114 تقييم) ولا نظير لدينا','حركة تُفقد للمنافس'),
 ('٤','تغذية ذاتية محتملة','فرع درب ثانٍ في نفس الحي (4.9★/755)','جزء من التراجع قد يكون داخلياً'),
 ('٥','أصل غير مُستثمَر','تغطية غير الوقود ١٣٣٪ — المصاريف مغطاة بالكامل','كل لتر إضافي ربح شبه صافٍ')]
for n,obs,ev,imp in diag:
    put(ws,r,1,n,BOLD(11,ORANGE),al='center'); put(ws,r,2,obs,BOLD(9.5),wrap=True)
    put(ws,r,3,ev,BLUE_F,wrap=True); put(ws,r,4,imp,Font(name=F,size=9.5,color=RED),wrap=True)
    ws.row_dimensions[r].height=36; r+=1
band(ws,d0,r-1,4)

# ═══════════ 5) خطط الإدارات ═══════════
ws=sh(wb,'خطط الإدارات')
r=banner(ws,'خطط الإدارات الأربع','المبادرات والمؤشرات لكل إدارة — المسؤول والحالة تُعبَّأ عند الاعتماد',6)
r=thead(ws,r,['الإدارة','الهدف','المبادرات','المؤشرات','المسؤول','الحالة'],[22,28,52,30,20,14])
d0=r
dvp=DataValidation(type='list',formula1='"لم تبدأ,جارية,منجزة,مؤجّلة"',allow_blank=True); ws.add_data_validation(dvp)
plans=[('⛽ مبيعات الوقود','وقف التراجع واستعادة 4.5 مليون لتر/شهر',
  'تأمين 305 ألف لتر متعاقد · ضمان توفّر الديزل · مسار سريع للشاحنات والحافلات · تكثيف التشغيل حول ذروة 17:00 والخميس',
  'اللترات · نسبة الحجم المتعاقد · حصة الديزل · زمن الخدمة'),
 ('🏪 التأجير','سدّ فجوة المرساة ورفع مبيعات المستأجرين',
  'تأجير الوحدة الشاغرة لمرساة قهوة عالمية (أولوية قصوى) · مراجعة تسعير المستأجرين فوق ١٥٪ · تحويل العقود إلى ثابت + نسبة من المبيعات',
  'الإشغال · مبيعات المستأجر · نسبة الإيجار · مبيعات/م²'),
 ('🤝 الشراكات','تحويل الجوار التجاري إلى عقود',
  'إغلاق ميداء (تواصل جارٍ) · فتح المراعي وi كومكس · اتفاقية سيارة اب كمنصّة تفتح محفظة مكاتب التأجير · ريتاج مورّد تموين للمستأجرين',
  'اللترات المتعاقدة · عدد العقود · قيمة خط الأنابيب المُغلق'),
 ('📣 التسويق','حصة التطبيق 0.6٪ → ٥٪ وتحويل التقييم إلى حركة',
  'تفعيل التطبيق عند المضخة بحافز لموظف المحطة · عروض حول ذروة 17:00 والخميس · تكرار نموذج «المذاق المغربي» (+9.5٪ بتمويل الشريك) · حملات موسم العمرة',
  'حصة التطبيق · المستخدمون النشطون · أثر الحملة على الزيارات')]
for d_,goal,init,kpis_ in plans:
    put(ws,r,1,d_,BOLD(10),wrap=True); put(ws,r,2,goal,BOLD(9.5),wrap=True)
    put(ws,r,3,init,BLUE_F,wrap=True); put(ws,r,4,kpis_,Font(name=F,size=9,color=GREY),wrap=True)
    put(ws,r,5,'',BLUE_F,bg='FFFFCC'); put(ws,r,6,'',BLUE_F,bg='FFFFCC',al='center'); dvp.add(ws.cell(r,6))
    ws.row_dimensions[r].height=62; r+=1
r+=1
put(ws,r,1,'ميزانية الترويج',BOLD(11)); r+=1
bh=thead(ws,r,['البند','الحصة','','','',''],[22,28,52,30,20,14]); r=bh
for b in S['marketing']['budget']:
    put(ws,r,1,b[0],BOLD(9.5)); put(ws,r,2,(b[1] or 0)/100,BLUE_F,PCT)
    for j in range(3,7): put(ws,r,j,'')
    r+=1
put(ws,r,1,'المجموع',BOLD(),bg=ORANGE_L); put(ws,r,2,f'=SUM(B{bh}:B{r-1})',BOLD(),PCT,bg=ORANGE_L)
for j in range(3,7): put(ws,r,j,'',bg=ORANGE_L)
ws.conditional_formatting.add(f'B{r}:B{r}',CellIsRule(operator='notEqual',formula=['1'],fill=fill(RED_L),font=Font(name=F,size=10,bold=True,color=RED)))
ws.conditional_formatting.add(f'B{bh}:B{r-1}',DataBarRule(start_type='num',start_value=0,end_type='num',end_value=1,color=ORANGE))

# ═══════════ 6) المستهدفات ═══════════
ws=sh(wb,'المستهدفات')
r=banner(ws,'المستهدفات — ١٢ شهراً','مرتبطة بنموذج المحرّكات · الفجوة تُحسب تلقائياً',5)
r=thead(ws,r,['المؤشر','الأساس','المستهدف','الفجوة','ملاحظة'],[32,20,20,20,44])
d0=r
def trow(lbl,base_,tgt,note,fmt=NUM,gap=True):
    global r
    put(ws,r,1,lbl,BOLD(9.5)); put(ws,r,2,base_,BLACK_F if isinstance(base_,str) else BLUE_F,fmt)
    put(ws,r,3,tgt,BLACK_F if isinstance(tgt,str) else BLUE_F,fmt)
    put(ws,r,4,f'=IFERROR(C{r}-B{r},"")' if gap else '—',BLACK_F,fmt)
    put(ws,r,5,note,Font(name=F,size=9,color=GREY),wrap=True); r+=1
trow('حجم الوقود (لتر/شهر)',f"={MODEL}!B{A['liters']}",f"={MODEL}!B{A['target']}",'٥٢٪ من الفجوة تُسدّ بعقود مؤمَّنة')
trow('الحجم المتعاقد (أساطيل)',0,f"={MODEL}!B{A['secured']}",'من صفر إلى حجم مؤمَّن — جوهر التحوّل')
trow('حصة التطبيق',0.006,0.05,'تمكيني — شرط لقياس واستهداف الباقي',PCT)
trow('الوحدات الشاغرة',1,0,'مرساة القهوة العالمية')
trow('الصافي (ر.س/شهر)',f"={MODEL}!B{A['net']}",f"={MODEL}!B{A['newnet']}",'‎+١٧٪',MONEY)
trow('الربح لكل زيارة (ر.س)',f"={MODEL}!B{A['ppv0']}",f"={MODEL}!B{A['ppv1']}",'المؤشر المشترك للإدارات الأربع',DEC2)
trow('التقييم',4.8,4.8,'حماية الأصل التسويقي الأقوى',DEC2)
band(ws,d0,r-1,5)
r+=1
put(ws,r,1,'نسبة الإيجار للمستأجرين المتجاوزين (الصحّي ≤ ١٥٪)',BOLD(11)); r+=1
th=thead(ws,r,['المستأجر','مبيعات آخر شهر','الإيجار','نسبة الإيجار','الإجراء'],[32,20,20,20,44]); r=th
for t in S['tenants']['items']:
    if t['status']!='مؤجّرة' or not t.get('salesM'): continue
    put(ws,r,1,t['name'],BOLD(9.5)); put(ws,r,2,(t['salesM'] or [0])[-1],BLUE_F,MONEY)
    put(ws,r,3,t.get('rent',0),BLUE_F,MONEY); put(ws,r,4,f'=IFERROR(C{r}/B{r},"")',BLACK_F,PCT)
    put(ws,r,5,f'=IF(D{r}>0.15,"مراجعة التسعير","سليم")',BLACK_F,wrap=True); r+=1
tl=r-1; band(ws,th,tl,5)
ws.conditional_formatting.add(f'D{th}:D{tl}',CellIsRule(operator='greaterThan',formula=['0.15'],fill=fill(RED_L),font=Font(name=F,size=10,bold=True,color=RED)))
ws.conditional_formatting.add(f'D{th}:D{tl}',CellIsRule(operator='lessThanOrEqual',formula=['0.15'],fill=fill(GREEN_L),font=Font(name=F,size=10,bold=True,color=GREEN)))

# ═══════════ 7) خارطة الطريق ═══════════
ws=sh(wb,'خارطة الطريق')
r=banner(ws,'خارطة الطريق — أربعة أرباع','المسؤول والحالة تُعبَّآن عند الاعتماد',6)
r=thead(ws,r,['الربع','التركيز','المخرجات','المسؤول','الحالة','ملاحظات'],[14,22,58,20,14,30])
d0=r
dvr=DataValidation(type='list',formula1='"لم تبدأ,جارية,منجزة,مؤجّلة"',allow_blank=True); ws.add_data_validation(dvr)
road=[('الربع ١','🔒 تأمين الحجم','إغلاق عقد ميداء · إطلاق تفعيل التطبيق عند المضخة بحافز · بدء التفاوض مع المراعي وi كومكس · طرح الوحدة الشاغرة لمرساة القهوة'),
 ('الربع ٢','📈 التوسّع','توقيع سيارة اب والتكاسي · افتتاح مرساة القهوة · مراجعة تسعير المستأجرين المتجاوزين · أول حملة مشتركة ممولة من شريك'),
 ('الربع ٣','💰 التسييل','باقات «تعبئة + وجبة» عبر التطبيق · بيع وصول التطبيق للمستأجرين · عروض مخصّصة بالشريحة'),
 ('الربع ٤','🔁 الترسيخ','تجديد عقود الأساطيل بحجوم أعلى · قياس الربح/الزيارة · تعميم النموذج على منافذ الممر')]
for q,foc,out in road:
    put(ws,r,1,q,BOLD(10,GOLD),al='center'); put(ws,r,2,foc,BOLD(9.5),wrap=True)
    put(ws,r,3,out,BLUE_F,wrap=True); put(ws,r,4,'',BLUE_F,bg='FFFFCC')
    put(ws,r,5,'',BLUE_F,bg='FFFFCC',al='center'); dvr.add(ws.cell(r,5))
    put(ws,r,6,'',BLUE_F); ws.row_dimensions[r].height=52; r+=1
band(ws,d0,r-1,3)

# ═══════════ 8) المخاطر ═══════════
ws=sh(wb,'المخاطر')
r=banner(ws,'سجل المخاطر','الاحتمال والتخفيف — يُراجَع ربعياً',5)
r=thead(ws,r,['المخاطرة','الاحتمال','الأثر','التخفيف','المالك'],[36,14,14,52,20])
d0=r
dvl=DataValidation(type='list',formula1='"منخفض,متوسط,عالٍ"',allow_blank=True); ws.add_data_validation(dvl)
risks=[('استمرار تراجع الحركة العابرة','متوسط','عالٍ','العقود تُثبّت جزءاً من الحجم بمعزل عن الحركة'),
 ('التغذية الذاتية مع فرع درب الثاني','متوسط','متوسط','إدارة الحيّ كمحفظة واحدة لا كمحطتين متنافستين'),
 ('تعذّر جذب مرساة قهوة عالمية','متوسط','متوسط','بديل: مشغّل قهوة مختصة قوي محلياً بعقد نسبة من المبيعات'),
 ('بطء تبنّي التطبيق','عالٍ','عالٍ','حافز مباشر لموظف المحطة + قيمة فورية للعميل عند التفعيل'),
 ('ضغط سعري من الدريس','عالٍ','متوسط','المنافسة على الخدمة والمرافق والمرساة — لا على سعر مقنَّن')]
for rk,p,i,mit in risks:
    put(ws,r,1,rk,BOLD(9.5),wrap=True)
    put(ws,r,2,p,BLUE_F,al='center'); dvl.add(ws.cell(r,2))
    put(ws,r,3,i,BLUE_F,al='center'); dvl.add(ws.cell(r,3))
    put(ws,r,4,mit,Font(name=F,size=9.5,color=GREEN),wrap=True); put(ws,r,5,'',BLUE_F,bg='FFFFCC')
    ws.row_dimensions[r].height=34; r+=1
band(ws,d0,r-1,5)
for col in ('B','C'):
    ws.conditional_formatting.add(f'{col}{d0}:{col}{r-1}',FormulaRule(formula=[f'${col}{d0}="عالٍ"'],fill=fill(RED_L),font=Font(name=F,size=10,bold=True,color=RED)))
    ws.conditional_formatting.add(f'{col}{d0}:{col}{r-1}',FormulaRule(formula=[f'${col}{d0}="متوسط"'],fill=fill(GOLD_L),font=Font(name=F,size=10,color=GOLD)))

# ═══════════ 9) القرارات المطلوبة ═══════════
ws=sh(wb,'القرارات المطلوبة')
r=banner(ws,'القرارات المطلوبة من الإدارة التنفيذية','تُعتمد أو تُعدَّل في الاجتماع',5)
r=thead(ws,r,['#','القرار','الأثر إن اعتُمد','القرار (اعتماد/تعديل/تأجيل)','ملاحظات'],[7,44,44,26,30])
d0=r
dvd=DataValidation(type='list',formula1='"اعتماد,تعديل,تأجيل"',allow_blank=True); ws.add_data_validation(dvd)
dec=[('١','اعتماد الأولوية: عقود الأساطيل أولاً — وتفويض التفاوض بخصم حجمي','تأمين ٥٢٪ من فجوة الحجم'),
 ('٢','اعتماد حافز تفعيل التطبيق لموظفي المحطة','يفتح قياس واستهداف كل ما بعده'),
 ('٣','اعتماد نموذج الإيجار «ثابت + نسبة من مبيعات المستأجر»','يربط دخلنا بنجاح المستأجر ويضمن استدامته'),
 ('٤','تسمية مالك المنفذ (P&L Owner) للعمرة الجديدة','مسؤول واحد عن ربح المحطة الكامل'),
 ('٥','اعتماد مستهدف حصة التطبيق (٥٪) والربح لكل زيارة','مؤشرات المتابعة المشتركة')]
for n,d_,imp in dec:
    put(ws,r,1,n,BOLD(11,ORANGE),al='center'); put(ws,r,2,d_,BOLD(9.5),wrap=True)
    put(ws,r,3,imp,Font(name=F,size=9.5,color=GREEN),wrap=True)
    put(ws,r,4,'',BLUE_F,bg='FFFFCC',al='center'); dvd.add(ws.cell(r,4))
    put(ws,r,5,'',BLUE_F); ws.row_dimensions[r].height=38; r+=1
band(ws,d0,r-1,3)

wb.move_sheet('الملخص التنفيذي',offset=-(len(wb.sheetnames)-1))
wb.save(OUT)
print('saved:',OUT)
print('sheets:',[s.title for s in wb.worksheets])
