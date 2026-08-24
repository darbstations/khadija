# -*- coding: utf-8 -*-
"""ورقة الأداء المالي — من تقرير يوليو ٢٠٢٦ (النسخة المصححة، ١٣ صفحة)"""
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule
import worker_model as W
import brand as B

SHEET = "الأداء المالي — يوليو ٢٠٢٦"

# قائمة الدخل التراكمية — ألف ريال (صفحة ٦)
PL = [
    ("إيرادات الوقود",            1173578, 1219570,  411451),
    ("تكلفة المبيعات",           -1133905, -1177404, -391474),
    ("هامش المساهمة — الوقود",      39673,    42166,   19976),
    ("صافي الإيراد العقاري",        31465,    33138,   24887),
    ("إهلاك أصول حق الاستخدام",    -15347,   -15999,  -12391),
    ("تكلفة التمويل",              -15525,   -14569,   -9828),
    ("هامش المساهمة — العقارات",      593,     2570,    2669),
    ("مصاريف التشغيل",             -28704,   -27760,  -13962),
    ("مجمل الربح",                  11563,    16976,    8683),
    ("عمومية وإدارية + بيع وتوزيع", -13958,   -10927,  -12063),
    ("فوائد القروض",                -3043,    -1653,    -459),
    ("الزكاة",                       -120,     -110,    -101),
    ("صافي الدخل",                  -5535,     4285,   -3941),
]
BOLDROWS = {"هامش المساهمة — الوقود", "مجمل الربح", "صافي الدخل"}

# أداء نماذج العمل (صفحة ٨) — بعد التحميل المركزي
MODELS = [
    # النموذج، محطات، لتر مليون، إيراد، هامش وقود، بيع وتوزيع، عمومية، فوائد، صافي، قبل التحميل، خاسرة
    ("استثماري", 6,   91.8, 174309, 10765,  -597,  -2921,  -761,  2722,  9505, 0),
    ("إيجاري",   17,  77.7, 146564,  8787,  -597,  -2921,  -761, -7038, -1290, 11),
    ("تشغيلي",   26, 141.2, 268915, 15984,  -717,  -3505,  -913, -1869,  6864, 1),
    ("امتياز",   92, 350.8, 583790,  4136,  -478,  -2337,  -609,   609,  3296, 6),
]

NET = {"استثماري": "الوحيد الرابح — وأعلى هامش لكل لتر",
       "إيجاري": "الخسارة من إهلاك حق الاستخدام والتمويل لا من التشغيل",
       "تشغيلي": "يربح قبل التحميل ويخسر بعده",
       "امتياز": "نصف اللترات وعُشر الهامش — ١٫١٨ هللة/لتر"}


def build_financial(wb, idx=None):
    NC = 11
    ws = wb.create_sheet(SHEET) if idx is None else wb.create_sheet(SHEET, idx)
    W.setup(ws, [4, 30, 14, 14, 13, 14, 12, 12, 12, 12, 40], freeze="A5")
    W.title(ws, "الأداء المالي — يناير إلى يوليو ٢٠٢٦",
            "من تقرير درب التجاري (النسخة المصححة) · ألف ريال سعودي", NC)
    W.rule(ws, 3, NC)

    # ── ① قائمة الدخل
    r = W.band(ws, 4, NC, "① قائمة الدخل التراكمية")
    W.header(ws, r, ["#", "البند", "الفعلي", "الموازنة", "الفرق", "العام الماضي"]
             + [""] * 4 + ["ملاحظة"])
    for i, (lbl, act, bud, ly) in enumerate(PL):
        rr = r + 1 + i
        ws.cell(rr, 1, i + 1)
        c = ws.cell(rr, 2, lbl)
        c.font = Font(name=W.F, size=10, bold=lbl in BOLDROWS, color=B.INK)
        c.alignment = W.RGT
        for j, v in ((3, act), (4, bud), (5, act - bud), (6, ly)):
            cc = ws.cell(rr, j, v)
            cc.number_format = "#,##0;(#,##0)" if j != 5 else "+#,##0;(#,##0)"
        for j in range(1, NC + 1):
            ws.cell(rr, j).border = W.BOX
            if j != 2: ws.cell(rr, j).font = W.BLACK
            if j != 11: ws.cell(rr, j).alignment = W.CTR
        if lbl in BOLDROWS:
            for j in range(1, NC + 1): ws.cell(rr, j).fill = W.FS
        ws.row_dimensions[rr].height = 19
    pend = r + len(PL)
    ws.conditional_formatting.add(f"E{r+1}:E{pend}", CellIsRule(
        operator="lessThan", formula=["0"], fill=W.BAD,
        font=Font(name=W.F, size=10, color=B.D_BAD)))
    ws.conditional_formatting.add(f"E{r+1}:E{pend}", CellIsRule(
        operator="greaterThan", formula=["0"], fill=W.OK,
        font=Font(name=W.F, size=10, color=B.D_GOOD)))
    r = pend + 2

    # ── ② نماذج العمل
    r = W.band(ws, r, NC, "② أداء نماذج العمل — قبل التحميل المركزي وبعده")
    W.header(ws, r, ["#", "النموذج", "محطات", "لتر (مليون)", "الإيراد",
                     "هامش الوقود", "هللة/لتر", "قبل التحميل", "التحميل",
                     "صافي الربح", "القراءة"])
    m0 = r + 1
    for i, (nm, st, vol, rev, fm, sd, ga, fin, net, pre, lossy) in enumerate(MODELS):
        rr = m0 + i
        load = sd + ga + fin
        vals = [i + 1, nm, st, vol, rev, fm, fm * 1000 / (vol * 1e6) * 100,
                pre, load, net]
        fmts = [None, None, "0", "#,##0.0", "#,##0", "#,##0", "0.00",
                "+#,##0;(#,##0)", "(#,##0)", "+#,##0;(#,##0)"]
        for j, (v, fm_) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(rr, j, v); c.border = W.BOX; c.alignment = W.CTR
            c.font = W.BOLD if j == 2 else W.BLACK
            if fm_: c.number_format = fm_
        ws.cell(rr, NC, NET[nm]).alignment = W.WRAP
        ws.cell(rr, NC).border = W.BOX
        ws.row_dimensions[rr].height = 24
    mend = m0 + len(MODELS) - 1
    tot = mend + 1
    ws.cell(tot, 2, "الإجمالي").font = W.BOLD
    for j, col in ((3, 1), (4, 2), (5, 3), (6, 4), (8, 9), (10, 8)):
        ws.cell(tot, j, f"=SUM({chr(64+j)}{m0}:{chr(64+j)}{mend})")
    ws.cell(tot, 9, f"=SUM(I{m0}:I{mend})")
    ws.cell(tot, 7, f"=IFERROR(F{tot}*1000/(D{tot}*1000000)*100,\"\")").number_format = "0.00"
    for j in range(1, NC + 1):
        ws.cell(tot, j).border = W.BOX; ws.cell(tot, j).fill = W.FS
        ws.cell(tot, j).font = W.BOLD; ws.cell(tot, j).alignment = W.CTR
    for j, fmt in ((3, "0"), (4, "#,##0.0"), (5, "#,##0"), (6, "#,##0"),
                   (8, "+#,##0;(#,##0)"), (9, "(#,##0)"), (10, "+#,##0;(#,##0)")):
        ws.cell(tot, j).number_format = fmt
    ws.conditional_formatting.add(f"J{m0}:J{mend}", CellIsRule(
        operator="lessThan", formula=["0"], fill=W.BAD,
        font=Font(name=W.F, size=10, bold=True, color=B.D_BAD)))
    ws.conditional_formatting.add(f"J{m0}:J{mend}", CellIsRule(
        operator="greaterThan", formula=["0"], fill=W.OK,
        font=Font(name=W.F, size=10, bold=True, color=B.D_GOOD)))
    r = tot + 2

    # ── ③ نقطة التعادل والتحميل
    r = W.band(ws, r, NC, "③ نقطة التعادل — أين تقف الشركة")
    k0 = r
    KEY = [
        ("مجمل الربح", "=C14", "#,##0", "من قائمة الدخل أعلاه"),
        ("التحميل المركزي", f"=-I{tot}", "#,##0", "بيع وتوزيع + عمومية + فوائد محمّلة"),
        ("التحميل ÷ مجمل الربح", f"=C{k0+1}/C{k0}", "0%",
         "التحميل يتجاوز مجمل الربح — هنا أصل الخسارة"),
        ("المحطات النشطة", "146", "0", "خطة يوليو ١٩٨ · هدف ديسمبر ٢٣٢"),
        ("مجمل الربح لكل محطة", f"=C{k0}/C{k0+3}", "#,##0.0", "ألف ريال / ٧ أشهر"),
        ("محطات التعادل المطلوبة", f"=ROUND(C{k0+1}/C{k0+4},0)", "0",
         "عند مجمل الربح الحالي لكل محطة — الموازنة نفسها (١٩٨) كانت دون التعادل"),
    ]
    for i, (lbl, f_, fmt, note) in enumerate(KEY):
        rr = k0 + i
        ws.cell(rr, 2, lbl).font = W.BOLD; ws.cell(rr, 2).alignment = W.RGT
        c = ws.cell(rr, 3, f_); c.number_format = fmt
        c.font = W.BLACK; c.fill = W.CALC; c.alignment = W.CTR; c.border = W.BOX
        ws.cell(rr, 2).border = W.BOX
        ws.cell(rr, NC, note).alignment = W.WRAP
        ws.row_dimensions[rr].height = 20
    ws.conditional_formatting.add(f"C{k0+2}", CellIsRule(
        operator="greaterThan", formula=["1"], fill=W.BAD,
        font=Font(name=W.F, size=10, bold=True, color=B.D_BAD)))
    r = k0 + len(KEY) + 2

    W.bullets(ws, r, NC, [
        ("ما تقوله الأرقام", [
            "المحطات تربح +١٨٫٤ مليون قبل التحميل المركزي — والشركة تخسر (٥٫٥) بعده",
            "الامتياز ٥٣٪ من اللترات و١٠٪ من الهامش: ١٫١٨ هللة/لتر مقابل ١١٫٣ لبقية النماذج",
            "محطة استثمارية واحدة تعادل ٦٩ محطة امتياز في صافي الربح",
            "الفوائد ٦٫٦ أضعاف العام الماضي وتجاوز الموازنة ٨٤٪ — أسرع بند نمواً",
        ], W.NAVY),
        ("ما يحتاج تحققاً قبل أي قرار", [
            "مفتاح توزيع العمومية غير مُفسَّر: ٦ محطات استثمارية تحمل ما تحمله ١٧ إيجارية",
            "خسارة العقار قد تكون محاسبية لا نقدية — تمويل IFRS 16 مُقدَّم التحميل والعقود جديدة",
            "رفع هامش الامتياز نقطة يعطي ٥٫٨ مليون، لكن هامش المُمتاز نفسه غير معروض",
            "فجوة ٦٬٧٣٨ بين صافي المحطات ومجمل الربح ما زالت مفتوحة في التقرير",
        ], "C00000"),
    ])
    return ws
