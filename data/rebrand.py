# -*- coding: utf-8 -*-
"""إعادة تلوين الأوراق القديمة إلى هوية درب — دون المساس بالصيغ أو التنسيقات الرقمية"""
from copy import copy
from openpyxl.styles import Font, PatternFill, Side, Border
import brand as B

# ── خرائط التحويل: القديم ← الهوية
FILL = {
    "1F3864": B.BGRAY,      # رأس الجدول الكحلي
    "D9E2F3": B.T_BAND,     # صف الإجمالي
    "FFFF00": B.T_GOLD,     # خانة إدخال
    "DDEBF7": B.T_ORANGE,   # شريط القسم
    "C6E0B4": B.T_GOOD,
    "FFE699": B.T_GOLD,
    "F8CBAD": B.T_BAD,
    "F2F2F2": B.T_NEUTRAL,
    "E2EFDA": B.T_GOOD,
    "FCE4D6": B.T_BAD,
}
FONT = {
    "1F3864": B.BGRAY,      # عناوين
    "44546A": B.INK2,       # نص ثانوي
    "0000FF": B.BLUE,       # خط الإدخال
    "008000": B.GOOD,       # روابط الأوراق
    "C00000": B.D_BAD,
    "375623": B.D_GOOD,
    "833C00": B.D_GOLD,
    "000000": B.INK,
}
BAND_FILLS = {"DDEBF7"}     # شريط القسم — نصّه يصير برتقالياً


def _rgb(o):
    try:
        v = o.rgb
        return str(v)[-6:].upper() if isinstance(v, str) else None
    except Exception:
        return None


def rebrand(ws):
    """يعيد تلوين كل خلايا الورقة ويوحّد الخط والحدود"""
    thin = Side(style="thin", color=B.LINE2)
    changed = 0
    for row in ws.iter_rows():
        for c in row:
            # ① التعبئة
            band = False
            f = c.fill
            if f is not None and f.patternType == "solid":
                cur = _rgb(f.fgColor)
                if cur in FILL:
                    band = cur in BAND_FILLS
                    c.fill = PatternFill("solid", fgColor=FILL[cur]); changed += 1
            # ② الخط — يُبنى جديداً حفاظاً على الحجم والسماكة
            ft = c.font
            if ft is not None:
                cur = _rgb(ft.color)
                col = ft.color
                if band and ft.bold:
                    col = B.ORANGE
                elif cur in FONT:
                    col = FONT[cur]
                elif cur is None or cur == "FFFFFF":
                    col = ft.color
                c.font = Font(name=B.FONT, size=ft.size, bold=ft.bold, italic=ft.italic,
                              underline=ft.underline, color=col)
            # ③ الحدود — لون الهوية الفاتح
            bd = c.border
            if bd is not None and any([bd.left.style, bd.right.style, bd.top.style, bd.bottom.style]):
                c.border = Border(
                    left=thin if bd.left.style else copy(bd.left),
                    right=thin if bd.right.style else copy(bd.right),
                    top=thin if bd.top.style else copy(bd.top),
                    bottom=thin if bd.bottom.style else copy(bd.bottom))
    # ④ قواعد التنسيق الشرطي تحمل ألوانها الخاصة
    for rng in list(ws.conditional_formatting):
        for rule in rng.rules:
            dxf = getattr(rule, "dxf", None)
            if dxf is None:
                continue
            if dxf.fill is not None:
                for attr in ("bgColor", "fgColor"):
                    o = getattr(dxf.fill, attr, None)
                    cur = _rgb(o) if o is not None else None
                    if cur in FILL:
                        setattr(dxf.fill, attr, PatternFill("solid", fgColor=FILL[cur]).fgColor)
            if dxf.font is not None and dxf.font.color is not None:
                cur = _rgb(dxf.font.color)
                if cur in FONT:
                    dxf.font.color.rgb = "00" + FONT[cur]
    return changed


def rebrand_all(wb, skip=()):
    """يعيد تلوين كل الأوراق عدا ما بُني أصلاً بالهوية"""
    total = {}
    for ws in wb:
        if ws.title in skip:
            continue
        total[ws.title] = rebrand(ws)
    return total
