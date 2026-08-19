# -*- coding: utf-8 -*-
"""ورقة السيطرة الميدانية — تُسند دعوى «من يسيطر» بقياس القرب حول محطاتنا الخمس"""
import json, re, collections, pathlib
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
import worker_model as W

SHEET = "السيطرة الميدانية"
RIVALS = "المنافسون"
BRANDS = [("Aldrees", "الدريس"), ("الدريس", "الدريس"), ("NAFT", "نفط"), ("نفط", "نفط"),
          ("ساسكو", "ساسكو"), ("SASCO", "ساسكو"), ("Petrogen", "بتروجين"),
          ("بتروجين", "بتروجين"), ("اومكو", "أومكو"), ("OMCO", "أومكو"),
          ("ADNOC", "أدنوك"), ("توتال", "توتال")]


def brand(n):
    for k, v in BRANDS:
        if k.lower() in n.lower():
            return v
    return "مستقلة/أخرى"


def metres(s):
    return int(re.sub(r"[^\d]", "", str(s)) or 0)


def five(path=None):
    p = path or pathlib.Path(__file__).parent / "five.json"
    return json.load(open(p, encoding="utf-8"))


def build_control(wb, ST, idx=None):
    NC = 11
    ws = wb.create_sheet(SHEET) if idx is None else wb.create_sheet(SHEET, idx)
    W.setup(ws, [4, 18, 10, 13, 13, 14, 10, 11, 10, 10, 36], freeze="A4")
    W.title(ws, "السيطرة الميدانية — من يسيطر فعلاً حول محطاتنا",
            "قياس القرب من خرائط جوجل (يوليو ٢٠٢٦) — خمس محطات مكة محل الدراسة", NC)

    # ── ① العلامات حول محطاتنا
    r = W.band(ws, 3, NC, "① العلامات المنافسة حول محطاتنا — من عيّنة أقرب المحطات")
    cnt = collections.Counter(); dist = collections.defaultdict(list)
    per = {}
    for s in ST:
        c = collections.Counter()
        for x in s["competitors"]:
            b = brand(x["name"]); cnt[b] += 1; c[b] += 1; dist[b].append(metres(x["dist"]))
        per[s["code"]] = c
    total = sum(cnt.values())
    W.header(ws, r, ["#", "العلامة", "مرات", "الحصة", "أقرب مسافة", "متوسط المسافة"]
             + [""] * 4 + ["الدلالة"])
    NOTE = {
        "الدريس": "أقرب منافس في محطتين من خمس — الانتشار ليس دعوى سوقية بل واقع مقيس",
        "مستقلة/أخرى": "سوق مجزّأ حول محطاتنا — لا علامة واحدة تحتكر الجوار",
        "بتروجين": "حضور محدود لكنه بتقييم مرتفع — منافس على الخدمة لا على العدد",
        "ساسكو": "حضور متفرّق وبعيد نسبياً — الضغط منها أقل مما توحي حصتها الوطنية",
        "أومكو": "حضور هامشي بعيد",
        "نفط": "تابعة لساسكو — تُحسب ضمن ثقلها الفعلي",
    }
    for i, (b, n) in enumerate(cnt.most_common()):
        rr = r + 1 + i
        ds = dist[b]
        vals = [i + 1, b, n, n / total, min(ds), sum(ds) // len(ds)]
        fmts = [None, None, "#,##0", "0%", '#,##0" م"', '#,##0" م"']
        for j, (v, fm) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(rr, j, v); c.border = W.BOX; c.alignment = W.CTR
            c.font = W.BOLD if j == 2 else W.BLACK
            if fm: c.number_format = fm
        ws.cell(rr, NC, NOTE.get(b, "")).alignment = W.WRAP
        for j in range(1, NC + 1):
            ws.cell(rr, j).border = W.BOX
        ws.row_dimensions[rr].height = 20
    bend = r + len(cnt)
    ws.conditional_formatting.add(f"D{r+1}:D{bend}", ColorScaleRule(
        start_type="min", start_color="FFFFFF", end_type="max", end_color="F8CBAD"))
    sasco = cnt.get("ساسكو", 0) + cnt.get("نفط", 0)
    ws.cell(bend + 1, 2, f"ساسكو ونفط علامة واحدة بعد الاستحواذ: {sasco} مرة "
                         f"({sasco/total*100:.0f}٪) — وتبقى دون نصف حضور الدريس").font = W.SMALL
    ws.cell(bend + 1, 2).alignment = W.RGT
    ws.cell(bend + 2, 2, f"العيّنة {total} محطة (أقرب ١٠ لكل محطة) من أصل "
                         f"{sum(s['nComp'] for s in ST)} محطة مرصودة ضمن ٥ كم").font = W.SMALL
    ws.cell(bend + 2, 2).alignment = W.RGT
    r = bend + 4

    # ── ② المنافسة لكل محطة
    r = W.band(ws, r, NC, "② المنافسة حول كل محطة — القرب مقابل النمو")
    W.header(ws, r, ["#", "المحطة", "الكود", "منافسون ≤٥كم", "أقرب منافس", "مسافته",
                     "تقييمنا", "متوسطهم", "الفارق", "النمو", "القراءة"])
    READ = {
        "MK017": "أبعد منافس وأقل كثافة — والوحيدة النامية",
        "MK007": "جوار شبه خالٍ — تراجعها موسمي لا تنافسي، وقد تعافت",
        "MK019": "منافس على ٤٨٨ م يعترض العميل قبلنا — يطابق ما تأكد ميدانياً",
        "MK023": "أعلى كثافة في المجموعة (٢٨ محطة) ومنافس على ٤١١ م",
        "MK002": "منافس على ١٨١ م — لكن السبب الأول تغيّر مسار الطريق لا المنافسة",
    }
    order = sorted(ST, key=lambda s: -metres(s["nearDist"]))
    for i, s in enumerate(order):
        rr = r + 1 + i
        g = float(re.sub(r"[^\d\-+.]", "", s["growth"])) / 100
        vals = [i + 1, s["name"], s["code"], s["nComp"], s["nearName"],
                metres(s["nearDist"]), s["rating"], s["compAvg"],
                s["rating"] - s["compAvg"], g]
        fmts = [None, None, None, "#,##0", None, '#,##0" م"', "0.0", "0.0",
                "+0.0;-0.0", "+0%;-0%"]
        for j, (v, fm) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(rr, j, v); c.border = W.BOX; c.alignment = W.CTR
            c.font = W.BOLD if j == 2 else W.BLACK
            if fm: c.number_format = fm
        ws.cell(rr, NC, READ.get(s["code"], "")).alignment = W.WRAP
        for j in range(1, NC + 1):
            ws.cell(rr, j).border = W.BOX
        ws.row_dimensions[rr].height = 22
    send = r + len(ST)
    ws.conditional_formatting.add(f"F{r+1}:F{send}", CellIsRule(
        operator="lessThan", formula=["500"], fill=W.BAD,
        font=Font(name=W.F, size=10, bold=True, color="C00000")))
    ws.conditional_formatting.add(f"J{r+1}:J{send}", CellIsRule(
        operator="greaterThan", formula=["0"], fill=W.OK,
        font=Font(name=W.F, size=10, bold=True, color="375623")))
    ws.conditional_formatting.add(f"J{r+1}:J{send}", CellIsRule(
        operator="lessThan", formula=["0"], fill=W.BAD,
        font=Font(name=W.F, size=10, color="C00000")))
    ws.conditional_formatting.add(f"I{r+1}:I{send}", CellIsRule(
        operator="greaterThan", formula=["0"], fill=W.OK,
        font=Font(name=W.F, size=10, bold=True, color="375623")))
    r = send + 2

    # ── ③ مصفوفة العلامات لكل محطة
    r = W.band(ws, r, NC, "③ أي علامة عند أي محطة")
    blist = [b for b, _ in cnt.most_common()]
    W.header(ws, r, ["#", "المحطة"] + blist + [""] * (NC - 2 - len(blist)))
    for i, s in enumerate(order):
        rr = r + 1 + i
        ws.cell(rr, 1, i + 1); ws.cell(rr, 2, s["name"]).font = W.BOLD
        for j, b in enumerate(blist, 3):
            ws.cell(rr, j, per[s["code"]].get(b, 0)).number_format = '0;;"—"'
        for j in range(1, NC + 1):
            ws.cell(rr, j).border = W.BOX; ws.cell(rr, j).alignment = W.CTR
            if j != 2: ws.cell(rr, j).font = W.BLACK
        ws.row_dimensions[rr].height = 19
    mend = r + len(ST)
    ws.conditional_formatting.add(
        f"C{r+1}:{chr(ord('C') + len(blist) - 1)}{mend}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                       end_type="max", end_color="F8CBAD"))
    r = mend + 2

    W.bullets(ws, r, NC, [
        ("ما يغيّره القياس في دعوى السيطرة", [
            "الدريس ٤٠٪ من أقرب المنافسين حولنا — أي ضعف حصتها الوطنية (١٨.٩٥٪) تقريباً",
            "وهي صاحبة أقرب منافس في الفردوس (٤١١ م) والشرايع (٤٨٨ م) — المحطتان الهابطتان",
            "ساسكو حضورها حول محطاتنا أقل بكثير مما توحي حصتها الوطنية — الضغط منها أخف",
            "ثلث الجوار مستقلات — فالسوق حولنا مجزّأ ولا تحتكره علامة واحدة",
        ], W.NAVY),
        ("ما لا يغيّره", [
            "درب أعلى تقييماً من كل منافسيها في المواقع الخمسة — الخسارة ليست في جودة الخدمة",
            "المعيصم سببها الأول تغيّر مسار الطريق لا المنافسة — القرب عامل مساعد لا سبباً",
            "النورية جوارها شبه خالٍ وتراجعها موسمي — لا يُقرأ ضمن ضغط المنافسة",
        ], W.NAVY),
        ("حدود هذا القياس", [
            "العيّنة أقرب ١٠ محطات لكل موقع — لا كل المرصود ضمن ٥ كم",
            "خرائط جوجل لا ترصد كل المحطات ولا تُحدّث فوراً — الأرقام أرضية لا سقف",
            "القرب لا يقيس الحجم — محطة قريبة صغيرة أخف أثراً من بعيدة كبيرة",
        ], "C00000"),
    ])
    return ws


# ══════════════════════════════════════════════════════════════════
ROWS = [
    ("الدريس", "وقود · أفراد", "١٬٣٨٣ محطة · ١٨٫٩٥٪",
     "الانتشار وسرعة ضمّ المستقلين",
     "المنافسة على الخدمة والربح لكل زيارة لا على السعر",
     "٤٠٪ من أقرب المنافسين حول محطاتنا — وأقرب منافس في الفردوس (٤١١ م) والشرايع (٤٨٨ م)"),
    ("ساسكو", "وقود · أفراد", "~٦٨٣ محطة · تشمل نفط",
     "النمو بالاستحواذ · وتطبيق أفراد ناضج",
     "تحسين تجربة المحطة وربط التطبيق بالمستأجرين",
     "ساسكو ونفط مجتمعتين ١٢٪ فقط حول محطاتنا — الضغط أخف مما توحي الحصة الوطنية"),
    ("بترو آب", "وقود · شركات", "٥٠٠ ألف مركبة · ١٠ آلاف منشأة",
     "ملكية العلاقة مع المنشأة بلا أصول",
     "الاشتراك كنقطة قبول + التعاقد المباشر مع أساطيل الحج والعمرة", ""),
    ("واعي (الدريس)", "وقود · شركات", "شبكة الدريس ١٬٣٨٣ محطة",
     "شريحة في فوهة المضخة تثبّت الأسطول",
     "خصم هللة لكل لتر يفوق ميزة الأداة التقنية", ""),
    ("سيارة آب", "وقود · شركات", "٢٬٥٠٠+ محطة قبول",
     "منصة وشبكة معاً بعد استحواذ ٢٠٢٦",
     "شراكة قائمة — رفع أساطيلنا عليها", ""),
    ("درب باي", "وقود · شركات", "مرخّصة من ساما — ١١ نوفمبر ٢٠٢٥",
     "لا تبني شبكة قبول — تركب على شبكة Visa فتُقبل في كل محطة",
     "لا سبيل للاشتراك معها — الردّ بالتعاقد المباشر مع الأسطول قبل أن تصله",
     "تشابه اسم مع علامتنا في القناة نفسها — والقضية منظورة"),
    ("Fleet Plus (كفرات بلس)", "إكسسوارات وخدمات · شركات", "٥٬٠٠٠+ مزوّد خدمة",
     "أصلها خدمات سيارات ثم أضافت الوقود — تنافس على الوحدة المؤجّرة قبل اللتر",
     "شراكة لا مواجهة — عرض مغاسلنا ومراكز خدمتنا كنقاط تنفيذ ضمن شبكتها", ""),
    ("الوسطاء العقاريون", "عقار", "مجزّأ — لا سلسلة مسيطرة",
     "ملكية قائمة المستأجرين الباحثين",
     "التحالف لا المنافسة — عمولة مقابل إشغال", ""),
    ("محلات التجزئة", "إكسسوارات", "مجزّأ — متاجر ومعارض",
     "المخزون والخبرة وعلاقة المورّد",
     "التعاقد لا المنافسة — مساحة مقابل نسبة", ""),
]


def build_rivals(wb, idx=None):
    """يعيد بناء ورقة المنافسين — بعد إضافة درب باي وإعادة تصنيف Fleet Plus"""
    NC = 7
    ws = wb.create_sheet(RIVALS) if idx is None else wb.create_sheet(RIVALS, idx)
    W.setup(ws, [4, 20, 20, 26, 34, 40, 44], freeze="A4")
    W.title(ws, "المنافسون حسب المنفذ", "من يسيطر · وبماذا · وردّنا — مع الشاهد الميداني", NC)
    W.header(ws, 3, ["#", "المنافس", "المنفذ والقناة", "الحجم", "مصدر السيطرة",
                     "الردّ", "شاهد ميداني — خمس محطات مكة"])
    for i, (nm, ch, sz, src, ans, ev) in enumerate(ROWS):
        rr = 4 + i
        for j, v in enumerate([i + 1, nm, ch, sz, src, ans, ev], 1):
            c = ws.cell(rr, j, v); c.border = W.BOX
            c.font = W.BOLD if j == 2 else W.BLACK
            c.alignment = W.CTR if j in (1, 2, 3) else W.WRAP
        if nm == "درب باي":
            for j in range(1, NC + 1):
                ws.cell(rr, j).fill = W.BAD
            ws.cell(rr, 2).font = Font(name=W.F, size=10, bold=True, color="C00000")
        elif ev:
            ws.cell(rr, NC).fill = W.WARN
        ws.row_dimensions[rr].height = 30
    r = 4 + len(ROWS) + 1
    W.bullets(ws, r, NC, [
        ("القراءة", [
            "في قناة الأفراد المنافس مشغّل محطات — والسلاح تجربة الخدمة",
            "في قناة الشركات المنافس منصة أو أداة — والسلاح السعر والعقد",
            "في العقار والإكسسوارات لا مسيطر — والسلاح الشراكة لا المواجهة",
        ], W.NAVY),
        ("منافسان يحتاجان تصنيفاً مختلفاً", [
            "درب باي وحدها لا يمكن التفاوض معها — لا تحتاج شبكتنا لأنها تُقبل عبر Visa",
            "Fleet Plus أصلها خدمات سيارات لا وقود — فالأولى شراكتها على الوحدات المؤجّرة",
        ], "C00000"),
    ])
    return ws
