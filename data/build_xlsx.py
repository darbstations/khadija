# -*- coding: utf-8 -*-
"""ملف إكسل تنفيذي لبيانات وخطط القسم التجاري — لوحة تحكم + رسوم + تنسيق شرطي"""
import json, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

D = json.load(io.open('data.json', encoding='utf-8'))
ST = D['stations']; MONTHS = D['platform']['months']; IG = D['platform']['integration']
NM = len(MONTHS); NS = len(ST)
OUT = 'القسم-التجاري-البيانات-والخطط.xlsx'

# ── هوية بصرية ──
F = 'Arial'
DARK   = '1B1C1E'   # رأس داكن
ORANGE = 'F18A2B'
ORANGE_L = 'FDEBD8'
GREEN  = '2E8B57'; GREEN_L = 'E3F4EA'
RED    = 'C0392B'; RED_L   = 'FBE6E4'
GOLD   = 'B8860B'; GOLD_L  = 'FFF6DC'
SKY    = '2874A6'; SKY_L   = 'E4F0F8'
GREY   = '6B7280'; BAND    = 'F7F8F9'
INK    = '1F2937'

fill = lambda c: PatternFill('solid', fgColor=c)
HDR_FILL, DARK_FILL = fill(ORANGE), fill(DARK)
BLUE_F  = Font(name=F, size=10, color='0000C0')            # مُدخل
BLACK_F = Font(name=F, size=10, color=INK)                 # معادلة
LINK_F  = Font(name=F, size=10, color=GREEN)               # رابط ورقة
BOLD    = lambda s=10, c=INK: Font(name=F, size=s, bold=True, color=c)
hair = Side(style='hair', color='C8CCD0')
BOX = Border(left=hair, right=hair, top=hair, bottom=hair)
MONEY='#,##0;(#,##0);"–"'; NUM='#,##0;(#,##0);"–"'; DEC2='#,##0.00;(#,##0.00);"–"'; PCT='0.0%'

def sheet(wb, name, cover=False):
    ws = wb.create_sheet(name) if wb.sheetnames != ['Sheet'] or wb['Sheet'].title != 'Sheet' else wb.active
    if ws.title == 'Sheet': ws.title = name
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DARK if cover else ORANGE
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    return ws

def banner(ws, text, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text); c.font = Font(name=F, size=16, bold=True, color='FFFFFF')
    c.fill = DARK_FILL; c.alignment = Alignment(horizontal='right', vertical='center', indent=1)
    ws.row_dimensions[1].height = 34
    for j in range(2, ncols+1): ws.cell(1, j).fill = DARK_FILL
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, sub); c.font = Font(name=F, size=9, italic=True, color=GREY)
    c.alignment = Alignment(horizontal='right', vertical='center', indent=1)
    ws.row_dimensions[2].height = 18
    return 4

def thead(ws, row, headers, widths):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = Font(name=F, size=9.5, bold=True, color='FFFFFF')
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 30
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row+1, 1)
    return row+1

def put(ws, r, c, v, font=None, fmt=None, wrap=False, bg=None, al='right'):
    cell = ws.cell(r, c, v)
    cell.font = font or BLACK_F
    cell.border = BOX
    cell.alignment = Alignment(horizontal=al, vertical='center', wrap_text=wrap, indent=1 if al=='right' else 0)
    if fmt: cell.number_format = fmt
    if bg: cell.fill = fill(bg)
    return cell

def band(ws, r0, r1, ncols):
    for r in range(r0, r1+1):
        if (r-r0) % 2:
            for c in range(1, ncols+1):
                if ws.cell(r, c).fill.fgColor.rgb in (None, '00000000'): ws.cell(r, c).fill = fill(BAND)

def kpi(ws, row, col, span, value, label, color=ORANGE, fmt=None):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c = ws.cell(row, col, value); c.font = Font(name=F, size=18, bold=True, color=color)
    c.alignment = Alignment(horizontal='center', vertical='center')
    if fmt: c.number_format = fmt
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+span-1)
    l = ws.cell(row+1, col, label); l.font = Font(name=F, size=8.5, color=GREY)
    l.alignment = Alignment(horizontal='center', vertical='top')
    top = Side(style='thick', color=color)
    for j in range(col, col+span):
        ws.cell(row, j).border = Border(top=top, left=hair, right=hair)
        ws.cell(row+1, j).border = Border(left=hair, right=hair, bottom=hair)
        ws.cell(row, j).fill = fill('FFFFFF'); ws.cell(row+1, j).fill = fill('FFFFFF')
    ws.row_dimensions[row].height = 30; ws.row_dimensions[row+1].height = 16

def clean_labels(ch, val=False, pct=False):
    """تسميات نظيفة: القيمة أو النسبة فقط — بلا اسم سلسلة/فئة/مفتاح."""
    d = DataLabelList()
    d.showVal = val; d.showPercent = pct
    d.showSerName = False; d.showCatName = False
    d.showLegendKey = False; d.showBubbleSize = False
    ch.dLbls = d
    return ch

def style_chart(ch, title, w=14.5, h=7.5):
    ch.title = title; ch.width = w; ch.height = h; ch.style = 2
    return ch

wb = Workbook()

# ══════════════ 0) ورقة الحسابات (مصدر الرسوم) ══════════════
calc = wb.active; calc.title = 'حسابات'; calc.sheet_state = 'hidden'
calc['A1'] = 'المحطة'; calc['B1'] = 'صافي الربح'; calc['C1'] = 'الزيارات'; calc['D1'] = 'الربح/الزيارة'
for j, s in enumerate(ST):
    f = s['finance']
    exp = f['salaries']+f['oprent']+f['utils']+f['maint']
    net = f['margin']*f['liters'] + f['units']*f['occ']/100*f['rent'] + f['store'] - exp
    calc.cell(2+j, 1, s['station']['name']); calc.cell(2+j, 2, round(net))
    calc.cell(2+j, 3, f.get('visits', 0)); calc.cell(2+j, 4, round(net/f['visits'], 2) if f.get('visits') else 0)
# مزيج الوقود للمحفظة
calc['F1'] = 'المنتج'; calc['G1'] = 'لتر/شهر'
agg = {}
for s in ST:
    for m in s['fuel']['mix']:
        agg[m[0]] = agg.get(m[0], 0) + (m[1] or 0)/100*s['finance']['liters']
for i, (k, v) in enumerate(sorted(agg.items(), key=lambda x: -x[1])):
    calc.cell(2+i, 6, k); calc.cell(2+i, 7, round(v))
NPROD = len(agg)
# الاتجاه الشهري
calc['I1'] = 'الشهر'
for j, s in enumerate(ST): calc.cell(1, 10+j, s['meta']['code'])
for i, mo in enumerate(MONTHS):
    calc.cell(2+i, 9, mo)
    for j, s in enumerate(ST): calc.cell(2+i, 10+j, s['fuel']['monthly'][i] if i < len(s['fuel']['monthly']) else 0)
# خط الأنابيب حسب القناة
calc['N1'] = 'القناة'; calc['O1'] = 'القيمة'
ch_agg = {}
for s in ST:
    for o in s['pipeline']['items']: ch_agg[o['channel']] = ch_agg.get(o['channel'], 0) + (o['value'] or 0)
for i, (k, v) in enumerate(sorted(ch_agg.items(), key=lambda x: -x[1])):
    calc.cell(2+i, 14, k); calc.cell(2+i, 15, v)
NCH = len(ch_agg)
# مبيعات المستأجرين الشهرية (إجمالي)
calc['Q1'] = 'الشهر'; calc['R1'] = 'مبيعات المستأجرين'
for i in range(NM):
    calc.cell(2+i, 17, MONTHS[i])
    calc.cell(2+i, 18, sum((t.get('salesM') or [0]*NM)[i] for s in ST for t in s['tenants']['items'] if t['status']=='مؤجّرة'))

TOT_VIS = sum(s['finance'].get('visits',0) for s in ST)
TOT_NET = sum(calc.cell(2+j, 2).value for j in range(NS))
TOT_PIPE = sum(ch_agg.values())
TOT_FLEET = sum(f['liters'] for s in ST for f in s['fleet']['items'])
TOT_UNITS = sum(t.get('count',1) for s in ST for t in s['tenants']['items'])
VAC_UNITS = sum(t.get('count',1) for s in ST for t in s['tenants']['items'] if t['status']!='مؤجّرة')

# ══════════════ 1) لوحة التحكم ══════════════
ws = sheet(wb, 'لوحة التحكم', cover=True)
for i, w in enumerate([3,11,11,11,11,11,11,11,11,11,11,11,3], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
r = banner(ws, 'القسم التجاري — درب · لوحة التحكم', 'مؤشرات المحفظة والمنافذ · كل الأرقام محسوبة من أوراق البيانات', 13)

kpi(ws, r, 2, 2, NS, 'عدد المنافذ', ORANGE, NUM)
kpi(ws, r, 4, 2, TOT_VIS, 'الزيارات / شهر', SKY, NUM)
kpi(ws, r, 6, 2, TOT_NET, 'صافي المحفظة (ر.س/شهر)', GREEN, MONEY)
kpi(ws, r, 8, 2, round(TOT_NET/TOT_VIS, 2) if TOT_VIS else 0, 'الربح لكل زيارة (ر.س)', ORANGE, DEC2)
kpi(ws, r, 10, 3, TOT_PIPE, 'خط أنابيب المبيعات (ر.س/شهر)', GOLD, MONEY)
r += 3
kpi(ws, r, 2, 2, TOT_FLEET, 'حجم الأساطيل (لتر/شهر)', SKY, NUM)
kpi(ws, r, 4, 2, TOT_UNITS, 'إجمالي الوحدات', ORANGE, NUM)
kpi(ws, r, 6, 2, VAC_UNITS, 'وحدات شاغرة (فرصة)', RED, NUM)
kpi(ws, r, 8, 2, sum(agg.values()), 'إجمالي اللترات / شهر', GREEN, NUM)
kpi(ws, r, 10, 3, TOT_VIS, 'أثر +١ ر.س لكل زيارة (ر.س/شهر)', GOLD, MONEY)
ws.cell(r, 10).comment = Comment('رفع غير الوقود بريال واحد لكل زيارة على مستوى المحفظة.', 'القسم التجاري')
r += 3

ch = BarChart(); ch.type = 'col'
ch.add_data(Reference(calc, min_col=2, min_row=1, max_row=1+NS), titles_from_data=True)
ch.set_categories(Reference(calc, min_col=1, min_row=2, max_row=1+NS))
clean_labels(ch, val=True)
ch.y_axis.numFmt = '#,##0'; ch.legend = None; ch.gapWidth = 60
style_chart(ch, 'صافي الربح لكل منفذ (ر.س/شهر)')
ws.add_chart(ch, f'B{r}')

pie = PieChart()
pie.add_data(Reference(calc, min_col=7, min_row=1, max_row=1+NPROD), titles_from_data=True)
pie.set_categories(Reference(calc, min_col=6, min_row=2, max_row=1+NPROD))
clean_labels(pie, pct=True)
style_chart(pie, 'مزيج منتجات الوقود — المحفظة', 12, 7.5)
ws.add_chart(pie, f'H{r}')
r += 16

ln = LineChart()
ln.add_data(Reference(calc, min_col=10, max_col=9+NS, min_row=1, max_row=1+NM), titles_from_data=True)
ln.set_categories(Reference(calc, min_col=9, min_row=2, max_row=1+NM))
ln.y_axis.numFmt = '0.0'
style_chart(ln, 'اتجاه مبيعات الوقود الشهري (مليون لتر)')
ws.add_chart(ln, f'B{r}')

bh = BarChart(); bh.type = 'bar'
bh.add_data(Reference(calc, min_col=15, min_row=1, max_row=1+NCH), titles_from_data=True)
bh.set_categories(Reference(calc, min_col=14, min_row=2, max_row=1+NCH))
clean_labels(bh, val=True)
bh.x_axis.numFmt = '#,##0'; bh.legend = None; bh.gapWidth = 60
style_chart(bh, 'خط أنابيب المبيعات حسب القناة (ر.س/شهر)', 12, 7.5)
ws.add_chart(bh, f'H{r}')
r += 16

lt = LineChart()
lt.add_data(Reference(calc, min_col=18, min_row=1, max_row=1+NM), titles_from_data=True)
lt.set_categories(Reference(calc, min_col=17, min_row=2, max_row=1+NM))
lt.y_axis.numFmt = '#,##0'; lt.legend = None
style_chart(lt, 'إجمالي مبيعات المستأجرين شهرياً (ر.س)', 27, 7.5)
ws.add_chart(lt, f'B{r}')
r += 16
put(ws, r, 2, 'التقديرات المالية والتشغيلية للاعتماد — تُستبدل بأرقام المالية وسجل الأصول.', Font(name=F, size=9, italic=True, color=GOLD), bg=GOLD_L)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)

# ══════════════ 2) بطاقة الأداء المشتركة ══════════════
ws = sheet(wb, 'بطاقة الأداء المشتركة')
r = banner(ws, 'بطاقة الأداء المشتركة', 'المؤشرات التي تُقاس عليها الإدارات الأربع معاً · مرتبطة بورقة «النموذج المالي»', 8)
r = thead(ws, r, ['الرمز','المنفذ','الزيارات/شهر','الربح/الزيارة','حصة التطبيق','المستهدف','تغطية غير الوقود','ربح المنفذ'],
          [10,28,15,15,14,12,17,17])
sc0 = r
app_t = (IG.get('appTarget') or 5)/100.0
for j, s in enumerate(ST):
    col = get_column_letter(3+j)
    app = next((p[1] for p in s['payment'] if 'تطبيق' in p[0]), 0)
    put(ws, r, 1, s['meta']['code'], BOLD(), al='center')
    put(ws, r, 2, s['station']['name'], BLUE_F)
    put(ws, r, 3, f"='النموذج المالي'!{col}{{VIS}}", LINK_F, NUM)
    put(ws, r, 4, f"='النموذج المالي'!{col}{{PPV}}", LINK_F, DEC2)
    put(ws, r, 5, (app or 0)/100.0, BLUE_F, PCT)
    put(ws, r, 6, app_t, BLUE_F, PCT)
    put(ws, r, 7, f"=IFERROR('النموذج المالي'!{col}{{NF}}/'النموذج المالي'!{col}{{EXP}},\"\")", LINK_F, PCT)
    put(ws, r, 8, f"='النموذج المالي'!{col}{{NET}}", LINK_F, MONEY)
    r += 1
sc1 = r-1
put(ws, r, 1, '', bg=ORANGE_L); put(ws, r, 2, 'المحفظة', BOLD(11), bg=ORANGE_L)
put(ws, r, 3, f'=SUM(C{sc0}:C{sc1})', BOLD(11), NUM, bg=ORANGE_L)
put(ws, r, 4, f'=IFERROR(H{r}/C{r},"")', BOLD(11), DEC2, bg=ORANGE_L)
for c in (5,6,7): put(ws, r, c, '', bg=ORANGE_L)
put(ws, r, 8, f'=SUM(H{sc0}:H{sc1})', BOLD(11), MONEY, bg=ORANGE_L)
ws.row_dimensions[r].height = 22
SCORE_TOTAL = r
ws.conditional_formatting.add(f'C{sc0}:C{sc1}', DataBarRule(start_type='num', start_value=0, end_type='max', color=SKY))
ws.conditional_formatting.add(f'D{sc0}:D{sc1}', ColorScaleRule(start_type='min', start_color='FBE6E4', end_type='max', end_color='CDEBDA'))
ws.conditional_formatting.add(f'E{sc0}:E{sc1}', CellIsRule(operator='lessThan', formula=[f'$F${sc0}'], fill=fill(RED_L), font=Font(name=F, size=10, bold=True, color=RED)))
ws.conditional_formatting.add(f'E{sc0}:E{sc1}', CellIsRule(operator='greaterThanOrEqual', formula=[f'$F${sc0}'], fill=fill(GREEN_L), font=Font(name=F, size=10, bold=True, color=GREEN)))
ws.conditional_formatting.add(f'G{sc0}:G{sc1}', CellIsRule(operator='greaterThanOrEqual', formula=['1'], fill=fill(GREEN_L), font=Font(name=F, size=10, bold=True, color=GREEN)))
ws.conditional_formatting.add(f'G{sc0}:G{sc1}', CellIsRule(operator='lessThan', formula=['1'], fill=fill(GOLD_L), font=Font(name=F, size=10, bold=True, color=GOLD)))
ws.conditional_formatting.add(f'H{sc0}:H{sc1}', DataBarRule(start_type='num', start_value=0, end_type='max', color=GREEN))
r += 2
put(ws, r, 2, 'أثر رفع غير الوقود ١ ر.س لكل زيارة', BOLD(10, GOLD), bg=GOLD_L)
put(ws, r, 3, f'=C{SCORE_TOTAL}', BOLD(11, GOLD), MONEY, bg=GOLD_L)
put(ws, r, 4, 'ر.س/شهر — المحفظة', Font(name=F, size=9, italic=True, color=GREY), bg=GOLD_L)
for c in (5,6,7,8): put(ws, r, c, '', bg=GOLD_L)

# ══════════════ 3) النموذج المالي ══════════════
ws = sheet(wb, 'النموذج المالي')
r = banner(ws, 'النموذج المالي لكل منفذ (شهرياً)', 'الأزرق = مُدخل يُعدَّل · الأسود = معادلة · عدّل أي مُدخل ليعاد الحساب', 3+NS)
hr = r
put(ws, hr, 1, 'البند', Font(name=F, size=9.5, bold=True, color='FFFFFF'), bg=ORANGE, al='center')
put(ws, hr, 2, 'الوحدة', Font(name=F, size=9.5, bold=True, color='FFFFFF'), bg=ORANGE, al='center')
for j, s in enumerate(ST):
    put(ws, hr, 3+j, f"{s['meta']['code']}\n{s['station']['name']}", Font(name=F, size=9.5, bold=True, color='FFFFFF'), bg=ORANGE, al='center').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
put(ws, hr, 3+NS, 'المحفظة', Font(name=F, size=9.5, bold=True, color='FFFFFF'), bg=DARK, al='center')
ws.row_dimensions[hr].height = 34
ws.column_dimensions['A'].width = 32; ws.column_dimensions['B'].width = 13
for j in range(NS+1): ws.column_dimensions[get_column_letter(3+j)].width = 18
ws.freeze_panes = ws.cell(hr+1, 3)
r = hr+1
def sec(label, color=ORANGE_L):
    global r
    put(ws, r, 1, label, BOLD(10), bg=color)
    for j in range(NS+2): put(ws, r, 2+j, '', bg=color)
    ws.row_dimensions[r].height = 20; r += 1
def inrow(label, unit, key, fmt=MONEY):
    global r
    put(ws, r, 1, label); put(ws, r, 2, unit, Font(name=F, size=8.5, color=GREY), al='center')
    for j, s in enumerate(ST): put(ws, r, 3+j, s['finance'].get(key, 0), BLUE_F, fmt)
    put(ws, r, 3+NS, '', bg='FFFFFF'); r += 1; return r-1
def frow(label, unit, tmpl, fmt=MONEY, bold=False, bg=None, total=True):
    global r
    fnt = BOLD() if bold else BLACK_F
    put(ws, r, 1, label, fnt, bg=bg); put(ws, r, 2, unit, Font(name=F, size=8.5, color=GREY), al='center', bg=bg)
    for j in range(NS): put(ws, r, 3+j, tmpl.format(c=get_column_letter(3+j)), fnt, fmt, bg=bg)
    if total:
        cols = [get_column_letter(3+j) for j in range(NS)]
        put(ws, r, 3+NS, '=' + '+'.join(f'{c}{r}' for c in cols), BOLD(), fmt, bg=bg or 'F0F1F2')
    else:
        put(ws, r, 3+NS, '', bg=bg or 'FFFFFF')
    r += 1; return r-1
sec('⛽ الوقود')
R_LIT = inrow('اللترات', 'لتر/شهر', 'liters', NUM)
R_MRG = inrow('هامش الوقود', 'ر.س/لتر', 'margin', DEC2)
R_FI  = frow('دخل الوقود', 'ر.س', '={c}'+str(R_LIT)+'*{c}'+str(R_MRG), bold=True)
sec('🏪 التأجير والخدمات')
R_UN  = inrow('عدد الوحدات', 'وحدة', 'units', NUM)
R_OC  = inrow('نسبة الإشغال', '%', 'occ', NUM)
R_RT  = inrow('متوسط الإيجار للوحدة', 'ر.س', 'rent')
R_STO = inrow('دخل المتجر/الخدمات', 'ر.س', 'store')
R_RI  = frow('دخل التأجير', 'ر.س', '={c}'+str(R_UN)+'*{c}'+str(R_OC)+'/100*{c}'+str(R_RT), bold=True)
R_NF  = frow('إجمالي غير الوقود', 'ر.س', '={c}'+str(R_RI)+'+{c}'+str(R_STO), bold=True)
sec('💸 المصاريف')
R_SAL = inrow('رواتب', 'ر.س', 'salaries')
R_OPR = inrow('إيجار/تشغيل', 'ر.س', 'oprent')
R_UTL = inrow('كهرباء وماء', 'ر.س', 'utils')
R_MNT = inrow('صيانة وأخرى', 'ر.س', 'maint')
R_EXP = frow('إجمالي المصاريف', 'ر.س', '=SUM({c}'+str(R_SAL)+':{c}'+str(R_MNT)+')', bold=True)
sec('🏭 التشغيل')
R_WRK = inrow('عدد العمّال', 'عامل', 'workers', NUM)
R_TNK = inrow('سعة الخزان', 'لتر', 'tankCapacity', NUM)
R_PMP = inrow('عدد المضخات', 'مضخة', 'pumps', NUM)
R_VIS = inrow('الزيارات', 'زيارة/شهر', 'visits', NUM)
sec('📊 النتائج', GREEN_L)
R_INC = frow('إجمالي الدخل', 'ر.س', '={c}'+str(R_FI)+'+{c}'+str(R_NF), bold=True)
R_NET = frow('صافي الربح/الخسارة', 'ر.س', '={c}'+str(R_INC)+'-{c}'+str(R_EXP), bold=True, bg=ORANGE_L)
R_PPV = frow('الربح لكل زيارة', 'ر.س', '=IFERROR({c}'+str(R_NET)+'/{c}'+str(R_VIS)+',"")', DEC2, bold=True, total=False)
frow('تكلفة التشغيل/لتر', 'ر.س', '=IFERROR({c}'+str(R_EXP)+'/{c}'+str(R_LIT)+',"")', '#,##0.000', total=False)
frow('تكلفة الرواتب/عامل', 'ر.س', '=IFERROR({c}'+str(R_SAL)+'/{c}'+str(R_WRK)+',"")', MONEY, total=False)
R_COV = frow('تغطية غير الوقود للمصاريف', '%', '=IFERROR({c}'+str(R_NF)+'/{c}'+str(R_EXP)+',"")', PCT, total=False)
frow('أيام تغطية الخزان', 'يوم', '=IFERROR({c}'+str(R_TNK)+'/({c}'+str(R_LIT)+'/30),"")', DEC2, total=False)
frow('لتر لكل مضخة', 'لتر/شهر', '=IFERROR({c}'+str(R_LIT)+'/{c}'+str(R_PMP)+',"")', NUM, total=False)
frow('نقطة التعادل', 'لتر', '=IFERROR(MAX(0,{c}'+str(R_EXP)+'-{c}'+str(R_NF)+')/{c}'+str(R_MRG)+',"")', NUM, total=False)
cols = [get_column_letter(3+j) for j in range(NS)]
put(ws, R_PPV, 3+NS, f'=IFERROR({get_column_letter(3+NS)}{R_NET}/({"+".join(f"{c}{R_VIS}" for c in cols)}),"")', BOLD(), DEC2, bg='F0F1F2')
rng = f'{get_column_letter(3)}{R_NET}:{get_column_letter(2+NS)}{R_NET}'
ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['0'], fill=fill(RED_L), font=Font(name=F, size=10, bold=True, color=RED)))
ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=['0'], font=Font(name=F, size=10, bold=True, color=GREEN)))

# اربط بطاقة الأداء بأرقام الصفوف الفعلية
sc = wb['بطاقة الأداء المشتركة']
for rr in range(1, sc.max_row+1):
    for cc in range(1, sc.max_column+1):
        v = sc.cell(rr, cc).value
        if isinstance(v, str) and '{' in v:
            sc.cell(rr, cc).value = v.replace('{VIS}', str(R_VIS)).replace('{PPV}', str(R_PPV))\
                .replace('{NF}', str(R_NF)).replace('{EXP}', str(R_EXP)).replace('{NET}', str(R_NET))

# ══════════════ 4) بطاقات المنافذ ══════════════
ws = sheet(wb, 'بطاقات المنافذ')
r = banner(ws, 'بطاقات المنافذ', 'بيانات الموقع والتشغيل لكل محطة', 12)
r = thead(ws, r, ['الرمز','اسم المحطة','المدينة','نمط الموقع','الحجم','المساحة','خدمة ذاتية','المضخات','سعة الخزان','العمّال','الزيارات/شهر','الوصف'],
          [10,26,12,24,14,14,12,11,14,10,13,58])
d0 = r
dv = DataValidation(type='list', formula1='"لا,نعم,الاثنتان"', allow_blank=True); ws.add_data_validation(dv)
for s in ST:
    m, f = s['meta'], s['finance']
    put(ws, r, 1, m['code'], BOLD(), al='center'); put(ws, r, 2, s['station']['name'], BLUE_F, wrap=True)
    put(ws, r, 3, m['city'], BLUE_F, al='center'); put(ws, r, 4, m['type'], BLUE_F, wrap=True)
    put(ws, r, 5, m['size'], BLUE_F, al='center'); put(ws, r, 6, m['area'], BLUE_F, al='center')
    put(ws, r, 7, m.get('selfService','لا'), BLUE_F, al='center'); dv.add(ws.cell(r, 7))
    put(ws, r, 8, f.get('pumps',0), BLUE_F, NUM, al='center'); put(ws, r, 9, f.get('tankCapacity',0), BLUE_F, NUM)
    put(ws, r,10, f.get('workers',0), BLUE_F, NUM, al='center'); put(ws, r,11, f.get('visits',0), BLUE_F, NUM)
    put(ws, r,12, s['station']['tagline'], BLUE_F, wrap=True)
    ws.row_dimensions[r].height = 44; r += 1
band(ws, d0, r-1, 12); ws.auto_filter.ref = f'A{d0-1}:L{r-1}'
put(ws, r+1, 1, 'أرقام التشغيل تقديرات مبدئية — تُعتمد من سجل الأصول والموارد البشرية.', Font(name=F, size=9, italic=True, color=GOLD), bg=GOLD_L)
ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=12)

# ══════════════ 5) شرائح العملاء ══════════════
ws = sheet(wb, 'شرائح العملاء')
r = banner(ws, 'شرائح العملاء', 'مَن نخدمه في كل منفذ · ما يحتاجه · العرض المقدّم له', 6)
r = thead(ws, r, ['الرمز','الشريحة','الحصة/الوزن','ملف الشريحة','ما تحتاجه','العرض المقدّم'], [10,24,20,44,38,44])
d0 = r
for s in ST:
    for sg in s['segments']:
        put(ws, r, 1, s['meta']['code'], al='center'); put(ws, r, 2, f"{sg['ic']} {sg['t']}", BOLD(), wrap=True)
        put(ws, r, 3, sg.get('share',''), BLUE_F, wrap=True); put(ws, r, 4, sg.get('profile',''), BLUE_F, wrap=True)
        put(ws, r, 5, sg.get('need',''), BLUE_F, wrap=True); put(ws, r, 6, sg.get('offer',''), BLUE_F, wrap=True)
        ws.row_dimensions[r].height = 42; r += 1
band(ws, d0, r-1, 6); ws.auto_filter.ref = f'A{d0-1}:F{r-1}'

# ══════════════ 6) المحيط التنافسي ══════════════
ws = sheet(wb, 'المحيط التنافسي')
r = banner(ws, 'المحيط التنافسي وبطاقة السيناريو', 'موقع كل محطة على الطريق · المنافسون · الميزة والفجوة', 4)
for i, w in enumerate([24,44,30,30], 1): ws.column_dimensions[get_column_letter(i)].width = w
for s in ST:
    put(ws, r, 1, f"{s['meta']['code']} — {s['station']['name']}", Font(name=F, size=11, bold=True, color='FFFFFF'), bg=DARK)
    for j in range(2, 5): put(ws, r, j, '', bg=DARK)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); ws.row_dimensions[r].height = 24; r += 1
    rows = [('الطريق', s['scenario']['road']), ('المحطة السابقة', s['scenario']['before']),
            ('المحطة التالية', s['scenario']['after']), ('كثافة المنافسة', s['scenario']['density']),
            ('العلامات في المحطة', ' | '.join(f"{a}: {b}" for a, b in s['us'])),
            ('المنافسون', ' | '.join(f"{a}: {b}" for a, b in s['them'])),
            ('براندات مقترحة', ' | '.join(f"{a} ({b})" for a, b in s['suggested']))]
    for k, v in rows:
        put(ws, r, 1, k, BOLD(9.5), bg=ORANGE_L); put(ws, r, 2, v, BLUE_F, wrap=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4); ws.row_dimensions[r].height = 30; r += 1
    put(ws, r, 1, '✅ الميزة التنافسية', BOLD(9.5, GREEN), bg=GREEN_L)
    put(ws, r, 2, s['advantage'], Font(name=F, size=10, color=GREEN), wrap=True, bg=GREEN_L)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4); ws.row_dimensions[r].height = 30; r += 1
    put(ws, r, 1, '⚠️ الفجوة', BOLD(9.5, RED), bg=RED_L)
    put(ws, r, 2, s['gap'], Font(name=F, size=10, color=RED), wrap=True, bg=RED_L)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4); ws.row_dimensions[r].height = 30; r += 2

# ══════════════ 7) وحدات التأجير ══════════════
ws = sheet(wb, 'وحدات التأجير')
r = banner(ws, 'وحدات التأجير — مكوّنات المحطة', 'كل وحدة: النوع · العدد · الخدمة · النوع الفرعي · الحالة', 9)
r = thead(ws, r, ['الرمز','اسم الوحدة/المستأجر','نوع الوحدة','العدد','قسم الخدمة','النوع الفرعي','الحالة','الإيجار/شهر','المساحة م²'],
          [10,30,16,9,20,20,12,15,12])
d0 = r
dv_u = DataValidation(type='list', formula1='"كشك,محل تجاري,دريف ثرو,سوبر ماركت,عربة/كارت,أخرى"', allow_blank=True); ws.add_data_validation(dv_u)
dv_s = DataValidation(type='list', formula1='"مؤجّرة,شاغرة"', allow_blank=True); ws.add_data_validation(dv_s)
for s in ST:
    for t in s['tenants']['items']:
        put(ws, r, 1, s['meta']['code'], al='center'); put(ws, r, 2, t['name'], BLUE_F, wrap=True)
        put(ws, r, 3, t['unitType'], BLUE_F, al='center'); dv_u.add(ws.cell(r, 3))
        put(ws, r, 4, t.get('count',1), BLUE_F, NUM, al='center')
        put(ws, r, 5, t['serviceCat'], BLUE_F, al='center'); put(ws, r, 6, t['subType'], BLUE_F, al='center')
        put(ws, r, 7, t['status'], BLUE_F, al='center'); dv_s.add(ws.cell(r, 7))
        put(ws, r, 8, t.get('rent',0), BLUE_F, MONEY); put(ws, r, 9, t.get('area',0), BLUE_F, NUM, al='center')
        r += 1
d1 = r-1
band(ws, d0, d1, 9); ws.auto_filter.ref = f'A{d0-1}:I{d1}'
put(ws, r, 1, 'الإجمالي', BOLD(), bg=ORANGE_L)
for c in (2,3,5,6,7): put(ws, r, c, '', bg=ORANGE_L)
put(ws, r, 4, f'=SUM(D{d0}:D{d1})', BOLD(), NUM, bg=ORANGE_L, al='center')
put(ws, r, 8, f'=SUM(H{d0}:H{d1})', BOLD(), MONEY, bg=ORANGE_L)
put(ws, r, 9, f'=SUM(I{d0}:I{d1})', BOLD(), NUM, bg=ORANGE_L, al='center')
ws.conditional_formatting.add(f'G{d0}:G{d1}', FormulaRule(formula=[f'$G{d0}="شاغرة"'], fill=fill(RED_L), font=Font(name=F, size=10, bold=True, color=RED)))
ws.conditional_formatting.add(f'G{d0}:G{d1}', FormulaRule(formula=[f'$G{d0}="مؤجّرة"'], fill=fill(GREEN_L), font=Font(name=F, size=10, color=GREEN)))

# ══════════════ 8) مبيعات المستأجرين ══════════════
ws = sheet(wb, 'مبيعات المستأجرين')
r = banner(ws, 'مبيعات المستأجرين الشهرية', 'نسبة الإيجار = الإيجار ÷ مبيعات آخر شهر · الصحّي ≤ ١٥٪', 3+NM+4)
r = thead(ws, r, ['الرمز','المستأجر','قسم الخدمة'] + MONTHS + ['الإيجار/شهر','المساحة م²','نسبة الإيجار','مبيعات/م²'],
          [10,26,18] + [14]*NM + [14,12,14,13])
d0 = r
cL = get_column_letter(3+NM); cR = get_column_letter(4+NM); cA = get_column_letter(5+NM)
for s in ST:
    for t in s['tenants']['items']:
        if t['status'] != 'مؤجّرة': continue
        put(ws, r, 1, s['meta']['code'], al='center'); put(ws, r, 2, t['name'], BLUE_F, wrap=True)
        put(ws, r, 3, t['serviceCat'], BLUE_F, al='center')
        sm = t.get('salesM') or [0]*NM
        for i in range(NM): put(ws, r, 4+i, sm[i] if i < len(sm) else 0, BLUE_F, MONEY)
        put(ws, r, 4+NM, t.get('rent',0), BLUE_F, MONEY); put(ws, r, 5+NM, t.get('area',0), BLUE_F, NUM, al='center')
        put(ws, r, 6+NM, f'=IFERROR({cR}{r}/{cL}{r},"")', BLACK_F, PCT)
        put(ws, r, 7+NM, f'=IFERROR({cL}{r}/{cA}{r},"")', BLACK_F, NUM)
        r += 1
d1 = r-1
band(ws, d0, d1, 7+NM); ws.auto_filter.ref = f'A{d0-1}:{get_column_letter(7+NM)}{d1}'
put(ws, r, 1, '', bg=ORANGE_L); put(ws, r, 2, 'الإجمالي', BOLD(), bg=ORANGE_L); put(ws, r, 3, '', bg=ORANGE_L)
for i in range(NM):
    cl = get_column_letter(4+i); put(ws, r, 4+i, f'=SUM({cl}{d0}:{cl}{d1})', BOLD(), MONEY, bg=ORANGE_L)
put(ws, r, 4+NM, f'=SUM({cR}{d0}:{cR}{d1})', BOLD(), MONEY, bg=ORANGE_L)
put(ws, r, 5+NM, f'=SUM({cA}{d0}:{cA}{d1})', BOLD(), NUM, bg=ORANGE_L, al='center')
put(ws, r, 6+NM, f'=IFERROR({cR}{r}/{cL}{r},"")', BOLD(), PCT, bg=ORANGE_L)
put(ws, r, 7+NM, '', bg=ORANGE_L)
rr = f'{get_column_letter(6+NM)}{d0}:{get_column_letter(6+NM)}{d1}'
ws.conditional_formatting.add(rr, CellIsRule(operator='greaterThan', formula=['0.15'], fill=fill(RED_L), font=Font(name=F, size=10, bold=True, color=RED)))
ws.conditional_formatting.add(rr, CellIsRule(operator='lessThanOrEqual', formula=['0.15'], fill=fill(GREEN_L), font=Font(name=F, size=10, bold=True, color=GREEN)))
for i in range(NM):
    cl = get_column_letter(4+i)
    ws.conditional_formatting.add(f'{cl}{d0}:{cl}{d1}', DataBarRule(start_type='num', start_value=0, end_type='max', color=ORANGE))

# ══════════════ 9) مبيعات الوقود ══════════════
ws = sheet(wb, 'مبيعات الوقود')
r = banner(ws, 'مبيعات الوقود — مزيج المنتجات', 'لتر المنتج = الحصة × إجمالي اللترات · الربح = لتر المنتج × الهامش', 7)
r = thead(ws, r, ['الرمز','المنتج','الحصة %','إجمالي اللترات/شهر','هامش ر.س/لتر','لتر المنتج/شهر','ربح المنتج/شهر'],
          [10,18,12,20,15,19,19])
d0 = r
dv_p = DataValidation(type='list', formula1='"بنزين 91,بنزين 95,بنزين 98,ديزل"', allow_blank=True); ws.add_data_validation(dv_p)
for s in ST:
    f = s['finance']
    for mx in s['fuel']['mix']:
        put(ws, r, 1, s['meta']['code'], al='center')
        put(ws, r, 2, mx[0], BLUE_F, al='center'); dv_p.add(ws.cell(r, 2))
        put(ws, r, 3, (mx[1] or 0)/100.0, BLUE_F, PCT); put(ws, r, 4, f.get('liters',0), BLUE_F, NUM)
        put(ws, r, 5, f.get('margin',0), BLUE_F, DEC2)
        put(ws, r, 6, f'=C{r}*D{r}', BLACK_F, NUM); put(ws, r, 7, f'=F{r}*E{r}', BLACK_F, MONEY)
        r += 1
d1 = r-1
band(ws, d0, d1, 7); ws.auto_filter.ref = f'A{d0-1}:G{d1}'
ws.conditional_formatting.add(f'G{d0}:G{d1}', DataBarRule(start_type='num', start_value=0, end_type='max', color=GREEN))
put(ws, r, 1, '', bg=ORANGE_L); put(ws, r, 2, 'الإجمالي', BOLD(), bg=ORANGE_L)
for c in (3,4,5): put(ws, r, c, '', bg=ORANGE_L)
put(ws, r, 6, f'=SUM(F{d0}:F{d1})', BOLD(), NUM, bg=ORANGE_L)
put(ws, r, 7, f'=SUM(G{d0}:G{d1})', BOLD(), MONEY, bg=ORANGE_L)
r += 2
put(ws, r, 1, 'الاتجاه الشهري (مليون لتر)', BOLD(11)); r += 1
r = thead(ws, r, ['الرمز'] + MONTHS, [10] + [14]*NM)
m0 = r
for s in ST:
    put(ws, r, 1, s['meta']['code'], BOLD(), al='center')
    for i, v in enumerate(s['fuel']['monthly'][:NM]): put(ws, r, 2+i, v, BLUE_F, DEC2)
    r += 1
band(ws, m0, r-1, 1+NM)
for i in range(NM):
    cl = get_column_letter(2+i)
    ws.conditional_formatting.add(f'{cl}{m0}:{cl}{r-1}', DataBarRule(start_type='num', start_value=0, end_type='max', color=SKY))

# ══════════════ 10) فرص B2B ══════════════
ws = sheet(wb, 'فرص B2B')
r = banner(ws, 'فرص المبيعات — الجوار التجاري (B2B)', 'الجهات المجاورة كأهداف بيع · القناة والقيمة والحالة والأولوية', 7)
r = thead(ws, r, ['الرمز','الجهة','النشاط / الوصف','قناة البيع','القيمة ر.س/شهر','الحالة','الأولوية'], [10,30,50,20,17,14,12])
d0 = r
dv_st = DataValidation(type='list', formula1='"فرصة,تواصل,عرض مقدّم,متعاقد"', allow_blank=True); ws.add_data_validation(dv_st)
dv_pr = DataValidation(type='list', formula1='"عالية,متوسطة,منخفضة"', allow_blank=True); ws.add_data_validation(dv_pr)
for s in ST:
    for o in s['pipeline']['items']:
        put(ws, r, 1, s['meta']['code'], al='center'); put(ws, r, 2, o['name'], BOLD(9.5), wrap=True)
        put(ws, r, 3, o['type'], BLUE_F, wrap=True); put(ws, r, 4, o['channel'], BLUE_F, al='center')
        put(ws, r, 5, o['value'], BLUE_F, MONEY)
        put(ws, r, 6, o['status'], BLUE_F, al='center'); dv_st.add(ws.cell(r, 6))
        put(ws, r, 7, o['pri'], BLUE_F, al='center'); dv_pr.add(ws.cell(r, 7))
        ws.row_dimensions[r].height = 30; r += 1
d1 = r-1
band(ws, d0, d1, 7); ws.auto_filter.ref = f'A{d0-1}:G{d1}'
ws.conditional_formatting.add(f'E{d0}:E{d1}', DataBarRule(start_type='num', start_value=0, end_type='max', color=GOLD))
ws.conditional_formatting.add(f'F{d0}:F{d1}', FormulaRule(formula=[f'$F{d0}="متعاقد"'], fill=fill(GREEN_L), font=Font(name=F, size=10, bold=True, color=GREEN)))
ws.conditional_formatting.add(f'G{d0}:G{d1}', FormulaRule(formula=[f'$G{d0}="عالية"'], fill=fill(ORANGE_L), font=Font(name=F, size=10, bold=True, color=GOLD)))
put(ws, r, 1, '', bg=ORANGE_L); put(ws, r, 2, 'إجمالي خط الأنابيب', BOLD(), bg=ORANGE_L)
for c in (3,4,6,7): put(ws, r, c, '', bg=ORANGE_L)
put(ws, r, 5, f'=SUM(E{d0}:E{d1})', BOLD(11), MONEY, bg=ORANGE_L)
put(ws, r+2, 2, 'المتعاقد', BOLD()); put(ws, r+2, 5, f'=SUMIF(F{d0}:F{d1},"متعاقد",E{d0}:E{d1})', BLACK_F, MONEY)
put(ws, r+3, 2, 'فرص مفتوحة', BOLD()); put(ws, r+3, 5, f'=E{r}-E{r+2}', BLACK_F, MONEY)

# ══════════════ 11) عقود الأساطيل ══════════════
ws = sheet(wb, 'عقود الأساطيل')
r = banner(ws, 'عقود وقود الأساطيل (B2B)', 'الفئات وآليات التعاقد والجهات المستهدفة والحجم والربح المتوقّع', 9)
r = thead(ws, r, ['الرمز','فئة الأسطول','المنتج','آلية التعاقد','الجهات المستهدفة','الحجم لتر/شهر','هامش ر.س/لتر','ربح متوقّع/شهر','الحالة'],
          [10,26,15,58,48,16,14,17,13])
d0 = r
dv_f = DataValidation(type='list', formula1='"فرصة,تواصل,عرض مقدّم,متعاقد"', allow_blank=True); ws.add_data_validation(dv_f)
for s in ST:
    mg = s['finance'].get('margin', 0)
    for f_ in s['fleet']['items']:
        put(ws, r, 1, s['meta']['code'], al='center'); put(ws, r, 2, f_['seg'], BOLD(9.5), wrap=True)
        put(ws, r, 3, f_['product'], BLUE_F, al='center'); put(ws, r, 4, f_['mechanism'], BLUE_F, wrap=True)
        put(ws, r, 5, f_['targets'], BLUE_F, wrap=True); put(ws, r, 6, f_['liters'], BLUE_F, NUM)
        put(ws, r, 7, mg, BLUE_F, DEC2); put(ws, r, 8, f'=F{r}*G{r}', BLACK_F, MONEY)
        put(ws, r, 9, f_['status'], BLUE_F, al='center'); dv_f.add(ws.cell(r, 9))
        ws.row_dimensions[r].height = 56; r += 1
d1 = r-1
band(ws, d0, d1, 9); ws.auto_filter.ref = f'A{d0-1}:I{d1}'
ws.conditional_formatting.add(f'H{d0}:H{d1}', DataBarRule(start_type='num', start_value=0, end_type='max', color=SKY))
put(ws, r, 1, '', bg=ORANGE_L); put(ws, r, 2, 'الإجمالي', BOLD(), bg=ORANGE_L)
for c in (3,4,5,7,9): put(ws, r, c, '', bg=ORANGE_L)
put(ws, r, 6, f'=SUM(F{d0}:F{d1})', BOLD(), NUM, bg=ORANGE_L)
put(ws, r, 8, f'=SUM(H{d0}:H{d1})', BOLD(11), MONEY, bg=ORANGE_L)

# ══════════════ 12) المكس التسويقي ══════════════
ws = sheet(wb, 'المكس التسويقي')
r = banner(ws, 'المكس التسويقي (٧Ps)', 'حسب نمط الموقع · سعر الوقود مقنَّن فالمكس على المنتج والمكان والناس', 7)
r = thead(ws, r, ['الرمز','العنصر','القرار الاستراتيجي'], [10,22,96])
d0 = r
for s in ST:
    for m in s['marketing']['mix']:
        put(ws, r, 1, s['meta']['code'], al='center'); put(ws, r, 2, f"{m['ic']} {m['p']}", BOLD(9.5), al='center')
        put(ws, r, 3, m['d'], BLUE_F, wrap=True); ws.row_dimensions[r].height = 40; r += 1
band(ws, d0, r-1, 3); ws.auto_filter.ref = f'A{d0-1}:C{r-1}'
r += 1
put(ws, r, 1, 'توزيع ميزانية الترويج', BOLD(11)); r += 1
r = thead(ws, r, ['الرمز','البند','الحصة %'], [10,38,14])
for s in ST:
    b0 = r
    for b in s['marketing']['budget']:
        put(ws, r, 1, s['meta']['code'], al='center'); put(ws, r, 2, b[0], BLUE_F)
        put(ws, r, 3, (b[1] or 0)/100.0, BLUE_F, PCT); r += 1
    ws.conditional_formatting.add(f'C{b0}:C{r-1}', DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color=ORANGE))
    put(ws, r, 1, '', bg=ORANGE_L); put(ws, r, 2, 'المجموع (يجب أن يساوي ١٠٠٪)', BOLD(), bg=ORANGE_L)
    put(ws, r, 3, f'=SUM(C{b0}:C{r-1})', BOLD(), PCT, bg=ORANGE_L)
    ws.conditional_formatting.add(f'C{r}:C{r}', CellIsRule(operator='notEqual', formula=['1'], fill=fill(RED_L), font=Font(name=F, size=10, bold=True, color=RED)))
    r += 2
put(ws, r, 1, 'الحملات', BOLD(11)); r += 1
c0 = thead(ws, r, ['الرمز','الحملة','النوع','المموّل','التكلفة ر.س','الأثر %','الحالة'], [10,38,24,16,15,12,14]); r = c0
dv_c = DataValidation(type='list', formula1='"مقترحة,جارية,منفّذة,مؤجّلة"', allow_blank=True); ws.add_data_validation(dv_c)
dv_fu = DataValidation(type='list', formula1='"التسويق,الشريك,مشترك,المستأجر"', allow_blank=True); ws.add_data_validation(dv_fu)
for s in ST:
    for c in s['marketing']['campaigns']:
        put(ws, r, 1, s['meta']['code'], al='center'); put(ws, r, 2, c['name'], BLUE_F, wrap=True)
        put(ws, r, 3, c['type'], BLUE_F, wrap=True)
        put(ws, r, 4, c['funder'], BLUE_F, al='center'); dv_fu.add(ws.cell(r, 4))
        put(ws, r, 5, c['cost'], BLUE_F, MONEY); put(ws, r, 6, (c.get('lift') or 0)/100.0, BLUE_F, PCT)
        put(ws, r, 7, c['status'], BLUE_F, al='center'); dv_c.add(ws.cell(r, 7))
        r += 1
c1 = r-1
band(ws, c0, c1, 7)
ws.conditional_formatting.add(f'D{c0}:D{c1}', FormulaRule(formula=[f'OR($D{c0}="الشريك",$D{c0}="المستأجر")'], fill=fill(GREEN_L), font=Font(name=F, size=10, bold=True, color=GREEN)))
put(ws, r, 2, 'إجمالي تكلفة الحملات', BOLD(), bg=ORANGE_L)
for c in (1,3,4,6,7): put(ws, r, c, '', bg=ORANGE_L)
put(ws, r, 5, f'=SUM(E{c0}:E{c1})', BOLD(), MONEY, bg=ORANGE_L)
put(ws, r+1, 2, 'منها على ميزانية التسويق', BOLD()); put(ws, r+1, 5, f'=SUMIF(D{c0}:D{c1},"التسويق",E{c0}:E{c1})', BLACK_F, MONEY)
put(ws, r+2, 2, 'مموّل من الشركاء/المستأجرين', BOLD(10, GREEN), bg=GREEN_L)
put(ws, r+2, 5, f'=E{r}-E{r+1}', BOLD(10, GREEN), MONEY, bg=GREEN_L)

# ══════════════ 13) خطة التنفيذ ══════════════
ws = sheet(wb, 'خطة التنفيذ')
r = banner(ws, 'خطة التنفيذ — التكامل بين الإدارات', 'الركائز · دورة النمو · قواعد فضّ التعارض · التسلسل التنفيذي', 6)
put(ws, r, 1, IG.get('principle',''), Font(name=F, size=13, bold=True, color='FFFFFF'), bg=ORANGE, al='center')
for j in range(2, 7): put(ws, r, j, '', bg=ORANGE)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); ws.row_dimensions[r].height = 30; r += 1
put(ws, r, 1, IG.get('unit',''), Font(name=F, size=10, color=INK), wrap=True, bg=GOLD_L)
for j in range(2, 7): put(ws, r, j, '', bg=GOLD_L)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); ws.row_dimensions[r].height = 30; r += 2
put(ws, r, 1, 'الركائز الهيكلية', BOLD(11)); r += 1
r = thead(ws, r, ['الركيزة','التفصيل'], [32,90])
p0 = r
for p in IG['pillars']:
    put(ws, r, 1, p['t'], BOLD(9.5), wrap=True); put(ws, r, 2, p['d'], BLUE_F, wrap=True)
    ws.row_dimensions[r].height = 32; r += 1
band(ws, p0, r-1, 2); r += 1
put(ws, r, 1, 'دورة النمو', BOLD(11)); r += 1
r = thead(ws, r, ['#','المرحلة','الأثر'], [7,32,72])
y0 = r
for i, c in enumerate(IG['cycle'], 1):
    put(ws, r, 1, i, BOLD(), al='center'); put(ws, r, 2, f"{c['ic']} {c['t']}", BOLD(9.5))
    put(ws, r, 3, c['d'], BLUE_F, wrap=True); r += 1
band(ws, y0, r-1, 3); r += 1
put(ws, r, 1, 'قواعد فضّ التعارض', BOLD(11)); r += 1
r = thead(ws, r, ['التعارض','القاعدة الحاكمة'], [48,74])
for a, b in IG['rules']:
    put(ws, r, 1, a, Font(name=F, size=9.5, color=RED), wrap=True, bg=RED_L)
    put(ws, r, 2, b, Font(name=F, size=9.5, color=GREEN), wrap=True, bg=GREEN_L)
    ws.row_dimensions[r].height = 32; r += 1
r += 1
put(ws, r, 1, 'التسلسل التنفيذي', BOLD(11)); r += 1
r = thead(ws, r, ['المرحلة','التركيز','المبادرات','المبرّر','المسؤول','الحالة'], [18,22,50,34,22,16])
ph0 = r
dv_ph = DataValidation(type='list', formula1='"لم تبدأ,جارية,منجزة,مؤجّلة"', allow_blank=True); ws.add_data_validation(dv_ph)
for p in IG['phases']:
    put(ws, r, 1, p['p'], BOLD(9.5, GOLD), al='center'); put(ws, r, 2, p['t'], BOLD(9.5))
    put(ws, r, 3, p['d'], BLUE_F, wrap=True); put(ws, r, 4, p['why'], Font(name=F, size=9, color=GREY), wrap=True)
    put(ws, r, 5, '', BLUE_F, bg='FFFFCC'); put(ws, r, 6, '', BLUE_F, bg='FFFFCC', al='center'); dv_ph.add(ws.cell(r, 6))
    ws.row_dimensions[r].height = 40; r += 1
band(ws, ph0, r-1, 4)
put(ws, r+1, 1, 'الخلايا الصفراء (المسؤول/الحالة) تُعبَّأ عند الاعتماد.', Font(name=F, size=9, italic=True, color=GOLD), bg=GOLD_L)
ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=6)

# ══════════════ 14) دليل الملف ══════════════
ws = sheet(wb, 'دليل الملف')
r = banner(ws, 'دليل الملف', 'فهرس الأوراق ومصادر البيانات ودليل الألوان', 3)
r = thead(ws, r, ['الورقة','المحتوى','المصدر'], [26,64,32])
g0 = r
for name, desc, src in [
 ('لوحة التحكم','مؤشرات المحفظة ورسوم: صافي كل منفذ · مزيج الوقود · الاتجاه الشهري · خط الأنابيب','محسوبة'),
 ('بطاقة الأداء المشتركة','المؤشرات المشتركة لكل منفذ + المحفظة (مرتبطة بالنموذج المالي)','النموذج المالي'),
 ('النموذج المالي','الدخل والمصاريف → الصافي · الربح/الزيارة · تكلفة اللتر · نقطة التعادل','المالية + التشغيل'),
 ('بطاقات المنافذ','الموقع والنوع والمساحة والتشغيل (مضخات/خزان/عمّال/خدمة ذاتية)','سجل الأصول + خرائط جوجل'),
 ('شرائح العملاء','لكل شريحة: ملفها وما تحتاجه والعرض المقدّم','تحليلات درب + دراسة الموقع'),
 ('المحيط التنافسي','بطاقة السيناريو والمنافسون والميزة والفجوة','خرائط جوجل — سحب فعلي'),
 ('وحدات التأجير','كل وحدة: النوع والعدد والخدمة والنوع الفرعي والحالة','جرد المحطة'),
 ('مبيعات المستأجرين','مبيعات كل مستأجر شهرياً + نسبة الإيجار ومبيعات/م²','تقارير المستأجرين'),
 ('مبيعات الوقود','مزيج المنتجات ولتر وربح كل منتج + الاتجاه الشهري','لوحة تحليلات درب'),
 ('فرص B2B','الجوار التجاري كأهداف بيع: القناة والقيمة والحالة','خرائط جوجل + المسح الميداني'),
 ('عقود الأساطيل','الفئات وآلية التعاقد والجهات المستهدفة والحجم والربح','إدارة الشراكات'),
 ('المكس التسويقي','٧Ps + توزيع ميزانية الترويج + الحملات ومموّلوها','إدارة التسويق'),
 ('خطة التنفيذ','الركائز ودورة النمو وقواعد فضّ التعارض والتسلسل','استراتيجية التكامل')]:
    put(ws, r, 1, name, BOLD(9.5), wrap=True); put(ws, r, 2, desc, BLACK_F, wrap=True)
    put(ws, r, 3, src, Font(name=F, size=9, color=GREY), wrap=True); ws.row_dimensions[r].height = 26; r += 1
band(ws, g0, r-1, 3); r += 2
put(ws, r, 1, 'دليل الألوان', BOLD(11)); r += 1
r = thead(ws, r, ['اللون','المعنى','ملاحظة'], [26,64,32])
for c_, meaning, note, fnt in [
    ('أزرق','مُدخل يدوي — عدّله','الأرقام والنصوص القابلة للتعديل', BLUE_F),
    ('أسود','معادلة محسوبة — لا تُعدَّل','تتحدّث تلقائياً مع المُدخلات', BLACK_F),
    ('أخضر','رابط لورقة أخرى','يقرأ من النموذج المالي', LINK_F),
    ('أصفر','يحتاج تعبئة عند الاعتماد','المسؤول والحالة في خطة التنفيذ', Font(name=F, size=10, color=GOLD))]:
    put(ws, r, 1, c_, fnt, al='center'); put(ws, r, 2, meaning, BLACK_F); put(ws, r, 3, note, Font(name=F, size=9, color=GREY))
    r += 1
r += 1
put(ws, r, 1, 'الأرقام المالية والتشغيلية تقديرات نموذجية للاعتماد — تُستبدل بأرقام المالية وسجل الأصول.', Font(name=F, size=9.5, bold=True, color=GOLD), bg=GOLD_L)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

wb.move_sheet('دليل الملف', offset=-(len(wb.sheetnames)-2))
wb.move_sheet('لوحة التحكم', offset=-(len(wb.sheetnames)-1))
wb.save(OUT)
print('saved:', OUT)
print('sheets:', [s.title for s in wb.worksheets if s.sheet_state == 'visible'])
