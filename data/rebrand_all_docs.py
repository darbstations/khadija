# -*- coding: utf-8 -*-
"""إعادة تلوين ملفات الإكسل القديمة إلى هوية درب
   يُشغَّل من جذر المستودع:  PYTHONPATH=data python3 data/rebrand_all_docs.py"""
import sys
from openpyxl import load_workbook
import rebrand as R

FILES = sys.argv[1:] or ["docs/executive-plan.xlsx"]
for f in FILES:
    wb = load_workbook(f)
    done = R.rebrand_all(wb)
    wb.save(f)
    n = sum(done.values())
    print(f"{f}: {len(wb.sheetnames)} ورقة · {n} تغيير لوني")
