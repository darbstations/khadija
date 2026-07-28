# -*- coding: utf-8 -*-
"""ملف إكسل احترافي موحّد: استراتيجية + مستهدفات 2026 + لوحة + جدول مستقل لكل إدارة + سجل مشاريع.
   دلّات حقيقية · RTL · مرتّب. المصدر: platform/app/kpi_data.py"""
import os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "platform"))
from app import kpi_data as K

ORANGE="F47A21"; NAVY="58595B"; BLUE="808285"; STEEL="A7A9AC"; GREEN_IN="C6EFCE"
GOLD="FDE3D1"; WHITE="FFFFFF"; LIGHT="F2F3F5"; RED="C00000"; RED_FILL="F8D7DA"
thin=Side(style="thin", color="D9D9D9"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
FMT={"pct":"0%","int":"#,##0","num1":"0.0","rial":'#,##0 "ر.س"'}
PERSPS=["مالي","العملاء","العمليات الداخلية","التعلّم والنمو"]
def F_(sz=10,b=False,color="222222"): return Font(name="Tajawal",size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
def C(ws,r,c,v=None,*,f=None,fillc=None,al="right",fmt=None,lock=True,wrap=False,border=True):
    cell=ws.cell(r,c)
    if v is not None: cell.value=v
    cell.font=f or F_()
    if fillc: cell.fill=fill(fillc)
    cell.alignment=Alignment(horizontal=al,vertical="center",wrap_text=wrap)
    if fmt: cell.number_format=fmt
    cell.protection=Protection(locked=lock)
    if border: cell.border=BORDER
    return cell
# مؤشرات «المعالم» (مهام لمرة واحدة) — تُعلَّم بعلامة مميزة بدل أولوية الأداء
_MILESTONE_KW=["اكتمال المستندات","اكتمال أرشفة","أرشفة الملفات","اعتماد الهيكل","اكتمال الأوصاف",
  "ترسيخ القيم","أتمتة عمليات الموارد","تفعيل الأنظمة الأساسية"]
def is_milestone(nm): return any(k in nm for k in _MILESTONE_KW)
def prio_label(nm): return "🏁 معلم" if is_milestone(nm) else K.priority_label(K.priority(nm))

# ===== حوكمة المؤشرات: توليد بطاقة التعريف لكل مؤشر =====
MGR_TITLE={"exec":"الرئيس التنفيذي","franchise":"مدير الامتياز","operations":"مدير التشغيل",
 "investment":"مدير الاستثمار","realestate":"مدير العقار","digital":"مدير التقنية",
 "hr":"مدير الموارد البشرية","marketing":"مدير التسويق","quality":"مدير الجودة",
 "legal":"المدير القانوني","finance":"المدير المالي",
 "projects":"مدير إدارة المشاريع","supplychain":"مدير سلاسل الإمداد","maintenance":"مدير الصيانة والأتمتة"}
# أوصاف معتمدة (من ملفات الإدارات) للموارد/الإمداد/التقنية
DESC_OVERRIDE={
 "نسبة الالتزام بالتوطين والامتثال الحكومي دون مخالفات":"يقيس التزام الموارد البشرية بمتطلبات التوطين ونطاقات والمنصات الحكومية ذات العلاقة، مع متابعة المخالفات والغرامات النظامية المرتبطة بالموارد البشرية.",
 "نسبة الالتزام بنظام العمل واللوائح وخطة القوى العاملة المعتمدة":"يقيس الالتزام بنظام العمل واللائحة الداخلية، وعدم تنفيذ أي توظيف أو استحداث أو تعديل وظيفي خارج الهيكل أو خطة القوى العاملة أو الصلاحيات المعتمدة.",
 "نسبة إنجاز معاملات شؤون الموظفين ضمن المدة المستهدفة SLA":"يقيس سرعة إنجاز معاملات الموظفين والإدارات، مثل التعاريف، الإجازات، المباشرات، الإخلاءات، الإنذارات، تحديث البيانات، والتعديلات الوظيفية.",
 "نسبة الالتزام بتجديد الوثائق واكتمال ملفات الموظفين":"يقيس الالتزام بتجديد الإقامات، رخص العمل، العقود، التأمينات، الشهادات الصحية، مع اكتمال ودقة ملفات الموظفين ورقيًا أو إلكترونيًا.",
 "نسبة جاهزية وتغطية عمالة المحطات حسب خطة التشغيل والافتتاحات":"يقيس قدرة الموارد البشرية على توفير وتغطية عمالة المحطات وفق خطة التشغيل والافتتاحات، بما يشمل التوظيف في الوقت المناسب، التدريب قبل المباشرة، إدارة الإجازات، وتوفير البدلاء.",
 "متوسط مدة إغلاق الشواغر المعتمدة":"يقيس متوسط المدة من تاريخ اعتماد الاحتياج الوظيفي حتى قبول المرشح أو مباشرته، بما يعكس سرعة الاستقطاب وتلبية احتياج الإدارات.",
 "نسبة نجاح التعيينات الجديدة خلال فترة التجربة":"يقيس جودة الاستقطاب والاختيار من خلال نسبة الموظفين الجدد الذين اجتازوا فترة التجربة بنجاح واستمروا في العمل دون عدم ملاءمة.",
 "نسبة كفاءة تكلفة القوى العاملة":"يقيس مدى ضبط تكلفة القوى العاملة مقارنة بالميزانية المعتمدة، بما يشمل الرواتب، البدلات، العمل الإضافي، الاستقدام، التدريب، التأمين، السكن، النقل، والتكاليف غير المباشرة.",
 "نسبة دقة الرواتب والمستحقات الشهرية":"يقيس مدى خلو الرواتب والمستحقات من الأخطاء أو التأخير، مثل الرواتب، البدلات، الخصومات، الإجازات، نهاية الخدمة، التأمينات، والعمل الإضافي.",
 "معدل دوران الموظفين":"يقيس نسبة مغادرة الموظفين مقارنة بمتوسط عدد الموظفين خلال الفترة، ويساعد في كشف مشاكل الاستقرار الوظيفي أو بيئة العمل أو جودة التوظيف.",
 "نسبة تنفيذ برامج التدريب والتأهيل المعتمدة":"يقيس مدى تنفيذ برامج التدريب والتأهيل، خصوصًا التدريب المرتبط بعمالة المحطات، السلامة، خدمة العملاء، والمهارات الوظيفية الأساسية.",
 "نسبة إغلاق تقييمات الأداء وخطط الإحلال للوظائف الحرجة":"يقيس إكمال تقييمات الأداء ضمن المدة المحددة، ومدى توفر بدلاء أو مرشحين داخليين للوظائف الحرجة والقيادية لدعم استمرارية الأعمال.",
 "نسبة تلبية طلبات المحطات من الوقود Fill Rate":"يقيس قدرة سلاسل الإمداد على تلبية طلبات المحطات من الوقود بالكميات المطلوبة وفي الوقت المناسب دون نقص يؤثر على التشغيل أو المبيعات.",
 "عدد انقطاعات التوريد المؤثرة شهريًا":"يقيس عدد حالات انقطاع أو تعثر توريد الوقود أو المواد الحرجة التي أثرت على التشغيل أو المبيعات أو جاهزية المحطات.",
 "نسبة تغطية مخزون الوقود بالأيام":"يقيس عدد الأيام التي يغطيها مخزون الوقود المتوفر بناءً على متوسط الاستهلاك اليومي، بهدف المحافظة على الحد الآمن من المخزون.",
 "نسبة توفر المواد المخزنية والحرجة":"يقيس توفر المواد المخزنية والمواد التشغيلية الحرجة التي يؤثر عدم توفرها على التشغيل أو الصيانة أو جاهزية المحطات.",
 "نسبة المواد الراكدة من المخزون":"يقيس نسبة المواد غير المتحركة أو الراكدة مقارنة بإجمالي المخزون، بهدف تقليل الهدر وتجميد رأس المال في مواد غير مستخدمة.",
 "دقة أوامر الشراء ومستنداتها":"يقيس مدى اكتمال وصحة أوامر الشراء ومستنداتها، مثل طلب الشراء، عروض الأسعار، التعميد، أمر الشراء، الاستلام، الفاتورة، والمطابقة النظامية.",
 "نسبة إنجاز أوامر الشراء خلال المدة المستهدفة":"يقيس سرعة معالجة أوامر الشراء من تاريخ اعتماد الطلب حتى إصدار أمر الشراء أو التعميد وفق المدة المستهدفة المعتمدة.",
 "نسبة الالتزام بزمن التسليم والتوريد للمحطات والإدارات":"يقيس مدى الالتزام بتسليم احتياجات المحطات والإدارات في الوقت المتفق عليه بعد اعتماد الطلب أو أمر الشراء.",
 "نسبة جاهزية الإمداد للمحطات الجديدة وخطط الافتتاح":"يقيس مدى جاهزية احتياجات المحطات الجديدة قبل التشغيل، مثل الوقود الأولي، المواد التشغيلية، المستلزمات، الزي، أدوات التشغيل، وأي احتياجات معتمدة ضمن خطة الافتتاح.",
 "نسبة الالتزام بجدول النقل واستخدام الأسطول بكفاءة":"يقيس التزام النقل بالجدول المعتمد، ومدى كفاءة استخدام الأسطول المملوك مقارنة بإجمالي رحلات النقل، بما يدعم ضبط التكلفة واستمرارية التوريد.",
 "كفاءة إدارة الموردين":"يقيس أداء الموردين من حيث الالتزام بالمواعيد، جودة المواد أو الخدمة، الاستجابة، معالجة الملاحظات، اكتمال الوثائق، والالتزام بالعقود أو الأسعار المعتمدة.",
 "مستوى رضا الإدارات عن خدمات سلاسل الإمداد":"يقيس رضا الإدارات والمحطات عن سرعة الاستجابة، جودة التوريد، وضوح التواصل، الالتزام بالطلبات، ودعم الاحتياجات التشغيلية.",
 "نسبة توفر الأنظمة والخدمات التقنية المركزية":"يقيس استمرارية عمل الأنظمة والخدمات التقنية المركزية التي تخدم الشركة ككل، مثل ERP، البريد الإلكتروني، الشبكة المركزية، السيرفرات، قواعد البيانات، الإنترنت الرئيسي، وأنظمة المركز الرئيسي.",
 "نسبة معالجة الأعطال وطلبات الدعم الفني ضمن المدة المستهدفة SLA":"يقيس كفاءة تقنية المعلومات في إغلاق بلاغات الأعطال وطلبات الدعم الفني ضمن المدد المعتمدة، حسب أولوية البلاغ وتأثيره على العمل.",
 "نسبة استقرار التشغيل التقني للمحطات":"يقيس استقرار الأنظمة والخدمات التقنية في المحطات، والتأكد من عدم وجود أعطال مؤثرة على البيع، المخزون، الربط، التقارير، أو متابعة التشغيل، مثل POS وCashIn وATG والشبكة والإنترنت بالمحطة.",
 "نسبة إنجاز مشاريع التحول الرقمي والمشاريع التقنية حسب الخطة":"يقيس مدى إنجاز محفظة مشاريع التحول الرقمي والمشاريع التقنية وفق الخطة الزمنية والأولويات المعتمدة.",
 "نسبة تسليم المشاريع التقنية ضمن الوقت والميزانية ونطاق العمل المعتمد":"يقيس انضباط تنفيذ المشاريع التقنية من حيث الالتزام بالجدول الزمني، والميزانية، ونطاق العمل المعتمد، والحد من التوسع غير المعتمد في النطاق.",
 "نسبة الالتزام بإدارة صلاحيات المستخدمين والحسابات":"يقيس مدى الالتزام بإنشاء وتعديل وإيقاف صلاحيات المستخدمين وفق طلبات معتمدة، مع إيقاف حسابات المنتهية خدماتهم ومنع الصلاحيات الزائدة عن الحاجة.",
 "نسبة نجاح النسخ الاحتياطي واختبارات استعادة البيانات":"يقيس انتظام ونجاح النسخ الاحتياطي للأنظمة والبيانات المهمة، مع اختبار قابلية استعادة البيانات عند الحاجة.",
 "نسبة الامتثال لضوابط الأمن السيبراني الأساسية":"يقيس مدى تطبيق الضوابط الأساسية للأمن السيبراني، مثل التحديثات، مضادات الفيروسات، كلمات المرور، سجلات الدخول، حماية الأجهزة، ومراقبة محاولات الدخول غير المصرح.",
 "نسبة دقة وتكامل البيانات بين الأنظمة الرقمية":"يقيس مدى صحة وتطابق البيانات بين الأنظمة المختلفة، مثل ERP وPOS وCashIn وATG وإيجار والأنظمة الأخرى، وتقليل الفروقات الناتجة عن أخطاء التكامل أو الترحيل.",
 "نسبة الالتزام بتحديث وتوثيق أصول تقنية المعلومات والتراخيص":"يقيس مدى اكتمال وتحديث سجل أصول تقنية المعلومات، بما يشمل الأجهزة، السيرفرات، نقاط البيع، أجهزة الشبكات، أجهزة المستخدمين، التراخيص، الاشتراكات، الضمانات، والعهد التقنية.",
 "نسبة تنفيذ الصيانة الوقائية للأجهزة والبنية التقنية":"يقيس مدى تنفيذ الصيانة الوقائية لأجهزة الشبكة، نقاط البيع، السيرفرات، الطابعات، الكاميرات، وأجهزة المحطات حسب الخطة المعتمدة.",
 "مستوى رضا الإدارات والمستخدمين عن خدمات تقنية المعلومات":"يقيس رضا الإدارات والمستخدمين عن سرعة الاستجابة، جودة الحلول، وضوح التواصل، واستقرار الخدمات التقنية المقدمة.",
}
# طريقة القياس التفصيلية لمؤشرات الصيانة (من ملف الإدارة) — تُعرض كوصف/طريقة قياس
MAINT_METHODS={
 "نسبة الجاهزية الفنية للأصول والمعدات (Availability)":"إجمالي ساعات جاهزية الأصول والمعدات الحرجة ÷ إجمالي ساعات التشغيل المخططة لها × 100",
 "نسبة تنفيذ الصيانة الوقائية في موعدها (PM)":"عدد أوامر الصيانة الوقائية المنفذة ضمن الموعد المحدد ÷ إجمالي أوامر الصيانة الوقائية المستحقة × 100",
 "نسبة إغلاق بلاغات الصيانة ضمن اتفاقية الخدمة (SLA)":"عدد بلاغات الصيانة المغلقة ضمن الزمن المحدد حسب الأولوية ÷ إجمالي البلاغات المغلقة × 100",
 "متوسط زمن الاستجابة للبلاغات الطارئة":"مجموع الدقائق من تسجيل البلاغ الطارئ حتى بدء المعالجة الفعلية ÷ إجمالي البلاغات الطارئة",
 "متوسط زمن إصلاح الأعطال (MTTR)":"إجمالي ساعات الإصلاح الفعلية للأعطال المغلقة ÷ إجمالي عدد الأعطال المغلقة",
 "معدل تكرار الأعطال":"عدد الأعطال المتكررة لنفس الأصل وبنفس السبب ÷ إجمالي عدد الأعطال × 100",
 "نسبة استمرارية تشغيل المحطات دون توقف فني":"عدد المحطات دون توقف كلي بسبب عطل فني ضمن مسؤولية الصيانة ÷ إجمالي المحطات العاملة × 100",
 "نسبة التشغيل السليم لأنظمة الأتمتة (Uptime)":"عدد حالات الأتمتة العاملة والمتصلة بقراءات صحيحة ومحدثة ÷ إجمالي حالات الأتمتة المطلوبة × 100",
 "نسبة استعادة أنظمة الأتمتة ضمن الزمن المحدد":"عدد أعطال الأتمتة المستعادة ضمن الزمن المعتمد حسب الأولوية ÷ إجمالي أعطال الأتمتة المغلقة × 100",
 "كفاءة تكاليف الصيانة (الالتزام بالميزانية)":"(1 − |الفرق بين المصروف الفعلي والميزانية| ÷ الميزانية المعتمدة) × 100",
}
def round_weights_pct(weights):
    """تقريب الأوزان إلى نِسب مئوية صحيحة مجموعها 100٪ (largest remainder) بحد أدنى 1٪."""
    raw=[w*100 for w in weights]; floor=[int(x) for x in raw]; rem=[x-int(x) for x in raw]
    for i in range(len(floor)):
        if floor[i]==0: floor[i]=1; rem[i]=0
    diff=100-sum(floor)
    order=sorted(range(len(raw)),key=lambda i:rem[i],reverse=(diff>0))
    i=0
    while diff!=0 and order:
        idx=order[i%len(order)]
        if diff>0: floor[idx]+=1; diff-=1
        elif floor[idx]>1: floor[idx]-=1; diff+=1
        i+=1
    return [f/100 for f in floor]
def kpi_source(nm,axis,key=None):
    # مصدر القياس حسب الإدارة أولاً (للإدارتين الجديدتين ذات النظام المتخصّص)
    if key=="supplychain": return "نظام سلاسل الإمداد والمخزون (WMS/ERP)"
    if key=="projects":    return "نظام إدارة المشاريع (PMO/Primavera)"
    if key=="maintenance": return "نظام إدارة الصيانة (CMMS) وأنظمة الأتمتة (SCADA)"
    if key=="digital":     return "أنظمة تقنية المعلومات (ERP/ITSM/SIEM)"
    if key=="hr":
        if any(k in nm for k in ["رواتب","الرواتب","الأجور","مستحقات","تكلفة القوى العاملة","تعويضات"]):
            return "نظام الرواتب (Payroll) / حماية الأجور"
        return "نظام الموارد البشرية (HRMS)"
    t=nm+" "+axis
    for kws,src in [
     (["إيراد","ربح","هامش","تحصيل","DSO","سداد","موازنة","تدفق","نقدي","العائد","ROE","مستحق","مورد","تدقيق","ميزاني","دين","ذمم","إقفال"],"نظام المحاسبة / ERP (D365)"),
     (["قوقل","Google","خرائط"],"تقييمات خرائط قوقل"),
     (["رضا","CSAT","eNPS","استبيان"],"استبيان دوري"),
     (["تانكي","الولاء","متابع","وصول","تفاعل","حمل","MAU"],"أنظمة التسويق الرقمي / تانكي"),
     (["الرواتب","رواتب","الأجور","WPS","حماية الأجور","تعويضات"],"نظام الرواتب (Payroll) / حماية الأجور"),
     (["تدريب","سعودة","استبقاء","توظيف","موظف","شواغر","دوران","هيكل","أوصاف","الغياب","الترقيات","تقييمات"],"نظام الموارد البشرية (HRMS)"),
     (["مشروع","تحول رقمي","أتمتة","سيبراني","أنظمة","تكامل","Uptime"],"مكتب إدارة المشاريع (PMO)"),
     (["جودة","ملاحظات","مخاطر","مطابقة","تفتيش","QMS","امتثال"],"نظام إدارة الجودة (QMS)"),
     (["قضايا","مذكرة","مذكرات","استشار","صياغة","تحقيق","تحقيقات","التوصيات الصادرة"],"نظام إدارة القضايا والتحقيقات"),
     (["عقد","عقود","امتياز","استثمار","تأجير","إيجار","إشغال","استحواذ"],"نظام إدارة العقود (CRM)"),
     (["محطة","محطات","مبيعات","لتر","مضخ","صيانة","سلامة","تشجير","وقود","تعبئة"],"نظام تشغيل المحطات"),
    ]:
        if any(k in t for k in kws): return src
    return "سجلات الإدارة"
def kpi_freq(nm,ttxt):
    if any(k in nm for k in ["eNPS","رضا","CSAT","استبيان","تدقيق","مراجعة الوثائق"]): return "ربعي"
    if "سنوي" in (ttxt or "") or any(k in nm for k in ["ROE","العائد على حقوق","نمو الإيرادات السنوي","الموازنة"]): return "ربعي"
    return "شهري"
def kpi_formula(fmt,unit,nm):
    if "نسبة" in nm or fmt=="pct": return "(المتحقّق ÷ الإجمالي) × 100٪"
    if fmt=="rial": return "إجمالي القيمة (ر.س) خلال الفترة"
    if any(u in unit for u in ["يوم","أسبوع","ساعة","دقيقة","سنة"]): return f"متوسط {unit} خلال الفترة"
    if "نجمة" in unit: return "متوسط التقييم (من 5)"
    if "نقطة" in unit: return "٪ المروّجين − ٪ المنتقدين"
    return "إجمالي العدد خلال الفترة"
import re as _re
def _core(nm): return _re.sub(r'\s*\([^)]*\)','',nm).strip()   # إزالة اللواحق بين قوسين
def measure_method(nm,unit,pol,fmt,agg):
    """طريقة قياس تفصيلية خاصة بكل مؤشر (بسط ÷ مقام) — تُشتق من صيغة الاسم والوحدة."""
    n=nm.strip(); c=_core(n); lead=lambda p:c[len(p):].strip()
    if n.startswith("نسبة"):
        s=lead("نسبة"); return f"قيمة/عدد المتحقّق من «{s}» ÷ الإجمالي المستهدف لنفس البند × 100٪"
    if n.startswith("معدل"):
        s=lead("معدل"); return f"عدد حالات «{s}» خلال الفترة ÷ إجمالي الحالات القابلة للقياس × 100٪"
    if n.startswith("متوسط"):
        s=lead("متوسط")
        if "دقيقة" in unit: return f"مجموع دقائق «{s}» لكل الحالات ÷ عدد الحالات"
        if "ساعة"  in unit: return f"مجموع ساعات «{s}» ÷ عدد الحالات المغلقة"
        if "يوم"   in unit: return f"مجموع أيام «{s}» ÷ عدد الحالات"
        if "هللة"  in unit: return f"مجموع «{s}» (هللة/لتر) ÷ عدد الوحدات"
        return f"مجموع قيم «{s}» ÷ عدد المفردات المقيسة"
    if n.startswith("إجمالي"):
        s=lead("إجمالي")
        if fmt=="rial": return f"مجموع قيمة «{s}» بالريال خلال الفترة"
        if "لتر" in unit: return f"مجموع لترات «{s}» خلال الفترة"
        return f"مجموع «{s}» خلال الفترة"
    if n.startswith("عدد"):
        cum="تراكمياً حتى نهاية الفترة" if (agg=="LAST" or "تراكمي" in n) else "المُسجّلة خلال الفترة"
        return f"إجمالي عدد «{lead('عدد')}» {cum}"
    if fmt=="rial" or unit=="ر.س": return f"إجمالي قيمة «{c}» بالريال خلال الفترة"
    if "لتر" in unit: return f"مجموع لترات «{c}» خلال الفترة"
    if "eNPS" in n or "نقطة" in unit: return "٪ الموظفين المروّجين − ٪ المنتقدين (مقياس eNPS)"
    if "نجمة" in unit or "درجة" in unit: return f"متوسط تقييم «{c}» على مقياسه المعتمد"
    if "يوم" in unit: return f"متوسط عدد أيام «{c}» خلال الفترة"
    if "ساعة" in unit: return f"متوسط عدد ساعات «{c}» خلال الفترة"
    if fmt=="pct": return f"القيمة المتحقّقة من «{c}» ÷ المستهدف × 100٪"
    if fmt=="int": return f"إجمالي عدد «{c}» المُسجّل خلال الفترة"
    return f"قياس «{c}» مقابل مستهدفه المعتمد خلال الفترة"
_OUTCOME=["إيراد","ربح","هامش","العائد","ROE","رضا","CSAT","eNPS","نمو","حصة","إشغال","تحصيل","مبيعات",
 "استبقاء","قيمة","تقييم","شكاو","الوصول","EBITDA","ربحية","جاهزية","Uptime","سلامة","امتثال","التزام",
 "حل الشكاو","دقة","مطابقة","إغلاق","فعالية","استدامة","تجديد","انخفاض","خفض","سعودة","توطين","مخاطر",
 "إنجاز محفظة","معالجة","حوادث"]
_ACTIVITY=["عدد ","زيارات","حمل","دورات","تنفيذ","إنجاز التقارير","اكتمال المستندات","اكتمال الأوصاف",
 "أرشفة","تدريب","إعداد","مذكرة","معارض","اعتماد الهيكل"]
def kpi_type(nm):
    if any(k in nm for k in _ACTIVITY): return "⚙️ نشاط"
    if any(k in nm for k in _OUTCOME): return "🎯 نتيجة"
    return "🔧 تشغيلي"
def kpi_def(nm,axis,pillar):
    return f"يقيس {nm} ضمن محور «{axis}»، ويعكس مساهمة الإدارة في ركيزة {pillar}."
def kpi_owner(key,axis): return MGR_TITLE.get(key,"مدير الإدارة")
def kpi_approver(src,key):
    if key=="exec": return "اللجنة التنفيذية"
    if "المحاسبة" in src: return "الإدارة المالية"
    return MGR_TITLE.get(key,"مدير الإدارة")

# ===== المستوى 4: حدود مخصّصة · خط أساس · تنبؤ · انحراف =====
COMPLETION=0.5   # نسبة اكتمال السنة (للتنبؤ بنهاية السنة)
def thresholds(nm):
    crit=["Uptime","جاهزية","سلامة","سيبراني","مطابقة","أرامكو","امتثال","دقة التعبئة"]
    soft=["رضا","CSAT","eNPS","تفاعل","متابع","الوصول"]
    if any(k in nm for k in crit): return (1.0,0.97)
    if any(k in nm for k in soft): return (0.90,0.75)
    return (1.0,0.85)
def status_of(a,nm):
    if a is None: return "⏳ بانتظار هدف"
    gr,yl=thresholds(nm)
    if a>=gr: return "✅ محقق"
    if a>=yl: return "🟡 قريب"
    return "🔴 تحت الهدف"
def baseline_of(nm,pol,act):
    if act is None: return None
    frac=(int(hashlib.md5((nm+'|b').encode('utf-8')).hexdigest(),16)%1000)/1000.0
    gap=0.08+frac*0.12
    return round(act*(1-gap),3) if pol=="↑" else round(act*(1+gap),3)
def previous_of(act,base):
    return None if (act is None or base is None) else round((act+base)/2,3)
def forecast_of(agg,act):
    if act is None: return None
    return round(act/max(COMPLETION,0.01),3) if agg=="SUM" else act
def growth_vs_base(act,base):
    return None if (act is None or not base) else round(act/base-1,4)

def rtl(ws): ws.sheet_view.rightToLeft=True; ws.sheet_view.showGridLines=False

# بيانات تجريبية واقعية (deterministic) لتعبئة الفراغات — تُستبدل بالأرقام الفعلية لاحقاً
import hashlib
DEMO=True
def eff_actual(nm,pol,tgt):
    if nm in K.SEED_ACTUALS: return K.SEED_ACTUALS[nm]      # رقم حقيقي من الملفات
    if not DEMO or tgt is None or tgt==0: return None        # اترك بلا رقم (أهداف «صفر» أو غير محدّدة)
    frac=(int(hashlib.md5(nm.encode("utf-8")).hexdigest(),16)%1000)/1000.0
    factor=(0.85+frac*0.33) if pol=="↓" else (0.72+frac*0.33)  # تشتيت لإظهار أخضر/أصفر/أحمر
    val=tgt*factor
    return round(val) if tgt>=1000 else round(val,1) if tgt>=10 else round(val,3)

wb=Workbook(); wb.remove(wb.active)
LOGO=os.path.join(os.path.dirname(os.path.abspath(__file__)),"platform/app/static/darb_logo.png")
def logo(ws,anchor="B1",w=150):
    try:
        im=XLImage(LOGO); im.width=w; im.height=int(w/3.83); ws.add_image(im,anchor); ws.row_dimensions[1].height=46
    except Exception: pass

# ============ جداول الإدارات (كل إدارة ورقة مستقلة) ============
# أعمدة: A# B محور C مؤشر D وحدة E قطبية F أولوية G وزن% H مستهدف I نص المستهدف J ركيزة K مشروع L منظور M المُحقَّق N نسبة التحقيق O الحالة
HEAD=["#","المحور","المؤشر","الوحدة","القطبية","الأولوية","الوزن %","المستهدف","نص المستهدف","الركيزة","المشروع","منظور BSC","المُحقَّق","نسبة التحقيق","الحالة","الوصف / طريقة القياس","مصدر البيانات","دورية القياس"]
WID=[4,16,40,9,7,12,8,12,16,12,22,15,12,12,14,48,24,11]
DEPT_RANGES=[]; ROWMAP={}; DEPT_DETAILS={}
INLINE_AGG=bool(os.environ.get("INLINE_AGG"))   # الخيار (ب): إدراج المؤشر الجامع داخل أوراق الإدارات
def safe_title(name): return name[:31]
for dname,key,records in K.DEPARTMENTS:
    ws=wb.create_sheet(safe_title(dname)); rtl(ws)
    for i,w in enumerate(WID,1): ws.column_dimensions[get_column_letter(i)].width=w
    last=get_column_letter(len(HEAD))
    ws.merge_cells(f"A1:{last}1"); C(ws,1,1,f"إدارة {dname} · مؤشرات الأداء 2026",f=F_(13,True,WHITE),fillc=NAVY,al="center",border=False); ws.row_dimensions[1].height=24
    ws.merge_cells(f"A2:{last}2"); C(ws,2,1,"🟩 عمود «المُحقَّق» أرقام تجريبية — استبدلها بالفعلية · نسبة التحقيق والحالة دلّات تُحسب تلقائياً",f=F_(9,False,NAVY),fillc=GOLD,al="center",border=False)
    for c,h in enumerate(HEAD,1): C(ws,3,c,h,f=F_(9,True,WHITE),fillc=BLUE,al="center")
    ws.row_dimensions[3].height=28
    dv=DataValidation(type="decimal",operator="greaterThanOrEqual",formula1="0",allow_blank=True); ws.add_data_validation(dv)
    weights=round_weights_pct(K.kpi_weights(records)); start=4
    ROWMAP[safe_title(dname)]={}; DEPT_DETAILS[safe_title(dname)]=[]
    r=start; seen_axes=set()
    def _agg_formula(dn,ax):
        JJ="'قاعدة المؤشرات'!$J:$J"; KK="'قاعدة المؤشرات'!$K:$K"; BB="'قاعدة المؤشرات'!$B:$B"; LL="'قاعدة المؤشرات'!$L:$L"
        return f'=IFERROR(SUMIFS({JJ},{BB},"{dn}",{LL},"{ax}")/SUMIFS({KK},{BB},"{dn}",{LL},"{ax}"),"")'
    for i,(axis,nm,unit,pol,agg,tgt,ttxt,fmt,pillar,project) in enumerate(records):
        # الخيار (ب): إدراج صف «جامع» عند أول ظهور لكل محور
        if INLINE_AGG and axis and axis not in seen_axes:
            seen_axes.add(axis); n=sum(1 for rr in records if rr[0]==axis and rr[1])
            C(ws,r,2,axis,f=F_(9,True,WHITE),fillc=STEEL,al="right",wrap=True)
            C(ws,r,3,f"🔷 جامع: {axis} ({n})",f=F_(9,True,WHITE),fillc=STEEL,al="right",wrap=True)
            for c in (1,4,5,6,7,8,9,10,11,12,13,16,17,18): C(ws,r,c,None,fillc=STEEL)
            ws.cell(r,14).value=_agg_formula(dname,axis); C(ws,r,14,f=F_(9,True,WHITE),fillc=STEEL,fmt="0%",al="center")
            ws.cell(r,15).value=f'=IF($N{r}="","🔷 —",IF($N{r}>=1,"🔷 ✅",IF($N{r}>=0.85,"🔷 🟡","🔷 🔴")))'
            C(ws,r,15,f=F_(9,True,WHITE),fillc=STEEL,al="center"); ws.row_dimensions[r].height=22; r+=1
        ROWMAP[safe_title(dname)][i]=r; DEPT_DETAILS[safe_title(dname)].append(r)
        C(ws,r,1,i+1,f=F_(9,True),al="center")
        blank = not (nm and str(nm).strip())   # صف قالب فارغ للإدخال اليدوي
        if blank:
            for c in (2,3,16,17): C(ws,r,c,None,fillc=GREEN_IN,al="right",lock=False,wrap=True)
            for c in (4,5,6,7,8,9,10,11,12,18): C(ws,r,c,None,fillc=GREEN_IN,al="center",lock=False)
            C(ws,r,7,None,fillc=GREEN_IN,fmt="0%",al="center",lock=False)
            C(ws,r,13,None,fillc=GREEN_IN,al="center",lock=False); dv.add(ws.cell(r,13))
        else:
            C(ws,r,2,axis,f=F_(9),al="right",wrap=True)
            C(ws,r,3,nm,f=F_(9),al="right",wrap=True)
            C(ws,r,4,unit,al="center")
            C(ws,r,5,pol,al="center")
            C(ws,r,6,prio_label(nm),f=F_(8),al="center")
            C(ws,r,7,weights[i],fmt="0%",al="center")
            C(ws,r,8,tgt if tgt is not None else "",fmt=FMT[fmt] if tgt is not None else None,al="center")
            C(ws,r,9,ttxt,f=F_(8,color="666666"),al="center",wrap=True)
            C(ws,r,10,pillar,f=F_(9,color=ORANGE),al="center")
            C(ws,r,11,project,f=F_(8),al="right",wrap=True)
            C(ws,r,12,K.perspective(nm),f=F_(8),al="center")
            ach_v=eff_actual(nm,pol,tgt)
            C(ws,r,13,ach_v if ach_v is not None else None,fillc=GREEN_IN,fmt=FMT[fmt],al="center",lock=False)
            dv.add(ws.cell(r,13))
        # نسبة التحقيق — مع سقف 100% (MIN) لمنع تضخّم المتوسط
        ws.cell(r,14).value=(f'=IF($H{r}="","",IF($M{r}="","",MIN(1,IF($H{r}=0,IF($M{r}<=0,1,0),'
            f'IF($E{r}="↓",IFERROR($H{r}/$M{r},0),IFERROR($M{r}/$H{r},0))))))')
        C(ws,r,14,f=F_(9,True,NAVY),fmt="0%",al="center")
        _gr,_yl=thresholds(nm)   # حدود تنبيه مخصّصة لكل مؤشر
        ws.cell(r,15).value=f'=IF($N{r}="","⏳ بانتظار هدف",IF($N{r}>={_gr},"✅ محقق",IF($N{r}>={_yl},"🟡 قريب","🔴 تحت الهدف")))'
        C(ws,r,15,f=F_(9,True),al="center")
        if not blank:
            desc=DESC_OVERRIDE.get(nm) or MAINT_METHODS.get(nm) or measure_method(nm,unit,pol,fmt,agg)   # وصف/طريقة قياس تفصيلية لكل مؤشر
            C(ws,r,16,desc,f=F_(8,color="555555"),al="right",wrap=True)
            C(ws,r,17,kpi_source(nm,axis,key),f=F_(8,color="555555"),al="right",wrap=True)  # مصدر البيانات
            C(ws,r,18,kpi_freq(nm,ttxt),f=F_(8),al="center")  # دورية القياس
        ws.row_dimensions[r].height=26
        r+=1
    end=r-1
    ws.auto_filter.ref=f"A3:{last}{end}"
    ws.freeze_panes="A4"
    DEPT_RANGES.append((dname,safe_title(dname),start,end))

# ============ أهم المؤشرات (Top 5 لكل إدارة) — لوحة المتابعة التنفيذية (قيَم حيّة) ============
top=wb.create_sheet("أهم المؤشرات (Top 5)"); rtl(top)
for i,w in enumerate([4,48,18,12,13,14,8],1): top.column_dimensions[get_column_letter(i)].width=w
top.merge_cells("A1:G1"); C(top,1,1,"أهم 5 مؤشرات رئيسية لكل إدارة — لوحة المتابعة التنفيذية (قيَم حيّة مرتبطة بأوراق الإدارات)",f=F_(13,True,WHITE),fillc=NAVY,al="center",border=False); top.row_dimensions[1].height=26
for c,h in enumerate(["#","المؤشر","المستهدف","المُحقَّق","نسبة التحقيق","الحالة","الوزن"],1): C(top,2,c,h,f=F_(9,True,WHITE),fillc=BLUE,al="center")
top.row_dimensions[2].height=24
tr=3
for (dname,key,records),(dn2,sh,start,end) in zip(K.DEPARTMENTS,DEPT_RANGES):
    weights=round_weights_pct(K.kpi_weights(records))
    idx=sorted(range(len(records)),key=lambda i:(weights[i], -i),reverse=True)[:5]  # أعلى 5 وزناً
    dw=int(round(K.DEPT_WEIGHTS.get(key,0)*100))
    top.merge_cells(f"A{tr}:G{tr}")
    C(top,tr,1,f"▾ {dname} · وزن الإدارة {dw}%",f=F_(10,True,WHITE),fillc=ORANGE,al="right"); top.row_dimensions[tr].height=22; tr+=1
    for n,i in enumerate(idx,1):
        r=ROWMAP[sh][i]
        C(top,tr,1,n,f=F_(9,True),al="center")
        top.cell(tr,2).value=f"='{sh}'!C{r}"; C(top,tr,2,f=F_(9),al="right",wrap=True)
        top.cell(tr,3).value=f"='{sh}'!I{r}"; C(top,tr,3,f=F_(8,color="666666"),al="center",wrap=True)
        top.cell(tr,4).value=f"='{sh}'!M{r}"; C(top,tr,4,f=F_(9),al="center")
        top.cell(tr,5).value=f"='{sh}'!N{r}"; C(top,tr,5,f=F_(9,True,NAVY),fmt="0%",al="center")
        top.cell(tr,6).value=f"='{sh}'!O{r}"; C(top,tr,6,f=F_(9,True),al="center")
        top.cell(tr,7).value=f"='{sh}'!G{r}"; C(top,tr,7,f=F_(8),fmt="0%",al="center")
        top.row_dimensions[tr].height=22; tr+=1
top.freeze_panes="A3"

# ============ قاعدة المؤشرات (تجميع — للوحة والاستراتيجية) ============
cons=wb.create_sheet("قاعدة المؤشرات"); rtl(cons)
for i,w in enumerate([4,20,42,16,24,16,12,14,10],1): cons.column_dimensions[get_column_letter(i)].width=w
for c,h in enumerate(["#","الإدارة","المؤشر","الركيزة","المشروع","منظور BSC","نسبة التحقيق","الحالة","الوزن","و×ن","الوزن المتاح","المحور"],1):
    C(cons,1,c,h,f=F_(9,True,WHITE),fillc=BLUE,al="center")
cr=2; seq=1
for dname,sh,s,e in DEPT_RANGES:
    for rr in DEPT_DETAILS[sh]:   # الصفوف التفصيلية فقط (تُستبعد صفوف الجامع لمنع المرجع الدائري)
        C(cons,cr,1,seq,f=F_(8),al="center")
        C(cons,cr,2,dname,f=F_(8),al="right")
        C(cons,cr,3,f"='{sh}'!C{rr}",f=F_(8),al="right")
        C(cons,cr,4,f"='{sh}'!J{rr}",f=F_(8),al="center")
        C(cons,cr,5,f"='{sh}'!K{rr}",f=F_(8),al="right")
        C(cons,cr,6,f"='{sh}'!L{rr}",f=F_(8),al="center")
        C(cons,cr,7,f"='{sh}'!N{rr}",f=F_(8,True,NAVY),fmt="0%",al="center")
        C(cons,cr,8,f"='{sh}'!O{rr}",f=F_(8),al="center")
        C(cons,cr,9,f"='{sh}'!G{rr}",f=F_(8),fmt="0%",al="center")
        # أعمدة الترجيح: و×ن = الوزن×التحقيق (للمؤشرات المُعبّأة فقط) · الوزن المتاح = الوزن إن وُجد رقم
        C(cons,cr,10,f'=IF($G{cr}="",0,$G{cr}*$I{cr})',f=F_(8),al="center")
        C(cons,cr,11,f'=IF($G{cr}="",0,$I{cr})',f=F_(8),al="center")
        C(cons,cr,12,f"='{sh}'!B{rr}",f=F_(8),al="right")  # المحور (لحساب المؤشر الجامع)
        cr+=1; seq+=1
CE=cr-1
ACH=f"'قاعدة المؤشرات'!$G$2:$G${CE}"; DEPC=f"'قاعدة المؤشرات'!$B$2:$B${CE}"
PILC=f"'قاعدة المؤشرات'!$D$2:$D${CE}"; PROJC=f"'قاعدة المؤشرات'!$E$2:$E${CE}"
PERC=f"'قاعدة المؤشرات'!$F$2:$F${CE}"; STC=f"'قاعدة المؤشرات'!$H$2:$H${CE}"
WXN=f"'قاعدة المؤشرات'!$J$2:$J${CE}"; WAV=f"'قاعدة المؤشرات'!$K$2:$K${CE}"  # و×ن · الوزن المتاح
AXISC=f"'قاعدة المؤشرات'!$L$2:$L${CE}"  # المحور

# ============ المؤشرات الجامعة (متوسط موزون حيّ لكل محور — الخيار (أ): ورقة مستقلة) ============
aggw=wb.create_sheet("المؤشرات الجامعة"); rtl(aggw)
for i,w in enumerate([4,20,32,10,13,15,46],1): aggw.column_dimensions[get_column_letter(i)].width=w
aggw.merge_cells("A1:G1"); C(aggw,1,1,"المؤشرات الجامعة — مؤشر واحد لكل محور (متوسط موزون حيّ من مؤشراته التفصيلية · لا يُحسب في الإجمالي)",f=F_(13,True,WHITE),fillc=NAVY,al="center",border=False); aggw.row_dimensions[1].height=26
for c,h in enumerate(["#","الإدارة","المحور (مؤشر جامع)","عدد المؤشرات","نسبة التحقيق (موزون)","الحالة","المؤشرات التفصيلية المشمولة"],1): C(aggw,2,c,h,f=F_(9,True,WHITE),fillc=BLUE,al="center")
aggw.row_dimensions[2].height=28
gr_=3; gseq=1; _cur=None
for dname,key,records in K.DEPARTMENTS:
    axes=[]
    for rec in records:
        if rec[0] and rec[0] not in axes: axes.append(rec[0])
    for ax in axes:
        subs=[rec[1] for rec in records if rec[0]==ax and rec[1]]
        C(aggw,gr_,1,gseq,f=F_(8),al="center")
        C(aggw,gr_,2,dname,f=F_(8),al="right")
        C(aggw,gr_,3,ax,f=F_(9,True,NAVY),al="right",wrap=True)
        C(aggw,gr_,4,len(subs),f=F_(8),al="center")
        aggw.cell(gr_,5).value=(f'=IFERROR(SUMIFS({WXN},{DEPC},"{dname}",{AXISC},"{ax}")/'
                                f'SUMIFS({WAV},{DEPC},"{dname}",{AXISC},"{ax}"),"")')
        C(aggw,gr_,5,f=F_(10,True,ORANGE),fmt="0%",al="center")
        aggw.cell(gr_,6).value=f'=IF($E{gr_}="","⏳ بانتظار هدف",IF($E{gr_}>=1,"✅ محقق",IF($E{gr_}>=0.85,"🟡 قريب","🔴 تحت الهدف")))'
        C(aggw,gr_,6,f=F_(9,True),al="center")
        C(aggw,gr_,7," · ".join(subs),f=F_(8,color="666666"),al="right",wrap=True)
        aggw.row_dimensions[gr_].height=26; gr_+=1; gseq+=1
aggw.freeze_panes="A3"; aggw.auto_filter.ref=f"A2:G{gr_-1}"

# ============ الاستراتيجية والمستهدفات ============
ws=wb.create_sheet("الاستراتيجية والمستهدفات"); rtl(ws)
for i,w in enumerate([3,26,40,30,14],1): ws.column_dimensions[get_column_letter(i)].width=w
logo(ws,"B1")
ws.merge_cells("B2:E2"); C(ws,2,2,"درب · الاستراتيجية ومستهدفات 2026 — قطاع المحطات والعقار",f=F_(15,True,WHITE),fillc=NAVY,al="center",border=False); ws.row_dimensions[2].height=26
for c,h in enumerate(["","الركيزة (5 سنوات) / الهدف","الوصف","مستهدف 2026"],1): C(ws,4,c,h,f=F_(10,True,WHITE),fillc=BLUE,al="center")
strat=[("النمو","ركيزة","التوسع في المواقع وزيادة الوصول وتنمية الإيراد.",""),
 ("التوسع في المواقع","","افتتاح وتشغيل محطات جديدة.","التشغيل 85 · الامتياز 150 محطة"),
 ("الامتياز والعقود","","تنمية الإيراد عبر الامتياز والاستثمار.","الامتياز 167 · الاستثمار 190 عقد"),
 ("زيادة وصول العملاء","","رفع المبيعات والتغطية الجغرافية.","718.8M لتر · 13 منطقة · 59 مدينة"),
 ("الابتكار","ركيزة","مشاريع جديدة وتحول تقني وهوية وتانكي.",""),
 ("التحول التقني","","أتمتة 164 محطة · D365 · جاهزية الأنظمة.","أتمتة 90% · جاهزية 99%"),
 ("ساحات درب وتانكي","","تشغيل ساحات درب ونمو تطبيق تانكي (التسويق).","تأجير ساحة ≥80% · تانكي تصاعدي"),
 ("الاستدامة","ركيزة","جودة التشغيل وتجربة العميل والموظفين.",""),
 ("تجربة العميل وجودة التشغيل","","رضا وأمن سيبراني وجاهزية.","CSAT 90% · جاهزية 99%"),
 ("منظومة التأجير","","الإشغال والتحصيل واستدامة المستأجرين.","إشغال ≥60% · 246 علامة"),
 ("الاستثمار في الموظفين","","التوطين والتدريب والاستبقاء.","6 دورات · استبقاء 90% · سعودة"),]
r=5
for nm,kind,desc,tgt in strat:
    isr=kind=="ركيزة"
    C(ws,r,2,nm,f=F_(11,True,WHITE if isr else NAVY),fillc=BLUE if isr else None,al="right",wrap=True)
    C(ws,r,3,desc,f=F_(9,False,WHITE if isr else "222222"),fillc=BLUE if isr else None,al="right",wrap=True)
    C(ws,r,4,tgt,f=F_(9,True,WHITE if isr else ORANGE),fillc=BLUE if isr else None,al="right",wrap=True)
    C(ws,r,1,"",fillc=BLUE if isr else None); ws.row_dimensions[r].height=24; r+=1
r+=1; C(ws,r,2,"الخارطة التنفيذية — المشاريع (أداء حي)",f=F_(12,True,WHITE),fillc=ORANGE,al="center"); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5); r+=1
for c,h in zip([2,3,4,5],["المشروع","الركيزة","عدد المؤشرات","الأداء (حي)"]): C(ws,r,c,h,f=F_(10,True,WHITE),fillc=BLUE,al="center")
r+=1
for p in K.PROJECTS:
    C(ws,r,2,p,f=F_(9),al="right",wrap=True); C(ws,r,3,K.PROJECT_PILLAR.get(p,""),f=F_(9),al="center")
    C(ws,r,4,f'=COUNTIF({PROJC},"{p}")',f=F_(9,True),al="center")
    C(ws,r,5,f'=IFERROR(AVERAGEIFS({ACH},{PROJC},"{p}"),"—")',f=F_(9,True,NAVY),fmt="0%",al="center"); r+=1

# خريطة وزن كل إدارة (ميزان الإدارات)
DW_BY_NAME={name:K.DEPT_WEIGHTS.get(key,0) for name,key,_ in K.DEPARTMENTS}
EXEC_NAME="الإدارة التنفيذية"

# ============ اللوحة الموجزة ============
ws=wb.create_sheet("اللوحة الموجزة"); rtl(ws)
for i,w in enumerate([3,28,14,16,16],1): ws.column_dimensions[get_column_letter(i)].width=w
for col in ("H","I","J"): ws.column_dimensions[col].hidden=True   # أعمدة مساعدة للترجيح
ws.merge_cells("B2:E2"); C(ws,2,2,"اللوحة الموجزة — إنجاز موزون (وزن المؤشر × وزن الإدارة) · سقف 100%",f=F_(13,True,WHITE),fillc=NAVY,al="center",border=False); ws.row_dimensions[2].height=26
C(ws,4,2,"الإنجاز العام (موزون)",f=F_(11,True,WHITE),fillc=BLUE,al="center")
hr=7
drng=f"H{hr+2}:H{hr+1+len(DEPT_RANGES)}"; wrng=f"I{hr+2}:I{hr+1+len(DEPT_RANGES)}"; hasrng=f"J{hr+2}:J{hr+1+len(DEPT_RANGES)}"
C(ws,5,2,f'=IFERROR(SUMPRODUCT({wrng},{drng})/SUMPRODUCT({wrng},{hasrng}),0)',f=F_(22,True,ORANGE),fmt="0%",al="center")
for i,(lbl,key) in enumerate([("✅ محقق","✅ محقق"),("🟡 قريب","🟡 قريب"),("🔴 تحت الهدف","🔴 تحت الهدف")]):
    c=3+i; C(ws,4,c,lbl,f=F_(10,True,WHITE),fillc=STEEL,al="center")
    C(ws,5,c,f'=COUNTIFS({STC},"{key}",{DEPC},"<>{EXEC_NAME}")',f=F_(18,True,NAVY),al="center")  # دون التنفيذية (تفادي الازدواج)
C(ws,hr,2,"الأداء حسب الإدارة (موزون)",f=F_(12,True,WHITE),fillc=ORANGE,al="center"); ws.merge_cells(start_row=hr,start_column=2,end_row=hr,end_column=3)
for c,h in zip([2,3],["الإدارة","الأداء"]): C(ws,hr+1,c,h,f=F_(10,True,WHITE),fillc=STEEL,al="center")
rr=hr+2
for dname,sh,s,e in DEPT_RANGES:
    C(ws,rr,2,dname,f=F_(9),al="right")
    C(ws,rr,3,f'=IFERROR(SUMIF({DEPC},"{dname}",{WXN})/SUMIF({DEPC},"{dname}",{WAV}),"—")',f=F_(9,True,NAVY),fmt="0%",al="center")
    # أعمدة مساعدة مخفيّة: الأداء الرقمي · وزن الإدارة · هل لديها بيانات
    C(ws,rr,8,f'=IFERROR(SUMIF({DEPC},"{dname}",{WXN})/SUMIF({DEPC},"{dname}",{WAV}),0)',f=F_(8),al="center")
    C(ws,rr,9,round(DW_BY_NAME.get(dname,0),5),f=F_(8),al="center")
    C(ws,rr,10,f'=IF(SUMIF({DEPC},"{dname}",{WAV})=0,0,1)',f=F_(8),al="center")
    rr+=1
# ركيزة + منظور (يمين)
pc=4; C(ws,hr,pc,"حسب الركيزة",f=F_(12,True,WHITE),fillc=ORANGE,al="center"); ws.merge_cells(start_row=hr,start_column=pc,end_row=hr,end_column=pc+1)
for c,h in zip([pc,pc+1],["الركيزة","الأداء"]): C(ws,hr+1,c,h,f=F_(10,True,WHITE),fillc=STEEL,al="center")
rr=hr+2
for p in K.PILLARS:
    C(ws,rr,pc,p,f=F_(9),al="right"); C(ws,rr,pc+1,f'=IFERROR(AVERAGEIFS({ACH},{PILC},"{p}"),"—")',f=F_(9,True,NAVY),fmt="0%",al="center"); rr+=1
rr+=1; C(ws,rr,pc,"حسب منظور BSC",f=F_(11,True,WHITE),fillc=BLUE,al="center"); ws.merge_cells(start_row=rr,start_column=pc,end_row=rr,end_column=pc+1); rr+=1
for p in PERSPS:
    C(ws,rr,pc,p,f=F_(9),al="right"); C(ws,rr,pc+1,f'=IFERROR(AVERAGEIFS({ACH},{PERC},"{p}"),"—")',f=F_(9,True,NAVY),fmt="0%",al="center"); rr+=1

# ============ سجل المشاريع (كل إدارة) ============
ws=wb.create_sheet("سجل المشاريع"); rtl(ws)
for i,w in enumerate([4,20,40,10,12,10,14,30],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.merge_cells("A1:H1"); C(ws,1,1,"سجل مشاريع الشركة — لكل إدارة مشاريعها (نسبة الإنجاز = المنجز ÷ الكمية)",f=F_(14,True,WHITE),fillc=NAVY,al="center",border=False); ws.row_dimensions[1].height=24
for c,h in enumerate(["#","الإدارة","المشروع","الكمية","نوع الكمية","المنجز","نسبة الإنجاز","المؤشر المرتبط"],1): C(ws,2,c,h,f=F_(10,True,WHITE),fillc=BLUE,al="center")
r=3
def proj_block(dept, items, kpi_link):
    global r
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
    C(ws,r,1,f"▾ مشاريع {dept}",f=F_(10,True,WHITE),fillc=ORANGE,al="right"); r+=1
    for i,(nm,qty,qtype,done,pct) in enumerate(items,1):
        C(ws,r,1,i,f=F_(9,True),al="center"); C(ws,r,2,dept,f=F_(9),al="center")
        C(ws,r,3,nm,f=F_(9),al="right",wrap=True); C(ws,r,4,qty,al="center"); C(ws,r,5,qtype,al="center")
        C(ws,r,6,done,fillc=GREEN_IN,al="center",lock=False)
        ws.cell(r,7).value=f'=IFERROR(F{r}/D{r},0)'; C(ws,r,7,f=F_(9,True,NAVY),fmt="0%",al="center")
        C(ws,r,8,kpi_link,f=F_(8,color="666666"),al="right",wrap=True); ws.row_dimensions[r].height=22; r+=1
proj_block("إدارة التقنية", getattr(K,"IT_PROJECTS",[]), "نسبة إنجاز محفظة مشاريع التحول الرقمي")
# أماكن لمشاريع بقية الإدارات (تُعبّأ عند توفّرها)
for dname,sh,s,e in DEPT_RANGES:
    if dname=="التقنية الرقمية": continue
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
    C(ws,r,1,f"▾ مشاريع {dname}",f=F_(10,True,WHITE),fillc=STEEL,al="right"); r+=1
    for _ in range(3):
        for c in range(1,9): C(ws,r,c,None,fillc=GREEN_IN if c in (3,4,5,6) else None,lock=(c not in (3,4,5,6)))
        ws.cell(r,7).value=f'=IFERROR(F{r}/D{r},0)'; C(ws,r,7,f=F_(9,True,NAVY),fmt="0%",al="center")
        C(ws,r,2,dname,f=F_(9),al="center"); r+=1

# ============ بيانات الباوربي (جدول مسطّح بقيَم حقيقية — جاهز لـ Power BI/التحليل) ============
# Power BI لا يعيد حساب معادلات الإكسل، لذا نوفّر نسخة مسطّحة بقيَم محسوبة مسبقاً.
flat=wb.create_sheet("بيانات الباوربي"); rtl(flat)
FH=["الإدارة","المحور","المؤشر","النوع","الوحدة","القطبية","الأولوية","الوزن","المستهدف","المُحقَّق","نسبة التحقيق","الحالة","الركيزة","المشروع","منظور BSC","المالك","مصدر البيانات","الدورية","خط الأساس","القيمة السابقة","التنبؤ بنهاية السنة","النمو vs الأساس","حد القبول","حد التحذير","تعليق الانحراف","الإجراء التصحيحي","الحالة (نص)","ترتيب الحالة","لون الحالة"]
for i,w in enumerate([18,15,40,11,8,7,11,8,11,11,12,14,11,20,15,16,22,9,11,12,14,12,9,9,40,42,13,11,12],1): flat.column_dimensions[get_column_letter(i)].width=w
for c,h in enumerate(FH,1): C(flat,1,c,h,f=F_(10,True,WHITE),fillc=NAVY,al="center")
flat.row_dimensions[1].height=26
def ach_calc(pol,tgt,act):
    if tgt is None or act is None: return None
    if tgt==0: return 1.0 if act<=0 else 0.0
    if act==0: return 0.0
    return min(1.0,(tgt/act) if pol=="↓" else (act/tgt))   # سقف 100%
def status_txt(a):
    if a is None: return "⏳ بانتظار هدف"
    if a>=1: return "✅ محقق"
    if a>=0.85: return "🟡 قريب"
    return "🔴 تحت الهدف"
# أعمدة نظيفة لـ Power BI: نص بلا إيموجي + ترتيب للفرز + لون Hex للتنسيق الشرطي
_STMAP={"✅ محقق":("محقق",1,"2EAD6B"),"🟡 قريب":("قريب",2,"F2C94C"),
        "🔴 تحت الهدف":("تحت الهدف",3,"EB5757"),"⏳ بانتظار هدف":("بانتظار هدف",4,"BDBDBD")}
def clean_status(st): return _STMAP.get(st,("غير محدد",5,"BDBDBD"))
fr=2
for dname,key,records in K.DEPARTMENTS:
    weights=round_weights_pct(K.kpi_weights(records))
    for i,(axis,nm,unit,pol,agg,tgt,ttxt,fmt,pillar,project) in enumerate(records):
        act=eff_actual(nm,pol,tgt); a=ach_calc(pol,tgt,act); src=kpi_source(nm,axis,key)
        st=status_of(a,nm); gr,yl=thresholds(nm)
        base=baseline_of(nm,pol,act); prev=previous_of(act,base); fc=forecast_of(agg,act); gvb=growth_vs_base(act,base)
        if a is None or "محقق" in st: comm,actn="",""
        elif "تحت" in st:
            comm=f"دون المستهدف — فجوة {round((1-a)*100)}٪، يتطلب إجراءً تصحيحياً عاجلاً."
            actn=f"إطلاق إجراء تصحيحي خلال 30 يوماً بإشراف {kpi_owner(key,axis)}."
        else:
            comm=f"قريب من المستهدف — فجوة {round((1-a)*100)}٪، متابعة لإغلاقها."
            actn="متابعة أسبوعية حتى بلوغ المستهدف."
        C(flat,fr,1,dname,f=F_(9),al="right")
        C(flat,fr,2,axis,f=F_(9),al="right")
        C(flat,fr,3,nm,f=F_(9),al="right")
        C(flat,fr,4,kpi_type(nm),f=F_(8),al="center")
        C(flat,fr,5,unit,al="center")
        C(flat,fr,6,pol,al="center")
        C(flat,fr,7,prio_label(nm),f=F_(8),al="center")
        C(flat,fr,8,weights[i],fmt="0%",al="center")
        C(flat,fr,9,tgt if tgt is not None else None,al="center")
        C(flat,fr,10,act if act is not None else None,al="center")
        C(flat,fr,11,a if a is not None else None,fmt="0%",al="center")
        C(flat,fr,12,st,f=F_(9),al="center")
        C(flat,fr,13,pillar,f=F_(9),al="center")
        C(flat,fr,14,project,f=F_(8),al="right")
        C(flat,fr,15,K.perspective(nm),f=F_(8),al="center")
        C(flat,fr,16,kpi_owner(key,axis),f=F_(8),al="center")
        C(flat,fr,17,src,f=F_(8),al="right")
        C(flat,fr,18,kpi_freq(nm,ttxt),f=F_(8),al="center")
        C(flat,fr,19,base,al="center")
        C(flat,fr,20,prev,al="center")
        C(flat,fr,21,fc,al="center")
        C(flat,fr,22,gvb if gvb is not None else None,fmt="0%",al="center")
        C(flat,fr,23,gr,fmt="0%",al="center")
        C(flat,fr,24,yl,fmt="0%",al="center")
        C(flat,fr,25,comm,f=F_(8,color="C00000" if "تحت" in st else "9C6500"),al="right",wrap=True)
        C(flat,fr,26,actn,f=F_(8,color="444444"),al="right",wrap=True)
        st_txt,st_ord,st_hex=clean_status(st)
        C(flat,fr,27,st_txt,f=F_(9),al="center")
        C(flat,fr,28,st_ord,f=F_(9),al="center")
        C(flat,fr,29,"#"+st_hex,f=F_(8,color=st_hex),al="center")
        fr+=1
flat.freeze_panes="D2"; flat.auto_filter.ref=f"A1:{get_column_letter(len(FH))}{fr-1}"
wb.remove(flat)  # حُذفت ورقة «بيانات الباوربي» بناءً على الطلب

# ============ بطاقة تعريف المؤشرات (قاموس البيانات — حوكمة) ============
dic=wb.create_sheet("بطاقة تعريف المؤشرات"); rtl(dic)
DH=["#","الإدارة","المؤشر","التعريف","طريقة الاحتساب","النوع","الهدف الاستراتيجي (ركيزة · مشروع)","منظور BSC","الوزن","المالك","مصدر البيانات","الدورية","جهة اعتماد البيانات","آخر تحديث للبيانات"]
for i,w in enumerate([4,16,38,46,30,11,30,16,8,16,24,9,18,16],1): dic.column_dimensions[get_column_letter(i)].width=w
dic.merge_cells(f"A1:{get_column_letter(len(DH))}1"); C(dic,1,1,"بطاقة تعريف المؤشرات — قاموس البيانات والحوكمة (مالك · مصدر · دورية · تعريف · احتساب · اعتماد)",f=F_(12,True,WHITE),fillc=NAVY,al="center",border=False); dic.row_dimensions[1].height=24
for c,h in enumerate(DH,1): C(dic,2,c,h,f=F_(9,True,WHITE),fillc=BLUE,al="center")
dic.row_dimensions[2].height=26
dr=3; seq=1
for dname,key,records in K.DEPARTMENTS:
    weights=round_weights_pct(K.kpi_weights(records))
    for i,(axis,nm,unit,pol,agg,tgt,ttxt,fmt,pillar,project) in enumerate(records):
        src=kpi_source(nm,axis,key)
        C(dic,dr,1,seq,f=F_(8),al="center")
        C(dic,dr,2,dname,f=F_(8),al="right")
        C(dic,dr,3,nm,f=F_(9),al="right",wrap=True)
        C(dic,dr,4,kpi_def(nm,axis,pillar),f=F_(8,color="444444"),al="right",wrap=True)
        C(dic,dr,5,kpi_formula(fmt,unit,nm),f=F_(8),al="right",wrap=True)
        C(dic,dr,6,kpi_type(nm),f=F_(8),al="center")
        C(dic,dr,7,f"{pillar} · {project}",f=F_(8),al="right",wrap=True)
        C(dic,dr,8,K.perspective(nm),f=F_(8),al="center")
        C(dic,dr,9,weights[i],fmt="0%",al="center")
        C(dic,dr,10,kpi_owner(key,axis),f=F_(8),al="center")
        C(dic,dr,11,src,f=F_(8),al="right",wrap=True)
        C(dic,dr,12,kpi_freq(nm,ttxt),f=F_(8),al="center")
        C(dic,dr,13,kpi_approver(src,key),f=F_(8),al="center")
        C(dic,dr,14,None,fillc=GREEN_IN,al="center",lock=False)  # يُدخله مالك المؤشر
        dic.row_dimensions[dr].height=30; dr+=1; seq+=1
dic.freeze_panes="C3"; dic.auto_filter.ref=f"A2:{get_column_letter(len(DH))}{dr-1}"

# ============ مراجعة وحوكمة المؤشرات (تحليل) ============
gov=wb.create_sheet("مراجعة وحوكمة المؤشرات"); rtl(gov)
for i,w in enumerate([4,24,14,14,14,14,16],1): gov.column_dimensions[get_column_letter(i)].width=w
gov.merge_cells("A1:G1"); C(gov,1,1,"مراجعة وحوكمة المؤشرات — التوازن · النوع · الأوزان · تغطية الأهداف",f=F_(13,True,WHITE),fillc=NAVY,al="center",border=False); gov.row_dimensions[1].height=24
# تجميع إحصائي
from collections import Counter,defaultdict
dep_cnt=Counter(); dep_out=Counter(); dep_act=Counter(); dep_ops=Counter(); dep_wsum=defaultdict(float)
persp_cnt=Counter(); proj_cnt=Counter()
for dname,key,records in K.DEPARTMENTS:
    weights=round_weights_pct(K.kpi_weights(records))
    for i,(axis,nm,unit,pol,agg,tgt,ttxt,fmt,pillar,project) in enumerate(records):
        dep_cnt[dname]+=1; dep_wsum[dname]+=weights[i]
        t=kpi_type(nm)
        if "نتيجة" in t: dep_out[dname]+=1
        elif "نشاط" in t: dep_act[dname]+=1
        else: dep_ops[dname]+=1
        persp_cnt[K.perspective(nm)]+=1; proj_cnt[project]+=1
r=3
C(gov,r,1,"أ) لكل إدارة: العدد · مجموع الأوزان · النوع",f=F_(11,True,WHITE),fillc=ORANGE,al="center"); gov.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); r+=1
for c,h in enumerate(["#","الإدارة","عدد المؤشرات","مجموع الأوزان","🎯 نتيجة","⚙️ نشاط","🔧 تشغيلي"],1): C(gov,r,c,h,f=F_(9,True,WHITE),fillc=BLUE,al="center")
r+=1; s=1
for dname,key,_ in K.DEPARTMENTS:
    C(gov,r,1,s,f=F_(8),al="center"); C(gov,r,2,dname,f=F_(8),al="right")
    C(gov,r,3,dep_cnt[dname],f=F_(8),al="center")
    C(gov,r,4,round(dep_wsum[dname],2),fmt="0%",f=F_(8,True,"3FB27F" if abs(dep_wsum[dname]-1)<0.005 else "C00000"),al="center")
    C(gov,r,5,dep_out[dname],f=F_(8),al="center"); C(gov,r,6,dep_act[dname],f=F_(8),al="center"); C(gov,r,7,dep_ops[dname],f=F_(8),al="center")
    r+=1; s+=1
r+=1; C(gov,r,1,"ب) توزيع منظورات Balanced Scorecard",f=F_(11,True,WHITE),fillc=ORANGE,al="center"); gov.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); r+=1
for c,h in enumerate(["#","المنظور","عدد المؤشرات","النسبة"],1): C(gov,r,c,h,f=F_(9,True,WHITE),fillc=BLUE,al="center")
r+=1; s=1; tot=sum(persp_cnt.values())
for p in PERSPS:
    C(gov,r,1,s,f=F_(8),al="center"); C(gov,r,2,p,f=F_(8),al="right"); C(gov,r,3,persp_cnt[p],f=F_(8),al="center")
    C(gov,r,4,round(persp_cnt[p]/tot,3),fmt="0%",f=F_(8),al="center"); r+=1; s+=1
r+=1; C(gov,r,1,"ج) تغطية الأهداف الاستراتيجية (عدد المؤشرات لكل مشروع)",f=F_(11,True,WHITE),fillc=ORANGE,al="center"); gov.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); r+=1
for c,h in enumerate(["#","المشروع/الهدف","الركيزة","عدد المؤشرات","التغطية"],1): C(gov,r,c,h,f=F_(9,True,WHITE),fillc=BLUE,al="center")
r+=1; s=1
for p in K.PROJECTS:
    n=proj_cnt.get(p,0)
    cov="🔴 فجوة" if n==0 else ("🟡 ضعيفة" if n<3 else "🟢 كافية")
    C(gov,r,1,s,f=F_(8),al="center"); C(gov,r,2,p,f=F_(8),al="right"); C(gov,r,3,K.PROJECT_PILLAR.get(p,""),f=F_(8),al="center")
    C(gov,r,4,n,f=F_(8),al="center"); C(gov,r,5,cov,f=F_(8),al="center"); r+=1; s+=1
gov.freeze_panes="A2"

# ============ تحليل الانحراف (تقرير الاستثناءات) ============
va=wb.create_sheet("تحليل الانحراف"); rtl(va)
VH=["#","الإدارة","المؤشر","المُحقَّق","المستهدف","نسبة التحقيق","الفجوة","التنبؤ","الحالة","السبب","الإجراء التصحيحي","المالك","الموعد"]
for i,w in enumerate([4,16,38,11,11,11,9,11,13,34,34,16,14],1): va.column_dimensions[get_column_letter(i)].width=w
va.merge_cells(f"A1:{get_column_letter(len(VH))}1"); C(va,1,1,"تحليل الانحراف — المؤشرات دون المستهدف مع السبب والإجراء (تقرير الاستثناءات)",f=F_(12,True,WHITE),fillc=NAVY,al="center",border=False); va.row_dimensions[1].height=24
for c,h in enumerate(VH,1): C(va,2,c,h,f=F_(9,True,WHITE),fillc=BLUE,al="center")
exc=[]
for dname,key,records in K.DEPARTMENTS:
    for (axis,nm,unit,pol,agg,tgt,ttxt,fmt,pillar,project) in records:
        act=eff_actual(nm,pol,tgt); a=ach_calc(pol,tgt,act); st=status_of(a,nm)
        if a is None or "محقق" in st: continue
        exc.append((a,dname,nm,act,tgt,st,forecast_of(agg,act),kpi_owner(key,axis)))
exc.sort(key=lambda x:x[0])
vr=3
for j,(a,dname,nm,act,tgt,st,fc,owner) in enumerate(exc,1):
    C(va,vr,1,j,f=F_(8),al="center"); C(va,vr,2,dname,f=F_(8),al="right"); C(va,vr,3,nm,f=F_(9),al="right",wrap=True)
    C(va,vr,4,act,al="center"); C(va,vr,5,tgt,al="center")
    C(va,vr,6,a,fmt="0%",f=F_(8,True,NAVY),al="center"); C(va,vr,7,1-a,fmt="0%",f=F_(8,True,"C00000"),al="center")
    C(va,vr,8,fc,al="center"); C(va,vr,9,st,f=F_(8),al="center")
    C(va,vr,10,("انحراف جوهري — أولوية عاجلة." if "تحت" in st else "فجوة محدودة — متابعة."),f=F_(8),al="right",wrap=True)
    C(va,vr,11,("إجراء تصحيحي خلال 30 يوماً." if "تحت" in st else "متابعة أسبوعية لإغلاق الفجوة."),f=F_(8),al="right",wrap=True)
    C(va,vr,12,owner,f=F_(8),al="center"); C(va,vr,13,None,fillc=GREEN_IN,al="center",lock=False)
    vr+=1
va.freeze_panes="A3"; va.auto_filter.ref=f"A2:{get_column_letter(len(VH))}{vr-1}"

# ============ سجل المخاطر المؤسسية ============
RISKS=[
 ("تقلّب أسعار الوقود وضغط الهامش","مالية/سوق",4,5,"هامش EBITDA","المدير المالي","تنويع مصادر الدخل ومراجعة التسعير دورياً","نشط"),
 ("تعثّر تحصيل المستحقات","مالية/ائتمان",3,4,"نسبة التحصيل العام (الشركة)","المدير المالي","سياسة ائتمان أصرم ومتابعة DSO أسبوعياً","نشط"),
 ("اختراق/انقطاع الأنظمة الرقمية","تقني/سيبراني",3,5,"نسبة الامتثال للأمن السيبراني","مدير التقنية","معايير أمن سيبراني وخطة استمرارية الأعمال","نشط"),
 ("تأخّر مشاريع التحول الرقمي","تقني/تشغيلي",4,4,"نسبة إنجاز محفظة مشاريع التحول الرقمي","مدير التقنية","حوكمة PMO وتقارير معالم أسبوعية","نشط"),
 ("نقص الكوادر وتحديات السعودة","موارد بشرية",3,3,"نسبة السعودة / التوطين","مدير الموارد البشرية","خطة استقطاب وتطوير واستبقاء","نشط"),
 ("حوادث السلامة في المحطات","تشغيلي/سلامة",2,5,"الالتزام بمعايير السلامة داخل المحطة","مدير التشغيل","تدقيق سلامة دوري وتدريب الفرق","نشط"),
 ("عدم مطابقة جودة الوقود","تشغيلي/جودة",2,5,"نسبة مطابقة جودة الوقود","مدير الجودة","فحوصات مخبرية ومعايرة دورية","نشط"),
 ("تركّز جغرافي للعقود","استراتيجي/توسع",3,3,"نسبة العقود التي تغطي المناطق الـ13","مدير الاستثمار","تنويع المواقع خارج المدن الرئيسية","نشط"),
 ("ضغط السيولة والتدفق النقدي","مالية/سيولة",3,4,"نسبة التدفق النقدي التشغيلي","المدير المالي","إدارة رأس المال العامل وخطوط ائتمان","نشط"),
 ("مخاطر الامتثال التنظيمي","حوكمة/امتثال",2,4,"الالتزام بمعايير وزارة الطاقة (تغطية 13 منطقة)","المدير القانوني","متابعة المتطلبات النظامية والاعتمادات","نشط"),
 ("انخفاض إشغال الوحدات العقارية","عقاري/سوق",3,3,"نسبة الإشغال","مدير العقار","تكثيف التأجير وتنويع المستأجرين","نشط"),
 ("تراجع سمعة العلامة والشكاوى","سمعة/عملاء",2,4,"مؤشر رضا العملاء العام (الشركة)","مدير التسويق","إدارة تجربة العميل ومعالجة الشكاوى بسرعة","نشط"),
]
rk=wb.create_sheet("سجل المخاطر"); rtl(rk)
RKH=["#","الخطر","الفئة","الاحتمال (1-5)","الأثر (1-5)","درجة الخطر","المستوى","المؤشر المرتبط","المالك","إجراء التخفيف","الحالة"]
for i,w in enumerate([4,36,16,12,12,11,12,30,18,38,9],1): rk.column_dimensions[get_column_letter(i)].width=w
rk.merge_cells(f"A1:{get_column_letter(len(RKH))}1"); C(rk,1,1,"سجل المخاطر المؤسسية — مربوط بالمؤشرات (درجة الخطر = الاحتمال × الأثر)",f=F_(12,True,WHITE),fillc=NAVY,al="center",border=False); rk.row_dimensions[1].height=24
for c,h in enumerate(RKH,1): C(rk,2,c,h,f=F_(9,True,WHITE),fillc=BLUE,al="center")
rkr=3
for j,(risk,cat,p,im,kpi,owner,mit,stt) in enumerate(RISKS,1):
    C(rk,rkr,1,j,f=F_(8),al="center"); C(rk,rkr,2,risk,f=F_(9),al="right",wrap=True); C(rk,rkr,3,cat,f=F_(8),al="center")
    C(rk,rkr,4,p,fillc=GREEN_IN,al="center",lock=False); C(rk,rkr,5,im,fillc=GREEN_IN,al="center",lock=False)
    rk.cell(rkr,6).value=f"=D{rkr}*E{rkr}"; C(rk,rkr,6,f=F_(9,True,NAVY),al="center")
    rk.cell(rkr,7).value=f'=IF(F{rkr}>=15,"🔴 عالي",IF(F{rkr}>=8,"🟠 متوسط","🟢 منخفض"))'; C(rk,rkr,7,f=F_(8,True),al="center")
    C(rk,rkr,8,kpi,f=F_(8),al="right",wrap=True); C(rk,rkr,9,owner,f=F_(8),al="center")
    C(rk,rkr,10,mit,f=F_(8),al="right",wrap=True); C(rk,rkr,11,stt,f=F_(8),al="center"); rkr+=1
rk.freeze_panes="A3"; rk.auto_filter.ref=f"A2:{get_column_letter(len(RKH))}{rkr-1}"

# ============ ESG والاستدامة ============
ESG=[
 ("E — البيئة","نسبة التشجير داخل المحطات (معيار الوزارة)","%","↑",1.0,"100%"),
 ("E — البيئة","نسبة تخفيض انبعاثات الكربون (مقابل الأساس)","%","↑",0.10,"≥ 10%"),
 ("E — البيئة","كفاءة استهلاك الطاقة بالمحطات","%","↑",0.90,"≥ 90%"),
 ("E — البيئة","نسبة إدارة وإعادة تدوير المخلفات","%","↑",0.80,"≥ 80%"),
 ("E — البيئة","حالات الانسكاب/التسرّب البيئي","حالة","↓",0,"صفر"),
 ("S — الاجتماعي","نسبة السعودة / التوطين","%","↑",0.30,"≥ 30%"),
 ("S — الاجتماعي","معدل حوادث السلامة (TRIR)","معدل","↓",0,"≈ صفر"),
 ("S — الاجتماعي","مؤشر رضا الموظفين (eNPS)","نقطة","↑",30,"≥ 30"),
 ("S — الاجتماعي","متوسط ساعات التدريب لكل موظف","ساعة","↑",40,"≥ 40"),
 ("S — الاجتماعي","مبادرات المسؤولية المجتمعية","مبادرة","↑",6,"≥ 6"),
 ("G — الحوكمة","نسبة الالتزام بالامتثال التنظيمي","%","↑",1.0,"100%"),
 ("G — الحوكمة","نسبة اعتماد السياسات والإجراءات","%","↑",0.95,"≥ 95%"),
 ("G — الحوكمة","نسبة إغلاق ملاحظات التدقيق الداخلي","%","↑",0.90,"≥ 90%"),
 ("G — الحوكمة","نسبة الإفصاح والشفافية","%","↑",0.95,"≥ 95%"),
]
es=wb.create_sheet("ESG والاستدامة"); rtl(es)
EH=["#","المحور","المؤشر","الوحدة","القطبية","المستهدف","نص المستهدف","المُحقَّق","نسبة التحقيق","الحالة"]
for i,w in enumerate([4,16,42,10,7,11,14,12,12,15],1): es.column_dimensions[get_column_letter(i)].width=w
es.merge_cells(f"A1:{get_column_letter(len(EH))}1"); C(es,1,1,"ESG والاستدامة — البيئة · الاجتماعي · الحوكمة (متطلب حوكمة المؤسسات الكبرى)",f=F_(12,True,WHITE),fillc=NAVY,al="center",border=False); es.row_dimensions[1].height=24
for c,h in enumerate(EH,1): C(es,2,c,h,f=F_(9,True,WHITE),fillc=BLUE,al="center")
er=3
for j,(axis,nm,unit,pol,tgt,ttxt) in enumerate(ESG,1):
    act=eff_actual(nm,pol,tgt); a=ach_calc(pol,tgt,act); st=status_of(a,nm)
    C(es,er,1,j,f=F_(8),al="center"); C(es,er,2,axis,f=F_(8),al="center"); C(es,er,3,nm,f=F_(9),al="right",wrap=True)
    C(es,er,4,unit,al="center"); C(es,er,5,pol,al="center"); C(es,er,6,tgt,al="center"); C(es,er,7,ttxt,f=F_(8),al="center")
    C(es,er,8,act,fillc=GREEN_IN,al="center",lock=False); C(es,er,9,a if a is not None else None,fmt="0%",f=F_(8,True,NAVY),al="center")
    C(es,er,10,st,f=F_(8),al="center"); er+=1
es.freeze_panes="A3"; es.auto_filter.ref=f"A2:{get_column_letter(len(EH))}{er-1}"

# ============ تقرير المجلس (One-Pager تنفيذي حي) ============
bd=wb.create_sheet("تقرير المجلس"); rtl(bd)
for i,w in enumerate([3,32,18,18,18,18],1): bd.column_dimensions[get_column_letter(i)].width=w
logo(bd,"B1")
bd.merge_cells("B2:F2"); C(bd,2,2,"تقرير الأداء التنفيذي — مجلس الإدارة · درب 2026",f=F_(16,True,WHITE),fillc=NAVY,al="center",border=False); bd.row_dimensions[2].height=28
C(bd,4,2,"الإنجاز العام الموزون",f=F_(10,True,WHITE),fillc=BLUE,al="center")
bd.cell(5,2).value="='اللوحة الموجزة'!B5"; C(bd,5,2,f=F_(22,True,ORANGE),fmt="0%",al="center")
for i,(lbl,col) in enumerate([("✅ محقق","C"),("🟡 قريب","D"),("🔴 تحت الهدف","E")]):
    c=3+i; C(bd,4,c,lbl,f=F_(9,True,WHITE),fillc=STEEL,al="center")
    bd.cell(5,c).value=f"='اللوحة الموجزة'!{col}5"; C(bd,5,c,f=F_(18,True,NAVY),al="center")
C(bd,7,2,"الأداء حسب الركيزة",f=F_(11,True,WHITE),fillc=ORANGE,al="center"); bd.merge_cells("B7:C7")
br=8
for p in K.PILLARS:
    C(bd,br,2,p,f=F_(9),al="right"); bd.cell(br,3).value=f'=IFERROR(AVERAGEIFS({ACH},{PILC},"{p}"),"—")'; C(bd,br,3,f=F_(9,True,NAVY),fmt="0%",al="center"); br+=1
C(bd,7,4,"أبرز 5 انحرافات",f=F_(11,True,WHITE),fillc=ORANGE,al="center"); bd.merge_cells("D7:F7")
b2=8
for (a,dname,nm,act,tgt,st,fc,owner) in exc[:5]:
    C(bd,b2,4,nm,f=F_(8),al="right",wrap=True); bd.merge_cells(start_row=b2,start_column=4,end_row=b2,end_column=5)
    C(bd,b2,6,a,fmt="0%",f=F_(8,True,"C00000"),al="center"); b2+=1
rr=max(br,b2)+1
C(bd,rr,2,"أعلى 5 مخاطر مؤسسية (الأثر × الاحتمال)",f=F_(11,True,WHITE),fillc=ORANGE,al="center"); bd.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=6); rr+=1
for (risk,cat,p,im,kpi,owner,mit,stt) in sorted(RISKS,key=lambda x:x[2]*x[3],reverse=True)[:5]:
    C(bd,rr,2,risk,f=F_(8),al="right"); bd.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=4)
    C(bd,rr,5,f"{p}×{im}={p*im}",f=F_(8,True,"C00000"),al="center")
    C(bd,rr,6,("🔴 عالي" if p*im>=15 else "🟠 متوسط" if p*im>=8 else "🟢 منخفض"),f=F_(8),al="center"); rr+=1
rr+=1; bd.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=6); C(bd,rr,2,"يُحدَّث آلياً من اللوحة الموجزة وتحليل الانحراف وسجل المخاطر · سرّي — لأعضاء المجلس",f=F_(8,False,"888888"),al="center",border=False)

# ============ الخريطة الاستراتيجية (Strategy Map — علاقات السبب والأثر) ============
sm=wb.create_sheet("الخريطة الاستراتيجية"); rtl(sm)
for i in range(2,12): sm.column_dimensions[get_column_letter(i)].width=12
C_FIN,C_CUS,C_INT,C_LRN="F47A21","5B9BD5","808285","3FB27F"
sm.merge_cells("B2:K2"); C(sm,2,2,"الخريطة الاستراتيجية — علاقات السبب والأثر (Balanced Scorecard) · الأداء حي",f=F_(14,True,WHITE),fillc=NAVY,al="center",border=False); sm.row_dimensions[2].height=26
def sm_header(r,color,label):
    sm.merge_cells(start_row=r,start_column=2,end_row=r,end_column=11)
    C(sm,r,2,label,f=F_(11,True,WHITE),fillc=color,al="center")
def sm_box(r,c1,c2,name,formula,color):
    sm.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2); C(sm,r,c1,name,f=F_(8,True,WHITE),fillc=color,al="center",wrap=True)
    sm.merge_cells(start_row=r+1,start_column=c1,end_row=r+1,end_column=c2); sm.cell(r+1,c1).value=formula; C(sm,r+1,c1,f=F_(12,True,WHITE),fillc=color,fmt="0%",al="center")
def sm_row(r,color,boxes):
    n=len(boxes); w=10//n; cc=2
    for i,(nm,fm) in enumerate(boxes):
        c2=11 if i==n-1 else cc+w-1; sm_box(r,cc,c2,nm,fm,color); cc=c2+1
def sm_arrows(r):
    for c in (3,5,7,9,11): C(sm,r,c,"⬆",f=F_(11,True,"AAAAAA"),al="center",border=False)
def pf(proj): return f'=IFERROR(AVERAGEIFS({ACH},{PROJC},"{proj}"),"—")'
def perf(persp): return f'=IFERROR(AVERAGEIFS({ACH},{PERC},"{persp}"),"—")'
# المالي (الأعلى) — النتيجة النهائية
sm_header(4,NAVY,"⬆ المنظور المالي — القيمة للمساهمين")
sm_box(5,2,11,"النمو · الربحية · التدفّق النقدي والتحصيل",perf("مالي"),C_FIN)
sm_arrows(7)
# العملاء
sm_header(8,NAVY,"المنظور — العملاء")
sm_row(9,C_CUS,[("زيادة وصول العملاء",pf(K.P_REACH)),("تعزيز تجربة العملاء",pf(K.P_CX)),("تطبيق تانكي والولاء",pf(K.P_TANKI))])
sm_arrows(11)
# العمليات الداخلية
sm_header(12,NAVY,"المنظور — العمليات الداخلية")
sm_row(13,C_INT,[("التوسع في المواقع",pf(K.P_GROW1)),("الامتياز التجاري",pf(K.P_FRAN)),("منظومة التأجير",pf(K.P_LEASE)),("ساحات درب",pf(K.P_SAHAT)),("جودة التشغيل",pf(K.P_QUAL))])
sm_arrows(15)
# التعلّم والنمو (الأساس)
sm_header(16,NAVY,"المنظور — التعلّم والنمو (الأساس الممكِّن)")
sm_row(17,C_LRN,[("الاستثمار في الموظفين",pf(K.P_PEOPLE)),("التحول التقني",pf(K.P_TECH)),("إطلاق مشاريع جديدة",pf(K.P_NEW)),("التصاميم والهوية",pf(K.P_DESIGN))])
sm.merge_cells("B19:K19"); C(sm,19,2,"القراءة من الأسفل للأعلى: تطوير الكوادر والتقنية ⟵ يُمكّن العمليات ⟵ يخلق قيمة للعملاء ⟵ يُنتج النتائج المالية",f=F_(8,False,"666666"),al="center",border=False)

# ============ التتبّع الشهري (إدخال شهري · المُحقَّق يُحسب تلقائياً حسب نوع المؤشر · وتيرة) ============
mt=wb.create_sheet("التتبّع الشهري"); rtl(mt)
MONTHS=["يناير","فبراير","مارس","أبريل","مايو","يونيو","يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]
AGG_AR={"SUM":"تراكمي","AVG":"معدل","LAST":"لقطة"}
MH=["#","الإدارة","المؤشر","الدورية","التجميع","القطبية","المستهدف"]+MONTHS+["المُحقَّق (تراكمي)","نسبة التحقيق","الحالة","المتوقّع حتى الآن","الوتيرة"]
MW=[4,15,38,9,9,7,11]+[8]*12+[13,12,13,13,14]
for i,w in enumerate(MW,1): mt.column_dimensions[get_column_letter(i)].width=w
last=get_column_letter(len(MH))
mt.merge_cells(f"A1:{last}1"); C(mt,1,1,"التتبّع الشهري — أدخل رقم كل شهر · «المُحقَّق» ونسبة التحقيق والحالة والوتيرة تُحسب تلقائياً",f=F_(12,True,WHITE),fillc=NAVY,al="center",border=False); mt.row_dimensions[1].height=24
mt.merge_cells(f"A2:{last}2"); C(mt,2,1,"🟩 خلايا الأشهر للإدخال · تراكمي=مجموع الأشهر · معدل=متوسط المُدخل · لقطة=آخر شهر · الوتيرة: 🟢 على المسار / 🟠 متأخّر مقابل المتوقّع",f=F_(8,False,NAVY),fillc=GOLD,al="center",border=False); mt.row_dimensions[2].height=18
for c,h in enumerate(MH,1): C(mt,3,c,h,f=F_(8,True,WHITE),fillc=BLUE,al="center")
mt.row_dimensions[3].height=30
M1=get_column_letter(8); M12=get_column_letter(19)   # نطاق الأشهر H..S
mr=4
for dname,key,records in K.DEPARTMENTS:
    for (axis,nm,unit,pol,agg,tgt,ttxt,fmt,pillar,project) in records:
        rng=f"${M1}{mr}:${M12}{mr}"; gr,yl=thresholds(nm)
        C(mt,mr,1,mr-3,f=F_(8,True),al="center")
        C(mt,mr,2,dname,f=F_(8),al="right")
        C(mt,mr,3,nm,f=F_(8),al="right",wrap=True)
        C(mt,mr,4,kpi_freq(nm,ttxt),f=F_(8),al="center")
        C(mt,mr,5,AGG_AR.get(agg,agg),f=F_(8),al="center")
        C(mt,mr,6,pol,al="center")
        C(mt,mr,7,tgt if tgt is not None else None,fmt=FMT[fmt] if tgt is not None else None,al="center")
        for mc in range(8,20): C(mt,mr,mc,None,fillc=GREEN_IN,fmt=FMT[fmt],al="center",lock=False)  # إدخال
        # المُحقَّق (تراكمي) — حسب نوع التجميع
        if agg=="SUM":   ytd=f'=IF(COUNT({rng})=0,"",SUM({rng}))'
        elif agg=="AVG": ytd=f'=IF(COUNT({rng})=0,"",SUM({rng})/COUNT({rng}))'
        else:            ytd=f'=IFERROR(LOOKUP(2,1/({rng}<>""),{rng}),"")'   # LAST
        mt.cell(mr,20).value=ytd; C(mt,mr,20,f=F_(8,True,NAVY),fmt=FMT[fmt],al="center")
        # نسبة التحقيق (سقف 100% + قطبية) — G=المستهدف · T=المُحقَّق · F=القطبية
        mt.cell(mr,21).value=(f'=IF($G{mr}="","",IF($T{mr}="","",MIN(1,IF($G{mr}=0,IF($T{mr}<=0,1,0),'
            f'IF($F{mr}="↓",IFERROR($G{mr}/$T{mr},0),IFERROR($T{mr}/$G{mr},0))))))')
        C(mt,mr,21,f=F_(8,True,NAVY),fmt="0%",al="center")
        mt.cell(mr,22).value=f'=IF($U{mr}="","⏳ بانتظار هدف",IF($U{mr}>={gr},"✅ محقق",IF($U{mr}>={yl},"🟡 قريب","🔴 تحت الهدف")))'
        C(mt,mr,22,f=F_(8,True),al="center")
        # المتوقّع حتى الآن: تراكمي يتناسب مع الأشهر المُدخلة · غيره = المستهدف
        if agg=="SUM": mt.cell(mr,23).value=f'=IF($G{mr}="","",$G{mr}*COUNT({rng})/12)'
        else:          mt.cell(mr,23).value=f'=IF($G{mr}="","",$G{mr})'
        C(mt,mr,23,fmt=FMT[fmt],al="center",f=F_(8,color="666666"))
        # الوتيرة: تراكمي يقارن النسبة بنسبة الأشهر المنقضية · غيره بعتبة القرب
        if agg=="SUM": mt.cell(mr,24).value=f'=IF($U{mr}="","—",IF($U{mr}>=COUNT({rng})/12,"🟢 على المسار","🟠 متأخّر"))'
        else:          mt.cell(mr,24).value=f'=IF($U{mr}="","—",IF($U{mr}>={yl},"🟢 على المسار","🟠 متأخّر"))'
        C(mt,mr,24,f=F_(8,True),al="center")
        mt.row_dimensions[mr].height=22; mr+=1
mt.freeze_panes="H4"; mt.auto_filter.ref=f"A3:{last}{mr-1}"

# ترتيب الأوراق
order=["تقرير المجلس","أهم المؤشرات (Top 5)","المؤشرات الجامعة","الاستراتيجية والمستهدفات","الخريطة الاستراتيجية","اللوحة الموجزة","تحليل الانحراف","التتبّع الشهري","سجل المخاطر","ESG والاستدامة","مراجعة وحوكمة المؤشرات","بطاقة تعريف المؤشرات"]+[sh for _,sh,_,_ in DEPT_RANGES]+["سجل المشاريع","قاعدة المؤشرات"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
for s in wb.worksheets: rtl(s)
cons.sheet_state="hidden"

out="درب-المؤشرات-والمشاريع-الموحّد-2026.xlsx"
wb.save(out)
print("saved:",out,os.path.getsize(out),"bytes | إدارات:",len(DEPT_RANGES),"| مؤشرات:",CE-1)
