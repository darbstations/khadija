#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
تحديث مصنّف خديجة ليتوافق مع المرجع الرسمي لمؤشرات الإدارة.

ماذا يفعل هذا السكربت (داخل نفس مصنّف الإكسل):
  1) يضيف ورقة جديدة «📋 مؤشرات الإدارة» تحوي القائمة الرسمية المعتمدة (5 مؤشرات)
     منقولة حرفياً من ملف الإدارة، مع عمود يربط كل مؤشر بما يقابله في اللوحة الآلية.
  2) يوائم «طريقة القياس» لمؤشّرَين فقط في اللوحة الآلية ليطابقا تعريف الإدارة:
       • توفر الوقود (KPI 1): المقام = (عدد نقاط المنتج × عدد أيام الشهر الفعلي)
         بدل الثابت 9000، مع إدخال «عدد نقاط المنتج» في خلية قابلة للتعديل ($F$38).
       • فواقد الوقود (KPI 6): القيمة = إجمالي الفاقد الشهري ÷ عدد الردود الشهرية
         بدل متوسط القيم المُطبّعة لكل 33 ألف لتر.
  3) المؤشرات الثلاثة الأخرى المشتركة (3، 4، 8) مطابقة أصلاً فلا تتغيّر دوالها.
  4) المؤشرات الثمانية غير المدرجة في قائمة الإدارة تبقى كما هي دون مساس.

كل التنسيقات (RTL، التنسيق الشرطي، التحقق من القوائم) محفوظة عبر openpyxl.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

WB_PATH  = "workbook/khadija-supply-chain-kpis-2026.xlsx"
SRC_PATH = "workbook/source-management-kpis.xlsx"

DASH   = "📊 ملخص آلي + KPIs"
OUTAGE = "🛢️ انقطاعات الوقود"
RETURN = "💧 الردود والفاقد"

# ---- هوية بصرية مطابقة للوحة ----
ORANGE, GRAY, CREAM = "FFC04E0E", "FF3D3D3D", "FFFEF0E6"
WHITE, YELLOW       = "FFFFFFFF", "FFFFF9D6"
FONT = "DIN Next LT Arabic"
thin = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def f_kpi1(m: int) -> str:
    """توفر الوقود = 1 − (إجمالي مدد الانقطاعات ÷ (عدد نقاط المنتج × أيام الشهر))."""
    return (
        f"=IFERROR(1-(SUMIFS('{OUTAGE}'!F5:F104,"
        f"'{OUTAGE}'!B5:B104,\">=\"&DATE(2026,{m},1),"
        f"'{OUTAGE}'!B5:B104,\"<\"&DATE(2026,{m}+1,1))"
        f"/($F$38*DAY(DATE(2026,{m}+1,1)-1))),1)"
    )


def f_kpi6(m: int) -> str:
    """فواقد الوقود = إجمالي كمية الفاقد الشهري ÷ عدد الردود الشهرية."""
    return (
        f"=IFERROR(SUMIFS('{RETURN}'!G5:G304,"
        f"'{RETURN}'!B5:B304,\">=\"&DATE(2026,{m},1),"
        f"'{RETURN}'!B5:B304,\"<\"&DATE(2026,{m}+1,1))"
        f"/COUNTIFS('{RETURN}'!B5:B304,\">=\"&DATE(2026,{m},1),"
        f"'{RETURN}'!B5:B304,\"<\"&DATE(2026,{m}+1,1)),0)"
    )


def main() -> None:
    wb = openpyxl.load_workbook(WB_PATH)
    try:
        wb.calculation.fullCalcOnLoad = True  # إجبار Excel على إعادة الحساب عند الفتح
    except Exception:
        pass

    dash = wb[DASH]

    # (2) تحديث دوال المؤشرين 1 و 6 لكل الأشهر (الأعمدة G..R = 7..18)
    for m in range(1, 13):
        col = 6 + m
        dash.cell(row=7,  column=col).value = f_kpi1(m)   # KPI 1
        dash.cell(row=18, column=col).value = f_kpi6(m)   # KPI 6

    # خلية إدخال «عدد نقاط المنتج» المستخدمة في توفر الوقود
    dash.merge_cells("A37:R37")
    h = dash["A37"]
    h.value = "⚙️ مُدخلات الحساب اليدوية (حسب طريقة قياس الإدارة)"
    h.fill = PatternFill("solid", fgColor=ORANGE)
    h.font = Font(name=FONT, bold=True, size=11, color=WHITE)
    h.alignment = Alignment(horizontal="right", vertical="center")

    dash.merge_cells("A38:E38")
    lbl = dash["A38"]
    lbl.value = "عدد نقاط المنتج (محطات × منتجات) — لحساب «توفر الوقود»"
    lbl.font = Font(name=FONT, bold=True, size=9, color=GRAY)
    lbl.alignment = Alignment(horizontal="right", vertical="center")
    inp = dash["F38"]
    inp.value = 300  # = 9000 / 30 (يحافظ على مقياس الثابت الأصلي؛ عدّله حسب الواقع)
    inp.fill = PatternFill("solid", fgColor=YELLOW)
    inp.font = Font(name=FONT, bold=True, size=11, color="FF1565C0")
    inp.alignment = Alignment(horizontal="center", vertical="center")
    inp.border = BORDER
    dash.merge_cells("A39:R39")
    note = dash["A39"]
    note.value = ("ℹ️ الافتراضي 300 = 9000 ÷ 30 (مكافئ للثابت القديم في الأشهر ذات 30 يوماً). "
                  "غيّره ليعكس عدد (المحطات × المنتجات) الفعلي.")
    note.font = Font(name=FONT, italic=True, size=8, color=GRAY)
    note.alignment = Alignment(horizontal="right", vertical="center")

    # (1) ورقة «مؤشرات الإدارة» المرجعية — تُنقل القيم حرفياً من ملف الإدارة
    src = openpyxl.load_workbook(SRC_PATH, data_only=True)["ورقة1"]
    rows = []
    for r in range(6, 11):
        rows.append([
            src.cell(r, 1).value,  # م
            src.cell(r, 2).value,  # اسم المؤشر
            src.cell(r, 3).value,  # وحدة القياس
            src.cell(r, 4).value,  # الهدف
            (src.cell(r, 5).value or "").strip(),  # دورية القياس
            src.cell(r, 6).value,  # المحور
            src.cell(r, 7).value,  # طريقة القياس
        ])
    mapping = [
        "KPI 1 — حُدِّثت طريقة الحساب",
        "KPI 6 — حُدِّثت طريقة الحساب",
        "KPI 8 — مطابقة (دون تغيير)",
        "KPI 3 — مطابقة (دون تغيير)",
        "KPI 4 — مطابقة (دون تغيير)",
    ]

    ref = wb.create_sheet("📋 مؤشرات الإدارة", 1)
    ref.sheet_view.rightToLeft = True
    headers = ["م", "اسم المؤشر (KPI)", "وحدة القياس", "الهدف",
               "دورية القياس", "المحور", "طريقة القياس", "يقابله في اللوحة"]
    widths = [5, 34, 12, 22, 12, 16, 62, 22]
    for i, w in enumerate(widths, start=1):
        ref.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # العنوان
    ref.merge_cells("A1:H1")
    t = ref["A1"]
    t.value = "مؤشرات الأداء (KPI's) — سلاسل الإمداد · درب الوقود"
    t.fill = PatternFill("solid", fgColor=ORANGE)
    t.font = Font(name=FONT, bold=True, size=15, color=WHITE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ref.row_dimensions[1].height = 30
    # العنوان الفرعي
    ref.merge_cells("A2:H2")
    s = ref["A2"]
    s.value = "🏛️ المرجع الرسمي المعتمد من الإدارة · 5 مؤشرات · دورية شهرية"
    s.fill = PatternFill("solid", fgColor=CREAM)
    s.font = Font(name=FONT, size=10, color=GRAY)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ref.row_dimensions[2].height = 20

    # رأس الجدول
    for c, htxt in enumerate(headers, start=1):
        cell = ref.cell(row=3, column=c, value=htxt)
        cell.fill = PatternFill("solid", fgColor=GRAY)
        cell.font = Font(name=FONT, bold=True, size=9, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ref.row_dimensions[3].height = 26

    # صفوف المؤشرات
    for i, (data, mp) in enumerate(zip(rows, mapping)):
        r = 4 + i
        fill = WHITE if i % 2 else CREAM
        values = list(data) + [mp]
        for c, v in enumerate(values, start=1):
            cell = ref.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = BORDER
            bold = c in (2, 8)
            cell.font = Font(name=FONT, bold=bold, size=9, color=GRAY)
            if c in (2, 7):  # اسم المؤشر + طريقة القياس → محاذاة يمين مع التفاف
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ref.row_dimensions[r].height = 54
    ref.freeze_panes = "A4"

    # تذييل توضيحي
    ref.merge_cells("A10:H10")
    f = ref["A10"]
    f.value = ("ℹ️ هذه القائمة هي مصدر الحقيقة الرسمي. في اللوحة الآلية حُدِّثت طريقة حساب "
               "«توفر الوقود» و«فواقد الوقود» لتطابق العمود «طريقة القياس» أعلاه؛ والمؤشرات "
               "الأخرى تبقى كما هي.")
    f.font = Font(name=FONT, italic=True, size=8, color=GRAY)
    f.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ref.row_dimensions[10].height = 32

    wb.save(WB_PATH)
    print("saved:", WB_PATH)


if __name__ == "__main__":
    main()
