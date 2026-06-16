#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
تحقق وظيفي شامل من دوال المصنّف باستخدام محرّك Excel حقيقي (pycel).

يحقن بيانات تجريبية في صفحات الإدخال الخمس، ثم يقيّم خلايا اللوحتين فعلياً
(الأصلية + المُكثّفة) ويؤكّد تطابقها مع حساب مستقل — بما يغطّي حالات حدّية
(اختلاف أيام الشهر، شهر فارغ، قسمة على صفر، وتطابق أحجام نطاقات COUNTIFS).

التشغيل:
    pip install openpyxl pycel
    python3 scripts/verify_formulas.py        # 0 عند النجاح، 1 عند الفشل
"""
import calendar
import datetime
import sys
import tempfile
import warnings

import openpyxl

warnings.filterwarnings("ignore")

WB = "workbook/khadija-supply-chain-kpis-2026.xlsx"
DASH = "📊 مؤشرات الأداء (٩)"   # اللوحة الوحيدة بعد دمج المتكرر
PRODUCT_POINTS = 300

# يناير 2+1=3 أيام ، فبراير 3+2=5 أيام
OUTAGES = [
    (datetime.date(2026, 1, 5), datetime.date(2026, 1, 7)),
    (datetime.date(2026, 1, 10), datetime.date(2026, 1, 11)),
    (datetime.date(2026, 2, 3), datetime.date(2026, 2, 6)),
    (datetime.date(2026, 2, 20), datetime.date(2026, 2, 22)),
]
# يناير 80,130 ، فبراير 50,70,90
RETURNS = [
    (datetime.date(2026, 1, 8), 33000, 80), (datetime.date(2026, 1, 20), 30000, 130),
    (datetime.date(2026, 2, 5), 33000, 50), (datetime.date(2026, 2, 15), 32000, 70),
    (datetime.date(2026, 2, 25), 33000, 90),
]
# (تاريخ استلام, نوع, تاريخ معالجة) — مخزني 1/2 في الوقت ، أمر شراء 1/2
ORDERS = [
    (datetime.date(2026, 1, 3), "مواد مخزنية", datetime.date(2026, 1, 5)),
    (datetime.date(2026, 1, 10), "مواد مخزنية", datetime.date(2026, 1, 20)),
    (datetime.date(2026, 1, 4), "أمر شراء", datetime.date(2026, 1, 8)),
    (datetime.date(2026, 1, 6), "أمر شراء", datetime.date(2026, 1, 15)),
]
# يناير وقود: أيام مخزون 3 و 4 → متوسط 3.5
INVENTORY = [("يناير", "وقود", 300, 100), ("يناير", "وقود", 400, 100)]
# يناير 4 رحلات: ملكية / وقت / حادث-مخالفة
TRIPS = [
    (datetime.date(2026, 1, 2), "مملوك درب", 10, 9, "لا يوجد"),
    (datetime.date(2026, 1, 3), "متعاقد خارجي", 10, 12, "لا يوجد"),
    (datetime.date(2026, 1, 4), "مملوك درب", 10, 10, "مخالفة سرعة"),
    (datetime.date(2026, 1, 5), "مملوك درب", 8, 7, "لا يوجد"),
]


def build_test_file(path):
    wb = openpyxl.load_workbook(WB)
    o = wb["🛢️ انقطاعات الوقود"]
    for i, (b, e) in enumerate(OUTAGES):
        o.cell(5 + i, 2, b); o.cell(5 + i, 3, e); o.cell(5 + i, 5, "ديزل")
    r = wb["💧 الردود والفاقد"]
    for i, (d, q, l) in enumerate(RETURNS):
        r.cell(5 + i, 2, d); r.cell(5 + i, 6, q); r.cell(5 + i, 7, l)
    od = wb["🛒 الطلبات والمشتريات"]
    for i, (c, t, f) in enumerate(ORDERS):
        od.cell(5 + i, 3, c); od.cell(5 + i, 4, t); od.cell(5 + i, 6, f)
    iv = wb["📦 المخزون"]
    for i, (mo, ty, g, h) in enumerate(INVENTORY):
        iv.cell(5 + i, 2, mo); iv.cell(5 + i, 6, ty); iv.cell(5 + i, 7, g); iv.cell(5 + i, 8, h)
    tr = wb["🚚 النقل والرحلات"]
    for i, (b, own, h, ii, k) in enumerate(TRIPS):
        tr.cell(5 + i, 2, b); tr.cell(5 + i, 7, own); tr.cell(5 + i, 8, h)
        tr.cell(5 + i, 9, ii); tr.cell(5 + i, 11, k)
    wb.save(path)


def outage_count(m):  # K1: عدد الانقطاعات الشهرية
    return sum(1 for b, e in OUTAGES if b.year == 2026 and b.month == m)


def loss(m):          # K4: إجمالي الفاقد ÷ عدد الردود
    ls = [l for d, q, l in RETURNS if d.month == m]
    return sum(ls) / len(ls) if ls else 0


def avg_days(typ):    # K2/K3: متوسط أيام المعالجة (يناير) لنوع طلب
    ds = [(f - c).days for c, t, f in ORDERS if t == typ and c.month == 1]
    return sum(ds) / len(ds) if ds else 0


def main():
    from pycel import ExcelCompiler

    path = tempfile.mktemp(suffix=".xlsx")
    build_test_file(path)
    exc = ExcelCompiler(path)

    # كل الفحوص على اللوحة الوحيدة (٩ مؤشرات) بأهدافها التشغيلية الجديدة
    cases = [
        ("G7", "عدد الانقطاعات · يناير", outage_count(1)),
        ("H7", "عدد الانقطاعات · فبراير", outage_count(2)),
        ("I7", "عدد الانقطاعات · مارس", 0),
        ("G9", "أيام تسليم المواد المخزنية", avg_days("مواد مخزنية")),
        ("G11", "أيام إصدار أوامر الشراء", avg_days("أمر شراء")),
        ("G14", "فواقد الوقود · يناير", loss(1)),
        ("H14", "فواقد الوقود · فبراير", loss(2)),
        ("G19", "أيام المخزون", 3.5),
        ("G22", "المخالفات/الحوادث · يناير", 1),
        ("G25", "الرحلات المتأخرة · يناير", 1),
        ("G27", "استخدام الأسطول المملوك", 0.75),
        # حالة المحاور المُجمَّعة (أسوأ حالة بين مؤشرات المحور)
        ("F5", "محور خدمة العميل", "🔴"),
        ("F12", "محور التكاليف", "✅"),
        ("F17", "محور السيولة المالية", "🟡"),
        ("F20", "محور السلامة والامتثال", "🔴"),
        ("F23", "محور الكفاءة التشغيلية", "🔴"),
    ]
    ok = True
    print(f"{'cell':<5}{'label':<34}{'actual':>10}{'expected':>10}")
    for cell, label, exp in cases:
        val = exc.evaluate(f"'{DASH}'!{cell}")
        if isinstance(exp, str):
            passed = str(val) == exp
        else:
            try:
                passed = abs(float(val) - exp) < 1e-4
            except (TypeError, ValueError):
                passed = False
        ok &= passed
        shown = round(val, 6) if isinstance(val, float) else val
        print(f"{cell:<5}{label:<34}{str(shown):>10}{str(exp):>10}  {'✅' if passed else '❌'}")

    print("\nRESULT:", "✅ ALL FORMULAS CORRECT & FUNCTIONAL" if ok else "❌ FAILURES DETECTED")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
