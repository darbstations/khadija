#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
مولّد النموذج المالي لبرنامج تانكي (نسخة v5) — بمعادلات حقيقية مترابطة.

ينطلق من الملف المصدر في data/tanki_model_source.xlsx ويطبّق:
  1) معدلات الكسب الموصى بها (50/75/100)
  2) هوامش درب الفعلية (13 هللة/لتر)
  3) إصلاح أخطاء معادلات عمود المؤشرات في قسم المحطات
  4) تحديث جداول المقارنة الثابتة في الملخص
  5) كتلة مدخلات "الاقتصاد الموسّع" (الديزل + الشركاء)
  6) ورقة جديدة «الاقتصاد الموسّع» بمعادلات LIVE تشير لخلايا السيناريوهات
  7) تفعيل إعادة الحساب الكامل عند الفتح

كل القيم المحسوبة في الورقة الجديدة معادلات حقيقية (=...) لا أرقاماً ثابتة.
"""
import os
import warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC = os.path.join(ROOT, "data", "tanki_model_source.xlsx")
OUT = os.path.join(ROOT, "tanki_financial_model_v5.xlsx")

SC = "السيناريوهات"  # اسم ورقة السيناريوهات (للمراجع)

YELLOW = PatternFill("solid", fgColor="FFF6D26B")
HEADER = PatternFill("solid", fgColor="FF16223C")
RTL = Alignment(horizontal="right", vertical="center", readingOrder=2, wrap_text=True)
THIN = Side(style="thin", color="FF23314F")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main() -> None:
    wb = openpyxl.load_workbook(SRC, data_only=False)
    S = wb[SC]

    # 1) النموذج الشفّاف: كل ريال = نقطة · النقطة = نص هللة (200 نقطة = ريال) = 0.5%
    S["C6"] = 200           # 200 نقطة = ريال (النقطة = نص هللة)
    S["C7"], S["C8"], S["C9"] = 1, 1, 1   # كل ريال = نقطة (المستويات بالمنافع)
    S["C10"] = 0.5          # تكلفة الولاء = 0.5 هللة/ريال (0.5%)
    # 2) هوامش درب الفعلية (~13 هللة/لتر)
    S["C26"], S["C27"] = 13, 13

    # 3) إصلاح + إعادة محاذاة عمود المؤشرات (كل خلية مدمجة F#:G# → نكتب على F#)
    S["F23"] = '=IF(C23+C24+C25>=150,"✅ منطقي","⚠️ تحقق")'
    S["F24"] = ""
    S["F25"] = ""
    S["F26"] = '=IF(C26>=5,"✅ معقول","⚠️ منخفض")'
    S["F27"] = '=IF(C27>=5,"✅ معقول","⚠️ منخفض")'
    S["F28"] = '=IF(C28>=1,"✅ معقول","⚠️ ضيق")'
    S["F29"] = ""
    S["F30"] = '=IF(C29+C30=1,"✅ صحيح","🔴 خطأ المجموع")'   # الإصلاح (كان C30+C31)
    S["F31"] = '=IF(C31>=50000,"✅ ممتاز",IF(C31>=30000,"⚠️ متوسط","🔴 منخفض"))'
    S["F32"] = ""
    S["A33"] = ("💡 الفلسفة · نموذج شفّاف: كل ريال = نقطة · النقطة = نص هللة (200 نقطة = ريال) = 0.5% "
                "· موحّد مع المتاجر · بلا تضخيم")

    # 4) الملخص التنفيذي — الخلايا النصية الثابتة
    M = wb["الملخص التنفيذي"]
    M["A6"], M["D6"], M["G6"] = "0.5% من كل ريال إنفاق", "0.5% + منافع للأعضاء النشطين", "0.5% + منافع VIP"
    M["D16"], M["E16"], M["F16"], M["G16"] = "1 نقطة", "1 نقطة", "1 نقطة", "2.3x"
    M["D17"], M["E17"], M["F17"], M["G17"] = "200 = ريال", "200 = ريال", "200 = ريال", "نقطة = نص هللة"
    M["D18"], M["E18"], M["F18"], M["G18"] = "0.50%", "0.50%", "0.50%", "1.2x"
    M["D19"], M["E19"], M["F19"] = "100 نقطة", "100 نقطة", "100 نقطة"
    # التوصيات الاستراتيجية — مواءمة مع النموذج (كل ريال = نقطة · نص هللة)
    M["A23"] = "1. اعتمد النموذج الشفّاف: كل ريال = نقطة · النقطة = نص هللة (200 نقطة = ريال) = 0.5%"
    M["A25"] = "3. هدية ترحيب 1,000 نقطة = 5 ريال · ترحيب بسيط وواضح"
    M["A26"] = "4. التسويق: 'كل ريال نقطة · أكثر كاش باك من ساسكو'"

    # 5) كتلة مدخلات الاقتصاد الموسّع (الديزل + الشركاء) — خلايا صفراء قابلة للتعديل
    S["A35"] = "🆕 الاقتصاد الموسّع — مدخلات"
    S["A35"].font = Font(bold=True, color="FFFFD54A")
    ext_inputs = [
        (36, "سعر لتر الديزل (ريال)", 1.79, "يناير 2026"),
        (37, "هامش الديزل (هللة/لتر)", 4, "ضعيف جداً"),
        (38, "كسب الديزل (نقطة/ريال)", 0.5, "كاش باك 0.25% (هامش ضعيف)"),
        (39, "كسب الشركاء (نقطة/ريال)", 8, "~4% ممولة من الشريك"),
        (40, "مساهمة الشريك لدرب (هللة/ريال)", 1, "ما يدفعه الشريك مقابل عملاء درب"),
        (41, "هدية الترحيب الموصى بها (نقطة)", 1000, "= 5 ريال تكلفة فعلية"),
    ]
    for r, label, val, desc in ext_inputs:
        S.cell(r, 1, "+")
        S.cell(r, 2, label).alignment = RTL
        c = S.cell(r, 3, val)
        c.fill = YELLOW
        c.font = Font(bold=True)
        S.cell(r, 4, desc).alignment = RTL

    # 6) ورقة جديدة «الاقتصاد الموسّع» — معادلات حقيقية تشير لخلايا السيناريوهات
    if "الاقتصاد الموسّع" in wb.sheetnames:
        del wb["الاقتصاد الموسّع"]
    E = wb.create_sheet("الاقتصاد الموسّع")
    E.sheet_view.rightToLeft = True

    def put(coord, value, bold=False, fill=None, money=False):
        cell = E[coord]
        cell.value = value
        cell.alignment = RTL
        cell.font = Font(bold=bold, color="FFE8EEFC")
        if fill:
            cell.fill = fill
        if money:
            cell.number_format = "#,##0.00"
        return cell

    put("A1", "درب · محرك الاقتصاد الموسّع — الديزل والشركاء (معادلات حيّة)", bold=True)
    put("A2", "كل القيم أدناه معادلات تشير لخلايا ورقة السيناريوهات — غيّر هناك وتتحدث هنا")

    # ----- الديزل -----
    put("A4", "⛽ الديزل", bold=True, fill=HEADER)
    diesel = [
        ("A5", "البند", "B5", "القيمة", "C5", "الوحدة"),
    ]
    put("A5", "البند", bold=True); put("B5", "القيمة", bold=True); put("C5", "الوحدة", bold=True)
    rows_d = [
        (6, "هامش الديزل/لتر", f"={SC}!C37", "هللة/لتر"),
        (7, "هامش الديزل %", f"={SC}!C37/({SC}!C36*100)", "%"),
        (8, "كسب الديزل", f"={SC}!C38", "نقطة/ريال"),
        (9, "كاش باك الديزل %", f"={SC}!C38/{SC}!C6", "%"),
        (10, "تكلفة درب على الديزل", f"={SC}!C38/{SC}!C6*100", "هللة/ريال"),
        (11, "صافي هامش درب على الديزل", f"={SC}!C37/{SC}!C36-({SC}!C38/{SC}!C6*100)", "هللة/ريال"),
    ]
    for r, label, formula, unit in rows_d:
        put(f"A{r}", label)
        put(f"B{r}", formula, bold=True, money=True)
        put(f"C{r}", unit)
    E["B7"].number_format = "0.00%"
    E["B9"].number_format = "0.00%"
    put("A12", "الجدوى", bold=True)
    put("B12", '=IF(B11>0.5,"✅ مربح",IF(B11>0,"⚠️ هامش ضيق","🔴 خسارة"))', bold=True)

    # ----- الشركاء -----
    put("A14", "☕ الشركاء (كافيه/مطعم)", bold=True, fill=HEADER)
    put("A15", "البند", bold=True); put("B15", "القيمة", bold=True); put("C15", "الوحدة", bold=True)
    rows_p = [
        (16, "كسب العميل عند الشريك", f"={SC}!C39", "نقطة/ريال"),
        (17, "كاش باك العميل %", f"={SC}!C39/{SC}!C6", "%"),
        (18, "تكلفة درب (الشريك يموّل)", "=0", "هللة/ريال"),
        (19, "مساهمة الشريك لدرب", f"={SC}!C40", "هللة/ريال"),
        (20, "صافي لدرب من الشريك", f"={SC}!C40-B18", "هللة/ريال"),
    ]
    for r, label, formula, unit in rows_p:
        put(f"A{r}", label)
        put(f"B{r}", formula, bold=True, money=True)
        put(f"C{r}", unit)
    E["B17"].number_format = "0.00%"

    # ----- المقارنة: ريال وقود مقابل ريال شريك (أثره على درب) -----
    put("A22", "🔬 أثر الريال على درب: وقود مقابل شريك", bold=True, fill=HEADER)
    put("A23", "البند", bold=True); put("B23", "ريال بنزين", bold=True); put("C23", "ريال شريك", bold=True)
    put("A24", "دخل/ربح درب (هللة)")
    put("B24", f"={SC}!C26/{SC}!C11", bold=True, money=True)   # ربح بيع الوقود
    put("C24", f"={SC}!C40", bold=True, money=True)            # مساهمة الشريك
    put("A25", "تكلفة الولاء على درب (هللة)")
    put("B25", f"={SC}!C7/{SC}!C6*100", bold=True, money=True)  # تكلفة كسب الأبيض
    put("C25", "=0", bold=True, money=True)
    put("A26", "الصافي على درب (هللة/ريال)")
    put("B26", "=B24-B25", bold=True, money=True)
    put("C26", "=C24-C25", bold=True, money=True)
    put("A27", "الخلاصة", bold=True)
    put("B27", '=IF(B26>0,"✅ تكلفة احتفاظ","🔴 خسارة")', bold=True)
    put("C27", '=IF(C26>0,"✅ ربح صافٍ","—")', bold=True)

    put("A29", "💡 الوقود يجذب (هامش رفيع) والشركاء يموّلون ويربّحون — أولوية درب: توقيع الشركاء")

    # عرض الأعمدة
    for col, w in {"A": 34, "B": 16, "C": 16}.items():
        E.column_dimensions[col].width = w

    # 6.5) ورقة جديدة «قنوات الاستبدال» — معادلات حقيقية
    if "قنوات الاستبدال" in wb.sheetnames:
        del wb["قنوات الاستبدال"]
    R = wb.create_sheet("قنوات الاستبدال")
    R.sheet_view.rightToLeft = True

    def rput(coord, value, bold=False, fill=None, fmt=None):
        c = R[coord]
        c.value = value
        c.alignment = RTL
        c.font = Font(bold=bold, color="FFE8EEFC")
        if fill:
            c.fill = fill
        if fmt:
            c.number_format = fmt
        return c

    rput("A1", "درب · قنوات استبدال النقاط — التكلفة على درب لكل قناة (معادلات حيّة)", bold=True)
    rput("A3", "إجمالي القيمة المستبدلة سنوياً (ريال)")
    rput("C3", 1000000, bold=True, fill=YELLOW, fmt="#,##0")

    # ترويسة الجدول
    headers = ["القناة", "من يموّلها", "النسبة %", "تكلفة درب/ريال %", "القيمة المستبدلة", "تكلفة درب"]
    for j, h in enumerate(headers):
        c = R.cell(5, j + 1, h)
        c.font = Font(bold=True, color="FFE8EEFC")
        c.fill = HEADER
        c.alignment = RTL

    channels = [
        ("⛽ خصم بنزين", "درب", 0.30, 1.00),
        ("🛍️ عروض المستأجرين", "المستأجر", 0.35, 0.00),
        ("🎁 شركات خارجية (جرير/أمازون/الفرسان)", "درب (خصم جملة)", 0.20, 0.97),
        ("📱 كرت شحن", "درب (عمولة المشغّل)", 0.15, 0.95),
    ]
    first, last = 6, 6 + len(channels) - 1
    for idx, (label, funded, mix, cost) in enumerate(channels):
        r = first + idx
        rput(f"A{r}", label)
        rput(f"B{r}", funded)
        rput(f"C{r}", mix, fill=YELLOW, fmt="0%")        # نسبة (مدخل)
        rput(f"D{r}", cost, fill=YELLOW, fmt="0%")        # تكلفة درب لكل ريال (مدخل)
        rput(f"E{r}", f"=$C$3*C{r}", fmt="#,##0.00")      # القيمة المستبدلة (معادلة)
        rput(f"F{r}", f"=E{r}*D{r}", bold=True, fmt="#,##0.00")  # تكلفة درب (معادلة)

    # الإجمالي
    tr = last + 1
    rput(f"A{tr}", "📊 الإجمالي", bold=True)
    rput(f"C{tr}", f"=SUM(C{first}:C{last})", bold=True, fmt="0%")
    rput(f"E{tr}", f"=SUM(E{first}:E{last})", bold=True, fmt="#,##0.00")
    rput(f"F{tr}", f"=SUM(F{first}:F{last})", bold=True, fmt="#,##0.00")

    # مؤشرات
    rput(f"A{tr+2}", "متوسط تكلفة الريال المستبدل")
    rput(f"C{tr+2}", f"=F{tr}/E{tr}", bold=True, fmt="0.0%")
    rput(f"A{tr+3}", "التوفير مقابل الكل-بنزين")
    rput(f"C{tr+3}", f"=$C$3-F{tr}", bold=True, fmt="#,##0.00")
    rput(f"A{tr+4}", "تحقق مجموع النسب")
    rput(f"C{tr+4}", f'=IF(C{tr}=1,"✅ 100%","⚠️ راجع النسب")', bold=True)
    rput(f"A{tr+6}", "💡 وجّه الاستبدال نحو عروض المستأجرين (تكلفتها صفر) بدل خصم البنزين")

    for col, w in {"A": 38, "B": 22, "C": 12, "D": 16, "E": 18, "F": 16}.items():
        R.column_dimensions[col].width = w

    # 6.6) ورقة «نموذج المحفظة» — Escrow بمعادلات حقيقية
    if "نموذج المحفظة" in wb.sheetnames:
        del wb["نموذج المحفظة"]
    W = wb.create_sheet("نموذج المحفظة")
    W.sheet_view.rightToLeft = True

    def wput(coord, value, bold=False, fill=None, fmt=None):
        c = W[coord]
        c.value = value
        c.alignment = RTL
        c.font = Font(bold=bold, color="FFE8EEFC")
        if fill:
            c.fill = fill
        if fmt:
            c.number_format = fmt
        return c

    wput("A1", "درب · نموذج المحفظة المموّلة بالكامل (Escrow) — معادلات حيّة", bold=True)
    wput("A2", "كل مُصدِر يودع نسبته على مبيعاته هو · النقاط واجهة فوق محفظة بالريال")

    wput("A4", "📥 تغذية المحفظة", bold=True, fill=HEADER)
    for j, h in enumerate(["المصدر", "قيمة الشراء (﷼)", "نسبة الإيداع %", "المودَع (﷼)"]):
        c = W.cell(5, j + 1, h); c.font = Font(bold=True, color="FFE8EEFC"); c.fill = HEADER; c.alignment = RTL
    sources = [("⛽ بنزين (درب)", 600, 0.5), ("☕ كافيه", 200, 4), ("🍔 مطعم", 150, 3), ("🛒 سوبرماركت", 400, 1)]
    f0 = 6
    for i, (label, amount, rate) in enumerate(sources):
        r = f0 + i
        wput(f"A{r}", label)
        wput(f"B{r}", amount, fill=YELLOW, fmt="#,##0")
        wput(f"C{r}", rate, fill=YELLOW, fmt="0.0")
        wput(f"D{r}", f"=B{r}*C{r}/100", bold=True, fmt="#,##0.00")
    fn = f0 + len(sources)  # صف الإجمالي
    wput(f"A{fn}", "📊 إجمالي التغذية", bold=True)
    wput(f"D{fn}", f"=SUM(D{f0}:D{fn-1})", bold=True, fmt="#,##0.00")

    wput(f"A{fn+2}", "رصيد المحفظة (﷼)", bold=True)
    wput(f"C{fn+2}", f"=D{fn}", bold=True, fmt="#,##0.00")
    wput(f"A{fn+3}", "يُعرض للعميل كنقاط")
    wput(f"C{fn+3}", f"=C{fn+2}*السيناريوهات!C6", bold=True, fmt="#,##0")

    rd = fn + 5
    wput(f"A{rd}", "🔁 محاكاة استبدال", bold=True, fill=HEADER)
    wput(f"A{rd+1}", "يُسحب من المحفظة (﷼)"); wput(f"C{rd+1}", 10, fill=YELLOW, fmt="#,##0.00")
    wput(f"A{rd+2}", "يضيف التاجر تنافساً (﷼)"); wput(f"C{rd+2}", 3, fill=YELLOW, fmt="#,##0.00")
    wput(f"A{rd+3}", "يحصل العميل (﷼)"); wput(f"C{rd+3}", f"=C{rd+1}+C{rd+2}", bold=True, fmt="#,##0.00")
    wput(f"A{rd+4}", "القيمة لكل ريال مسحوب"); wput(f"C{rd+4}", f"=C{rd+3}/C{rd+1}", bold=True, fmt="0.00")
    wput(f"A{rd+5}", "الرصيد بعد الاستبدال (﷼)"); wput(f"C{rd+5}", f"=C{fn+2}-C{rd+1}", bold=True, fmt="#,##0.00")

    bk = rd + 7
    wput(f"A{bk}", "🕳️ الأموال الخاملة (Breakage)", bold=True, fill=HEADER)
    wput(f"A{bk+1}", "نسبة الخمول %"); wput(f"C{bk+1}", 12, fill=YELLOW, fmt="0")
    wput(f"A{bk+2}", "مدة الصلاحية (شهر)"); wput(f"C{bk+2}", 24, fill=YELLOW, fmt="0")
    wput(f"A{bk+3}", "المتوقع خامل (﷼)"); wput(f"C{bk+3}", f"=C{fn+2}*C{bk+1}/100", bold=True, fmt="#,##0.00")
    wput(f"A{bk+4}", "المتوقع استبداله فعلاً (﷼)"); wput(f"C{bk+4}", f"=C{fn+2}-C{bk+3}", bold=True, fmt="#,##0.00")
    wput(f"A{bk+6}", "🔒 النظام لا يخسر: كل ريال يُستبدل كان مودَعاً مسبقاً · لا أحد يموّل وعود غيره")

    for col, w in {"A": 30, "B": 16, "C": 16, "D": 14}.items():
        W.column_dimensions[col].width = w

    # 7) إعادة حساب كاملة عند الفتح
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    wb.save(OUT)
    print("تم الحفظ:", OUT)


if __name__ == "__main__":
    main()
