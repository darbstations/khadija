#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
يبني لوحة «📊 مؤشرات الأداء (٩)» الموحّدة في مصنّف خديجة بدمج المؤشرات المتكرّرة
(التي تؤدي نفس الهدف)، ثم يحذف لوحة الـ١٣ القديمة فتصبح هذه هي اللوحة الوحيدة.

قرارات الدمج:
  • عدد الانقطاعات (2)  → مطوي داخل «توفر الوقود» (الشدّة بدل التكرار).
  • دوران المخزون (9)   → مطوي داخل «أيام المخزون» (متطابقان: الدوران = 30÷الأيام).
  • Fill Rate (5)       → محذوف (كان تجميعة لمؤشري المواد المخزنية + أوامر الشراء).
  • حوادث النقل (10)    → مطوية داخل «الامتثال والسلامة» (نسبة الرحلات بلا حادث/مخالفة).

النتيجة: ١٣ → ٩ مؤشرات نظيفة عبر المحاور الخمسة، مع حالة مُجمَّعة لكل محور، وخلية
إدخال محلية «عدد نقاط المنتج» (F31). النسخة ذات الـ١٣ محفوظة في سجل Git.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

WB = "workbook/khadija-supply-chain-kpis-2026.xlsx"
DASH, OUT, RET = "📊 ملخص آلي + KPIs", "🛢️ انقطاعات الوقود", "💧 الردود والفاقد"
ORD, INV, TRP = "🛒 الطلبات والمشتريات", "📦 المخزون", "🚚 النقل والرحلات"

ORANGE, ORANGE2, GRAY, CREAM = "FFC04E0E", "FFF3741F", "FF3D3D3D", "FFFEF0E6"
WHITE, YELLOW, BLUE = "FFFFFFFF", "FFFFF9D6", "FF1565C0"
F = "DIN Next LT Arabic"
MONTHS = ["يناير","فبراير","مارس","أبريل","مايو","يونيو","يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]
SHEET = "📊 مؤشرات الأداء (٩)"     # اسم اللوحة (صارت الوحيدة بعد دمج المتكرر)
INPUT_ROW = 31                      # صف خلية إدخال «عدد نقاط المنتج» (F31) داخل اللوحة


# آخر صف بيانات لكل ورقة (يجب أن تتطابق كل نطاقات COUNTIFS/SUMIFS في الحجم)
LASTROW = {OUT: 104, RET: 304, ORD: 204, INV: 304, TRP: 304}


def mrange(sheet, col, m):  # شرط الشهر على عمود تاريخ
    end = LASTROW[sheet]
    return (f"'{sheet}'!{col}5:{col}{end},\">=\"&DATE(2026,{m},1),"
            f"'{sheet}'!{col}5:{col}{end},\"<\"&DATE(2026,{m}+1,1)")


def fa_avail(m):   # توفر الوقود (يشير لخلية الإدخال المحلية F{INPUT_ROW})
    return (f"=IFERROR(1-(SUMIFS('{OUT}'!F5:F104,{mrange(OUT,'B',m)})"
            f"/($F${INPUT_ROW}*DAY(DATE(2026,{m}+1,1)-1))),1)")
def fa_stock(m):   # توفر المواد المخزنية ≤3 أيام
    base=mrange(ORD,'C',m)
    return (f"=IFERROR(COUNTIFS({base},'{ORD}'!D5:D204,\"مواد مخزنية\",'{ORD}'!I5:I204,\"*ضمن*\")"
            f"/MAX(1,COUNTIFS({base},'{ORD}'!D5:D204,\"مواد مخزنية\")),0)")
def fa_po(m):      # إصدار أوامر الشراء ≤5 أيام
    base=mrange(ORD,'C',m)
    return (f"=IFERROR(COUNTIFS({base},'{ORD}'!D5:D204,\"أمر شراء\",'{ORD}'!I5:I204,\"*ضمن*\")"
            f"/MAX(1,COUNTIFS({base},'{ORD}'!D5:D204,\"أمر شراء\")),0)")
def fa_loss(m):    # فواقد الوقود = إجمالي الفاقد ÷ عدد الردود
    return f"=IFERROR(SUMIFS('{RET}'!G5:G304,{mrange(RET,'B',m)})/COUNTIFS({mrange(RET,'B',m)}),0)"
def fa_invd(m):    # أيام مخزون الوقود
    return f"=IFERROR(AVERAGEIFS('{INV}'!I5:I304,'{INV}'!C5:C304,{m},'{INV}'!F5:F304,\"وقود\"),0)"
def fa_comp(m):    # الامتثال والسلامة = رحلات بلا حادث/مخالفة ÷ الكل
    return (f"=IFERROR(COUNTIFS({mrange(TRP,'B',m)},'{TRP}'!K5:K304,\"لا يوجد\")"
            f"/MAX(1,COUNTIFS({mrange(TRP,'B',m)})),0)")
def fa_ontime(m):  # التزام جدول النقل
    return (f"=IFERROR(COUNTIFS({mrange(TRP,'B',m)},'{TRP}'!J5:J304,\"*في الوقت*\")"
            f"/MAX(1,COUNTIFS({mrange(TRP,'B',m)})),0)")
def fa_fleet(m):   # استخدام الأسطول المملوك
    return (f"=IFERROR(COUNTIFS({mrange(TRP,'B',m)},'{TRP}'!G5:G304,\"مملوك درب\")"
            f"/MAX(1,COUNTIFS({mrange(TRP,'B',m)})),0)")

# (الرقم, الاسم, الوحدة, الهدف_قيمة, الهدف_عرض, الاتجاه, صيغة_الرقم, دالة_الفعلي)
KPIS_BY_AXIS = [
 ("👥 خدمة العميل", [
    (1,"توفر الوقود (يطوي عدد الانقطاعات)","%",0.995,"≥ 99.5%","high","0.00%",fa_avail),
    (2,"توفر المواد المخزنية (تسليم ≤ 3 أيام)","%",1,"100%","high","0.00%",fa_stock),
    (3,"إصدار أوامر الشراء (≤ 5 أيام عمل)","%",0.95,"≥ 95%","high","0.00%",fa_po),
 ]),
 ("💰 التكاليف", [
    (4,"فواقد الوقود لكل 33K لتر","لتر",100,"≤ 100","low","0.0",fa_loss),
    (5,"نسبة المواد التالفة (يدوي شهرياً)","%",0.02,"≤ 2%","low","0.00%",None),
 ]),
 ("🏦 السيولة المالية", [
    (6,"أيام مخزون الوقود (يطوي الدوران)","يوم",3,"≤ 3","low","0.0",fa_invd),
 ]),
 ("🛡️ السلامة والامتثال", [
    (7,"الامتثال والسلامة (يطوي عدد الحوادث)","%",0.95,"≥ 95%","high","0.00%",fa_comp),
 ]),
 ("⚙️ الكفاءة التشغيلية", [
    (8,"التزام جدول النقل (On-time)","%",0.95,"≥ 95%","high","0.00%",fa_ontime),
    (9,"نسبة استخدام الأسطول المملوك","%",0.6,"≥ 60%","high","0.00%",fa_fleet),
 ]),
]


def status_formula(p, a, direction):
    op = ">=" if direction == "high" else "<="
    mult = "0.85" if direction == "high" else "1.3"
    av = f"IFERROR(AVERAGEIF(G{a}:R{a},\">0\",G{a}:R{a}),0)"
    pl = f"AVERAGE(G{p}:R{p})"
    return (f"=IFERROR(IF({pl}=0,\"\",IF({av}{op}{pl},\"✅ \","
            f"IF({av}{op}{pl}*{mult},\"🟡 \",\"🔴 \"))"
            f"&TEXT(({av}-{pl})/{pl},\"+0.0%;-0.0%\")),\"\")")


def main():
    wb = openpyxl.load_workbook(WB)
    for nm in ("📊 لوحة مُكثّفة (٩ مؤشرات)", SHEET):
        if nm in wb.sheetnames:
            del wb[nm]
    ws = wb.create_sheet(SHEET, 0)        # اللوحة الوحيدة → أول تبويب
    ws.sheet_view.rightToLeft = True

    widths = [4,38,9,13,8,14] + [9]*12
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    def style(cell, fill=None, color=GRAY, bold=False, size=9, align="center", fmt=None, italic=False):
        if fill: cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name=F, bold=bold, size=size, color=color, italic=italic)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if fmt: cell.number_format = fmt

    # العنوان والعنوان الفرعي
    ws.merge_cells("A1:R1"); style(ws["A1"], ORANGE, WHITE, True, 15)
    ws["A1"] = "📊 مؤشرات أداء سلاسل الإمداد ٢٠٢٦ (٩ مؤشرات)"
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:R2"); style(ws["A2"], CREAM, GRAY, False, 9)
    ws["A2"] = ("دُمج المتكرر: عدد الانقطاعات←التوفر · الدوران←أيام المخزون · "
                "Fill Rate (حُذف) · حوادث النقل←الامتثال/السلامة")
    # شريط الأقسام
    ws.merge_cells("A3:F3"); style(ws["A3"], ORANGE, WHITE, True, 11, "center"); ws["A3"]="📌 المؤشر"
    ws.merge_cells("G3:R3"); style(ws["G3"], ORANGE, WHITE, True, 11, "center")
    ws["G3"]="📅 الأشهر — خطة (يدوي) vs فعلي (آلي)"
    # رأس الأعمدة
    hdr = ["#","المؤشر","الوحدة","الهدف","النوع","الانحراف"] + MONTHS
    for c,h in enumerate(hdr, start=1):
        style(ws.cell(4,c), GRAY, WHITE, True, 9); ws.cell(4,c).value=h
    ws.freeze_panes = "G6"

    r = 5
    for axis, kpis in KPIS_BY_AXIS:
        axis_row = r
        ws.merge_cells(f"A{r}:E{r}")
        style(ws.cell(r,1), ORANGE2, WHITE, True, 11, "right"); ws.cell(r,1).value="  ◀ "+axis
        ws.merge_cells(f"G{r}:R{r}"); style(ws.cell(r,7), ORANGE2)
        r += 1
        first = r
        for (num,name,unit,tval,tdisp,direction,fmt,fn) in kpis:
            p, a = r, r+1
            ws.merge_cells(f"A{p}:A{a}"); style(ws.cell(p,1), CREAM, GRAY, True, 10); ws.cell(p,1).value=num
            ws.merge_cells(f"B{p}:B{a}"); style(ws.cell(p,2), CREAM, GRAY, True, 9, "right"); ws.cell(p,2).value=name
            ws.merge_cells(f"C{p}:C{a}"); style(ws.cell(p,3), CREAM, GRAY, False, 9); ws.cell(p,3).value=unit
            # الهدف = الرقم الفعلي (بنفس صيغة خلايا الخطة) بدل نص ≥/≤
            ws.merge_cells(f"D{p}:D{a}"); style(ws.cell(p,4), CREAM, BLUE, True, 9, fmt=fmt); ws.cell(p,4).value=tval
            style(ws.cell(p,5), CREAM, GRAY, False, 9); ws.cell(p,5).value="خطة"
            style(ws.cell(a,5), WHITE, GRAY, False, 9); ws.cell(a,5).value="فعلي"
            style(ws.cell(p,6), CREAM); style(ws.cell(a,6), WHITE)
            ws.cell(a,6).value = status_formula(p,a,direction)
            for m in range(1,13):
                col = 6+m
                # خطة (يدوي)
                style(ws.cell(p,col), YELLOW, BLUE, False, 9, fmt=fmt); ws.cell(p,col).value=tval
                # فعلي (آلي أو يدوي للمواد التالفة)
                style(ws.cell(a,col), WHITE, GRAY, False, 9, fmt=fmt)
                if fn is not None:
                    ws.cell(a,col).value = fn(m)
            r += 2
        # حالة المحور الإجمالية في عمود F لصف المحور (أسوأ حالة بين مؤشراته)
        style(ws.cell(axis_row,6), ORANGE2, WHITE, True, 12)
        ws.cell(axis_row,6).value = (f'=IF(COUNTIF(F{first}:F{r-1},"*🔴*")>0,"🔴",'
                                     f'IF(COUNTIF(F{first}:F{r-1},"*🟡*")>0,"🟡","✅"))')

    # تذييل: ملاحظة الدمج + خلية إدخال «عدد نقاط المنتج» المحلية
    note_r = r+1
    ws.merge_cells(f"A{note_r}:R{note_r}")
    style(ws.cell(note_r,1), None, GRAY, False, 8, "right", italic=True)
    ws.cell(note_r,1).value = ("ℹ️ دُمجت ٤ مؤشرات متكررة ضمن هذه التسعة: عدد الانقطاعات←التوفر، "
                               "الدوران←أيام المخزون، Fill Rate (حُذف)، حوادث النقل←الامتثال/السلامة.")
    ws.merge_cells(f"A{INPUT_ROW}:E{INPUT_ROW}")
    style(ws.cell(INPUT_ROW,1), None, GRAY, True, 9, "right")
    ws.cell(INPUT_ROW,1).value = "عدد نقاط المنتج (محطات × منتجات) — لحساب «توفر الوقود»"
    style(ws.cell(INPUT_ROW,6), YELLOW, BLUE, True, 11)
    ws.cell(INPUT_ROW,6).value = 300

    # تنسيق شرطي لعمود الحالة (مطابق للأصل) — يشمل صفوف المحاور (تبدأ من F5)
    last = r
    colors = [("✅","FFE8F5E9","FF2E7D32"),("🟡","FFFFF8E1","FFF9A825"),("🔴","FFFFEBEE","FFC62828")]
    for emo,bg,fc in colors:
        dxf = DifferentialStyle(fill=PatternFill(bgColor=bg), font=Font(color=fc))
        ws.conditional_formatting.add(f"F5:F{last}",
            Rule(type="expression", formula=[f'ISNUMBER(SEARCH("{emo}",F5))'], dxf=dxf))

    # تحديث ربط ورقة مؤشرات الإدارة للترقيم الجديد، ثم حذف لوحة الـ١٣ (دُمجت في التسعة)
    patch_management_mapping(wb)
    if DASH in wb.sheetnames:
        del wb[DASH]
    wb.active = 0
    wb.save(WB)
    print("single 9-KPI dashboard built; sheets:", wb.sheetnames)


def patch_management_mapping(wb):
    """يحدّث عمود «يقابله في اللوحة» في ورقة مؤشرات الإدارة لترقيم اللوحة الجديدة (٩)."""
    name = "📋 مؤشرات الإدارة"
    if name not in wb.sheetnames:
        return
    ws = wb[name]
    mapping = ["K1 — حُدِّثت طريقة الحساب", "K4 — حُدِّثت طريقة الحساب",
               "K6 — مطابقة", "K2 — مطابقة", "K3 — مطابقة"]
    for i, txt in enumerate(mapping):
        ws.cell(4 + i, 8).value = txt   # العمود H، الصفوف 4..8


if __name__ == "__main__":
    main()
