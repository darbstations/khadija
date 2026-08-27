import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "تسعير الباقة"
ws.sheet_view.rightToLeft = True

ORANGE = "E07C16"; ORANGE_L = "FCE9D2"; INK = "26262A"; GREY = "6D6E70"
YELLOW = "FFF4CC"; GREEN_L = "E4F2E7"; HEAD = "F3EFEA"; WHITE = "FFFFFF"
thin = Side(style="thin", color="D9D7D3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell(r, c, v=None, bold=False, size=11, color=INK, fill=None, align="right",
         fmt=None, wrap=False, border_on=True):
    cc = ws.cell(row=r, column=c, value=v)
    cc.font = Font(name="Arial", bold=bold, size=size, color=color)
    cc.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill: cc.fill = PatternFill("solid", fgColor=fill)
    if fmt: cc.number_format = fmt
    if border_on: cc.border = border
    return cc

# widths
for col, w in {"A": 3, "B": 26, "C": 12, "D": 16, "E": 16, "F": 15, "G": 15}.items():
    ws.column_dimensions[col].width = w

SAR = '#,##0 "﷼"'
PCT = '0.0%'

# ---------- Title ----------
ws.merge_cells("B2:G2")
cell(2, 2, "درب · تانكي — نموذج تسعير الباقة", bold=True, size=16, color=ORANGE, align="right", border_on=False)
ws.merge_cells("B3:G3")
cell(3, 2, "باقة بقيمة ٤٩٥ ﷼ · عدّل الخلايا الصفراء وتُحدّث كل النتائج تلقائياً", size=10, color=GREY, align="right", border_on=False)

# ---------- Section 1: Inputs table ----------
cell(5, 2, "① المدخلات (المحتوى · القيمة · تكلفة الشراء من التاجر)", bold=True, size=12, color=WHITE, fill=ORANGE, border_on=False)
ws.merge_cells("B5:G5")

hdr = ["البند", "الكمية", "قيمة التجزئة/وحدة", "تكلفة الشراء/وحدة", "إجمالي القيمة", "إجمالي التكلفة"]
for j, h in enumerate(hdr):
    cell(6, 2 + j, h, bold=True, size=10, color=INK, fill=HEAD, align="center", wrap=True)

# rows: بنزين / غسلة / قهوة   (qty, retail/unit, cost/unit)  -- yellow = editable
rows = [
    ("رصيد بنزين",            1, 250, 235),   # fuel: cost ≈ value − ~6% fuel margin
    ("غسلة داخلي + خارجي",    7,  30,  15),
    ("كوب قهوة",             15,  18,   7),
]
r0 = 7
for i, (name, qty, retail, cost) in enumerate(rows):
    r = r0 + i
    cell(r, 2, name, bold=True, size=11)
    cell(r, 3, qty,    fill=YELLOW, align="center")
    cell(r, 4, retail, fill=YELLOW, align="center", fmt=SAR)
    cell(r, 5, cost,   fill=YELLOW, align="center", fmt=SAR)
    cell(r, 6, f"=C{r}*D{r}", align="center", fmt=SAR, fill=WHITE)          # total value
    cell(r, 7, f"=C{r}*E{r}", align="center", fmt=SAR, fill=WHITE)          # total cost
rlast = r0 + len(rows) - 1

# totals row
rt = rlast + 1
cell(rt, 2, "الإجمالي", bold=True, size=11, fill=ORANGE_L)
cell(rt, 3, "", fill=ORANGE_L); cell(rt, 4, "", fill=ORANGE_L); cell(rt, 5, "", fill=ORANGE_L)
cell(rt, 6, f"=SUM(F{r0}:F{rlast})", bold=True, align="center", fmt=SAR, fill=ORANGE_L)
cell(rt, 7, f"=SUM(G{r0}:G{rlast})", bold=True, align="center", fmt=SAR, fill=ORANGE_L)

# fuel note
cell(rlast + 2, 2, "ملاحظة: تكلفة البنزين ≈ القيمة − هامش الوقود (~٦٪). القهوة والغسلات تكلفتها = سعر شرائك من التاجر.",
     size=9, color=GREY, align="right", border_on=False)
ws.merge_cells(f"B{rlast+2}:G{rlast+2}")

# ---------- Section 2: price input ----------
rp = rlast + 4
cell(rp, 2, "② سعر الباقة للعميل", bold=True, size=12, color=WHITE, fill=ORANGE, border_on=False)
ws.merge_cells(f"B{rp}:C{rp}")
cell(rp, 4, 495, bold=True, size=13, fill=YELLOW, align="center", fmt=SAR)
ws.merge_cells(f"D{rp}:E{rp}")
cell(rp, 6, "← عدّله", size=10, color=GREY, align="right", border_on=False)
ws.merge_cells(f"F{rp}:G{rp}")
PRICE = f"D{rp}"
VAL = f"F{rt}"; COST = f"G{rt}"

# ---------- Section 3: Results ----------
rr = rp + 2
cell(rr, 2, "③ النتيجة", bold=True, size=12, color=WHITE, fill=ORANGE, border_on=False)
ws.merge_cells(f"B{rr}:G{rr}")

res = [
    ("القيمة الظاهرة للعميل",        f"={VAL}",                    SAR,  HEAD,  "إجمالي أسعار التجزئة"),
    ("سعر الباقة",                   f"={PRICE}",                  SAR,  HEAD,  "ما يدفعه العميل"),
    ("توفير العميل (﷼)",            f"={VAL}-{PRICE}",            SAR,  GREEN_L, "القيمة − السعر"),
    ("نسبة التوفير للعميل",          f"=({VAL}-{PRICE})/{VAL}",    PCT,  GREEN_L, "كم يحس أنه وفّر"),
    ("تكلفة درب الفعلية",            f"={COST}",                   SAR,  HEAD,  "ما تدفعه درب فعلاً"),
    ("ربح درب (﷼)",                 f"={PRICE}-{COST}",           SAR,  ORANGE_L, "السعر − التكلفة"),
    ("هامش ربح درب",                f"=({PRICE}-{COST})/{PRICE}", PCT,  ORANGE_L, "الربح ÷ السعر"),
    ("الرصيد المدفوع مقدماً (Float)", f"={PRICE}",                 SAR,  GREEN_L, "سيولة تجلس عندك"),
]
for i, (label, formula, fmt, fill, note) in enumerate(res):
    r = rr + 1 + i
    cell(r, 2, label, bold=True, size=11, fill=fill)
    ws.merge_cells(f"B{r}:C{r}")
    big = label.startswith("ربح") or label.startswith("هامش") or label.startswith("توفير") or label.startswith("نسبة")
    cell(r, 4, formula, bold=True, size=13 if big else 11,
         color=ORANGE if label.startswith(("ربح","هامش")) else INK,
         align="center", fmt=fmt, fill=fill)
    ws.merge_cells(f"D{r}:E{r}")
    cell(r, 6, note, size=9, color=GREY, align="right", fill=fill)
    ws.merge_cells(f"F{r}:G{r}")

# ---------- Section 4: sensitivity (profit vs coffee & wash cost) ----------
rs = rr + 1 + len(res) + 2
cell(rs, 2, "④ حساسية الربح — كم يتغيّر ربحك حسب أسعار شرائك من التجار", bold=True, size=12, color=WHITE, fill=ORANGE, border_on=False)
ws.merge_cells(f"B{rs}:G{rs}")
cell(rs+1, 2, "تكلفة القهوة/كوب ↓ · تكلفة الغسلة →", bold=True, size=9, color=INK, fill=HEAD, align="center", wrap=True)
wash_costs = [12, 15, 18]
coffee_costs = [5, 7, 9]
for j, wc in enumerate(wash_costs):
    cell(rs+1, 3 + j, f"غسلة {wc} ﷼", bold=True, size=9, fill=HEAD, align="center")
# fuel cost fixed from E7 (=235), qty from C
for i, cc in enumerate(coffee_costs):
    r = rs + 2 + i
    cell(r, 2, f"قهوة {cc} ﷼", bold=True, size=9, fill=HEAD, align="center")
    for j, wc in enumerate(wash_costs):
        # profit = price - (fuelCost*fuelQty + washQty*wc + coffeeQty*cc)
        formula = f"={PRICE}-(G7 + C8*{wc} + C9*{cc})"
        prof_fill = GREEN_L
        cell(r, 3 + j, formula, align="center", fmt=SAR, fill=prof_fill, size=10, bold=True)
cell(rs + 2 + len(coffee_costs), 2,
     "الخلية الوسطى = وضعك الحالي. كل ما فاوضت التاجر على سعر أقل، ارتفع ربحك.",
     size=9, color=GREY, align="right", border_on=False)
ws.merge_cells(f"B{rs+2+len(coffee_costs)}:G{rs+2+len(coffee_costs)}")

# freeze / gridlines off
ws.sheet_view.showGridLines = False

path = "/tmp/claude-0/-home-user-khadija/05ed6524-f029-51fd-96d9-884d1c665dcf/scratchpad/bundle_pricing.xlsx"
wb.save(path)
print("saved", path)

# quick recompute check (mirror formulas)
val = 1*250 + 7*30 + 15*18
cost = 1*235 + 7*15 + 15*7
price = 495
print(f"value={val} cost={cost} price={price} saving={val-price} ({(val-price)/val:.1%}) profit={price-cost} ({(price-cost)/price:.1%})")
