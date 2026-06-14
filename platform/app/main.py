# -*- coding: utf-8 -*-
import os, json
from fastapi import FastAPI, Request, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session

from .database import get_db
from . import models, compute
from .kpi_data import PILLARS, PROJECTS, PROJECT_PILLAR
from .auth import (current_user, verify_password, can_edit_definition, can_edit_value, can_eval, ROLE_LABEL)
from .eval_data import grade as eval_grade
from .seed import init_db

BASE = os.path.dirname(os.path.abspath(__file__))
app = FastAPI(title="منصة درب لمؤشرات الأداء")
app.add_middleware(SessionMiddleware, secret_key=os.environ.get("DARB_SECRET", "darb-secret-key-change-me"))
app.mount("/static", StaticFiles(directory=os.path.join(BASE, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE, "templates"))
templates.env.globals["fmt_value"] = compute.fmt_value
templates.env.globals["status_class"] = compute.status_class
templates.env.globals["MONTHS"] = compute.MONTHS
templates.env.globals["ROLE_LABEL"] = ROLE_LABEL

init_db()

from types import SimpleNamespace
from .database import SessionLocal
_s = SessionLocal()
templates.env.globals["DEPTS_NAV"] = [SimpleNamespace(id=d.id, name=d.name) for d in _s.query(models.Department).all()]
_s.close()

PERSPECTIVES = ["مالي","العملاء","العمليات الداخلية","التعلّم والنمو"]

def get_month(db):
    s = db.query(models.Setting).get("report_month")
    try: return int(s.value) if s and s.value else 0
    except: return 0

def kpi_calc(kpi, nmonths):
    vals = {v.month: v.actual for v in kpi.values}
    series = [vals.get(m) for m in range(1, 13)]
    y = compute.ytd(kpi.agg, series)
    ach = compute.achievement(kpi.target, kpi.polarity, y)
    fc = compute.forecast(kpi.agg, y, nmonths)
    return {"kpi": kpi, "series": series, "ytd": y, "ach": ach,
            "status": compute.status(ach), "sclass": compute.status_class(ach), "forecast": fc}

def all_calcs(db):
    nmonths = get_month(db)
    kpis = db.query(models.KPI).filter(models.KPI.level == "strategic").all()
    return [kpi_calc(k, nmonths) for k in kpis], nmonths

def avg(achs):
    nums = [a for a in achs if a is not None]
    return sum(nums)/len(nums) if nums else None

# ---------------- المصادقة ----------------
@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    return RedirectResponse("/dashboard")

@app.get("/login", response_class=HTMLResponse)
def login_form(request: Request):
    return templates.TemplateResponse(request, "login.html", {"request": request, "error": None})

@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    u = db.query(models.User).filter_by(username=username).first()
    if not u or not verify_password(password, u.pw_hash, u.salt):
        return templates.TemplateResponse(request, "login.html", {"request": request, "error": "بيانات الدخول غير صحيحة"})
    request.session["uid"] = u.id
    return RedirectResponse("/dashboard", status_code=302)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=302)

def require(request, db):
    u = current_user(request, db)
    return u

# ---------------- لوحة الإدارة التنفيذية ----------------
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    calcs, nmonths = all_calcs(db)
    achs = [c["ach"] for c in calcs]
    overall = avg(achs)
    counts = {"ok":0,"warn":0,"bad":0,"muted":0}
    for c in calcs: counts[c["sclass"]] += 1
    def rollup(keyfn, keys):
        out = []
        for k in keys:
            sub = [c["ach"] for c in calcs if keyfn(c)==k]
            out.append({"name": k, "ach": avg(sub), "n": len([c for c in calcs if keyfn(c)==k])})
        return out
    pillars = rollup(lambda c: c["kpi"].pillar, PILLARS)
    depts = db.query(models.Department).all()
    dept_roll = []
    for d in depts:
        sub = [c["ach"] for c in calcs if c["kpi"].department_id==d.id]
        dept_roll.append({"id": d.id, "name": d.name, "ach": avg(sub), "n": len(sub)})
    persp = rollup(lambda c: c["kpi"].perspective, PERSPECTIVES)
    projects = []
    for p in PROJECTS:
        sub=[c["ach"] for c in calcs if c["kpi"].project==p]
        projects.append({"name":p,"pillar":PROJECT_PILLAR.get(p,""),"ach":avg(sub),"n":len(sub)})
    chart = {
        "pillars": {"labels": [p["name"] for p in pillars], "data": [round((p["ach"] or 0)*100,1) for p in pillars]},
        "depts": {"labels": [d["name"] for d in dept_roll], "data": [round((d["ach"] or 0)*100,1) for d in dept_roll]},
        "persp": {"labels": [p["name"] for p in persp], "data": [round((p["ach"] or 0)*100,1) for p in persp]},
    }
    return templates.TemplateResponse(request, "dashboard.html", {"request": request, "u": u, "overall": overall,
        "counts": counts, "pillars": pillars, "depts": dept_roll, "persp": persp, "projects": projects,
        "nmonths": nmonths, "chart_json": json.dumps(chart, ensure_ascii=False), "total": len(calcs)})

# ---------------- تغذية مؤشرات إدارة ----------------
@app.get("/department/{dept_id}", response_class=HTMLResponse)
def department(dept_id: int, request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    d = db.query(models.Department).get(dept_id)
    if not d: return RedirectResponse("/dashboard")
    nmonths = get_month(db)
    rows = []
    for k in d.kpis:
        if k.level != "strategic":   # المؤشرات الفردية تظهر في صفحات الموظفين
            continue
        c = kpi_calc(k, nmonths)
        c["editable"] = can_edit_value(u, k)
        rows.append(c)
    any_edit = any(r["editable"] for r in rows)
    staff = db.query(models.User).filter_by(department_id=d.id, role="employee").all()
    return templates.TemplateResponse(request, "department.html", {"request": request, "u": u, "dept": d,
        "rows": rows, "any_edit": any_edit, "nmonths": nmonths, "staff": staff})

@app.post("/department/{dept_id}/save")
async def department_save(dept_id: int, request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    d = db.query(models.Department).get(dept_id)
    form = await request.form()
    for k in d.kpis:
        if not can_edit_value(u, k):   # حماية: لا تغذية بدون صلاحية
            continue
        for m in range(1, 13):
            field = f"v_{k.id}_{m}"
            if field in form:
                raw = (form.get(field) or "").strip().replace(",", "")
                val = None
                if raw != "":
                    try: val = float(raw)
                    except: val = None
                existing = db.query(models.KPIValue).filter_by(kpi_id=k.id, month=m).first()
                if val is None:
                    if existing: db.delete(existing)
                elif existing:
                    existing.actual = val
                else:
                    db.add(models.KPIValue(kpi_id=k.id, month=m, actual=val))
    db.commit()
    return RedirectResponse(f"/department/{dept_id}?saved=1", status_code=302)

# ---------------- مؤشرات الموظفين (فردي) ----------------
def emp_score(db, uid, nmonths):
    ks = db.query(models.KPI).filter_by(owner_user_id=uid, level="individual").all()
    achs = [kpi_calc(k, nmonths)["ach"] for k in ks]
    achs = [a for a in achs if a is not None]
    return (sum(achs)/len(achs) if achs else None), len(ks)

@app.get("/employees", response_class=HTMLResponse)
def employees(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    nmonths = get_month(db)
    emps = db.query(models.User).filter_by(role="employee").all()
    rows = []
    for e in emps:
        sc, n = emp_score(db, e.id, nmonths)
        if n == 0: continue
        rows.append({"u": e, "dept": e.department.name if e.department else "", "score": sc, "n": n})
    rows.sort(key=lambda r: (r["score"] is not None, r["score"] or 0), reverse=True)
    return templates.TemplateResponse(request, "employees.html", {"request": request, "u": u, "rows": rows, "nmonths": nmonths})

@app.get("/employee/{uid}", response_class=HTMLResponse)
def employee(uid: int, request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    e = db.query(models.User).get(uid)
    if not e: return RedirectResponse("/employees")
    nmonths = get_month(db)
    ks = db.query(models.KPI).filter_by(owner_user_id=uid, level="individual").order_by(models.KPI.id).all()
    rows = []
    for k in ks:
        c = kpi_calc(k, nmonths); c["editable"] = can_edit_value(u, k); rows.append(c)
    score, n = emp_score(db, uid, nmonths)
    any_edit = any(r["editable"] for r in rows)
    return templates.TemplateResponse(request, "employee.html", {"request": request, "u": u, "emp": e,
        "rows": rows, "any_edit": any_edit, "score": score, "nmonths": nmonths})

@app.post("/employee/{uid}/save")
async def employee_save(uid: int, request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    ks = db.query(models.KPI).filter_by(owner_user_id=uid, level="individual").all()
    form = await request.form()
    for k in ks:
        if not can_edit_value(u, k): continue
        for m in range(1, 13):
            f = f"v_{k.id}_{m}"
            if f in form:
                raw = (form.get(f) or "").strip().replace(",", "")
                val = None
                if raw != "":
                    try: val = float(raw)
                    except: val = None
                ex = db.query(models.KPIValue).filter_by(kpi_id=k.id, month=m).first()
                if val is None:
                    if ex: db.delete(ex)
                elif ex: ex.actual = val
                else: db.add(models.KPIValue(kpi_id=k.id, month=m, actual=val))
    db.commit()
    return RedirectResponse(f"/employee/{uid}?saved=1", status_code=302)

# ---------------- إدارة المؤشرات (أدمن: تعديل المستهدفات والأوزان) ----------------
@app.get("/admin/kpis", response_class=HTMLResponse)
def admin_kpis(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    if not can_edit_definition(u):
        return templates.TemplateResponse(request, "forbidden.html", {"request": request, "u": u})
    depts = db.query(models.Department).all()
    return templates.TemplateResponse(request, "admin_kpis.html", {"request": request, "u": u, "depts": depts})

@app.post("/admin/kpis/save")
async def admin_kpis_save(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u or not can_edit_definition(u):
        return RedirectResponse("/dashboard")
    form = await request.form()
    for k in db.query(models.KPI).all():
        for attr in ("target", "baseline"):
            f = f"{attr}_{k.id}"
            if f in form:
                raw = (form.get(f) or "").strip().replace(",", "")
                setattr(k, attr, float(raw) if raw != "" else None)
        wf = f"weight_{k.id}"
        if wf in form:
            raw = (form.get(wf) or "").strip().replace(",", "")
            k.weight = (float(raw)/100.0) if raw != "" else 0   # المُدخل نسبة مئوية
        tt = form.get(f"target_text_{k.id}")
        if tt is not None: k.target_text = tt
    db.commit()
    return RedirectResponse("/admin/kpis?saved=1", status_code=302)

# ---------------- الاستراتيجية / الخارطة ----------------
@app.get("/strategy", response_class=HTMLResponse)
def strategy(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    calcs, _ = all_calcs(db)
    def roll(keyfn, k):
        sub=[c["ach"] for c in calcs if keyfn(c)==k];
        n=[c for c in calcs if keyfn(c)==k]
        return avg(sub), len(n)
    pillars=[{"name":p, **dict(zip(("ach","n"), roll(lambda c:c["kpi"].pillar,p)))} for p in PILLARS]
    persp=[{"name":p, **dict(zip(("ach","n"), roll(lambda c:c["kpi"].perspective,p)))} for p in PERSPECTIVES]
    projects=[]
    for p in PROJECTS:
        a,n=roll(lambda c:c["kpi"].project,p)
        projects.append({"name":p,"pillar":PROJECT_PILLAR.get(p,""),"ach":a,"n":n,"gap": n==0})
    return templates.TemplateResponse(request, "strategy.html", {"request": request, "u": u,
        "pillars": pillars, "persp": persp, "projects": projects})

# ---------------- المبادرات والمخاطر ----------------
@app.get("/initiatives", response_class=HTMLResponse)
def initiatives(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    items = db.query(models.Initiative).all()
    return templates.TemplateResponse(request, "initiatives.html", {"request": request, "u": u, "items": items,
        "pillars": PILLARS, "projects": PROJECTS})

@app.post("/initiatives/save")
async def initiatives_save(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    form = await request.form()
    iid = form.get("id")
    def num(x):
        x=(x or "").strip().replace(",","")
        try: return float(x)
        except: return None
    if iid:
        it = db.query(models.Initiative).get(int(iid))
    else:
        it = models.Initiative(name=form.get("name") or "مبادرة")
        db.add(it)
    it.name=form.get("name") or it.name; it.pillar=form.get("pillar") or ""
    it.project=form.get("project") or ""; it.owner=form.get("owner") or ""
    it.start=form.get("start") or ""; it.end=form.get("end") or ""
    it.budget=num(form.get("budget"));
    p=num(form.get("progress")); it.progress=(p/100 if p and p>1 else (p or 0))
    it.notes=form.get("notes") or ""
    db.commit()
    return RedirectResponse("/initiatives", status_code=302)

@app.get("/risks", response_class=HTMLResponse)
def risks(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    items = db.query(models.Risk).all()
    return templates.TemplateResponse(request, "risks.html", {"request": request, "u": u, "items": items})

@app.post("/risks/save")
async def risks_save(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    form = await request.form()
    rid = form.get("id")
    def int0(x):
        try: return int(x)
        except: return 0
    it = db.query(models.Risk).get(int(rid)) if rid else None
    if not it:
        it = models.Risk(title=form.get("title") or "خطر"); db.add(it)
    it.title=form.get("title") or it.title; it.area=form.get("area") or ""
    it.likelihood=int0(form.get("likelihood")); it.impact=int0(form.get("impact"))
    it.response=form.get("response") or ""; it.owner=form.get("owner") or ""
    db.commit()
    return RedirectResponse("/risks", status_code=302)

# ---------------- تقييم الموظفين ----------------
@app.get("/evaluations", response_class=HTMLResponse)
def evaluations(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    forms = db.query(models.EvalForm).all()
    by_dept = {}
    for f in forms:
        by_dept.setdefault(f.department.name, []).append(f)
    evs = db.query(models.Evaluation).order_by(models.Evaluation.id.desc()).all()
    results = []
    for ev in evs:
        total = sum((s.achievement or 0) * (s.item.weight or 0) if (s.achievement or 0) <= 1
                    else s.item.weight for s in ev.scores)
        results.append({"ev": ev, "total": total, "grade": eval_grade(total)})
    return templates.TemplateResponse(request, "evaluations.html", {"request": request, "u": u,
        "by_dept": by_dept, "results": results, "can_eval": can_eval})

@app.get("/evaluation/{form_id}", response_class=HTMLResponse)
def evaluation_form(form_id: int, request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    f = db.query(models.EvalForm).get(form_id)
    if not f: return RedirectResponse("/evaluations")
    if not can_eval(u, f.department_id):
        return templates.TemplateResponse(request, "forbidden.html", {"request": request, "u": u})
    return templates.TemplateResponse(request, "evaluation_form.html", {"request": request, "u": u, "f": f})

@app.post("/evaluation/{form_id}/save")
async def evaluation_save(form_id: int, request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    f = db.query(models.EvalForm).get(form_id)
    if not u or not f or not can_eval(u, f.department_id):
        return RedirectResponse("/evaluations")
    form = await request.form()
    ev = models.Evaluation(form_id=f.id, employee_name=form.get("employee_name") or "",
        manager_name=form.get("manager_name") or u.full_name, quarter=form.get("quarter") or "",
        notes=form.get("notes") or "")
    db.add(ev); db.flush()
    for it in f.items:
        raw = (form.get(f"a_{it.id}") or "").strip().replace("%","")
        val = None
        if raw != "":
            try:
                val = float(raw)
                if val > 1.5: val = val/100.0    # سمح بإدخال 95 أو 0.95
            except: val = None
        db.add(models.EvalScore(evaluation_id=ev.id, item_id=it.id, achievement=val))
    db.commit()
    return RedirectResponse(f"/evaluation/view/{ev.id}", status_code=302)

@app.get("/evaluation/view/{eval_id}", response_class=HTMLResponse)
def evaluation_view(eval_id: int, request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    ev = db.query(models.Evaluation).get(eval_id)
    if not ev: return RedirectResponse("/evaluations")
    rows = []
    total = 0
    for s in ev.scores:
        ach = s.achievement
        score = (s.item.weight or 0) * min(ach, 1) if ach is not None else 0
        total += score
        rows.append({"item": s.item, "ach": ach, "score": score})
    return templates.TemplateResponse(request, "evaluation_view.html", {"request": request, "u": u,
        "ev": ev, "rows": rows, "total": total, "grade": eval_grade(total)})

# ---------------- تصدير التقارير ----------------
@app.get("/report/print", response_class=HTMLResponse)
def report_print(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    calcs, nmonths = all_calcs(db)
    overall = avg([c["ach"] for c in calcs])
    depts = db.query(models.Department).all()
    dept_roll = [{"name": d.name, "ach": avg([c["ach"] for c in calcs if c["kpi"].department_id==d.id])} for d in depts]
    pillars = [{"name": p, "ach": avg([c["ach"] for c in calcs if c["kpi"].pillar==p])} for p in PILLARS]
    return templates.TemplateResponse(request, "report_print.html", {"request": request, "u": u,
        "overall": overall, "depts": dept_roll, "pillars": pillars, "calcs": calcs, "nmonths": nmonths})

@app.get("/export/excel")
def export_excel(request: Request, db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    calcs, nmonths = all_calcs(db)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "تقرير الأداء"; ws.sheet_view.rightToLeft = True
    hdr = ["الإدارة","المؤشر","الوحدة","المستهدف","YTD","نسبة التحقيق","الحالة","الركيزة","المشروع"]
    nav = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="58595B")
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(1, c, h); cell.font = nav; cell.fill = fill; cell.alignment = Alignment(horizontal="center")
    for i, x in enumerate(calcs, 2):
        k = x["kpi"]
        ws.cell(i,1,k.department.name); ws.cell(i,2,k.name); ws.cell(i,3,k.unit)
        ws.cell(i,4,k.target_text); ws.cell(i,5, round(x["ytd"],2) if x["ytd"] is not None else "")
        ws.cell(i,6, round(x["ach"],4) if x["ach"] is not None else "").number_format = "0%"
        ws.cell(i,7,x["status"]); ws.cell(i,8,k.pillar); ws.cell(i,9,k.project)
    for col, w in zip("ABCDEFGHI", [20,42,10,18,14,12,14,14,22]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=darb-report.xlsx"})

# ---------------- إعداد شهر التقرير ----------------
@app.post("/settings/report-month")
async def set_month(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u or u.role not in ("admin","executive"):
        return RedirectResponse("/dashboard")
    form = await request.form()
    s = db.query(models.Setting).get("report_month")
    if not s: s = models.Setting(key="report_month"); db.add(s)
    s.value = (form.get("report_month") or "0")
    db.commit()
    return RedirectResponse("/dashboard", status_code=302)

# ---------------- رفع/استيراد من إكسل (أدمن) ----------------
@app.get("/upload", response_class=HTMLResponse)
def upload_form(request: Request, db: Session = Depends(get_db)):
    u = require(request, db)
    if not u: return RedirectResponse("/login")
    if not can_edit_definition(u):
        return templates.TemplateResponse(request, "forbidden.html", {"request": request, "u": u})
    return templates.TemplateResponse(request, "upload.html", {"request": request, "u": u, "msg": None})

@app.post("/upload", response_class=HTMLResponse)
async def upload_import(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    u = require(request, db)
    if not u or not can_edit_definition(u):
        return RedirectResponse("/dashboard")
    import openpyxl, io
    data = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    name_to_kpi = {k.name: k for k in db.query(models.KPI).all()}
    imported = 0
    for ws in wb.worksheets:
        if "المؤشرات" not in ws.title:  # أوراق المؤشرات فقط
            continue
        for row in ws.iter_rows(min_row=4, values_only=True):
            if len(row) < 20 or not row[2]:  # العمود C = اسم المؤشر
                continue
            k = name_to_kpi.get(str(row[2]).strip())
            if not k: continue
            for m in range(12):
                cell = row[8+m]  # I..T = الأشهر
                if isinstance(cell, (int, float)):
                    ex = db.query(models.KPIValue).filter_by(kpi_id=k.id, month=m+1).first()
                    if ex: ex.actual = float(cell)
                    else: db.add(models.KPIValue(kpi_id=k.id, month=m+1, actual=float(cell)))
                    imported += 1
    db.commit()
    return templates.TemplateResponse(request, "upload.html", {"request": request, "u": u,
        "msg": f"تم استيراد {imported} قيمة شهرية من الملف."})
